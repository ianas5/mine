#!/usr/bin/env python3
"""
Advanced Time-Phased Monte Carlo Cost Model — workbook generator.

Builds a professional, app-like Excel workbook (.xlsx) for cost estimation with:
  - 3-point estimates (3PE) on cost lines, with annual cost profiling
  - Risk register with probability-weighted, 3PE financial impacts + risk profiling
  - Annual inflation, discount rate / NPV, configurable iterations & confidence
  - A formula-driven Monte Carlo engine (runs in-sheet; press F9 to re-roll)
  - Results: percentiles, contingency table, annual cash flow & NPV by confidence
  - Sensitivity (tornado), Assumptions, Checks, and an executive Dashboard

No macros required: the engine is volatile-formula based. A parallel set of VBA
modules (see ../vba) can be imported to add a real "Run Simulation" button.

Usage:  python3 generate_advanced_model.py [iterations]
Output: AdvancedMonteCarloCostModel.xlsx
"""

import sys, math, random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, ScatterChart, LineChart, Reference, Series
from openpyxl.utils import get_column_letter as L

# ============================================================================
# 0. CONFIG + EXAMPLE DATA  (edit, then re-run)
# ============================================================================
# Positional args:  [iterations]  [years]  [start_year]  [currency]  [mode]
#   mode = "flex"  -> build for MAXY years and pre-hide the unused ones, so the
#                     duration can be changed in-Excel with the ApplySettings
#                     macro (one file covers 1..MAXY years).
#   mode = "fixed" -> build exactly [years] year columns (lighter file). default.
ITER         = int(sys.argv[1]) if len(sys.argv) > 1 else 4000
_years_arg   = int(sys.argv[2]) if len(sys.argv) > 2 else 5
START_YEAR   = int(sys.argv[3]) if len(sys.argv) > 3 else 2025
BASE_YEAR    = START_YEAR
CURRENCY     = sys.argv[4] if len(sys.argv) > 4 else "SAR"   # display only — live in Setup
FLEX         = len(sys.argv) > 5 and sys.argv[5].lower() in ("flex", "flexible", "1")
MAXY         = 30
ACTIVE_YEARS = max(1, min(MAXY, _years_arg))
N_YEARS      = MAXY if FLEX else ACTIVE_YEARS                # structural build size
OUTFILE      = "AdvancedMonteCarloCostModel.xlsx"
PROJECT    = "Sample Construction Project"
CONF_DEF   = "P80"
DISCOUNT   = 0.08
INFL_BASE  = 0.03                                # default annual inflation
# Money cells carry NO embedded currency symbol, so the currency is whatever you
# type in the Setup "Currency" cell (shown across the Dashboard). Change it live.
MONEY      = '#,##0'
MONEY0     = '#,##0'
PCT1       = '0.0%'
PCT0       = '0%'

# Annual allocation is described by a SHAPE so the model works for ANY duration:
#   even  = equal each year     early = front-loaded     late = back-loaded
#   bell  = peaks in the middle
def make_profile(Y, shape):
    if Y == 1: return [1.0]
    if shape == "early": w = [float(Y - i) for i in range(Y)]
    elif shape == "late": w = [float(i + 1) for i in range(Y)]
    elif shape == "bell":
        mid = (Y - 1) / 2.0
        w = [max(0.05, 1.0 - abs(i - mid) / (mid + 1)) for i in range(Y)]
    else: w = [1.0] * Y               # even
    s = sum(w)
    return [x / s for x in w]

# Cost lines: wbs, item, category, unit, qty, minU, mlU, maxU, dist, profile-shape
COST_LINES = [
    ("1.1","Site Preparation","Civil","LS",1,28000,38000,60000,"Triangular","early"),
    ("1.2","Foundations","Civil","LS",1,45000,55000,75000,"Triangular","early"),
    ("2.1","Superstructure","Structural","LS",1,90000,120000,170000,"Pert","bell"),
    ("3.1","MEP Systems","MEP","LS",1,60000,80000,115000,"Pert","late"),
    ("4.1","Finishes","Architectural","LS",1,30000,45000,70000,"Triangular","late"),
    ("5.1","Project Management","PM","LS",1,20000,28000,40000,"Normal","even"),
]
# Risks: id, name, category, prob, minI, mlI, maxI, dist, linked, owner, profile-shape
RISKS = [
    ("R1","Adverse ground conditions","Technical",0.30,20000,45000,90000,"Triangular","1.2","Civil Lead","early"),
    ("R2","Major design change","Commercial",0.40,15000,30000,60000,"Pert","2.1","Design Mgr","bell"),
    ("R3","Material price spike","Market",0.50,10000,25000,55000,"Triangular","3.1","Procurement","bell"),
    ("R4","Schedule delay","Delivery",0.25,20000,40000,80000,"Pert","5.1","PM","late"),
]
K, J, Y = len(COST_LINES), len(RISKS), N_YEARS
FY = [f"FY{y+1}" for y in range(Y)]
FYYEAR = [START_YEAR + y for y in range(Y)]
# Profiles built over the ACTIVE years, padded with zeros up to the build size,
# so inactive years contribute nothing until they are switched on.
def pad(p): return p + [0.0] * (Y - len(p))
COST_PROFILES = [pad(make_profile(ACTIVE_YEARS, c[9])) for c in COST_LINES]
RISK_PROFILES = [pad(make_profile(ACTIVE_YEARS, r[10])) for r in RISKS]
INFL_RATES = ([0.0] + [INFL_BASE] * (ACTIVE_YEARS - 1) + [0.0] * (Y - ACTIVE_YEARS))
INACTIVE = list(range(ACTIVE_YEARS, Y))           # year indexes to pre-hide (flex mode)

# ============================================================================
# 1. STYLES
# ============================================================================
NAVY="1F2A44"; ORANGE="FF6B35"; TEAL="00A88A"; RED="C0392B"
INPUT="FFF7DC"; HEAD=NAVY; GREYBG="F2F3F5"; WHITE="FFFFFF"
def F(**k): return Font(name="Calibri", **k)
def fill(c): return PatternFill("solid", fgColor=c)
thin=Side("thin", color="C9CDD6")
BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
C=Alignment(horizontal="center",vertical="center")
Rt=Alignment(horizontal="right",vertical="center")
Lt=Alignment(horizontal="left",vertical="center")
H_FONT=F(bold=True,size=11,color=WHITE)
TITLE=F(bold=True,size=16,color=ORANGE)
SUB=F(size=10,color="6C6C70")
LBL=F(bold=True,size=10,color="333333")
inputfill=fill(INPUT)

wb=Workbook()

def newsheet(name, tab=None):
    ws=wb.create_sheet(name)
    ws.sheet_view.showGridLines=False
    if tab: ws.sheet_properties.tabColor=tab
    return ws

def hcell(ws,r,c,v):
    x=ws.cell(r,c,v); x.font=H_FONT; x.fill=fill(HEAD); x.alignment=C; x.border=BOX; return x
def put(ws,r,c,v,fmt=None,font=None,al=None,bd=True,fillc=None):
    x=ws.cell(r,c,v)
    if fmt: x.number_format=fmt
    x.font=font or F(size=10)
    x.alignment=al or Lt
    if bd: x.border=BOX
    if fillc: x.fill=fill(fillc)
    return x
def ref(sheet,c,r): return f"'{sheet}'!${L(c)}${r}"
def rng(sheet,c1,r1,c2,r2): return f"'{sheet}'!${L(c1)}${r1}:${L(c2)}${r2}"

# ============================================================================
# 2. SETUP SHEET
# ============================================================================
S="Setup"; setup=newsheet(S, NAVY)
setup.cell(1,2,"MODEL SETUP").font=TITLE
setup.cell(2,2,"Yellow cells are inputs. These settings drive the whole model.").font=SUB
rows=[
 ("Project name",PROJECT,None),("Currency",CURRENCY,None),("Base year",BASE_YEAR,"0"),
 ("Start year",START_YEAR,"0"),("Number of years",ACTIVE_YEARS,"0"),
 ("Monte Carlo iterations",ITER,"#,##0"),("Confidence level",CONF_DEF,None),
 ("Default distribution","Triangular",None),("Annual inflation (default)",0.03,PCT1),
 ("Discount rate",DISCOUNT,PCT1),("VAT treatment","Excluded",None),
 ("Model version","1.0 (MVP)",None),("Prepared by","",None),("Date","",None),
]
r0=4
for i,(lbl,val,fmt) in enumerate(rows):
    r=r0+i
    put(setup,r,2,lbl,font=LBL,bd=False)
    c=put(setup,r,3,val,fmt=fmt,al=Rt,fillc=INPUT); c.font=F(size=10,bold=True)
CONF_CELL=ref(S,3,r0+6)          # "P80"
NYEARS_CELL=ref(S,3,r0+4)
ITER_CELL=ref(S,3,r0+5)
DISC_CELL=ref(S,3,r0+9)
START_CELL=ref(S,3,r0+3)         # Start year — drives all FY calendar labels (live)
def fylabel(y): return f'="FY{y+1} ("&TEXT({START_CELL}+{y},"0")&")"'
# numeric confidence helper
put(setup,r0+15,2,"Confidence (numeric)",font=F(size=9,italic=True,color="888888"),bd=False)
cn=put(setup,r0+15,3,f"=VALUE(MID({CONF_CELL},2,2))/100",fmt=PCT0,al=Rt); cn.font=F(size=9,italic=True,color="888888")
CONFP_CELL=ref(S,3,r0+15)
# dropdowns
dv_conf=DataValidation(type="list",formula1='"P50,P60,P70,P80,P90,P95"'); setup.add_data_validation(dv_conf); dv_conf.add(f"C{r0+6}")
dv_dist0=DataValidation(type="list",formula1='"Triangular,Pert,Normal"'); setup.add_data_validation(dv_dist0); dv_dist0.add(f"C{r0+7}")
dv_vat=DataValidation(type="list",formula1='"Included,Excluded"'); setup.add_data_validation(dv_vat); dv_vat.add(f"C{r0+10}")
setup.column_dimensions["B"].width=26; setup.column_dimensions["C"].width=24

# ============================================================================
# 3. COST LINES SHEET  (tbl_CostLines)
# ============================================================================
CL="Cost Lines"; cl=newsheet(CL, ORANGE)
cl.cell(1,1,"COST LINES — 3-Point Estimates").font=TITLE
cl.cell(2,1,"Enter Min / Most Likely / Max unit costs. Totals are computed. Pick a distribution and whether to include the line.").font=SUB
CLH=3; CLF=4
cols=["WBS","Cost Item","Category","Description","Unit","Quantity","Min Unit Cost",
      "Most Likely Unit Cost","Max Unit Cost","Min Total Cost","Most Likely Total Cost",
      "Max Total Cost","Distribution","Include?","Notes"]
for j,h in enumerate(cols): hcell(cl,CLH,1+j,h)
for i,(wbs,item,cat,unit,qty,mn,ml,mx,dist,prof) in enumerate(COST_LINES):
    r=CLF+i
    put(cl,r,1,wbs,al=C); put(cl,r,2,item); put(cl,r,3,cat); put(cl,r,4,f"{item} works"); put(cl,r,5,unit,al=C)
    put(cl,r,6,qty,fmt=MONEY0,al=Rt,fillc=INPUT)
    put(cl,r,7,mn,fmt=MONEY0,al=Rt,fillc=INPUT)
    put(cl,r,8,ml,fmt=MONEY0,al=Rt,fillc=INPUT)
    put(cl,r,9,mx,fmt=MONEY0,al=Rt,fillc=INPUT)
    put(cl,r,10,f"=F{r}*G{r}",fmt=MONEY0,al=Rt)   # Min Total
    put(cl,r,11,f"=F{r}*H{r}",fmt=MONEY0,al=Rt)   # ML Total
    put(cl,r,12,f"=F{r}*I{r}",fmt=MONEY0,al=Rt)   # Max Total
    put(cl,r,13,dist,al=C); put(cl,r,14,"Yes",al=C); put(cl,r,15,"")
CL_LAST=CLF+K-1
tbl=Table(displayName="tbl_CostLines",ref=f"A{CLH}:O{CL_LAST}")
tbl.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True)
cl.add_table(tbl)
dv=DataValidation(type="list",formula1='"Triangular,Pert,Normal"'); cl.add_data_validation(dv); dv.add(f"M{CLF}:M{CL_LAST}")
dv=DataValidation(type="list",formula1='"Yes,No"'); cl.add_data_validation(dv); dv.add(f"N{CLF}:N{CL_LAST}")
# CF: highlight when not Min<=ML<=Max
cl.conditional_formatting.add(f"G{CLF}:I{CL_LAST}",
    FormulaRule(formula=[f"OR($G{CLF}>$H{CLF},$H{CLF}>$I{CLF})"], fill=fill("FFC7CE")))
for col,w in zip("ABCDEFGHIJKLMNO",[7,20,13,18,7,9,13,15,13,13,15,13,12,9,16]):
    cl.column_dimensions[col].width=w
cl.freeze_panes="B4"

# ============================================================================
# 4. COST PROFILING SHEET  (tbl_CostProfile)
# ============================================================================
CP="Cost Profiling"; cp=newsheet(CP, ORANGE)
cp.cell(1,1,"COST PROFILING — annual % allocation").font=TITLE
cp.cell(2,1,"Each row must sum to 100%. Rows that do not are highlighted red. Yellow cells are inputs.").font=SUB
CPH=3; CPF=4; CP_FY0=3
hcell(cp,CPH,1,"WBS"); hcell(cp,CPH,2,"Cost Item")
for y in range(Y): hcell(cp,CPH,CP_FY0+y,FY[y])   # static (Excel table headers must be text)
hcell(cp,CPH,CP_FY0+Y,"Total %")
for i,(wbs,item,*_rest) in enumerate(COST_LINES):
    r=CPF+i; prof=COST_PROFILES[i]
    put(cp,r,1,f"='{CL}'!A{CLF+i}",al=C); put(cp,r,2,f"='{CL}'!B{CLF+i}")
    for y in range(Y): put(cp,r,CP_FY0+y,prof[y],fmt=PCT0,al=C,fillc=INPUT)
    tot=put(cp,r,CP_FY0+Y,f"=SUM({L(CP_FY0)}{r}:{L(CP_FY0+Y-1)}{r})",fmt=PCT0,al=C); tot.font=F(bold=True,size=10)
CP_LAST=CPF+K-1
tbl=Table(displayName="tbl_CostProfile",ref=f"A{CPH}:{L(CP_FY0+Y)}{CP_LAST}")
tbl.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True); cp.add_table(tbl)
cp.conditional_formatting.add(f"{L(CP_FY0+Y)}{CPF}:{L(CP_FY0+Y)}{CP_LAST}",
    FormulaRule(formula=[f"ABS({L(CP_FY0+Y)}{CPF}-1)>0.001"], fill=fill("FFC7CE")))
cp.column_dimensions["A"].width=8; cp.column_dimensions["B"].width=20
for y in range(Y+1): cp.column_dimensions[L(CP_FY0+y)].width=12
def costprof(k,y): return ref(CP,CP_FY0+y,CPF+k)

# ============================================================================
# 5. RISK REGISTER  (tbl_RiskRegister)
# ============================================================================
RR="Risk Register"; rr=newsheet(RR, RED)
rr.cell(1,1,"RISK REGISTER — probability-weighted 3PE impacts").font=TITLE
rr.cell(2,1,"Each risk occurs with its Probability. When it occurs, the financial impact is sampled from its distribution.").font=SUB
RRH=3; RRF=4
cols=["Risk ID","Risk Name","Category","Probability","Min Impact","Most Likely Impact",
      "Max Impact","Distribution","Linked Cost Item","Risk Owner","Mitigation","Status","Include?","Notes"]
for j,h in enumerate(cols): hcell(rr,RRH,1+j,h)
for i,(rid,name,cat,prob,mn,ml,mx,dist,linked,owner,prof) in enumerate(RISKS):
    r=RRF+i
    put(rr,r,1,rid,al=C); put(rr,r,2,name); put(rr,r,3,cat)
    put(rr,r,4,prob,fmt=PCT0,al=C,fillc=INPUT)
    put(rr,r,5,mn,fmt=MONEY0,al=Rt,fillc=INPUT)
    put(rr,r,6,ml,fmt=MONEY0,al=Rt,fillc=INPUT)
    put(rr,r,7,mx,fmt=MONEY0,al=Rt,fillc=INPUT)
    put(rr,r,8,dist,al=C); put(rr,r,9,linked,al=C); put(rr,r,10,owner)
    put(rr,r,11,"Mitigation plan"); put(rr,r,12,"Open",al=C); put(rr,r,13,"Yes",al=C); put(rr,r,14,"")
RR_LAST=RRF+J-1
tbl=Table(displayName="tbl_RiskRegister",ref=f"A{RRH}:N{RR_LAST}")
tbl.tableStyleInfo=TableStyleInfo(name="TableStyleMedium3",showRowStripes=True); rr.add_table(tbl)
dv=DataValidation(type="list",formula1='"Triangular,Pert,Normal"'); rr.add_data_validation(dv); dv.add(f"H{RRF}:H{RR_LAST}")
dv=DataValidation(type="list",formula1='"Yes,No"'); rr.add_data_validation(dv); dv.add(f"M{RRF}:M{RR_LAST}")
dv=DataValidation(type="list",formula1='"Open,Closed,Mitigated"'); rr.add_data_validation(dv); dv.add(f"L{RRF}:L{RR_LAST}")
rr.conditional_formatting.add(f"E{RRF}:G{RR_LAST}",
    FormulaRule(formula=[f"OR($E{RRF}>$F{RRF},$F{RRF}>$G{RRF})"], fill=fill("FFC7CE")))
for col,w in zip("ABCDEFGHIJKLMN",[8,24,12,11,12,15,12,12,13,13,16,9,9,14]):
    rr.column_dimensions[col].width=w
rr.freeze_panes="B4"

# ============================================================================
# 6. RISK PROFILING  (tbl_RiskProfile)
# ============================================================================
RP="Risk Profiling"; rp=newsheet(RP, RED)
rp.cell(1,1,"RISK PROFILING — annual % allocation of risk impact").font=TITLE
rp.cell(2,1,"Each row must sum to 100%. Rows that do not are highlighted red.").font=SUB
RPH=3; RPF=4; RP_FY0=3
hcell(rp,RPH,1,"Risk ID"); hcell(rp,RPH,2,"Risk Name")
for y in range(Y): hcell(rp,RPH,RP_FY0+y,FY[y])   # static (Excel table headers must be text)
hcell(rp,RPH,RP_FY0+Y,"Total %")
for i,rk in enumerate(RISKS):
    r=RPF+i; prof=RISK_PROFILES[i]
    put(rp,r,1,f"='{RR}'!A{RRF+i}",al=C); put(rp,r,2,f"='{RR}'!B{RRF+i}")
    for y in range(Y): put(rp,r,RP_FY0+y,prof[y],fmt=PCT0,al=C,fillc=INPUT)
    tot=put(rp,r,RP_FY0+Y,f"=SUM({L(RP_FY0)}{r}:{L(RP_FY0+Y-1)}{r})",fmt=PCT0,al=C); tot.font=F(bold=True,size=10)
RP_LAST=RPF+J-1
tbl=Table(displayName="tbl_RiskProfile",ref=f"A{RPH}:{L(RP_FY0+Y)}{RP_LAST}")
tbl.tableStyleInfo=TableStyleInfo(name="TableStyleMedium3",showRowStripes=True); rp.add_table(tbl)
rp.conditional_formatting.add(f"{L(RP_FY0+Y)}{RPF}:{L(RP_FY0+Y)}{RP_LAST}",
    FormulaRule(formula=[f"ABS({L(RP_FY0+Y)}{RPF}-1)>0.001"], fill=fill("FFC7CE")))
rp.column_dimensions["A"].width=8; rp.column_dimensions["B"].width=24
for y in range(Y+1): rp.column_dimensions[L(RP_FY0+y)].width=12
def riskprof(j,y): return ref(RP,RP_FY0+y,RPF+j)

# ============================================================================
# 7. INFLATION
# ============================================================================
INF="Inflation"; inf=newsheet(INF, TEAL)
inf.cell(1,1,"INFLATION").font=TITLE
inf.cell(2,1,"Cumulative factor compounds year by year. FY1 = base year (factor 1.0). Inflated cost = nominal × factor.").font=SUB
INFH=3; INFF=4
hcell(inf,INFH,1,"FY"); hcell(inf,INFH,2,"Year"); hcell(inf,INFH,3,"Inflation Rate"); hcell(inf,INFH,4,"Cumulative Factor")
for y in range(Y):
    r=INFF+y
    put(inf,r,1,FY[y],al=C); put(inf,r,2,f"={START_CELL}+{y}",fmt="0",al=C)
    put(inf,r,3,INFL_RATES[y],fmt=PCT1,al=C,fillc=INPUT)
    if y==0: put(inf,r,4,"=1",fmt="0.0000",al=C)
    else:    put(inf,r,4,f"=D{r-1}*(1+C{r})",fmt="0.0000",al=C)
for col,w in zip("ABCD",[8,8,14,16]): inf.column_dimensions[col].width=w
def inflfac(y): return ref(INF,4,INFF+y)

# ============================================================================
# 8. ENGINE  (formula Monte Carlo)
# ============================================================================
EN="Engine"; en=newsheet(EN, "888888")
en.cell(1,1,"MONTE CARLO ENGINE — calculation only. One row per iteration. Press F9 to re-roll.").font=F(bold=True,size=11,color=ORANGE)
EHDR=3; EF=4; ELAST=EF+ITER-1
c=1
ITER_C=c; c+=1
SAMP0=c; c+=K
IMP0=c; c+=J
YT0=c; c+=Y
BASE_C=c; c+=1
RISK_C=c; c+=1
TOT_C=c; c+=1
NPV_C=c; c+=1
HU0=c; c+=K          # hidden uniforms for cost lines
HRO0=c; c+=J         # hidden uniforms risk occurrence
HRI0=c; c+=J         # hidden uniforms risk impact
# headers
hcell(en,EHDR,ITER_C,"Iter")
for k in range(K): hcell(en,EHDR,SAMP0+k,COST_LINES[k][1][:12])
for j in range(J): hcell(en,EHDR,IMP0+j,RISKS[j][0])
for y in range(Y): hcell(en,EHDR,YT0+y,FY[y])
hcell(en,EHDR,BASE_C,"Base"); hcell(en,EHDR,RISK_C,"Risk"); hcell(en,EHDR,TOT_C,"Total"); hcell(en,EHDR,NPV_C,"NPV")
for k in range(K): hcell(en,EHDR,HU0+k,f"u{k}")
for j in range(J): hcell(en,EHDR,HRO0+j,f"ro{j}")
for j in range(J): hcell(en,EHDR,HRI0+j,f"ri{j}")

def dist_sample(u, mn, ml, mx, distcell):
    """Return formula body sampling a 3PE distribution from stored uniform u."""
    tri=(f"IF({u}<({ml}-{mn})/({mx}-{mn}),{mn}+SQRT({u}*({mx}-{mn})*({ml}-{mn})),"
         f"{mx}-SQRT((1-{u})*({mx}-{mn})*({mx}-{ml})))")
    pert=(f"_xlfn.BETA.INV({u},1+4*({ml}-{mn})/({mx}-{mn}),1+4*({mx}-{ml})/({mx}-{mn}),{mn},{mx})")
    norm=f"MAX(0,_xlfn.NORM.INV({u},{ml},({mx}-{mn})/6))"
    return f"IF({distcell}=\"Pert\",{pert},IF({distcell}=\"Normal\",{norm},{tri}))"

for i in range(ITER):
    r=EF+i
    en.cell(r,ITER_C,i+1)
    # hidden uniforms
    for k in range(K): en.cell(r,HU0+k,"=RAND()")
    for j in range(J): en.cell(r,HRO0+j,"=RAND()")
    for j in range(J): en.cell(r,HRI0+j,"=RAND()")
    # cost line samples (include-gated)
    for k in range(K):
        clr=CLF+k
        mn=ref(CL,10,clr); ml=ref(CL,11,clr); mx=ref(CL,12,clr)   # Min/ML/Max Total
        dc=ref(CL,13,clr); inc=ref(CL,14,clr)
        u=f"${L(HU0+k)}{r}"
        en.cell(r,SAMP0+k,f"=IF({inc}=\"Yes\",{dist_sample(u,mn,ml,mx,dc)},0)")
    # risk impacts (probability-gated occurrence + sampled impact)
    for j in range(J):
        rrr=RRF+j
        prob=ref(RR,4,rrr); mn=ref(RR,5,rrr); ml=ref(RR,6,rrr); mx=ref(RR,7,rrr)
        dc=ref(RR,8,rrr); inc=ref(RR,13,rrr)
        ro=f"${L(HRO0+j)}{r}"; ui=f"${L(HRI0+j)}{r}"
        occ=f"IF({ro}<={prob},{dist_sample(ui,mn,ml,mx,dc)},0)"
        en.cell(r,IMP0+j,f"=IF({inc}=\"Yes\",{occ},0)")
    # year totals (inflated): (sum cost*profile + sum risk*profile) * inflation factor
    for y in range(Y):
        cterm="+".join(f"${L(SAMP0+k)}{r}*{costprof(k,y)}" for k in range(K))
        rterm="+".join(f"${L(IMP0+j)}{r}*{riskprof(j,y)}" for j in range(J))
        en.cell(r,YT0+y,f"=(({cterm})+({rterm}))*{inflfac(y)}")
    # base / risk inflated totals
    base="+".join(f"(({'+'.join(f'${L(SAMP0+k)}{r}*{costprof(k,y)}' for k in range(K))})*{inflfac(y)})" for y in range(Y))
    risk="+".join(f"(({'+'.join(f'${L(IMP0+j)}{r}*{riskprof(j,y)}' for j in range(J))})*{inflfac(y)})" for y in range(Y))
    en.cell(r,BASE_C,f"={base}")
    en.cell(r,RISK_C,f"={risk}")
    en.cell(r,TOT_C,f"=SUM({L(YT0)}{r}:{L(YT0+Y-1)}{r})")
    npv="+".join(f"${L(YT0+y)}{r}/(1+{DISC_CELL})^{y+1}" for y in range(Y))
    en.cell(r,NPV_C,f"={npv}")
# hide helper columns
for cc in list(range(HU0,HU0+K))+list(range(HRO0,HRO0+J))+list(range(HRI0,HRI0+J)):
    en.column_dimensions[L(cc)].hidden=True
en.freeze_panes="B4"
# ranges
TOTrng=rng(EN,TOT_C,EF,TOT_C,ELAST)
NPVrng=rng(EN,NPV_C,EF,NPV_C,ELAST)
BASErng=rng(EN,BASE_C,EF,BASE_C,ELAST)
RISKrng=rng(EN,RISK_C,EF,RISK_C,ELAST)
def YTrng(y): return rng(EN,YT0+y,EF,YT0+y,ELAST)
def SAMPrng(k): return rng(EN,SAMP0+k,EF,SAMP0+k,ELAST)
def IMPrng(j): return rng(EN,IMP0+j,EF,IMP0+j,ELAST)
def PCT(r,p): return f"_xlfn.PERCENTILE.INC({r},{p})"

# ============================================================================
# 9. RESULTS
# ============================================================================
RS="Results"; rs=newsheet(RS, NAVY)
rs.cell(1,1,"RESULTS").font=TITLE
rs.cell(2,1,"All figures recalc live from the Engine. Press F9 to re-run the simulation.").font=SUB
# A. Total cost percentiles
rs.cell(4,1,"A.  Total Cost Percentiles").font=F(bold=True,size=12,color=ORANGE)
hcell(rs,5,1,"Percentile"); hcell(rs,5,2,"Total Cost")
ps=[("P10",0.1),("P30",0.3),("P50",0.5),("P60",0.6),("P70",0.7),("P80",0.8),("P90",0.9),("P95",0.95)]
for i,(lbl,p) in enumerate(ps):
    r=6+i; put(rs,r,1,lbl,al=C,font=LBL); put(rs,r,2,f"={PCT(TOTrng,p)}",fmt=MONEY,al=Rt)
# B. Contingency table
b0=6+len(ps)+2
rs.cell(b0-1,1,"B.  Contingency by Confidence Level").font=F(bold=True,size=12,color=ORANGE)
for j,h in enumerate(["Confidence","Total Cost","Base Cost","Contingency","Contingency %"]): hcell(rs,b0,1+j,h)
BASEML=f"SUM('{CL}'!K{CLF}:K{CL_LAST})"   # deterministic base = sum of ML totals
for i,(lbl,p) in enumerate([("P50",0.5),("P70",0.7),("P80",0.8),("P90",0.9)]):
    r=b0+1+i
    put(rs,r,1,lbl,al=C,font=LBL)
    put(rs,r,2,f"={PCT(TOTrng,p)}",fmt=MONEY,al=Rt)
    put(rs,r,3,f"={BASEML}",fmt=MONEY,al=Rt)
    put(rs,r,4,f"=B{r}-C{r}",fmt=MONEY,al=Rt)
    put(rs,r,5,f"=IFERROR(D{r}/C{r},0)",fmt=PCT1,al=C)
# C. Annual cash flow by confidence
c0=b0+1+4+2
rs.cell(c0-1,1,"C.  Annual Cash Flow by Confidence Level (inflated)").font=F(bold=True,size=12,color=ORANGE)
confcols=[("P50",0.5),("P70",0.7),("P80",0.8),("P90",0.9)]
hcell(rs,c0,1,"Year")
for j,(lbl,_) in enumerate(confcols): hcell(rs,c0,2+j,lbl)
for y in range(Y):
    r=c0+1+y; put(rs,r,1,fylabel(y),al=C,font=LBL)
    for j,(lbl,p) in enumerate(confcols): put(rs,r,2+j,f"={PCT(YTrng(y),p)}",fmt=MONEY0,al=Rt)
CF_FIRST=c0+1; CF_LAST=c0+Y
# D. NPV table
d0=c0+1+Y+2
rs.cell(d0-1,1,"D.  NPV by Confidence Level").font=F(bold=True,size=12,color=ORANGE)
for j,h in enumerate(["Confidence","Nominal Total","NPV"]): hcell(rs,d0,1+j,h)
for i,(lbl,p) in enumerate([("P50",0.5),("P70",0.7),("P80",0.8),("P90",0.9)]):
    r=d0+1+i; put(rs,r,1,lbl,al=C,font=LBL)
    put(rs,r,2,f"={PCT(TOTrng,p)}",fmt=MONEY,al=Rt)
    put(rs,r,3,f"={PCT(NPVrng,p)}",fmt=MONEY,al=Rt)
rs.column_dimensions["A"].width=16
for col in "BCDE": rs.column_dimensions[col].width=16

# ============================================================================
# 10. SENSITIVITY
# ============================================================================
SN="Sensitivity"; sn=newsheet(SN, NAVY)
sn.cell(1,1,"SENSITIVITY (TORNADO)").font=TITLE
sn.cell(2,1,"Spread (P90 − P10) of each driver's sampled value — the bigger the bar, the more it drives uncertainty.").font=SUB
SNH=4; SNF=5
for j,h in enumerate(["Driver","Type","P10","P90","Range (P90−P10)"]): hcell(sn,SNH,1+j,h)
drivers=[]
for k in range(K): drivers.append(("'%s'"%COST_LINES[k][1], "Cost", SAMPrng(k)))
for j in range(J): drivers.append(("'%s'"%RISKS[j][1], "Risk", IMPrng(j)))
# static order by reference range (computed below) — keep input order for MVP
for i,(name,typ,rg) in enumerate(drivers):
    r=SNF+i
    put(sn,r,1,name.strip("'")); put(sn,r,2,typ,al=C)
    put(sn,r,3,f"={PCT(rg,0.1)}",fmt=MONEY0,al=Rt)
    put(sn,r,4,f"={PCT(rg,0.9)}",fmt=MONEY0,al=Rt)
    put(sn,r,5,f"=D{r}-C{r}",fmt=MONEY0,al=Rt)
SN_LAST=SNF+len(drivers)-1
sn.column_dimensions["A"].width=24
for col in "BCDE": sn.column_dimensions[col].width=15
tor=BarChart(); tor.type="bar"; tor.title="Tornado — uncertainty drivers"; tor.legend=None
tor.add_data(Reference(sn,min_col=5,min_row=SNH,max_row=SN_LAST),titles_from_data=True)
tor.set_categories(Reference(sn,min_col=1,min_row=SNF,max_row=SN_LAST))
tor.height=8; tor.width=16; tor.x_axis.delete=False; tor.y_axis.delete=False
sn.add_chart(tor,"G4")

# ============================================================================
# 11. CHARTDATA (hidden) — histogram, cumulative, s-curve
# ============================================================================
CD="ChartData"; cd=newsheet(CD)
cd.sheet_state="hidden"
NB=30
cd.cell(1,1,"Bin"); cd.cell(1,2,"Low"); cd.cell(1,3,"High"); cd.cell(1,4,"Center"); cd.cell(1,5,"Count")
mn=f"MIN({TOTrng})"; mx=f"MAX({TOTrng})"
for b in range(NB):
    r=2+b
    cd.cell(r,1,b+1)
    cd.cell(r,2,f"={mn}+{b}*({mx}-{mn})/{NB}")
    cd.cell(r,3,f"={mn}+{b+1}*({mx}-{mn})/{NB}")
    cd.cell(r,4,f"=(B{r}+C{r})/2").number_format=MONEY0
    cd.cell(r,5,f"=COUNTIF({TOTrng},\"<\"&C{r})-COUNTIF({TOTrng},\"<\"&B{r})")
HIST_FIRST=2; HIST_LAST=1+NB
# cumulative curve points
cd.cell(1,7,"Cum %"); cd.cell(1,8,"Total Cost")
NP=21
for i in range(NP):
    r=2+i; p=i/(NP-1)
    cd.cell(r,7,p).number_format=PCT0
    cd.cell(r,8,f"={PCT(TOTrng,p)}").number_format=MONEY0
CUM_FIRST=2; CUM_LAST=1+NP
# S-curve by year: cumulative P50 and P80 spend
cd.cell(1,10,"Year"); cd.cell(1,11,"Cum P50"); cd.cell(1,12,"Cum P80")
for y in range(Y):
    r=2+y; cd.cell(r,10,f"={START_CELL}+{y}")
    p50=f"SUM({rng(RS,2,CF_FIRST,2,CF_FIRST+y)})"   # cumulative of Results C P50 column
    p80=f"SUM({rng(RS,4,CF_FIRST,4,CF_FIRST+y)})"   # P80 column (4th col -> col D index 4)
    cd.cell(r,11,f"={p50}").number_format=MONEY0
    cd.cell(r,12,f"={p80}").number_format=MONEY0
SC_FIRST=2; SC_LAST=1+Y

# ============================================================================
# 12. ASSUMPTIONS
# ============================================================================
AS="Assumptions"; a=newsheet(AS, "888888")
a.cell(1,1,"ASSUMPTIONS LOG").font=TITLE
AH=3; AF=4
for j,h in enumerate(["ID","Area","Assumption","Impact","Source","Owner","Date","Status"]): hcell(a,AH,1+j,h)
ass=[
 ("A1","Currency",f"All values are in {CURRENCY}.","Low","Setup","PM","","Confirmed"),
 ("A2","Tax","VAT treatment follows the Setup selection.","Med","Setup","Finance","","Confirmed"),
 ("A3","Correlation","Cost lines assumed independent unless correlation is added later.","Med","Method","Modeller","","Open"),
 ("A4","Risk","Risk impacts are financial only unless stated otherwise.","Med","Method","Risk","","Confirmed"),
 ("A5","Inflation","Inflation applied annually via cumulative factors.","Med","Inflation","Finance","","Confirmed"),
 ("A6","NPV",f"NPV uses the discount rate in Setup ({int(DISCOUNT*100)}%).","High","Setup","Finance","","Confirmed"),
 ("A7","Profiling","Cost & risk profile rows must equal 100%.","High","Checks","Modeller","","Confirmed"),
]
for i,row in enumerate(ass):
    r=AF+i
    for j,v in enumerate(row): put(a,r,1+j,v)
for col,w in zip("ABCDEFGH",[6,12,46,8,10,10,10,11]): a.column_dimensions[col].width=w

# ============================================================================
# 13. CHECKS
# ============================================================================
CK="Checks"; ck=newsheet(CK, RED)
ck.cell(1,1,"MODEL CHECKS").font=TITLE
ck.cell(2,1,"All checks must read PASS. The Dashboard shows a warning if any fail.").font=SUB
CKH=4; CKF=5
hcell(ck,CKH,1,"Check"); hcell(ck,CKH,2,"Result")
checks=[
 ("Cost profile rows sum to 100%",
  f"=IF(SUMPRODUCT(--(ABS({rng(CP,CP_FY0+Y,CPF,CP_FY0+Y,CP_LAST)}-1)>0.001))=0,\"PASS\",\"FAIL\")"),
 ("Risk profile rows sum to 100%",
  f"=IF(SUMPRODUCT(--(ABS({rng(RP,RP_FY0+Y,RPF,RP_FY0+Y,RP_LAST)}-1)>0.001))=0,\"PASS\",\"FAIL\")"),
 ("Cost lines: Min<=ML<=Max",
  f"=IF(SUMPRODUCT(--({rng(CL,7,CLF,7,CL_LAST)}>{rng(CL,8,CLF,8,CL_LAST)}))+"
  f"SUMPRODUCT(--({rng(CL,8,CLF,8,CL_LAST)}>{rng(CL,9,CLF,9,CL_LAST)}))=0,\"PASS\",\"FAIL\")"),
 ("Risks: Min<=ML<=Max",
  f"=IF(SUMPRODUCT(--({rng(RR,5,RRF,5,RR_LAST)}>{rng(RR,6,RRF,6,RR_LAST)}))+"
  f"SUMPRODUCT(--({rng(RR,6,RRF,6,RR_LAST)}>{rng(RR,7,RRF,7,RR_LAST)}))=0,\"PASS\",\"FAIL\")"),
 ("Risk probabilities in 0..100%",
  f"=IF(SUMPRODUCT(--({rng(RR,4,RRF,4,RR_LAST)}<0))+SUMPRODUCT(--({rng(RR,4,RRF,4,RR_LAST)}>1))=0,\"PASS\",\"FAIL\")"),
 ("Iterations > 0", f"=IF({ITER_CELL}>0,\"PASS\",\"FAIL\")"),
 ("Discount rate valid (0..50%)", f"=IF(AND({DISC_CELL}>=0,{DISC_CELL}<0.5),\"PASS\",\"FAIL\")"),
 ("Number of years > 0", f"=IF({NYEARS_CELL}>0,\"PASS\",\"FAIL\")"),
]
for i,(lbl,fm) in enumerate(checks):
    r=CKF+i; put(ck,r,1,lbl)
    x=put(ck,r,2,fm,al=C); x.font=F(bold=True,size=10)
CK_LAST=CKF+len(checks)-1
ck.cell(CK_LAST+2,1,"All checks").font=LBL
allok=ck.cell(CK_LAST+2,2,f"=IF(COUNTIF(B{CKF}:B{CK_LAST},\"FAIL\")=0,\"ALL OK\",\"CHECK FAILED\")")
allok.font=F(bold=True,size=11); allok.alignment=C
ALLOK_CELL=ref(CK,2,CK_LAST+2)
ck.conditional_formatting.add(f"B{CKF}:B{CK_LAST+2}",
    FormulaRule(formula=['ISNUMBER(SEARCH("FAIL",B5))'], fill=fill("FFC7CE"), font=F(color="9C0006",bold=True)))
ck.conditional_formatting.add(f"B{CKF}:B{CK_LAST+2}",
    FormulaRule(formula=['OR(B5="PASS",B5="ALL OK")'], fill=fill("C6EFCE"), font=F(color="006100",bold=True)))
ck.column_dimensions["A"].width=34; ck.column_dimensions["B"].width=16

# ============================================================================
# 14. DASHBOARD  (built last; references everything)
# ============================================================================
DB="Dashboard"; db=newsheet(DB, ORANGE); wb.move_sheet(DB, -(len(wb.sheetnames)-1))
for col in "ABCDEFGHIJ": db.column_dimensions[col].width=15
db.cell(1,1,"MONTE CARLO COST MODEL").font=F(bold=True,size=20,color=ORANGE)
db.cell(2,1,f"=\"{PROJECT}\"").font=F(bold=True,size=12)
db.cell(2,1,f"={ref(S,3,r0)}").font=F(bold=True,size=12)
nt=db.cell(3,1,f'="All figures in "&{ref(S,3,r0+1)}&" — change the Currency cell on the Setup sheet to switch."')
nt.font=SUB
# meta strip
meta=[("Currency",ref(S,3,r0+1)),("Base year",ref(S,3,r0+2)),("Years",NYEARS_CELL),
      ("Iterations",ITER_CELL),("Confidence",CONF_CELL),("Distribution",ref(S,3,r0+7))]
for i,(lbl,cell) in enumerate(meta):
    cc=1+i
    x=db.cell(4,cc,lbl); x.font=F(size=9,color="888888"); x.alignment=C
    y=db.cell(5,cc,f"={cell}"); y.font=F(bold=True,size=11); y.alignment=C; y.fill=fill(GREYBG); y.border=BOX
# run / status
run=db.cell(7,1,"▶  PRESS  F9  TO  RUN / RE-ROLL  SIMULATION"); run.font=F(bold=True,size=12,color=WHITE)
run.fill=fill(TEAL); run.alignment=C
db.merge_cells("A7:D7")
warn=db.cell(7,6,f"={ALLOK_CELL}"); warn.font=F(bold=True,size=12); warn.alignment=C; warn.fill=fill(GREYBG); warn.border=BOX
db.merge_cells("F7:J7")
db.conditional_formatting.add("F7:J7",FormulaRule(formula=['$F$7="ALL OK"'],fill=fill("C6EFCE"),font=F(color="006100",bold=True)))
db.conditional_formatting.add("F7:J7",FormulaRule(formula=['$F$7<>"ALL OK"'],fill=fill("FFC7CE"),font=F(color="9C0006",bold=True)))

# KPI cards
def kpi(row,col,label,formula,fmt=MONEY,color=NAVY):
    l=db.cell(row,col,label); l.font=F(size=9,color="888888"); l.alignment=Lt
    v=db.cell(row+1,col,formula); v.number_format=fmt; v.font=F(bold=True,size=14,color=color); v.alignment=Lt
    db.cell(row,col).fill=fill(GREYBG); db.cell(row+1,col).fill=fill(GREYBG)
    for rr in (row,row+1):
        for cc in (col,col+1):
            db.cell(rr,cc).fill=fill(GREYBG)
    db.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+1)
    db.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+1)

db.cell(9,1,"KEY OUTPUTS").font=F(bold=True,size=12,color=ORANGE)
ML=f"={BASEML}"
kpi(10,1,"Base Cost (ML)",f"={BASEML}")
kpi(10,3,"Expected Cost",f"=AVERAGE({BASErng})")
kpi(10,5,"Expected Risk",f"=AVERAGE({RISKrng})",color=RED)
kpi(10,7,"Total Expected Cost",f"=AVERAGE({TOTrng})",color=ORANGE)
kpi(13,1,"P50",f"={PCT(TOTrng,0.5)}")
kpi(13,3,"P70",f"={PCT(TOTrng,0.7)}")
kpi(13,5,"P80",f"={PCT(TOTrng,0.8)}")
kpi(13,7,"P90",f"={PCT(TOTrng,0.9)}")
kpi(16,1,"Recommended Contingency",f"={PCT(TOTrng,CONFP_CELL)}-{BASEML}",color=TEAL)
kpi(16,3,"Recommended Budget",f"={PCT(TOTrng,CONFP_CELL)}",color=TEAL)
kpi(16,5,"NPV @ selected confidence",f"={PCT(NPVrng,CONFP_CELL)}",color=NAVY)
kpi(16,7,"P95 (worst-ish)",f"={PCT(TOTrng,0.95)}",color=RED)

# Charts on dashboard
hist=BarChart(); hist.type="col"; hist.title="Total Cost Distribution"; hist.legend=None; hist.gapWidth=6
hist.add_data(Reference(cd,min_col=5,min_row=1,max_row=HIST_LAST),titles_from_data=True)
hist.set_categories(Reference(cd,min_col=4,min_row=HIST_FIRST,max_row=HIST_LAST))
hist.x_axis.delete=False; hist.y_axis.delete=False; hist.height=7.5; hist.width=13
db.add_chart(hist,"A20")

cum=ScatterChart(); cum.title="Cumulative Probability (S-Curve)"; cum.legend=None
cum.x_axis.title="Total cost"; cum.y_axis.title="Probability"
xv=Reference(cd,min_col=8,min_row=CUM_FIRST,max_row=CUM_LAST)
yv=Reference(cd,min_col=7,min_row=CUM_FIRST,max_row=CUM_LAST)
s=Series(yv,xv,title="S-curve"); s.smooth=True; cum.series.append(s)
cum.x_axis.delete=False; cum.y_axis.delete=False; cum.height=7.5; cum.width=13
db.add_chart(cum,"F20")

sc=LineChart(); sc.title="Cumulative Cash Flow by Year"; sc.style=12
sc.add_data(Reference(cd,min_col=11,min_row=1,max_col=12,max_row=SC_LAST),titles_from_data=True)
sc.set_categories(Reference(cd,min_col=10,min_row=SC_FIRST,max_row=SC_LAST))
sc.x_axis.delete=False; sc.y_axis.delete=False; sc.height=7.5; sc.width=13
db.add_chart(sc,"A36")

tor2=BarChart(); tor2.type="bar"; tor2.title="Sensitivity (Tornado)"; tor2.legend=None
tor2.add_data(Reference(sn,min_col=5,min_row=SNH,max_row=SN_LAST),titles_from_data=True)
tor2.set_categories(Reference(sn,min_col=1,min_row=SNF,max_row=SN_LAST))
tor2.x_axis.delete=False; tor2.y_axis.delete=False; tor2.height=7.5; tor2.width=13
db.add_chart(tor2,"F36")

db.sheet_view.showGridLines=False

# Flex mode: pre-hide the year columns/rows beyond the active duration so the
# file opens looking like an ACTIVE_YEARS project. The ApplySettings macro
# toggles these the same way when the user changes "Number of years".
if FLEX and INACTIVE:
    for y in INACTIVE:
        cp.column_dimensions[L(CP_FY0 + y)].hidden = True   # Cost Profiling FY col
        rp.column_dimensions[L(RP_FY0 + y)].hidden = True   # Risk Profiling FY col
        inf.row_dimensions[INFF + y].hidden = True          # Inflation FY row
        rs.row_dimensions[CF_FIRST + y].hidden = True       # Results cash-flow row
        en.column_dimensions[L(YT0 + y)].hidden = True      # Engine year col (calc)

# remove the default empty sheet, then order
if "Sheet" in wb.sheetnames: del wb["Sheet"]
order=["Dashboard","Setup","Cost Lines","Cost Profiling","Risk Register","Risk Profiling",
       "Inflation","Engine","Results","Sensitivity","Assumptions","Checks","ChartData"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.active=0

wb.save(OUTFILE)

# ============================================================================
# 15. INDEPENDENT PYTHON REFERENCE (sanity check)
# ============================================================================
def samp(dist,mn,ml,mx,rng_):
    if dist=="Normal": return max(0,rng_.gauss(ml,(mx-mn)/6))
    if dist=="Pert":
        a=1+4*(ml-mn)/(mx-mn); b=1+4*(mx-ml)/(mx-mn)
        return mn+rng_.betavariate(a,b)*(mx-mn)
    return rng_.triangular(mn,mx,ml)
rng_=random.Random(7)
fac=[1.0]
for y in range(1,Y): fac.append(fac[-1]*(1+INFL_RATES[y]))
totals=[]; npvs=[]
for _ in range(20000):
    cs=[samp(c[8],c[5],c[6],c[7],rng_) for c in COST_LINES]
    ri=[(samp(rk[7],rk[4],rk[5],rk[6],rng_) if rng_.random()<=rk[3] else 0) for rk in RISKS]
    yt=[]
    for y in range(Y):
        base=sum(cs[k]*COST_PROFILES[k][y] for k in range(K))
        rsk=sum(ri[j]*RISK_PROFILES[j][y] for j in range(J))
        yt.append((base+rsk)*fac[y])
    totals.append(sum(yt)); npvs.append(sum(yt[y]/(1+DISCOUNT)**(y+1) for y in range(Y)))
totals.sort(); npvs.sort()
baseml=sum(c[6] for c in COST_LINES)
def pc(a,p): return a[int(p*(len(a)-1))]
print(f"Saved {OUTFILE}  ({ITER:,} iterations, {K} cost lines, {J} risks, "
      f"{'FLEX build ' + str(Y) + ' yrs, active=' + str(ACTIVE_YEARS) if FLEX else str(Y) + ' years'})")
print("Reference (independent 20k Python run — Excel will be close):")
print(f"  Base (ML)            {CURRENCY} {baseml:,.0f}")
print(f"  Total expected       {CURRENCY} {sum(totals)/len(totals):,.0f}")
print(f"  P50 / P80 / P90      {CURRENCY} {pc(totals,.5):,.0f} / {pc(totals,.8):,.0f} / {pc(totals,.9):,.0f}")
print(f"  Contingency @P80     {CURRENCY} {pc(totals,.8)-baseml:,.0f}")
print(f"  NPV @P80             {CURRENCY} {pc(npvs,.8):,.0f}")
