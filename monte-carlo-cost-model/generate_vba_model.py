#!/usr/bin/env python3
"""
VBA-engine Monte Carlo Cost Model — workbook generator.

This builds a LEAN, macro-driven workbook. Unlike the formula version, the
simulation is NOT in sheet formulas — it runs inside the VBA engine
(vba/modEngine.bas, RunSimulation). That makes everything fully dynamic:

  • number of cost lines / risks  -> just add rows to the tables
  • number of iterations          -> a normal input cell on Setup
  • number of years               -> a normal input cell on Setup

The macro reads the input tables (any size), runs the Monte Carlo in memory, and
writes the Results / Sensitivity / Dashboard / charts. The workbook therefore
shows results only AFTER you run the macro (there is no F9 fallback).

Import the VBA modules and save as .xlsm — see vba/VBA_SETUP.md.

Usage:  python3 generate_vba_model.py
Out:    MonteCarloCostModel_VBA.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, ScatterChart, LineChart, Reference, Series
from openpyxl.utils import get_column_letter as L

# ---------------------------------------------------------------- config / data
OUTFILE   = "MonteCarloCostModel_VBA.xlsx"
PROJECT   = "Sample Construction Project"
CURRENCY  = "SAR"
BASE_YEAR = 2025
ACTIVE_Y  = 5            # years shown by default
MAXY      = 30           # year columns/rows built into the input tables
ITERS_DEF = 10000        # default iterations (a real input the macro reads)
CONF_DEF  = "P80"
DISCOUNT  = 0.08
INFL_BASE = 0.03

# cost lines: WBS, item, category, unit, qty, minU, mlU, maxU, dist, shape
COST_LINES = [
    ("1.1","Site Preparation","Civil","LS",1,28000,38000,60000,"Triangular","early"),
    ("1.2","Foundations","Civil","LS",1,45000,55000,75000,"Triangular","early"),
    ("2.1","Superstructure","Structural","LS",1,90000,120000,170000,"Pert","bell"),
    ("3.1","MEP Systems","MEP","LS",1,60000,80000,115000,"Pert","late"),
    ("4.1","Finishes","Architectural","LS",1,30000,45000,70000,"Triangular","late"),
    ("5.1","Project Management","PM","LS",1,20000,28000,40000,"Normal","even"),
]
RISKS = [
    ("R1","Adverse ground conditions","Technical",0.30,20000,45000,90000,"Triangular","Civil Lead","early"),
    ("R2","Major design change","Commercial",0.40,15000,30000,60000,"Pert","Design Mgr","bell"),
    ("R3","Material price spike","Market",0.50,10000,25000,55000,"Triangular","Procurement","bell"),
    ("R4","Schedule delay","Delivery",0.25,20000,40000,80000,"Pert","PM","late"),
]
NUM_LINES = max(1, int(os.environ.get("LINES", len(COST_LINES))))
NUM_RISKS = max(0, int(os.environ.get("RISKS", len(RISKS))))
while len(COST_LINES) < NUM_LINES:
    i=len(COST_LINES); COST_LINES.append((f"{i+1}.0",f"Cost Item {i+1}","—","LS",1,0,0,0,"Triangular","even"))
while len(RISKS) < NUM_RISKS:
    i=len(RISKS); RISKS.append((f"R{i+1}",f"Risk {i+1}","—",0.0,0,0,0,"Triangular","","even"))
COST_LINES=COST_LINES[:NUM_LINES]; RISKS=RISKS[:NUM_RISKS]
K, J = len(COST_LINES), len(RISKS)

def make_profile(Yactive, shape):
    if Yactive==1: return [1.0]
    if shape=="early": w=[float(Yactive-i) for i in range(Yactive)]
    elif shape=="late": w=[float(i+1) for i in range(Yactive)]
    elif shape=="bell":
        mid=(Yactive-1)/2.0; w=[max(0.05,1.0-abs(i-mid)/(mid+1)) for i in range(Yactive)]
    else: w=[1.0]*Yactive
    s=sum(w); return [x/s for x in w]
def pad(p): return p+[0.0]*(MAXY-len(p))
COST_PROFILES=[pad(make_profile(ACTIVE_Y,c[9])) for c in COST_LINES]
RISK_PROFILES=[pad(make_profile(ACTIVE_Y,r[9])) for r in RISKS]

# ---------------------------------------------------------------- styles
NAVY="1F2A44"; ORANGE="FF6B35"; TEAL="00A88A"; PURPLE="7C5CFC"; RED="C0392B"
INPUT="FFF7DC"; GREY="F2F3F5"; WHITE="FFFFFF"
MONEY='#,##0'; PCT0='0%'; PCT1='0.0%'
def Fn(**k): return Font(name="Calibri",**k)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side("thin",color="C9CDD6"); BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
Cn=Alignment("center",vertical="center"); Rt=Alignment("right",vertical="center"); Lt=Alignment("left",vertical="center")
TITLE=Fn(bold=True,size=16,color=ORANGE); SUB=Fn(size=10,color="6C6C70"); LBL=Fn(bold=True,size=10,color="333333")
HF=Fn(bold=True,size=11,color=WHITE)

wb=Workbook()
def sheet(name,tab=None):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
    if tab: ws.sheet_properties.tabColor=tab
    return ws
def hc(ws,r,c,v):
    x=ws.cell(r,c,v); x.font=HF; x.fill=fill(NAVY); x.alignment=Cn; x.border=BOX; return x
def put(ws,r,c,v,fmt=None,font=None,al=None,bd=True,fillc=None):
    x=ws.cell(r,c,v)
    if fmt:x.number_format=fmt
    x.font=font or Fn(size=10); x.alignment=al or Lt
    if bd:x.border=BOX
    if fillc:x.fill=fill(fillc)
    return x

# ================================================================ SETUP
S="Setup"; st=sheet(S,NAVY)
st.cell(1,2,"MODEL SETUP").font=TITLE
st.cell(2,2,"Yellow cells are inputs. Run the simulation from the Dashboard (or Alt+F8 ▸ RunSimulation).").font=SUB
srows=[("Project name",PROJECT,None),("Currency",CURRENCY,None),("Base year",BASE_YEAR,"0"),
 ("Start year",BASE_YEAR,"0"),("Number of years",ACTIVE_Y,"0"),("Iterations",ITERS_DEF,"#,##0"),
 ("Confidence level",CONF_DEF,None),("Discount rate",DISCOUNT,PCT1),("VAT treatment","Excluded",None),
 ("Prepared by","",None),("Date","",None)]
r0=4
for i,(lbl,val,fmt) in enumerate(srows):
    put(st,r0+i,2,lbl,font=LBL,bd=False)
    put(st,r0+i,3,val,fmt=fmt,al=Rt,fillc=INPUT).font=Fn(size=10,bold=True)
# confidence numeric helper
put(st,r0+12,2,"Confidence (numeric)",font=Fn(size=9,italic=True,color="888888"),bd=False)
put(st,r0+12,3,f"=VALUE(MID(C{r0+6},2,2))/100",fmt=PCT0,al=Rt).font=Fn(size=9,italic=True,color="888888")
DataValidation(type="list",formula1='"P50,P60,P70,P80,P90,P95"')  # placeholder to keep imports tidy
dv=DataValidation(type="list",formula1='"P50,P60,P70,P80,P90,P95"'); st.add_data_validation(dv); dv.add(f"C{r0+6}")
dv=DataValidation(type="list",formula1='"Included,Excluded"'); st.add_data_validation(dv); dv.add(f"C{r0+8}")
st.cell(r0+8,5,"↳ inflation is per year on the Inflation sheet").font=Fn(size=9,italic=True,color="888888")
st.column_dimensions["B"].width=24; st.column_dimensions["C"].width=20; st.column_dimensions["E"].width=36
# named-ish cells (documented in VBA): C8=years C9=iters C10=conf C11=disc C16=confP

# ================================================================ COST LINES
CL="Cost Lines"; cl=sheet(CL,ORANGE)
cl.cell(1,1,"COST LINES — 3-Point Estimates  (add rows freely; the macro reads them all)").font=TITLE
cols=["WBS","Cost Item","Category","Description","Unit","Quantity","Min Unit Cost","Most Likely Unit Cost",
      "Max Unit Cost","Min Total Cost","Most Likely Total Cost","Max Total Cost","Distribution","Include?","Notes"]
H=3; F0=4
for j,h in enumerate(cols): hc(cl,H,1+j,h)
for i,(wbs,item,cat,unit,qty,mn,ml,mx,dist,shape) in enumerate(COST_LINES):
    r=F0+i
    put(cl,r,1,wbs,al=Cn);put(cl,r,2,item);put(cl,r,3,cat);put(cl,r,4,f"{item} works");put(cl,r,5,unit,al=Cn)
    for cc,val in ((6,qty),(7,mn),(8,ml),(9,mx)): put(cl,r,cc,val,fmt=MONEY,al=Rt,fillc=INPUT)
    put(cl,r,10,f"=F{r}*G{r}",fmt=MONEY,al=Rt);put(cl,r,11,f"=F{r}*H{r}",fmt=MONEY,al=Rt);put(cl,r,12,f"=F{r}*I{r}",fmt=MONEY,al=Rt)
    put(cl,r,13,dist,al=Cn);put(cl,r,14,"Yes" if ml>0 else "No",al=Cn);put(cl,r,15,"")
CLL=F0+K-1
t=Table(displayName="tbl_CostLines",ref=f"A{H}:O{CLL}"); t.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True); cl.add_table(t)
dv=DataValidation(type="list",formula1='"Triangular,Pert,Normal"'); cl.add_data_validation(dv); dv.add(f"M{F0}:M{CLL+60}")
dv=DataValidation(type="list",formula1='"Yes,No"'); cl.add_data_validation(dv); dv.add(f"N{F0}:N{CLL+60}")
b=cl.cell(2,1,"➕  Add Cost Line   →  assign macro: AddCostLine"); b.font=Fn(bold=True,size=10,color=WHITE); b.fill=fill(TEAL); b.alignment=Cn; cl.merge_cells("A2:F2")
for col,w in zip("ABCDEFGHIJKLMNO",[7,20,12,16,6,9,12,15,12,12,14,12,12,9,14]): cl.column_dimensions[col].width=w
cl.freeze_panes="B4"

# ================================================================ COST PROFILING
CP="Cost Profiling"; cp=sheet(CP,ORANGE)
cp.cell(1,1,"COST PROFILING — % per year (rows must total 100%)").font=TITLE
hc(cp,3,1,"WBS"); hc(cp,3,2,"Cost Item")
for y in range(MAXY): hc(cp,3,3+y,f"FY{y+1}")
hc(cp,3,3+MAXY,"Total %")
for i in range(K):
    r=F0+i
    put(cp,r,1,f"='{CL}'!A{F0+i}",al=Cn); put(cp,r,2,f"='{CL}'!B{F0+i}")
    for y in range(MAXY): put(cp,r,3+y,COST_PROFILES[i][y],fmt=PCT0,al=Cn,fillc=INPUT)
    # Total % counts only the ACTIVE years (Setup C8) so hidden years never break it
    put(cp,r,3+MAXY,f"=SUM(C{r}:INDEX(C{r}:{L(2+MAXY)}{r},'{S}'!$C$8))",fmt=PCT0,al=Cn).font=Fn(bold=True,size=10)
CPL=F0+K-1
t=Table(displayName="tbl_CostProfile",ref=f"A3:{L(3+MAXY)}{CPL}"); t.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True); cp.add_table(t)
cp.conditional_formatting.add(f"{L(3+MAXY)}{F0}:{L(3+MAXY)}{CPL}",FormulaRule(formula=[f"ABS({L(3+MAXY)}{F0}-1)>0.001"],fill=fill("FFC7CE")))
cp.column_dimensions["A"].width=8; cp.column_dimensions["B"].width=20

# ================================================================ RISK REGISTER
RR="Risk Register"; rr=sheet(RR,RED)
rr.cell(1,1,"RISK REGISTER — probability-weighted 3PE impacts").font=TITLE
rcols=["Risk ID","Risk Name","Category","Probability","Min Impact","Most Likely Impact","Max Impact","Distribution","Risk Owner","Include?","Notes"]
for j,h in enumerate(rcols): hc(rr,H,1+j,h)
for i,(rid,name,cat,prob,mn,ml,mx,dist,owner,shape) in enumerate(RISKS):
    r=F0+i
    put(rr,r,1,rid,al=Cn);put(rr,r,2,name);put(rr,r,3,cat)
    put(rr,r,4,prob,fmt=PCT0,al=Cn,fillc=INPUT)
    for cc,val in ((5,mn),(6,ml),(7,mx)): put(rr,r,cc,val,fmt=MONEY,al=Rt,fillc=INPUT)
    put(rr,r,8,dist,al=Cn);put(rr,r,9,owner);put(rr,r,10,"Yes" if ml>0 else "No",al=Cn);put(rr,r,11,"")
RRL=F0+J-1
t=Table(displayName="tbl_RiskRegister",ref=f"A{H}:K{RRL}"); t.tableStyleInfo=TableStyleInfo(name="TableStyleMedium3",showRowStripes=True); rr.add_table(t)
dv=DataValidation(type="list",formula1='"Triangular,Pert,Normal"'); rr.add_data_validation(dv); dv.add(f"H{F0}:H{RRL+60}")
dv=DataValidation(type="list",formula1='"Yes,No"'); rr.add_data_validation(dv); dv.add(f"J{F0}:J{RRL+60}")
b=rr.cell(2,1,"➕  Add Risk   →  assign macro: AddRisk"); b.font=Fn(bold=True,size=10,color=WHITE); b.fill=fill(RED); b.alignment=Cn; rr.merge_cells("A2:F2")
for col,w in zip("ABCDEFGHIJK",[8,24,12,11,12,15,12,12,13,9,14]): rr.column_dimensions[col].width=w
rr.freeze_panes="B4"

# ================================================================ RISK PROFILING
RP="Risk Profiling"; rp=sheet(RP,RED)
rp.cell(1,1,"RISK PROFILING — % per year (rows must total 100%)").font=TITLE
hc(rp,3,1,"Risk ID"); hc(rp,3,2,"Risk Name")
for y in range(MAXY): hc(rp,3,3+y,f"FY{y+1}")
hc(rp,3,3+MAXY,"Total %")
for i in range(J):
    r=F0+i
    put(rp,r,1,f"='{RR}'!A{F0+i}",al=Cn); put(rp,r,2,f"='{RR}'!B{F0+i}")
    for y in range(MAXY): put(rp,r,3+y,RISK_PROFILES[i][y],fmt=PCT0,al=Cn,fillc=INPUT)
    put(rp,r,3+MAXY,f"=SUM(C{r}:INDEX(C{r}:{L(2+MAXY)}{r},'{S}'!$C$8))",fmt=PCT0,al=Cn).font=Fn(bold=True,size=10)
RPL=F0+J-1
t=Table(displayName="tbl_RiskProfile",ref=f"A3:{L(3+MAXY)}{RPL}"); t.tableStyleInfo=TableStyleInfo(name="TableStyleMedium3",showRowStripes=True); rp.add_table(t)
rp.conditional_formatting.add(f"{L(3+MAXY)}{F0}:{L(3+MAXY)}{RPL}",FormulaRule(formula=[f"ABS({L(3+MAXY)}{F0}-1)>0.001"],fill=fill("FFC7CE")))
rp.column_dimensions["A"].width=8; rp.column_dimensions["B"].width=24

# ================================================================ INFLATION
IN="Inflation"; inf=sheet(IN,TEAL)
inf.cell(1,1,"INFLATION — rate per year; cumulative factor compounds (FY1 = 1.0)").font=TITLE
hc(inf,3,1,"FY");hc(inf,3,2,"Year");hc(inf,3,3,"Inflation Rate");hc(inf,3,4,"Cumulative Factor")
for y in range(MAXY):
    r=F0+y
    put(inf,r,1,f"FY{y+1}",al=Cn); put(inf,r,2,f"='{S}'!$C${r0+3}+{y}",fmt="0",al=Cn)
    put(inf,r,3,(0.0 if y==0 else INFL_BASE),fmt=PCT1,al=Cn,fillc=INPUT)
    put(inf,r,4,("=1" if y==0 else f"=D{r-1}*(1+C{r})"),fmt="0.0000",al=Cn)
for col,w in zip("ABCD",[8,8,14,16]): inf.column_dimensions[col].width=w

# ================================================================ RESULTS (VBA writes values)
RS="Results"; rs=sheet(RS,NAVY)
rs.cell(1,1,"RESULTS").font=TITLE
rs.cell(2,1,"▶ Run the simulation (Dashboard button or Alt+F8 ▸ RunSimulation) to populate these.").font=Fn(size=10,italic=True,color=ORANGE)
rs.cell(4,1,"A. Total Cost Percentiles").font=Fn(bold=True,size=12,color=ORANGE)
hc(rs,5,1,"Percentile");hc(rs,5,2,"Total Cost")
for i,p in enumerate(["P10","P30","P50","P60","P70","P80","P90","P95"]):
    put(rs,6+i,1,p,al=Cn,font=LBL); put(rs,6+i,2,None,fmt=MONEY,al=Rt)
rs.cell(15,1,"B. Contingency by Confidence").font=Fn(bold=True,size=12,color=ORANGE)
for j,h in enumerate(["Confidence","Total Cost","Base Cost","Contingency","Contingency %"]): hc(rs,16,1+j,h)
for i,p in enumerate(["P50","P70","P80","P90"]):
    put(rs,17+i,1,p,al=Cn,font=LBL)
    for c in range(2,5): put(rs,17+i,c,None,fmt=MONEY,al=Rt)
    put(rs,17+i,5,None,fmt=PCT1,al=Cn)
rs.cell(22,1,"C. Annual Cash Flow by Confidence").font=Fn(bold=True,size=12,color=ORANGE)
for j,h in enumerate(["Year","P50","P70","P80","P90"]): hc(rs,23,1+j,h)
for y in range(MAXY):            # rows 24..53 (VBA fills first N)
    put(rs,24+y,1,None,al=Cn,font=LBL)
    for c in range(2,6): put(rs,24+y,c,None,fmt=MONEY,al=Rt)
rs.cell(56,1,"D. NPV by Confidence").font=Fn(bold=True,size=12,color=ORANGE)
for j,h in enumerate(["Confidence","Nominal Total","NPV"]): hc(rs,57,1+j,h)
for i,p in enumerate(["P50","P70","P80","P90"]):
    put(rs,58+i,1,p,al=Cn,font=LBL); put(rs,58+i,2,None,fmt=MONEY,al=Rt); put(rs,58+i,3,None,fmt=MONEY,al=Rt)
rs.column_dimensions["A"].width=16
for c in "BCDE": rs.column_dimensions[c].width=15

# ================================================================ SENSITIVITY (VBA writes)
SN="Sensitivity"; sn=sheet(SN,NAVY)
sn.cell(1,1,"SENSITIVITY (TORNADO)").font=TITLE
sn.cell(2,1,"P90 − P10 spread of each driver's sampled value (filled by the macro).").font=SUB
for j,h in enumerate(["Driver","Type","P10","P90","Range (P90−P10)"]): hc(sn,4,1+j,h)
SN_F=5; SN_MAX=SN_F+ (K+J) + 6   # room for drivers (+spares)
for r in range(SN_F,SN_MAX):
    for c in range(1,6): put(sn,r,c,None,fmt=(MONEY if c>=3 else None),al=(Rt if c>=3 else Lt))
sn.column_dimensions["A"].width=24
for c in "BCDE": sn.column_dimensions[c].width=15
tor=BarChart(); tor.type="bar"; tor.title="Tornado — uncertainty drivers"; tor.legend=None
tor.add_data(Reference(sn,min_col=5,min_row=4,max_row=SN_MAX-1),titles_from_data=True)
tor.set_categories(Reference(sn,min_col=1,min_row=SN_F,max_row=SN_MAX-1))
tor.height=9; tor.width=16; tor.x_axis.delete=False; tor.y_axis.delete=False
sn.add_chart(tor,"G4")

# ================================================================ CHARTDATA (hidden, VBA writes)
CD="ChartData"; cd=sheet(CD); cd.sheet_state="hidden"
NB=30; NP=21
cd.cell(1,1,"Center");cd.cell(1,2,"Count")
for r in range(2,2+NB): put(cd,r,1,None,fmt=MONEY,bd=False); put(cd,r,2,None,bd=False)
cd.cell(1,4,"Cost");cd.cell(1,5,"CumProb")
for r in range(2,2+NP): put(cd,r,4,None,fmt=MONEY,bd=False); put(cd,r,5,None,fmt=PCT0,bd=False)
cd.cell(1,7,"Year");cd.cell(1,8,"Cum P50");cd.cell(1,9,"Cum P80")
for r in range(2,2+MAXY):
    put(cd,r,7,None,bd=False); put(cd,r,8,None,fmt=MONEY,bd=False); put(cd,r,9,None,fmt=MONEY,bd=False)

# ================================================================ CHECKS
CK="Checks"; ck=sheet(CK,RED)
ck.cell(1,1,"MODEL CHECKS").font=TITLE
hc(ck,4,1,"Check");hc(ck,4,2,"Result")
checks=[
 ("Cost profile rows total 100%", f"=IF(SUMPRODUCT(--(ABS('Cost Profiling'!${L(3+MAXY)}${F0}:${L(3+MAXY)}${CPL}-1)>0.001))=0,\"PASS\",\"FAIL\")"),
 ("Risk profile rows total 100%", f"=IF(SUMPRODUCT(--(ABS('Risk Profiling'!${L(3+MAXY)}${F0}:${L(3+MAXY)}${RPL}-1)>0.001))=0,\"PASS\",\"FAIL\")"),
 ("Cost lines Min<=ML<=Max", f"=IF(SUMPRODUCT(--('Cost Lines'!$G${F0}:$G${CLL}>'Cost Lines'!$H${F0}:$H${CLL}))+SUMPRODUCT(--('Cost Lines'!$H${F0}:$H${CLL}>'Cost Lines'!$I${F0}:$I${CLL}))=0,\"PASS\",\"FAIL\")"),
 ("Risks Min<=ML<=Max", f"=IF(SUMPRODUCT(--('Risk Register'!$E${F0}:$E${RRL}>'Risk Register'!$F${F0}:$F${RRL}))+SUMPRODUCT(--('Risk Register'!$F${F0}:$F${RRL}>'Risk Register'!$G${F0}:$G${RRL}))=0,\"PASS\",\"FAIL\")"),
 ("Iterations > 0", f"=IF('{S}'!$C${r0+5}>0,\"PASS\",\"FAIL\")"),
 ("Years between 1 and 30", f"=IF(AND('{S}'!$C${r0+4}>=1,'{S}'!$C${r0+4}<=30),\"PASS\",\"FAIL\")"),
 ("Discount rate valid", f"=IF(AND('{S}'!$C${r0+7}>=0,'{S}'!$C${r0+7}<0.5),\"PASS\",\"FAIL\")"),
]
for i,(lbl,fm) in enumerate(checks):
    put(ck,5+i,1,lbl); put(ck,5+i,2,fm,al=Cn).font=Fn(bold=True,size=10)
CKL=5+len(checks)-1
put(ck,CKL+2,1,"All checks",font=LBL,bd=False)
put(ck,CKL+2,2,f"=IF(COUNTIF(B5:B{CKL},\"FAIL\")=0,\"ALL OK\",\"CHECK FAILED\")",al=Cn).font=Fn(bold=True,size=11)
ck.conditional_formatting.add(f"B5:B{CKL+2}",FormulaRule(formula=['ISNUMBER(SEARCH("FAIL",B5))'],fill=fill("FFC7CE"),font=Fn(color="9C0006",bold=True)))
ck.conditional_formatting.add(f"B5:B{CKL+2}",FormulaRule(formula=['OR(B5="PASS",B5="ALL OK")'],fill=fill("C6EFCE"),font=Fn(color="006100",bold=True)))
ck.column_dimensions["A"].width=30; ck.column_dimensions["B"].width=15
ALLOK=f"'{CK}'!$B${CKL+2}"

# ================================================================ DASHBOARD
DB="Dashboard"; db=sheet(DB,ORANGE)
for c in "ABCDEFGHIJ": db.column_dimensions[c].width=15
db.cell(1,1,"MONTE CARLO COST MODEL").font=Fn(bold=True,size=20,color=ORANGE)
db.cell(2,1,f"='{S}'!C{r0}").font=Fn(bold=True,size=12)
db.cell(3,1,f'="All figures in "&\'{S}\'!C{r0+1}').font=SUB
meta=[("Currency",f"='{S}'!C{r0+1}"),("Years",f"='{S}'!C{r0+4}"),("Iterations",f"='{S}'!C{r0+5}"),
      ("Confidence",f"='{S}'!C{r0+6}"),("Discount",f"='{S}'!C{r0+7}"),("Checks",f"={ALLOK}")]
for i,(lbl,fm) in enumerate(meta):
    db.cell(4,1+i,lbl).font=Fn(size=9,color="888888"); db.cell(4,1+i).alignment=Cn
    x=db.cell(5,1+i,fm); x.font=Fn(bold=True,size=11); x.alignment=Cn; x.fill=fill(GREY); x.border=BOX
# buttons
def button(rng,label,color):
    f=rng.split(":")[0]; b=db[f]; b.value=label; b.font=Fn(bold=True,size=11,color=WHITE); b.fill=fill(color); b.alignment=Cn; b.border=BOX; db.merge_cells(rng)
button("A7:C7","▶  Run Simulation",TEAL); button("D7:E7","📅  Apply Years",PURPLE); button("F7:G7","⬇  Export Report",ORANGE)
db.cell(8,1,"Buttons need the VBA modules (vba/). Save as .xlsm. See vba/VBA_SETUP.md.").font=SUB; db.merge_cells("A8:J8")
# KPI cells (VBA writes the value row beneath each label)
db.cell(10,1,"KEY OUTPUTS  (filled by the macro)").font=Fn(bold=True,size=12,color=ORANGE)
kpis=[("Base Cost (ML)",NAVY),("Expected Cost",ORANGE),("Expected Risk",RED),("Total Expected",ORANGE),
      ("P50",NAVY),("P70",NAVY),("P80",TEAL),("P90",NAVY),
      ("Contingency",TEAL),("Recommended Budget",TEAL),("NPV @ confidence",NAVY),("P95",RED)]
KPI_CELLS=[]
for i,(lbl,color) in enumerate(kpis):
    row=11+(i//4)*3; col=1+(i%4)*2
    db.cell(row,col,lbl).font=Fn(size=9,color="888888"); db.cell(row,col).fill=fill(GREY)
    v=db.cell(row+1,col,None); v.number_format=MONEY; v.font=Fn(bold=True,size=14,color=color); v.fill=fill(GREY)
    db.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+1)
    db.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+1)
    KPI_CELLS.append((row+1,col))
# charts (reference ChartData / Sensitivity that VBA fills)
hist=BarChart(); hist.type="col"; hist.title="Total Cost Distribution"; hist.legend=None; hist.gapWidth=6
hist.add_data(Reference(cd,min_col=2,min_row=1,max_row=1+NB),titles_from_data=True)
hist.set_categories(Reference(cd,min_col=1,min_row=2,max_row=1+NB)); hist.height=7.5; hist.width=13
hist.x_axis.delete=False; hist.y_axis.delete=False; db.add_chart(hist,"A20")
cur=ScatterChart(); cur.title="Cumulative Probability"; cur.legend=None; cur.x_axis.title="Cost"; cur.y_axis.title="Prob"
s=Series(Reference(cd,min_col=5,min_row=2,max_row=1+NP),Reference(cd,min_col=4,min_row=2,max_row=1+NP),title="S-curve"); s.smooth=True
cur.series.append(s); cur.height=7.5; cur.width=13; cur.x_axis.delete=False; cur.y_axis.delete=False; db.add_chart(cur,"F20")
scf=LineChart(); scf.title="Cumulative Cash Flow by Year"; scf.style=12
scf.add_data(Reference(cd,min_col=8,min_row=1,max_col=9,max_row=1+MAXY),titles_from_data=True)
scf.set_categories(Reference(cd,min_col=7,min_row=2,max_row=1+MAXY)); scf.height=7.5; scf.width=13
scf.x_axis.delete=False; scf.y_axis.delete=False; db.add_chart(scf,"A36")
tor2=BarChart(); tor2.type="bar"; tor2.title="Sensitivity (Tornado)"; tor2.legend=None
tor2.add_data(Reference(sn,min_col=5,min_row=4,max_row=SN_MAX-1),titles_from_data=True)
tor2.set_categories(Reference(sn,min_col=1,min_row=SN_F,max_row=SN_MAX-1)); tor2.height=7.5; tor2.width=13
tor2.x_axis.delete=False; tor2.y_axis.delete=False; db.add_chart(tor2,"F36")

# tidy: hide year columns/rows beyond the active default (ApplyYears macro toggles these)
for y in range(ACTIVE_Y, MAXY):
    cp.column_dimensions[L(3+y)].hidden=True
    rp.column_dimensions[L(3+y)].hidden=True
    inf.row_dimensions[F0+y].hidden=True

# order + save
if "Sheet" in wb.sheetnames: del wb["Sheet"]
order=[DB,S,CL,CP,RR,RP,IN,RS,SN,CK,CD]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.active=0
wb.save(OUTFILE)
print(f"Saved {OUTFILE}  ({K} cost lines, {J} risks, up to {MAXY} years, default {ITERS_DEF:,} iters)")
print("Import vba/modEngine.bas (+ others), save as .xlsm, run RunSimulation.")
