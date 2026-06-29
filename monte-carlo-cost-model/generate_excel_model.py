#!/usr/bin/env python3
"""
Generate a self-contained Monte Carlo Cost Model workbook (.xlsx).

The simulation runs INSIDE Excel using volatile RAND()-based formulas, so the
whole model re-rolls every time you press F9 / edit a cell. No macros, no
add-ins. Edit the input table (yellow cells) and everything downstream —
summary stats, percentiles, histogram and S-curve — recalculates live.

Supported per-item distributions: fixed, uniform, triangular, pert, normal,
lognormal. Each draw uses the inverse-transform method on a single stored
uniform (RAND), so the sampling is statistically correct (no double-RAND bug)
and works in Excel 2010+ (modern statistical functions are written with the
required _xlfn. prefix).

Run:  python3 generate_excel_model.py [N_iterations]
Out:  MonteCarloCostModel.xlsx
"""

import sys
import math
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Model definition  — edit this block, then re-run, to change the model.
# Each item: (name, distribution, p1, p2, p3)
#   fixed       -> p1 = value
#   uniform     -> p1 = min,  p2 = max
#   triangular  -> p1 = min,  p2 = most likely, p3 = max
#   pert        -> p1 = min,  p2 = most likely, p3 = max
#   normal      -> p1 = mean, p2 = std dev
#   lognormal   -> p1 = mean, p2 = std dev
# ----------------------------------------------------------------------------
ITEMS = [
    ("Labor",          "pert",       40000, 55000, 90000),
    ("Materials",      "triangular", 30000, 38000, 60000),
    ("Equipment",      "normal",     15000,  3000,  None),
    ("Subcontractors", "uniform",    20000, 35000,  None),
    ("Permits & Fees", "fixed",       5000,  None,  None),
    ("Risk Reserve",   "lognormal",   8000,  4000,  None),
]
BUDGET     = 185000      # set to None for no budget line
CONFIDENCE = 0.80        # recommended-budget confidence level
CURRENCY   = "$"
N          = int(sys.argv[1]) if len(sys.argv) > 1 else 5000   # simulation rows
OUTFILE    = "MonteCarloCostModel.xlsx"

MONEY_FMT = f'"{CURRENCY}"#,##0'
PCT_FMT   = '0.0%'

# ----------------------------------------------------------------------------
# Styling helpers
# ----------------------------------------------------------------------------
NAVY   = "1F2A44"
ORANGE = "FF6B35"
TEAL   = "00A88A"
YELLOW = "FFF4D6"
GREY   = "EFEFF2"
WHITE  = "FFFFFF"

def fill(hex_):  return PatternFill("solid", fgColor=hex_)
def bold(sz=11, color="1C1C1E"): return Font(name="Calibri", bold=True, size=sz, color=color)
def reg(sz=11, color="1C1C1E"):  return Font(name="Calibri", size=sz, color=color)

thin = Side(style="thin", color="D0D0D5")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

# ----------------------------------------------------------------------------
# Layout constants
# ----------------------------------------------------------------------------
HDR_ROW   = 4                      # input table header row
FIRST_IT  = HDR_ROW + 1            # first item row (5)
LAST_IT   = FIRST_IT + len(ITEMS) - 1

DRAW_C0   = 13                     # first draw column (M)
TOTAL_C   = DRAW_C0 + len(ITEMS)   # Total column
HELP_C0   = 27                     # first helper RAND column (AA)
SIM_HDR   = 4
SIM_FIRST = 5
SIM_LAST  = SIM_FIRST + N - 1

def col(i):  return get_column_letter(i)
TOTAL_RANGE = f"${col(TOTAL_C)}${SIM_FIRST}:${col(TOTAL_C)}${SIM_LAST}"

# Settings cells (right of the input table)
SET_COL_L, SET_COL_V = "H", "I"
ITER_CELL  = f"${SET_COL_V}$5"
BUD_CELL   = f"${SET_COL_V}$6"
CONF_CELL  = f"${SET_COL_V}$7"

# ----------------------------------------------------------------------------
# Formula builders (inverse-transform sampling on a stored uniform `u`)
# ----------------------------------------------------------------------------
def central_formula(dist, r):
    if dist == "uniform":    return f"=(C{r}+D{r})/2"
    if dist == "triangular": return f"=(C{r}+D{r}+E{r})/3"
    if dist == "pert":       return f"=(C{r}+4*D{r}+E{r})/6"
    return f"=C{r}"          # fixed, normal, lognormal

def draw_formula(dist, r, u):
    """r = input row of this item; u = helper cell holding RAND() for this row."""
    if dist == "fixed":
        return f"=$C${r}"
    if dist == "uniform":
        return f"=$C${r}+($D${r}-$C${r})*{u}"
    if dist == "triangular":
        fc = f"($D${r}-$C${r})/($E${r}-$C${r})"
        lo = f"$C${r}+SQRT({u}*($E${r}-$C${r})*($D${r}-$C${r}))"
        hi = f"$E${r}-SQRT((1-{u})*($E${r}-$C${r})*($E${r}-$D${r}))"
        return f"=IF({u}<{fc},{lo},{hi})"
    if dist == "pert":
        a = f"1+4*($D${r}-$C${r})/($E${r}-$C${r})"
        b = f"1+4*($E${r}-$D${r})/($E${r}-$C${r})"
        return f"=_xlfn.BETA.INV({u},{a},{b},$C${r},$E${r})"
    if dist == "normal":
        return f"=_xlfn.NORM.INV({u},$C${r},$D${r})"
    if dist == "lognormal":
        v  = f"LN(1+($D${r}/$C${r})^2)"
        mu = f"LN($C${r})-0.5*{v}"
        sg = f"SQRT({v})"
        return f"=_xlfn.LOGNORM.INV({u},{mu},{sg})"
    raise ValueError(dist)

# ----------------------------------------------------------------------------
# Build workbook
# ----------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Monte Carlo"

# --- Title -------------------------------------------------------------------
ws["A1"] = "MONTE CARLO COST MODEL"
ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=ORANGE)
ws["A2"] = (f"{N:,} iterations run live in-sheet. Edit the yellow input cells, "
            f"then press F9 to re-roll. Stats below update automatically.")
ws["A2"].font = reg(10, "6C6C70")

# --- Input table -------------------------------------------------------------
headers = ["Item", "Distribution", "P1", "P2", "P3", "Central"]
for j, h in enumerate(headers):
    c = ws.cell(row=HDR_ROW, column=1 + j, value=h)
    c.font = bold(11, WHITE); c.fill = fill(NAVY); c.alignment = CENTER; c.border = BORDER

param_hint = {
    "fixed": ("value", "", ""), "uniform": ("min", "max", ""),
    "triangular": ("min", "likely", "max"), "pert": ("min", "likely", "max"),
    "normal": ("mean", "std", ""), "lognormal": ("mean", "std", ""),
}
for i, (name, dist, p1, p2, p3) in enumerate(ITEMS):
    r = FIRST_IT + i
    ws.cell(row=r, column=1, value=name).font = reg()
    ws.cell(row=r, column=2, value=dist).font = reg()
    for cc, val in ((3, p1), (4, p2), (5, p3)):
        cell = ws.cell(row=r, column=cc, value=val)
        cell.fill = fill(YELLOW)              # editable inputs
        cell.number_format = MONEY_FMT
        cell.alignment = RIGHT; cell.border = BORDER
        cell.font = reg()
    fcell = ws.cell(row=r, column=6, value=central_formula(dist, r))
    fcell.number_format = MONEY_FMT; fcell.alignment = RIGHT; fcell.font = reg(11, "6C6C70")
    for cc in range(1, 7):
        ws.cell(row=r, column=cc).border = BORDER

base_row = LAST_IT + 1
ws.cell(row=base_row, column=1, value="Base estimate").font = bold()
bc = ws.cell(row=base_row, column=6, value=f"=SUM(F{FIRST_IT}:F{LAST_IT})")
bc.number_format = MONEY_FMT; bc.font = bold(); bc.alignment = RIGHT

# --- Settings box ------------------------------------------------------------
ws[f"{SET_COL_L}4"] = "SETTINGS"
ws[f"{SET_COL_L}4"].font = bold(11, WHITE); ws[f"{SET_COL_L}4"].fill = fill(NAVY)
ws[f"{SET_COL_V}4"].fill = fill(NAVY)
settings = [("Iterations", N, "0"), ("Budget", BUDGET, MONEY_FMT),
            ("Confidence", CONFIDENCE, "0%")]
for k, (lbl, val, fmt) in enumerate(settings):
    rr = 5 + k
    ws[f"{SET_COL_L}{rr}"] = lbl; ws[f"{SET_COL_L}{rr}"].font = reg()
    vc = ws[f"{SET_COL_V}{rr}"]; vc.value = val; vc.number_format = fmt
    vc.fill = fill(YELLOW); vc.alignment = RIGHT; vc.font = reg(); vc.border = BORDER

# --- Summary block -----------------------------------------------------------
sum_row = base_row + 3
ws.cell(row=sum_row, column=1, value="SUMMARY  (live — press F9 to re-roll)").font = bold(12, ORANGE)
TR = TOTAL_RANGE
summary = [
    ("Mean (expected)",          f"=AVERAGE({TR})",                     MONEY_FMT),
    ("Std deviation",            f"=_xlfn.STDEV.S({TR})",               MONEY_FMT),
    ("Minimum",                  f"=MIN({TR})",                          MONEY_FMT),
    ("Maximum",                  f"=MAX({TR})",                          MONEY_FMT),
    ("P10  (optimistic)",        f"=_xlfn.PERCENTILE.INC({TR},0.1)",    MONEY_FMT),
    ("P50  (median)",            f"=_xlfn.PERCENTILE.INC({TR},0.5)",    MONEY_FMT),
    ("P80  (conservative)",      f"=_xlfn.PERCENTILE.INC({TR},0.8)",    MONEY_FMT),
    ("P90",                      f"=_xlfn.PERCENTILE.INC({TR},0.9)",    MONEY_FMT),
    ("P95",                      f"=_xlfn.PERCENTILE.INC({TR},0.95)",   MONEY_FMT),
    (f"Recommended @ {int(CONFIDENCE*100)}%",
                                 f"=_xlfn.PERCENTILE.INC({TR},{CONF_CELL})", MONEY_FMT),
    ("Contingency vs base",      f"=_xlfn.PERCENTILE.INC({TR},{CONF_CELL})-F{base_row}", MONEY_FMT),
    (f"P(cost <= budget)",       f"=COUNTIF({TR},\"<=\"&{BUD_CELL})/{N}", PCT_FMT),
]
for k, (lbl, formula, fmt) in enumerate(summary):
    rr = sum_row + 1 + k
    lc = ws.cell(row=rr, column=1, value=lbl); lc.font = reg(); lc.border = BORDER
    vc = ws.cell(row=rr, column=2, value=formula)
    vc.number_format = fmt; vc.alignment = RIGHT; vc.font = bold(); vc.border = BORDER
    # highlight the two headline rows
    if lbl.startswith("Mean") or lbl.startswith("Recommended"):
        vc.font = bold(11, ORANGE)
    if lbl.startswith("P(cost"):
        vc.font = bold(11, TEAL)

# --- Simulation block (the engine) ------------------------------------------
sim_note = ws.cell(row=2, column=DRAW_C0,
                   value="↓ SIMULATION ENGINE — one row per iteration (safe to ignore / collapse)")
sim_note.font = reg(10, "9A9AA0")
# header
for i, (name, *_ ) in enumerate(ITEMS):
    h = ws.cell(row=SIM_HDR, column=DRAW_C0 + i, value=name)
    h.font = bold(10, WHITE); h.fill = fill(NAVY); h.alignment = CENTER
th = ws.cell(row=SIM_HDR, column=TOTAL_C, value="Total")
th.font = bold(10, WHITE); th.fill = fill(TEAL); th.alignment = CENTER

# rows: helper RAND + per-item inverse-transform draw + row total
for row in range(SIM_FIRST, SIM_LAST + 1):
    for i, (name, dist, *_ ) in enumerate(ITEMS):
        in_row = FIRST_IT + i
        u_cell = f"${col(HELP_C0 + i)}{row}"
        ws.cell(row=row, column=HELP_C0 + i, value="=RAND()")
        dc = ws.cell(row=row, column=DRAW_C0 + i, value=draw_formula(dist, in_row, u_cell))
        dc.number_format = "#,##0"
    tc = ws.cell(row=row, column=TOTAL_C,
                 value=f"=SUM({col(DRAW_C0)}{row}:{col(TOTAL_C-1)}{row})")
    tc.number_format = "#,##0"

# hide helper columns + collapse the engine visually
for i in range(len(ITEMS)):
    ws.column_dimensions[col(HELP_C0 + i)].hidden = True

# ----------------------------------------------------------------------------
# Charts sheet (histogram + S-curve)
# ----------------------------------------------------------------------------
cs = wb.create_sheet("Charts")
cs["A1"] = "DISTRIBUTION OF TOTAL COST"; cs["A1"].font = bold(14, ORANGE)

NB = 30
cs["A3"] = "Bin #"; cs["B3"] = "Low"; cs["C3"] = "High"; cs["D3"] = "Center"; cs["E3"] = "Count"
for c in "ABCDE":
    cs[f"{c}3"].font = bold(10, WHITE); cs[f"{c}3"].fill = fill(NAVY); cs[f"{c}3"].alignment = CENTER

mn = f"MIN('Monte Carlo'!{TR})"
mx = f"MAX('Monte Carlo'!{TR})"
for b in range(NB):
    r = 4 + b
    cs.cell(row=r, column=1, value=b + 1)
    cs.cell(row=r, column=2, value=f"={mn}+({b})*({mx}-{mn})/{NB}")       # low edge
    cs.cell(row=r, column=3, value=f"={mn}+({b+1})*({mx}-{mn})/{NB}")     # high edge
    cs.cell(row=r, column=4, value=f"=(B{r}+C{r})/2")                     # center
    # count in [low, high) via two legacy COUNTIFs (no array / no _xlfn needed)
    cnt = (f"=COUNTIF('Monte Carlo'!{TR},\"<\"&C{r})"
           f"-COUNTIF('Monte Carlo'!{TR},\"<\"&B{r})")
    cs.cell(row=r, column=5, value=cnt)
    for cc in range(1, 6):
        cs.cell(row=r, column=cc).number_format = "#,##0"

hist = BarChart()
hist.type = "col"; hist.title = "Total Cost Distribution"; hist.legend = None
hist.gapWidth = 8
data = Reference(cs, min_col=5, min_row=3, max_row=3 + NB)         # Count (with header)
cats = Reference(cs, min_col=4, min_row=4, max_row=3 + NB)         # Center
hist.add_data(data, titles_from_data=True)
hist.set_categories(cats)
hist.x_axis.title = "Total cost"; hist.y_axis.title = "Frequency"
hist.x_axis.delete = False; hist.y_axis.delete = False
hist.height = 9; hist.width = 18
cs.add_chart(hist, "G3")

# S-curve (cumulative probability)
cs["A40"] = "CUMULATIVE PROBABILITY (S-CURVE)"; cs["A40"].font = bold(14, TEAL)
cs["A42"] = "Cumulative %"; cs["B42"] = "Total cost"
for c in "AB":
    cs[f"{c}42"].font = bold(10, WHITE); cs[f"{c}42"].fill = fill(NAVY); cs[f"{c}42"].alignment = CENTER
NPTS = 21
for k in range(NPTS):
    r = 43 + k
    p = k / (NPTS - 1)
    cs.cell(row=r, column=1, value=p).number_format = "0%"
    cs.cell(row=r, column=2,
            value=f"=_xlfn.PERCENTILE.INC('Monte Carlo'!{TR},A{r})").number_format = "#,##0"

scurve = ScatterChart()
scurve.title = "S-Curve — probability cost is at or below"
scurve.x_axis.title = "Total cost"; scurve.y_axis.title = "Cumulative probability"
scurve.x_axis.delete = False; scurve.y_axis.delete = False
scurve.legend = None
xref = Reference(cs, min_col=2, min_row=43, max_row=42 + NPTS)     # cost
yref = Reference(cs, min_col=1, min_row=43, max_row=42 + NPTS)     # cumulative %
ser = Series(yref, xref, title="S-curve")
ser.smooth = True
scurve.series.append(ser)
scurve.height = 9; scurve.width = 18
cs.add_chart(scurve, "G40")

# ----------------------------------------------------------------------------
# How-to sheet
# ----------------------------------------------------------------------------
hw = wb.create_sheet("How to Use")
lines = [
    ("MONTE CARLO COST MODEL — HOW IT WORKS", bold(14, ORANGE)),
    ("", reg()),
    ("1. Go to the 'Monte Carlo' sheet.", bold()),
    ("   Edit the YELLOW cells in the input table (P1/P2/P3) and the Budget/Confidence.", reg()),
    ("   Each item samples from its distribution using the parameters you enter:", reg()),
    ("      fixed       P1 = value", reg()),
    ("      uniform     P1 = min,  P2 = max", reg()),
    ("      triangular  P1 = min,  P2 = most likely,  P3 = max", reg()),
    ("      pert        P1 = min,  P2 = most likely,  P3 = max", reg()),
    ("      normal      P1 = mean, P2 = std dev", reg()),
    ("      lognormal   P1 = mean, P2 = std dev", reg()),
    ("", reg()),
    ("2. Press F9 to re-roll all iterations. The SUMMARY and the charts update live.", bold()),
    ("   (The model is volatile: it also re-rolls on any edit.)", reg()),
    ("", reg()),
    ("3. Read the results:", bold()),
    ("   Mean       = expected total cost.", reg()),
    ("   P80        = there is an 80% chance the real cost is at or below this.", reg()),
    ("   Recommended @ confidence = the budget you should set to hit that confidence.", reg()),
    ("   Contingency = recommended budget minus the deterministic base estimate.", reg()),
    ("   P(cost <= budget) = probability you finish within the budget you entered.", reg()),
    ("", reg()),
    ("4. To add or remove cost items, re-run generate_excel_model.py after editing", reg()),
    ("   the ITEMS list at the top of the script (formulas are wired per item).", reg()),
    ("", reg()),
    ("Note: uses NORM.INV / LOGNORM.INV / BETA.INV / PERCENTILE.INC / STDEV.S", reg(10, "6C6C70")),
    ("      → requires Excel 2010 or newer (also works in LibreOffice Calc).", reg(10, "6C6C70")),
]
for k, (txt, f) in enumerate(lines):
    c = hw.cell(row=1 + k, column=1, value=txt); c.font = f
hw.column_dimensions["A"].width = 92

# ----------------------------------------------------------------------------
# Cosmetics
# ----------------------------------------------------------------------------
widths = {"A": 18, "B": 13, "C": 11, "D": 11, "E": 11, "F": 13, "G": 3,
          "H": 13, "I": 12}
for c, w in widths.items():
    ws.column_dimensions[c].width = w
ws.freeze_panes = "A3"
for c in "ABCDE":
    cs.column_dimensions[c].width = 13
ws.sheet_view.showGridLines = False
cs.sheet_view.showGridLines = False

wb.save(OUTFILE)

# ----------------------------------------------------------------------------
# Independent pure-Python reference run (sanity check, printed to console)
# ----------------------------------------------------------------------------
def sample(dist, p1, p2, p3, rng):
    if dist == "fixed":      return p1
    if dist == "uniform":    return p1 + (p2 - p1) * rng.random()
    if dist == "normal":     return rng.gauss(p1, p2)
    if dist == "lognormal":
        v = math.log(1 + (p2 / p1) ** 2); mu = math.log(p1) - v / 2
        return math.exp(rng.gauss(mu, math.sqrt(v)))
    if dist == "triangular": return rng.triangular(p1, p3, p2)
    if dist == "pert":
        a = 1 + 4 * (p2 - p1) / (p3 - p1); b = 1 + 4 * (p3 - p2) / (p3 - p1)
        return p1 + rng.betavariate(a, b) * (p3 - p1)

rng = random.Random(12345)
tot = []
for _ in range(20000):
    tot.append(sum(sample(d, p1, p2, p3, rng) for _, d, p1, p2, p3 in ITEMS))
tot.sort()
def pc(p): return tot[int(p * (len(tot) - 1))]
mean = sum(tot) / len(tot)
under = sum(1 for t in tot if t <= BUDGET) / len(tot) if BUDGET else 0

print(f"Saved {OUTFILE}  ({N:,} in-sheet iterations, {len(ITEMS)} items)")
print("Reference values (independent 20k Python run — Excel will be close):")
print(f"  Mean ~ {CURRENCY}{mean:,.0f}")
print(f"  P50  ~ {CURRENCY}{pc(0.50):,.0f}")
print(f"  P80  ~ {CURRENCY}{pc(0.80):,.0f}")
print(f"  P95  ~ {CURRENCY}{pc(0.95):,.0f}")
if BUDGET:
    print(f"  P(cost <= {CURRENCY}{BUDGET:,}) ~ {under*100:.1f}%")
