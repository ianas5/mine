"""Load and validate the PCCM driver contract.

Owns the Cost Line and Risk Register schemas. Fails loudly; never repairs.

Excel address bounds are checked with the SAME central validators the input
contract uses (``check_cell`` / ``check_row`` / ``check_column``), not a second,
weaker implementation.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml
from openpyxl.utils import get_column_letter

from .contract_loader import (
    EXCEL_MAX_COLUMN,
    ContractError,
    InputContract,
    check_column,
    check_row,
)

TABLE_NAME_RE = re.compile(r"^tbl[A-Z][A-Za-z0-9]*$")
KEY_RE = re.compile(r"^[a-z][a-z0-9_]*$")
VALID_TYPES = ("text", "integer", "decimal", "percentage")
VALID_VALIDATION_KINDS = ("list", "whole", "decimal")


class DriverContractError(ContractError):
    """Raised when the driver contract is invalid.

    Subclasses ContractError so a caller may catch either, and so the build entry
    point reports every specification failure the same way.
    """


def _checked(fn, *args):
    """Run a central bound validator, reporting failures as a driver-contract fault.

    The Excel-bound machinery is deliberately shared with the input contract rather
    than duplicated; only the reported exception type is specialised, so the user is
    told which specification file is at fault.
    """
    try:
        return fn(*args)
    except DriverContractError:
        raise
    except ContractError as error:
        raise DriverContractError(str(error)) from error


@dataclass(frozen=True)
class DriverColumn:
    key: str
    header: str
    type: str
    editable: bool
    required: bool
    number_format: str
    width: float
    validation: dict[str, Any] | None
    note: str | None


@dataclass(frozen=True)
class ConditionalFormat:
    """Presentation-only rule. Applies no constraint to what may be entered."""

    purpose: str
    target_column: str
    when_column: str
    equals: str


@dataclass(frozen=True)
class DriverRegister:
    key: str
    sheet: str
    table_name: str
    header_row: int
    first_column: str
    reserved_rows: int
    columns: list[DriverColumn]
    section: str
    section_row: int
    note: str | None
    note_row: int | None
    intro: str | None
    intro_row: int | None
    conditional_formatting: list[ConditionalFormat] = field(default_factory=list)

    @property
    def headers(self) -> list[str]:
        return [c.header for c in self.columns]

    @property
    def first_data_row(self) -> int:
        return self.header_row + 1

    @property
    def last_data_row(self) -> int:
        return self.header_row + self.reserved_rows

    @property
    def first_col_index(self) -> int:
        return _column_index(self.first_column)

    @property
    def last_column(self) -> str:
        return get_column_letter(self.first_col_index + len(self.columns) - 1)

    @property
    def ref(self) -> str:
        return f"{self.first_column}{self.header_row}:{self.last_column}{self.last_data_row}"

    def column_letter(self, index: int) -> str:
        return get_column_letter(self.first_col_index + index)

    def column_index_of(self, key: str) -> int:
        for index, column in enumerate(self.columns):
            if column.key == key:
                return index
        raise DriverContractError(f"{self.table_name}: no column with key {key!r}")

    def letter_of(self, key: str) -> str:
        return self.column_letter(self.column_index_of(key))

    def data_range(self, index: int) -> str:
        letter = self.column_letter(index)
        return f"{letter}{self.first_data_row}:{letter}{self.last_data_row}"


@dataclass(frozen=True)
class DriverContract:
    version: str
    registers: dict[str, DriverRegister]
    locked_schema: dict[str, list[str]]
    forbidden_headers: list[str]
    source_path: Path

    @property
    def all_registers(self) -> list[DriverRegister]:
        return list(self.registers.values())

    @property
    def sheets(self) -> set[str]:
        return {r.sheet for r in self.all_registers}

    def register_for_sheet(self, sheet: str) -> DriverRegister | None:
        for register in self.all_registers:
            if register.sheet == sheet:
                return register
        return None


# ---------------------------------------------------------------------------
def load_driver_contract(path: str | Path) -> DriverContract:
    path = Path(path)
    if not path.is_file():
        raise DriverContractError(f"driver contract not found: {path}")

    with path.open("r", encoding="utf-8") as handle:
        raw = yaml.safe_load(handle)
    if not isinstance(raw, dict):
        raise DriverContractError(f"{path}: contract root must be a mapping")

    where = str(path)
    version = _req_str(raw, "driver_contract_version", where)
    raw_registers = _req(raw, "registers", where)
    locked_schema = _req(raw, "locked_schema", where)
    forbidden = _req(raw, "forbidden_headers", where)

    if not isinstance(raw_registers, dict) or not raw_registers:
        raise DriverContractError(f"{where}: 'registers' must be a non-empty mapping")
    if not isinstance(forbidden, list):
        raise DriverContractError(f"{where}: 'forbidden_headers' must be a list")

    registers = {
        key: _parse_register(key, entry, path) for key, entry in raw_registers.items()
    }

    contract = DriverContract(
        version=version,
        registers=registers,
        locked_schema=locked_schema,
        forbidden_headers=[str(h) for h in forbidden],
        source_path=path,
    )

    _validate_unique(contract, path)
    _validate_locked_schema(contract, path)
    _validate_forbidden_headers(contract, path)
    _validate_ownership(contract, path)
    _validate_excel_bounds(contract, path)
    _validate_no_overlap(contract, path)
    _validate_conditional_formatting(contract, path)
    return contract


def _parse_register(key: str, entry: Any, path: Path) -> DriverRegister:
    where = f"{path}: register {key!r}"
    if not isinstance(entry, dict):
        raise DriverContractError(f"{where}: must be a mapping")

    table_name = _req_str(entry, "table_name", where)
    if not TABLE_NAME_RE.match(table_name):
        raise DriverContractError(f"{where}: table_name {table_name!r} must match tbl<PascalCase>")

    reserved = _req(entry, "reserved_rows", where)
    if not isinstance(reserved, int) or isinstance(reserved, bool) or reserved < 1:
        raise DriverContractError(
            f"{where}: reserved_rows must be a positive integer (initial capacity, "
            f"not a business maximum), got {reserved!r}"
        )

    raw_columns = _req(entry, "columns", where)
    if not isinstance(raw_columns, list) or not raw_columns:
        raise DriverContractError(f"{where}: columns must be a non-empty list")
    columns = [_parse_column(c, f"{where}: columns[{i}]") for i, c in enumerate(raw_columns)]

    cf = [
        _parse_conditional_format(c, f"{where}: conditional_formatting[{i}]")
        for i, c in enumerate(entry.get("conditional_formatting") or [])
    ]

    return DriverRegister(
        key=key,
        sheet=_req_str(entry, "sheet", where),
        table_name=table_name,
        header_row=_positive_int(entry, "header_row", where),
        first_column=_checked(check_column, _req_str(entry, "first_column", where), f"{where}: first_column"),
        reserved_rows=reserved,
        columns=columns,
        section=_req_str(entry, "section", where),
        section_row=_positive_int(entry, "section_row", where),
        note=entry.get("note"),
        note_row=entry.get("note_row"),
        intro=entry.get("intro"),
        intro_row=entry.get("intro_row"),
        conditional_formatting=cf,
    )


def _parse_column(entry: Any, where: str) -> DriverColumn:
    if not isinstance(entry, dict):
        raise DriverContractError(f"{where}: must be a mapping")

    key = _req_str(entry, "key", where)
    if not KEY_RE.match(key):
        raise DriverContractError(f"{where}: key {key!r} must be lower_snake_case")

    type_ = _req_str(entry, "type", where)
    if type_ not in VALID_TYPES:
        raise DriverContractError(f"{where}: type {type_!r} must be one of {VALID_TYPES}")

    for flag in ("editable", "required"):
        if not isinstance(entry.get(flag), bool):
            raise DriverContractError(f"{where}: {flag!r} must be a boolean")

    width = entry.get("width")
    if not isinstance(width, (int, float)) or isinstance(width, bool) or width <= 0:
        raise DriverContractError(f"{where}: width must be a positive number")

    validation = entry.get("validation")
    if validation is not None:
        _validate_validation_block(validation, where)

    return DriverColumn(
        key=key,
        header=_req_str(entry, "header", where),
        type=type_,
        editable=bool(entry["editable"]),
        required=bool(entry["required"]),
        number_format=_req_str(entry, "number_format", where),
        width=float(width),
        validation=validation,
        note=entry.get("note"),
    )


def _parse_conditional_format(entry: Any, where: str) -> ConditionalFormat:
    if not isinstance(entry, dict):
        raise DriverContractError(f"{where}: must be a mapping")
    return ConditionalFormat(
        purpose=_req_str(entry, "purpose", where),
        target_column=_req_str(entry, "target_column", where),
        when_column=_req_str(entry, "when_column", where),
        equals=_req_str(entry, "equals", where),
    )


# ---------------------------------------------------------------------------
# validation
# ---------------------------------------------------------------------------
def _validate_unique(contract: DriverContract, path: Path) -> None:
    tables = [r.table_name for r in contract.all_registers]
    duplicates = sorted({t for t in tables if tables.count(t) > 1})
    if duplicates:
        raise DriverContractError(f"{path}: duplicate driver table names: {duplicates}")

    sheets = [r.sheet for r in contract.all_registers]
    duplicates = sorted({s for s in sheets if sheets.count(s) > 1})
    if duplicates:
        raise DriverContractError(f"{path}: more than one register targets sheet(s): {duplicates}")

    for register in contract.all_registers:
        keys = [c.key for c in register.columns]
        duplicates = sorted({k for k in keys if keys.count(k) > 1})
        if duplicates:
            raise DriverContractError(
                f"{path}: {register.table_name} has duplicate column keys: {duplicates}"
            )
        headers = register.headers
        duplicates = sorted({h for h in headers if headers.count(h) > 1})
        if duplicates:
            raise DriverContractError(
                f"{path}: {register.table_name} has duplicate headers: {duplicates}"
            )


def _validate_locked_schema(contract: DriverContract, path: Path) -> None:
    for key, register in contract.registers.items():
        locked = contract.locked_schema.get(key)
        if locked is None:
            raise DriverContractError(f"{path}: locked_schema has no entry for register {key!r}")
        if not isinstance(locked, list):
            raise DriverContractError(f"{path}: locked_schema[{key!r}] must be a list")
        if register.headers != locked:
            missing = [h for h in locked if h not in register.headers]
            extra = [h for h in register.headers if h not in locked]
            detail = ""
            if missing:
                detail += f" missing {missing};"
            if extra:
                detail += f" unexpected {extra};"
            if not detail:
                detail = " column order differs;"
            raise DriverContractError(
                f"{path}: {register.table_name} schema drifted from the architecture lock."
                f"{detail}\n  locked:   {locked}\n  contract: {register.headers}"
            )


def _validate_forbidden_headers(contract: DriverContract, path: Path) -> None:
    forbidden = {h.strip().lower() for h in contract.forbidden_headers}
    for register in contract.all_registers:
        for column in register.columns:
            if column.header.strip().lower() in forbidden:
                raise DriverContractError(
                    f"{path}: {register.table_name} declares forbidden column "
                    f"{column.header!r}. Every entered risk is simulated (no Included "
                    "column), and any total is derived from sampled unit cost x "
                    "deterministic quantity (no user-input total)."
                )


def _validate_ownership(contract: DriverContract, path: Path) -> None:
    """ID columns are model-controlled and carry no user validation."""
    for register in contract.all_registers:
        identity = register.columns[0]
        if not identity.header.endswith("ID"):
            raise DriverContractError(
                f"{path}: {register.table_name} first column is {identity.header!r}; "
                "the identity column must come first"
            )
        if identity.editable:
            raise DriverContractError(
                f"{path}: {register.table_name} identity column {identity.header!r} is "
                "editable. Permanent identifiers are model-controlled and are never "
                "user-owned."
            )
        if identity.validation is not None:
            raise DriverContractError(
                f"{path}: {register.table_name} identity column {identity.header!r} "
                "declares data validation; a model-controlled column must not."
            )
        for column in register.columns[1:]:
            if not column.editable and column.validation is not None:
                raise DriverContractError(
                    f"{path}: {register.table_name} column {column.header!r} is "
                    "model-controlled but declares user data validation"
                )


def _validate_excel_bounds(contract: DriverContract, path: Path) -> None:
    for register in contract.all_registers:
        where = f"{path}: {register.table_name}"
        _checked(check_row, register.header_row, f"{where} header_row")
        _checked(check_row, register.last_data_row, f"{where} last data row (header_row + reserved_rows)")
        _checked(check_row, register.section_row, f"{where} section_row")
        if register.note_row is not None:
            _checked(check_row, register.note_row, f"{where} note_row")
        if register.intro_row is not None:
            _checked(check_row, register.intro_row, f"{where} intro_row")
        last_index = register.first_col_index + len(register.columns) - 1
        if last_index > EXCEL_MAX_COLUMN:
            raise DriverContractError(
                f"{where}: {len(register.columns)} columns starting at "
                f"{register.first_column} reach column index {last_index}, beyond the "
                f"Excel maximum of XFD ({EXCEL_MAX_COLUMN})"
            )


def _validate_no_overlap(contract: DriverContract, path: Path) -> None:
    """No driver table may collide with its own sheet's headings or another table."""
    claimed: dict[str, dict[str, str]] = {}

    def claim(sheet: str, address: str, owner: str) -> None:
        cells = claimed.setdefault(sheet, {})
        if address in cells:
            raise DriverContractError(
                f"{path}: cell {sheet}!{address} claimed by both {cells[address]} and {owner}"
            )
        cells[address] = owner

    for register in contract.all_registers:
        claim(register.sheet, f"B{register.section_row}", f"section of {register.table_name}")
        if register.note_row is not None:
            claim(register.sheet, f"B{register.note_row}", f"note of {register.table_name}")
        if register.intro_row is not None:
            claim(register.sheet, f"B{register.intro_row}", f"intro of {register.table_name}")
        for index in range(len(register.columns)):
            letter = register.column_letter(index)
            for row in range(register.header_row, register.last_data_row + 1):
                claim(register.sheet, f"{letter}{row}", register.table_name)


def _validate_conditional_formatting(contract: DriverContract, path: Path) -> None:
    for register in contract.all_registers:
        keys = {c.key for c in register.columns}
        for rule in register.conditional_formatting:
            for field_name in ("target_column", "when_column"):
                value = getattr(rule, field_name)
                if value not in keys:
                    raise DriverContractError(
                        f"{path}: {register.table_name} conditional formatting "
                        f"{field_name}={value!r} is not a column of this register"
                    )


def validate_against_input_contract(
    drivers: DriverContract, inputs: InputContract, path: Path | None = None
) -> None:
    """Cross-contract check: every driver list source must exist in the input contract."""
    path = path or drivers.source_path
    known = set(inputs.list_defined_names)
    for register in drivers.all_registers:
        for column in register.columns:
            if not column.validation or column.validation.get("kind") != "list":
                continue
            source = column.validation["source"]
            if source not in known:
                raise DriverContractError(
                    f"{path}: {register.table_name} column {column.header!r} validates "
                    f"against {source!r}, which is not a list defined name declared in "
                    f"{inputs.source_path.name}. Known: {sorted(known)}"
                )

    known_sheets = {r.sheet for r in drivers.all_registers}
    overlap = known_sheets & inputs.contract_sheets
    if overlap:
        raise DriverContractError(
            f"{path}: driver register(s) target sheet(s) already owned by the input "
            f"contract: {sorted(overlap)}"
        )


# ---------------------------------------------------------------------------
def _validate_validation_block(validation: Any, where: str) -> None:
    if not isinstance(validation, dict):
        raise DriverContractError(f"{where}: validation must be a mapping")
    kind = _req_str(validation, "kind", where)
    if kind not in VALID_VALIDATION_KINDS:
        raise DriverContractError(
            f"{where}: validation kind {kind!r} must be one of {VALID_VALIDATION_KINDS}"
        )
    if kind == "list":
        _req_str(validation, "source", where)
    else:
        _req_str(validation, "operator", where)
        _req_str(validation, "formula1", where)
        if validation["operator"] == "between":
            _req_str(validation, "formula2", where)


def _req(mapping: Any, key: str, where: str) -> Any:
    if not isinstance(mapping, dict) or key not in mapping:
        raise DriverContractError(f"{where}: missing required key {key!r}")
    return mapping[key]


def _req_str(mapping: Any, key: str, where: str) -> str:
    value = _req(mapping, key, where)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        value = str(value)
    if not isinstance(value, str) or not value.strip():
        raise DriverContractError(f"{where}: {key!r} must be a non-empty string, got {value!r}")
    return value


def _positive_int(mapping: dict[str, Any], key: str, where: str) -> int:
    value = _req(mapping, key, where)
    if not isinstance(value, int) or isinstance(value, bool) or value < 1:
        raise DriverContractError(f"{where}: {key!r} must be a positive integer, got {value!r}")
    return value


def _column_index(letter: str) -> int:
    index = 0
    for char in letter:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index
