"""Create the PCCM workbook skeleton from the manifest.

Responsibilities are deliberately split:
    build_workbook  - workbook and sheet creation, visibility, active sheet
    _apply_presentation - column widths, gridlines, freeze panes
    _populate_sheet - Phase 1 titles, sections and placeholders
    BuildMetadata   - the values stamped into the workbook at build time

No business rule, calculation or later-phase table schema belongs in this module.
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .spec_loader import SheetSpec, WorkbookSpec
from .styling import StyleBook

BUILDER_VERSION = "0.1.0"
DEFAULT_SHEET_TITLE = "Sheet"
TIMESTAMP_ENV_VAR = "PCCM_BUILD_TIMESTAMP"


@dataclass(frozen=True)
class BuildMetadata:
    """Build-time provenance.

    None of these values is a computational input. They record how the artifact
    was produced and are kept distinct from:
      * model_version   - the version of the model design itself
      * builder_version - the version of this build tooling
      * run metadata    - which belongs to a Monte Carlo run, not to a build
    """

    model_version: str
    build_phase: str
    builder_version: str
    build_timestamp: str
    manifest_version: str

    @classmethod
    def create(cls, spec: WorkbookSpec) -> "BuildMetadata":
        return cls(
            model_version=spec.model["model_version"],
            build_phase=spec.model["build_phase"],
            builder_version=BUILDER_VERSION,
            build_timestamp=resolve_build_timestamp(),
            manifest_version=spec.manifest_version,
        )

    def as_rows(self) -> list[tuple[str, str]]:
        return [
            ("PCCM Model Version", self.model_version),
            ("Build Phase", self.build_phase),
            ("Builder Version", self.builder_version),
            ("Build Timestamp (UTC)", self.build_timestamp),
            ("Source Manifest Version", self.manifest_version),
        ]


def resolve_build_timestamp() -> str:
    """UTC build timestamp, overridable for reproducible builds.

    Setting PCCM_BUILD_TIMESTAMP makes two builds of the same source produce
    structurally identical workbooks, which is what the reproducibility test
    relies on.
    """
    override = os.environ.get(TIMESTAMP_ENV_VAR)
    if override:
        return override
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def build_workbook(spec: WorkbookSpec) -> tuple[Workbook, BuildMetadata]:
    """Create the workbook described by *spec*."""
    styles = StyleBook(spec.presentation)
    metadata = BuildMetadata.create(spec)

    workbook = Workbook()
    default_sheet = workbook.active

    for sheet_spec in spec.sheets:
        worksheet = workbook.create_sheet(title=sheet_spec.name)
        _apply_presentation(worksheet, sheet_spec, styles)
        _populate_sheet(worksheet, sheet_spec, styles, metadata)

    # Remove openpyxl's default sheet only after the real sheets exist, so the
    # workbook is never momentarily empty.
    if default_sheet is not None and default_sheet.title == DEFAULT_SHEET_TITLE:
        workbook.remove(default_sheet)
    if DEFAULT_SHEET_TITLE in workbook.sheetnames:
        raise RuntimeError("the default openpyxl sheet was not removed cleanly")

    # Visibility is applied after creation so the active sheet can be set while
    # every sheet is still visible.
    workbook.active = workbook.sheetnames.index(spec.active_sheet)
    for sheet_spec in spec.sheets:
        workbook[sheet_spec.name].sheet_state = sheet_spec.visibility

    _apply_document_properties(workbook, spec, metadata)
    return workbook, metadata


def _apply_document_properties(
    workbook: Workbook, spec: WorkbookSpec, metadata: BuildMetadata
) -> None:
    properties = workbook.properties
    properties.title = spec.model["name"]
    properties.creator = f"{spec.model['short_name']} builder {metadata.builder_version}"
    properties.lastModifiedBy = properties.creator
    properties.category = metadata.build_phase
    properties.description = (
        f"{spec.model['short_name']} model version {metadata.model_version}; "
        f"manifest {metadata.manifest_version}; built {metadata.build_timestamp}."
    )
    # Deterministic document timestamps keep repeated builds comparable.
    fixed = datetime(2000, 1, 1, tzinfo=timezone.utc).replace(tzinfo=None)
    properties.created = fixed
    properties.modified = fixed


def _apply_presentation(
    worksheet: Worksheet, sheet_spec: SheetSpec, styles: StyleBook
) -> None:
    worksheet.sheet_view.showGridLines = sheet_spec.show_gridlines
    for column, width in sheet_spec.column_widths.items():
        worksheet.column_dimensions[column].width = width
    if sheet_spec.freeze_panes:
        worksheet.freeze_panes = sheet_spec.freeze_panes


def _populate_sheet(
    worksheet: Worksheet,
    sheet_spec: SheetSpec,
    styles: StyleBook,
    metadata: BuildMetadata,
) -> None:
    layout = styles.layout
    label_col = layout.label_column
    value_col = layout.value_column

    _write(worksheet, f"{label_col}{layout.title_row}", sheet_spec.title, styles.title)
    worksheet.row_dimensions[layout.title_row].height = styles.row_height("title")

    if sheet_spec.subtitle:
        _write(
            worksheet,
            f"{label_col}{layout.subtitle_row}",
            sheet_spec.subtitle,
            styles.subtitle,
        )
        worksheet.row_dimensions[layout.subtitle_row].height = styles.row_height("subtitle")

    # A thin rule under the header, drawn across the used column span.
    for column in _rule_columns(sheet_spec):
        worksheet[f"{column}{layout.rule_row}"].border = styles.rule
    worksheet.row_dimensions[layout.rule_row].height = styles.row_height("spacer")

    row = layout.body_start_row
    for block in sheet_spec.blocks:
        row = _write_block(worksheet, block, row, styles, metadata, label_col, value_col)
    return None


def _rule_columns(sheet_spec: SheetSpec) -> list[str]:
    if not sheet_spec.column_widths:
        return ["B"]
    return sorted(sheet_spec.column_widths, key=lambda c: (len(c), c))


def _write_block(
    worksheet: Worksheet,
    block: dict[str, Any],
    row: int,
    styles: StyleBook,
    metadata: BuildMetadata,
    label_col: str,
    value_col: str,
) -> int:
    if block["type"] == "note":
        _write(worksheet, f"{label_col}{row}", block["text"], styles.note)
        return row + 2

    _write(worksheet, f"{label_col}{row}", block["title"], styles.section)
    worksheet.row_dimensions[row].height = styles.row_height("section")
    row += 1

    if block.get("note"):
        _write(worksheet, f"{label_col}{row}", block["note"], styles.note)
        row += 1

    for entry in block.get("rows") or []:
        _write(worksheet, f"{label_col}{row}", entry["label"], styles.label)
        if "value" in entry and entry["value"] is not None:
            font = styles.value_locked if entry.get("locked") else styles.value
            _write(worksheet, f"{value_col}{row}", entry["value"], font)
        if entry.get("note"):
            _write(worksheet, f"F{row}", entry["note"], styles.note)
        row += 1

    for item in block.get("list") or []:
        _write(worksheet, f"{value_col}{row}", item, styles.list_item)
        row += 1

    if block.get("metadata"):
        for label, value in metadata.as_rows():
            _write(worksheet, f"{label_col}{row}", label, styles.label)
            _write(worksheet, f"{value_col}{row}", value, styles.value)
            row += 1

    return row + 1


def _write(worksheet: Worksheet, address: str, value: Any, font) -> None:
    cell = worksheet[address]
    cell.value = value
    cell.font = font
