"""Load and validate the PCCM calculation contract.

`spec/calc_contract.yaml` is the narrowly scoped fifth authority: it owns the
PHYSICAL SHAPE of the Phase-5 `_Calc` workspace and the numeric tolerance
constants, and nothing else. This loader is what makes "and nothing else"
enforceable rather than aspirational.

Two classes of assertion live here.

**Layout conformance.** The accepted anchors of `docs/phase5_plan.md` Revision E
section 16.3 are LOCKED CONSTANTS in this module, and the contract is checked
against them. The contract ENCODES the accepted layout; it does not get to choose
it. Every band is checked to be exactly as wide as its schema, no two bands
overlap, and nothing intrudes on the frozen Phase-4 rows.

**Authority conformance.** The hash mathematics is not permitted to appear in the
contract at all - the loader scans for it and fails - and every value the contract
borrows from another specification is declared as a reference whose locator is
resolved against the owning file.

Excel address bounds are checked with the SAME central validators the other
contracts use, not a fifth implementation.

Fails loudly; never repairs.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl.utils import column_index_from_string, get_column_letter

from .contract_loader import ContractError, InputContract, check_cell, check_column, check_row
from .driver_loader import DriverContract
from .spec_loader import WorkbookSpec
from .structure_loader import StructureContract

KEY_RE = re.compile(r"^[a-z][a-z0-9_]*$")
TABLE_NAME_RE = re.compile(r"^tbl[A-Z][A-Za-z0-9]*$")
VALID_VALUE_TYPES = ("text", "integer", "double", "timestamp", "enum")
VALID_DRIVER_KINDS = ("cost_line", "risk")


class CalcContractError(ContractError):
    """Raised when the calculation contract is invalid.

    Subclasses ContractError so the build entry point reports every specification
    failure the same way.
    """


def _checked(fn, *args):
    """Run a central bound validator, reporting failures as a calc-contract fault."""
    try:
        return fn(*args)
    except CalcContractError:
        raise
    except ContractError as error:
        raise CalcContractError(str(error)) from error


# ---------------------------------------------------------------------------
# The accepted Revision-E layout - LOCKED HERE, not in the contract
# ---------------------------------------------------------------------------
LOCKED_SHEET = "_Calc"

LOCKED_PHASE4_FIRST_ROW = 1
LOCKED_PHASE4_LAST_ROW = 11
LOCKED_PHASE4_CELLS = ("C10", "C11")
"""Phase-4 territory on `_Calc`: the title block, the "Permanent ID Counters"
section and the two persistent counters. Frozen by the Phase-4 source freeze."""

LOCKED_CALC_STATE_ROWS = (13, 20)
LOCKED_CALC_TOTALS_ROWS = (23, 32)
"""Fixed-height scalar blocks at disjoint row ranges, so neither can grow into the
other. `calc_state`'s eight rows are what make the success commit a single
contiguous `C13:C20` assignment."""


@dataclass(frozen=True)
class ColumnSchema:
    """One locked ListObject column of the accepted Revision-E design."""

    key: str
    header: str
    value_type: str
    number_format: str
    units: str
    applies_to: tuple[str, ...] | None


@dataclass(frozen=True)
class TableSchema:
    """One locked Phase-5 ListObject: where it sits and exactly what it holds."""

    first_column: str
    last_column: str
    header_row: int
    row_rule: str
    columns: tuple[ColumnSchema, ...]

    @property
    def width(self) -> int:
        return len(self.columns)


@dataclass(frozen=True)
class StateRow:
    """One locked `calc_state` row."""

    row: int
    group: str
    key: str
    label: str
    value_type: str
    number_format: str
    enum: str | None
    initial: Any


@dataclass(frozen=True)
class TotalRow:
    """One locked `calc_totals` row."""

    row: int
    key: str
    label: str
    value_type: str
    number_format: str
    units: str
    measure: str
    basis: str
    initial: Any


LOCKED_TABLES: dict[str, TableSchema] = {
    # -- tblCalcYears --------------------------------------------------------
    "tblCalcYears": TableSchema(
        first_column="H",
        last_column="J",
        header_row=15,
        row_rule="one row per applied project year",
        columns=(
            ColumnSchema("project_index", "Project Index", "integer", "0", "index, from 1", None),
            ColumnSchema("calendar_year", "Calendar Year", "integer", "0", "year", None),
            ColumnSchema(
                "discount_factor", "Discount Factor", "double", "0.000000", "dimensionless", None
            ),
        ),
    ),
    # -- tblCalcInflationFactors ---------------------------------------------
    "tblCalcInflationFactors": TableSchema(
        first_column="M",
        last_column="P",
        header_row=15,
        row_rule=(
            "one row per referenced inflation profile per factor year, spanning "
            "BaseYear .. LastProjectYear inclusive"
        ),
        columns=(
            ColumnSchema("inflation_profile", "Inflation Profile", "text", "@", "key", None),
            ColumnSchema("calendar_year", "Calendar Year", "integer", "0", "year", None),
            ColumnSchema("annual_rate", "Annual Rate", "double", "0.00%", "rate", None),
            ColumnSchema(
                "cumulative_inflation_factor",
                "Cumulative Inflation Factor",
                "double",
                "0.000000",
                "dimensionless",
                None,
            ),
        ),
    ),
    # -- tblCalcFX -----------------------------------------------------------
    "tblCalcFX": TableSchema(
        first_column="S",
        last_column="U",
        header_row=15,
        row_rule="one row per referenced currency",
        columns=(
            ColumnSchema("currency", "Currency", "text", "@", "key", None),
            ColumnSchema("fx_to_sar", "FX to SAR", "double", "0.000000", "SAR per unit", None),
            ColumnSchema("referenced_by", "Referenced By", "integer", "0", "driver count", None),
        ),
    ),
    # -- tblCalcDrivers ------------------------------------------------------
    # 21 columns. No column carries two meanings by Driver Kind: a field that does
    # not apply to a kind is BLANK, never zero and never reused. `applies_to` is
    # part of the locked design, not decoration - widening it would quietly permit
    # a Risk row to carry a Quantity, and narrowing it would break the column-sum
    # reconstruction of A, B, C and D from the audit rows alone.
    "tblCalcDrivers": TableSchema(
        first_column="X",
        last_column="AR",
        header_row=15,
        row_rule="one row per identified Cost Line and Risk",
        columns=(
            ColumnSchema("permanent_id", "Permanent ID", "text", "@", "key", ("cost_line", "risk")),
            ColumnSchema(
                "driver_kind", "Driver Kind", "text", "@", "Cost Line / Risk", ("cost_line", "risk")
            ),
            ColumnSchema("distribution", "Distribution", "text", "@", "name", ("cost_line", "risk")),
            ColumnSchema(
                "central_basis", "Central Basis", "text", "@", "ML / Midpoint", ("cost_line", "risk")
            ),
            ColumnSchema("currency", "Currency", "text", "@", "key", ("cost_line", "risk")),
            ColumnSchema(
                "fx_to_sar", "FX to SAR", "double", "0.000000", "SAR per unit", ("cost_line", "risk")
            ),
            ColumnSchema(
                "inflation_profile", "Inflation Profile", "text", "@", "key", ("cost_line", "risk")
            ),
            ColumnSchema("quantity", "Quantity", "double", "#,##0.00", "units", ("cost_line",)),
            ColumnSchema("probability", "Probability", "double", "0.0%", "fraction", ("risk",)),
            ColumnSchema(
                "central_value", "Central Value", "double", "#,##0.00", "source currency",
                ("cost_line",),
            ),
            ColumnSchema(
                "mean_value", "Mean Value", "double", "#,##0.00", "source currency",
                ("cost_line", "risk"),
            ),
            ColumnSchema(
                "knom", "Knom", "double", "0.000000", "SAR per source unit", ("cost_line", "risk")
            ),
            ColumnSchema(
                "kpv", "Kpv", "double", "0.000000", "SAR per source unit", ("cost_line", "risk")
            ),
            ColumnSchema(
                "deterministic_nominal", "Deterministic Nominal", "double", "#,##0.00", "SAR",
                ("cost_line",),
            ),
            ColumnSchema(
                "deterministic_pv", "Deterministic PV", "double", "#,##0.00", "SAR", ("cost_line",)
            ),
            ColumnSchema(
                "mean_basis_nominal", "Mean-Basis Nominal", "double", "#,##0.00", "SAR",
                ("cost_line",),
            ),
            ColumnSchema(
                "mean_basis_pv", "Mean-Basis PV", "double", "#,##0.00", "SAR", ("cost_line",)
            ),
            ColumnSchema(
                "uncertainty_mean_shift_nominal", "Uncertainty Mean Shift Nominal", "double",
                "#,##0.00", "SAR", ("cost_line",),
            ),
            ColumnSchema(
                "uncertainty_mean_shift_pv", "Uncertainty Mean Shift PV", "double", "#,##0.00",
                "SAR", ("cost_line",),
            ),
            ColumnSchema(
                "expected_risk_nominal", "Expected Risk Nominal", "double", "#,##0.00", "SAR",
                ("risk",),
            ),
            ColumnSchema(
                "expected_risk_pv", "Expected Risk PV", "double", "#,##0.00", "SAR", ("risk",)
            ),
        ),
    ),
    # -- tblCalcAnnual -------------------------------------------------------
    # Calendar Year is present so an annual audit row stands on its own, without a
    # join to tblCalcYears merely to learn which year it describes.
    "tblCalcAnnual": TableSchema(
        first_column="AU",
        last_column="BB",
        header_row=15,
        row_rule="one row per applied project year",
        columns=(
            ColumnSchema("project_index", "Project Index", "integer", "0", "index", None),
            ColumnSchema("calendar_year", "Calendar Year", "integer", "0", "year", None),
            ColumnSchema(
                "base_cost_nominal", "Base Cost Nominal", "double", "#,##0.00", "SAR", None
            ),
            ColumnSchema(
                "expected_risk_nominal", "Expected Risk Nominal", "double", "#,##0.00", "SAR", None
            ),
            ColumnSchema("total_nominal", "Total Nominal", "double", "#,##0.00", "SAR", None),
            ColumnSchema("base_cost_pv", "Base Cost PV", "double", "#,##0.00", "SAR", None),
            ColumnSchema(
                "expected_risk_pv", "Expected Risk PV", "double", "#,##0.00", "SAR", None
            ),
            ColumnSchema("total_pv", "Total PV", "double", "#,##0.00", "SAR", None),
        ),
    ),
}
"""The accepted Revision-E ListObject design, in full.

LOCKED HERE, NOT IN THE CONTRACT. The contract encodes this layout; the loader
holds the independent copy that the contract is checked against. Anchors alone are
not enough: a header, a number format, a unit or an `applies_to` set can be edited
without moving a single column, and every one of those is a design change. The
loader compares the whole schema, attribute by attribute, so any such edit fails
the build instead of quietly redefining what the audit table means.
"""

LOCKED_TABLE_ANCHORS: dict[str, tuple[str, str, int, int]] = {
    name: (schema.first_column, schema.last_column, schema.header_row, schema.width)
    for name, schema in LOCKED_TABLES.items()
}
"""table name -> (first column, last column, header row, column count).

Derived from LOCKED_TABLES so the anchors and the schemas cannot disagree with each
other. Column bands rather than vertical stacking: every table has an unbounded row
count, so stacking them would make a growing table collide with the block below.
Two-column gutters separate the bands so that a widened schema is caught by the
overlap assertion instead of silently overwriting a neighbour."""

LOCKED_FP_VERSION = 1

LOCKED_DERIVED_STATUS = ("NOT CALCULATED", "CURRENT", "STALE", "INVALID")
LOCKED_ATTEMPT_RESULT = ("NONE", "SUCCESS", "REFUSED", "FAILED")
"""The two orthogonal axes. `REFUSED` and `FAILED` are ATTEMPT results and must
never appear on the derived-status axis: a model is INVALID whether or not anyone
pressed Calculate, and a rolled-back write leaves it STALE, derived from the
inputs rather than forced from the attempt."""

LOCKED_CALC_STATE: tuple[StateRow, ...] = (
    StateRow(13, "snapshot", "last_successful_stamp", "Last Successful Stamp",
             "timestamp", "yyyy-mm-dd hh:mm:ss", None, None),
    StateRow(14, "snapshot", "last_successful_fingerprint", "Last Successful Fingerprint",
             "text", "@", None, None),
    StateRow(15, "snapshot", "fingerprint_version", "Fingerprint Version",
             "integer", "0", None, None),
    StateRow(16, "snapshot", "last_successful_applied_timeline",
             "Last Successful Applied Timeline", "text", "@", None, None),
    StateRow(17, "attempt", "last_attempt_result", "Last Attempt Result",
             "enum", "@", "attempt_result", "NONE"),
    StateRow(18, "attempt", "last_attempt_detail", "Last Attempt Detail",
             "text", "@", None, None),
    StateRow(19, "derived", "calculation_status", "Calculation Status (last evaluated)",
             "enum", "@", "derived_status", "NOT CALCULATED"),
    StateRow(20, "derived", "status_evaluated_at", "Status Evaluated At",
             "timestamp", "yyyy-mm-dd hh:mm:ss", None, None),
)
"""Row order is load-bearing, not cosmetic: `C13:C16` is the snapshot, `C17:C18`
the attempt and `C19:C20` the derived reading, and only that grouping makes the
success commit one contiguous `C13:C20` assignment.

`Fingerprint Version` is seeded BLANK. Writing the algorithm version at build time
would make a never-calculated workbook look as though it held a partial successful
snapshot. The two timestamp rows carry a real date format, not `@`: storing them as
text would make the audit trail unsortable and uncomparable."""

LOCKED_CALC_STATE_GROUPS = ("snapshot", "attempt", "derived")

LOCKED_CALC_TOTALS: tuple[TotalRow, ...] = (
    TotalRow(23, "a_nom", "Escalated Deterministic Base - Nominal",
             "double", "#,##0.00", "SAR", "A", "nominal", None),
    TotalRow(24, "a_pv", "Escalated Deterministic Base - PV",
             "double", "#,##0.00", "SAR", "A", "pv", None),
    TotalRow(25, "b_nom", "Uncertainty Mean Shift - Nominal",
             "double", "#,##0.00", "SAR", "B", "nominal", None),
    TotalRow(26, "b_pv", "Uncertainty Mean Shift - PV",
             "double", "#,##0.00", "SAR", "B", "pv", None),
    TotalRow(27, "c_nom", "Mean-Basis Base Cost - Nominal",
             "double", "#,##0.00", "SAR", "C", "nominal", None),
    TotalRow(28, "c_pv", "Mean-Basis Base Cost - PV",
             "double", "#,##0.00", "SAR", "C", "pv", None),
    TotalRow(29, "d_nom", "Expected Risk / EMV - Nominal",
             "double", "#,##0.00", "SAR", "D", "nominal", None),
    TotalRow(30, "d_pv", "Expected Risk / EMV - PV",
             "double", "#,##0.00", "SAR", "D", "pv", None),
    TotalRow(31, "e_nom", "Analytical Mean Total - Nominal",
             "double", "#,##0.00", "SAR", "E", "nominal", None),
    TotalRow(32, "e_pv", "Analytical Mean Total - PV",
             "double", "#,##0.00", "SAR", "E", "pv", None),
)
"""Ten headline values, each measure's Nominal row immediately followed by its PV
row. All SAR, all `#,##0.00`, none seeded: a seeded 0 would be a fabricated
result."""

LOCKED_DRIVER_HEADERS = tuple(c.header for c in LOCKED_TABLES["tblCalcDrivers"].columns)
"""21 columns, in order. Derived from the full locked schema above."""

LOCKED_ANNUAL_HEADERS = tuple(c.header for c in LOCKED_TABLES["tblCalcAnnual"].columns)
"""8 columns, in order. Derived from the full locked schema above."""

LOCKED_TOLERANCES: dict[str, Any] = {
    "profiling_sum_absolute": 1e-9,
    "identity_absolute_floor": 1e-6,
    "identity_relative_coefficient": 1e-12,
    "conditioning_scale_floor": 1.0,
    "fx_rate_strictly_positive": True,
    "growth_factor_strictly_positive": True,
}
"""A TOLERANCE EDIT IS A NUMERICAL-DESIGN CHANGE.

Each of these was argued for in the plan and must not pass merely because it is
still a positive number. Loosening `profiling_sum_absolute` from `1e-9` to `1e-3`
would let a profile that sums to 99.9% through as if it summed to 100%; loosening
`identity_absolute_floor` would let a real bookkeeping mismatch pass as rounding;
raising `conditioning_scale_floor` above 1 would widen every identity tolerance at
once. Exact equality is the only check that catches those."""

LOCKED_CONDITIONING_TERMS: dict[str, tuple[str, ...]] = {
    "i1": ("abs_a", "abs_b", "abs_c"),
    "i2": ("abs_c", "abs_d", "abs_e"),
    "i3a": ("sum_abs_annual_base", "abs_c"),
    "i3b": ("sum_abs_annual_risk", "abs_d"),
    "i3c": ("sum_abs_annual_total", "abs_e"),
    "i4a": ("sum_abs_annual_base", "abs_c"),
    "i4b": ("sum_abs_annual_risk", "abs_d"),
    "i4c": ("sum_abs_annual_total", "abs_e"),
}
"""Which absolute magnitudes each identity's conditioning scale sums.

The terms are the identity's OWN, and swapping them silently rescales the wrong
comparison: giving I1 (`A + B = C`) the terms of I2 would size its tolerance by
quantities that identity never touches. A generic "at least two terms" rule cannot
see that, so the exact sets are locked."""

LOCKED_AUTHORITY_REFERENCES: tuple[tuple[str, str, str], ...] = (
    ("distribution master list", "input_contract.yaml", "config_tables.distributions"),
    ("FX convention", "input_contract.yaml", "conventions.fx_convention"),
    ("Cost Line and Risk Register input schemas", "driver_contract.yaml", "registers"),
    (
        "permanent-ID prefixes, patterns and counter rules",
        "structure_contract.yaml",
        "identity.counters",
    ),
    ("applied timeline and structural limits", "structure_contract.yaml", "timeline"),
    (
        "_Calc sheet declaration, visibility and column widths",
        "workbook.yaml",
        "sheets",
    ),
)
"""(concept, owner, locator) - the COMPLETE required set, exactly.

Checking only that the references present resolve is not enough: a reference can
be deleted, and then the boundary it documented simply stops being declared. These
six are the authority boundaries Step 1 exists to protect, so the set is locked -
none missing, none extra, none duplicated, none renamed. The referenced VALUES
remain owned upstream; only the boundary metadata is locked here."""
# ---------------------------------------------------------------------------
# The authority boundary - hash mathematics may not appear in the contract
# ---------------------------------------------------------------------------
FORBIDDEN_HASH_LITERALS = ("131", "2147483647", "2147483629")
"""FP_BASE, FP_MOD_1 and FP_MOD_2. They belong to exactly one source -
`builder/pccm_builder/calc_fingerprint.py` - and a second hand-maintained copy in
YAML is precisely the drift the contract architecture exists to prevent. Only the
fingerprint VERSION NUMBER is a layout fact and belongs in the contract."""

FORBIDDEN_HASH_KEYS = (
    "fp_base",
    "fp_mod",
    "fp_mod_1",
    "fp_mod_2",
    "fp_init",
    "fp_init_1",
    "fp_init_2",
    "base",
    "modulus",
    "moduli",
    "initial_state",
    "recurrence",
    "polynomial",
    "hash_base",
)

_LITERAL_RE = {
    literal: re.compile(rf"(?<![0-9A-Za-z_.]){literal}(?![0-9A-Za-z_.])")
    for literal in FORBIDDEN_HASH_LITERALS
}


# ---------------------------------------------------------------------------
# value types
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class ScalarField:
    key: str
    row: int
    label: str
    value_type: str
    number_format: str
    initial: Any
    note: str | None
    group: str | None = None
    enum: str | None = None
    measure: str | None = None
    basis: str | None = None
    units: str | None = None


@dataclass(frozen=True)
class ScalarBlock:
    key: str
    label_column: str
    value_column: str
    note_column: str
    first_row: int
    last_row: int
    fields: tuple[ScalarField, ...]

    @property
    def rows(self) -> range:
        return range(self.first_row, self.last_row + 1)

    def field_by_key(self, key: str) -> ScalarField:
        for entry in self.fields:
            if entry.key == key:
                return entry
        raise CalcContractError(f"{self.key}: no field with key {key!r}")

    def value_range(self) -> str:
        return f"{self.value_column}{self.first_row}:{self.value_column}{self.last_row}"


@dataclass(frozen=True)
class TableColumn:
    key: str
    header: str
    value_type: str
    number_format: str
    units: str
    applies_to: tuple[str, ...] | None = None


@dataclass(frozen=True)
class CalcTable:
    key: str
    table_name: str
    header_row: int
    first_column: str
    last_column: str
    row_rule: str
    columns: tuple[TableColumn, ...]

    @property
    def first_column_index(self) -> int:
        return column_index_from_string(self.first_column)

    @property
    def last_column_index(self) -> int:
        return column_index_from_string(self.last_column)

    @property
    def band_width(self) -> int:
        return self.last_column_index - self.first_column_index + 1

    @property
    def headers(self) -> list[str]:
        return [c.header for c in self.columns]

    def column_letter(self, index: int) -> str:
        return get_column_letter(self.first_column_index + index)


@dataclass(frozen=True)
class Tolerances:
    profiling_sum_absolute: float
    identity_absolute_floor: float
    identity_relative_coefficient: float
    conditioning_scale_floor: float
    fx_rate_strictly_positive: bool
    growth_factor_strictly_positive: bool
    conditioning_terms: dict[str, tuple[str, ...]]


@dataclass(frozen=True)
class AuthorityReference:
    concept: str
    owner: str
    locator: str


@dataclass(frozen=True)
class CalcContract:
    version: str
    sheet: str
    required_visibility: str
    phase4_first_row: int
    phase4_last_row: int
    phase4_cells: tuple[str, ...]
    fingerprint_version: int
    derived_status_labels: tuple[str, ...]
    attempt_result_labels: tuple[str, ...]
    scalar_blocks: dict[str, ScalarBlock]
    tables: dict[str, CalcTable]
    tolerances: Tolerances
    authority_references: tuple[AuthorityReference, ...]
    source_path: Path

    @property
    def calc_state(self) -> ScalarBlock:
        return self.scalar_blocks["calc_state"]

    @property
    def calc_totals(self) -> ScalarBlock:
        return self.scalar_blocks["calc_totals"]

    @property
    def all_tables(self) -> list[CalcTable]:
        return list(self.tables.values())

    @property
    def table_names(self) -> list[str]:
        return [t.table_name for t in self.all_tables]

    def table_by_name(self, name: str) -> CalcTable:
        for table in self.all_tables:
            if table.table_name == name:
                return table
        raise CalcContractError(f"no Phase-5 table named {name!r}")

    @property
    def phase4_reserved_rows(self) -> range:
        return range(self.phase4_first_row, self.phase4_last_row + 1)

    def initial_values(self) -> dict[str, Any]:
        """`calc_state` value cell -> the value Stage A seeds it with.

        `None` means the cell is left genuinely BLANK, never `0` and never `""`.
        """
        block = self.calc_state
        return {f"{block.value_column}{f.row}": f.initial for f in block.fields}


# ---------------------------------------------------------------------------
# loading
# ---------------------------------------------------------------------------
def load_calc_contract(path: str | Path) -> CalcContract:
    path = Path(path)
    if not path.is_file():
        raise CalcContractError(f"calculation contract not found: {path}")

    raw_text = path.read_text(encoding="utf-8")
    raw = yaml.safe_load(raw_text)
    if not isinstance(raw, dict):
        raise CalcContractError(f"{path}: contract root must be a mapping")

    where = str(path)
    version = _req_str(raw, "calc_contract_version", where)
    sheet_block = _req(raw, "sheet", where)
    reservation = _req(raw, "phase4_reservation", where)
    fingerprint = _req(raw, "fingerprint", where)
    labels = _req(raw, "state_labels", where)
    raw_blocks = _req(raw, "scalar_blocks", where)
    raw_tables = _req(raw, "tables", where)

    contract = CalcContract(
        version=version,
        sheet=_req_str(sheet_block, "name", f"{where}: sheet"),
        required_visibility=_req_str(sheet_block, "required_visibility", f"{where}: sheet"),
        phase4_first_row=_positive_int(reservation, "first_row", f"{where}: phase4_reservation"),
        phase4_last_row=_positive_int(reservation, "last_row", f"{where}: phase4_reservation"),
        phase4_cells=tuple(
            str(cell) for cell in _req(reservation, "cells", f"{where}: phase4_reservation")
        ),
        fingerprint_version=_positive_int(fingerprint, "version", f"{where}: fingerprint"),
        derived_status_labels=tuple(
            str(v) for v in _req(labels, "derived_status", f"{where}: state_labels")
        ),
        attempt_result_labels=tuple(
            str(v) for v in _req(labels, "attempt_result", f"{where}: state_labels")
        ),
        scalar_blocks={
            key: _parse_scalar_block(key, entry, f"{where}: scalar_blocks.{key}")
            for key, entry in _as_mapping(raw_blocks, f"{where}: scalar_blocks").items()
        },
        tables={
            key: _parse_table(key, entry, f"{where}: tables.{key}")
            for key, entry in _as_mapping(raw_tables, f"{where}: tables").items()
        },
        tolerances=_parse_tolerances(_req(raw, "tolerances", where), where),
        authority_references=_parse_references(_req(raw, "authority_references", where), where),
        source_path=path,
    )

    _validate_authority_boundary(raw, raw_text, path)
    _validate_sheet_and_reservation(contract, path)
    _validate_scalar_blocks(contract, path)
    _validate_tables(contract, path)
    _validate_no_overlap(contract, path)
    _validate_state_labels(contract, path)
    _validate_fingerprint(contract, path)
    _validate_authority_reference_set(contract, path)
    _validate_tolerances(contract, path)
    _validate_excel_bounds(contract, path)
    return contract


# ---------------------------------------------------------------------------
# parsing
# ---------------------------------------------------------------------------
def _parse_scalar_block(key: str, entry: Any, where: str) -> ScalarBlock:
    entry = _as_mapping(entry, where)
    raw_fields = _req(entry, "fields", where)
    if not isinstance(raw_fields, list) or not raw_fields:
        raise CalcContractError(f"{where}: fields must be a non-empty list")
    return ScalarBlock(
        key=key,
        label_column=_checked(
            check_column, _req_str(entry, "label_column", where), f"{where}: label_column"
        ),
        value_column=_checked(
            check_column, _req_str(entry, "value_column", where), f"{where}: value_column"
        ),
        note_column=_checked(
            check_column, _req_str(entry, "note_column", where), f"{where}: note_column"
        ),
        first_row=_positive_int(entry, "first_row", where),
        last_row=_positive_int(entry, "last_row", where),
        fields=tuple(
            _parse_scalar_field(item, f"{where}: fields[{i}]")
            for i, item in enumerate(raw_fields)
        ),
    )


def _parse_scalar_field(entry: Any, where: str) -> ScalarField:
    entry = _as_mapping(entry, where)
    value_type = _req_str(entry, "value_type", where)
    if value_type not in VALID_VALUE_TYPES:
        raise CalcContractError(
            f"{where}: value_type {value_type!r} must be one of {VALID_VALUE_TYPES}"
        )
    if value_type == "enum" and not entry.get("enum"):
        raise CalcContractError(
            f"{where}: a field of value_type 'enum' must name the label set it draws from"
        )
    return ScalarField(
        key=_req_key(entry, "key", where),
        row=_positive_int(entry, "row", where),
        label=_req_str(entry, "label", where),
        value_type=value_type,
        number_format=_req_str(entry, "number_format", where),
        initial=entry.get("initial"),
        note=entry.get("note"),
        group=entry.get("group"),
        enum=entry.get("enum"),
        measure=entry.get("measure"),
        basis=entry.get("basis"),
        units=entry.get("units"),
    )


def _parse_table(key: str, entry: Any, where: str) -> CalcTable:
    entry = _as_mapping(entry, where)
    table_name = _req_str(entry, "table_name", where)
    if not TABLE_NAME_RE.match(table_name):
        raise CalcContractError(
            f"{where}: table_name {table_name!r} must match tbl<PascalCase>"
        )
    raw_columns = _req(entry, "columns", where)
    if not isinstance(raw_columns, list) or not raw_columns:
        raise CalcContractError(f"{where}: columns must be a non-empty list")
    return CalcTable(
        key=key,
        table_name=table_name,
        header_row=_positive_int(entry, "header_row", where),
        first_column=_checked(
            check_column, _req_str(entry, "first_column", where), f"{where}: first_column"
        ),
        last_column=_checked(
            check_column, _req_str(entry, "last_column", where), f"{where}: last_column"
        ),
        row_rule=_req_str(entry, "row_rule", where),
        columns=tuple(
            _parse_table_column(item, f"{where}: columns[{i}]")
            for i, item in enumerate(raw_columns)
        ),
    )


def _parse_table_column(entry: Any, where: str) -> TableColumn:
    entry = _as_mapping(entry, where)
    value_type = _req_str(entry, "value_type", where)
    if value_type not in VALID_VALUE_TYPES:
        raise CalcContractError(
            f"{where}: value_type {value_type!r} must be one of {VALID_VALUE_TYPES}"
        )
    applies = entry.get("applies_to")
    if applies is not None:
        if not isinstance(applies, list) or not applies:
            raise CalcContractError(f"{where}: applies_to must be a non-empty list when present")
        for kind in applies:
            if kind not in VALID_DRIVER_KINDS:
                raise CalcContractError(
                    f"{where}: applies_to entry {kind!r} must be one of {VALID_DRIVER_KINDS}"
                )
        applies = tuple(str(k) for k in applies)
    return TableColumn(
        key=_req_key(entry, "key", where),
        header=_req_str(entry, "header", where),
        value_type=value_type,
        number_format=_req_str(entry, "number_format", where),
        units=_req_str(entry, "units", where),
        applies_to=applies,
    )


def _parse_tolerances(entry: Any, where: str) -> Tolerances:
    where = f"{where}: tolerances"
    entry = _as_mapping(entry, where)
    terms = _as_mapping(_req(entry, "conditioning_terms", where), f"{where}: conditioning_terms")
    return Tolerances(
        profiling_sum_absolute=_positive_float(entry, "profiling_sum_absolute", where),
        identity_absolute_floor=_positive_float(entry, "identity_absolute_floor", where),
        identity_relative_coefficient=_positive_float(
            entry, "identity_relative_coefficient", where
        ),
        conditioning_scale_floor=_positive_float(entry, "conditioning_scale_floor", where),
        fx_rate_strictly_positive=_req_bool(entry, "fx_rate_strictly_positive", where),
        growth_factor_strictly_positive=_req_bool(entry, "growth_factor_strictly_positive", where),
        conditioning_terms={
            str(key): tuple(str(v) for v in value) for key, value in terms.items()
        },
    )


def _parse_references(entry: Any, where: str) -> tuple[AuthorityReference, ...]:
    where = f"{where}: authority_references"
    if not isinstance(entry, list) or not entry:
        raise CalcContractError(f"{where}: must be a non-empty list")
    return tuple(
        AuthorityReference(
            concept=_req_str(item, "concept", f"{where}[{i}]"),
            owner=_req_str(item, "owner", f"{where}[{i}]"),
            locator=_req_str(item, "locator", f"{where}[{i}]"),
        )
        for i, item in enumerate(entry)
    )


# ---------------------------------------------------------------------------
# validation
# ---------------------------------------------------------------------------
def _validate_authority_boundary(raw: dict[str, Any], raw_text: str, path: Path) -> None:
    """The hash mathematics may not appear in this file, in any form.

    Two independent scans, because they catch different mistakes. The KEY scan
    catches a well-meaning `fp_base: 131` addition. The RAW-TEXT literal scan also
    catches a commented-out constant, which a parsed-structure scan would miss and
    which a later editor would eventually uncomment.
    """
    offenders: list[str] = []

    def walk(node: Any, trail: str) -> None:
        if isinstance(node, dict):
            for key, value in node.items():
                name = str(key).lower()
                if name in FORBIDDEN_HASH_KEYS:
                    offenders.append(f"key {trail}{key}")
                walk(value, f"{trail}{key}.")
        elif isinstance(node, list):
            for index, value in enumerate(node):
                walk(value, f"{trail}[{index}].")

    walk(raw, "")

    for literal, pattern in _LITERAL_RE.items():
        if pattern.search(raw_text):
            offenders.append(f"literal {literal}")

    if offenders:
        raise CalcContractError(
            f"{path}: the fingerprint hash mathematics must not appear in the calculation "
            f"contract, but it does: {sorted(set(offenders))}. FP_BASE, FP_MOD_1, FP_MOD_2, "
            "FP_INIT_1 and FP_INIT_2 are owned by builder/pccm_builder/calc_fingerprint.py and "
            "are projected into VBA from there. Only the fingerprint VERSION NUMBER belongs in "
            "this file: which encoding produced a stored digest is a layout fact, how the "
            "encoding works is not."
        )


def _validate_sheet_and_reservation(contract: CalcContract, path: Path) -> None:
    if contract.sheet != LOCKED_SHEET:
        raise CalcContractError(
            f"{path}: Phase-5 blocks live on {LOCKED_SHEET!r}, not {contract.sheet!r}"
        )
    if contract.required_visibility != "hidden":
        raise CalcContractError(
            f"{path}: {LOCKED_SHEET} must be 'hidden', not {contract.required_visibility!r}. "
            "'veryHidden' would put the calculation record beyond an auditor's reach."
        )
    if (contract.phase4_first_row, contract.phase4_last_row) != (
        LOCKED_PHASE4_FIRST_ROW,
        LOCKED_PHASE4_LAST_ROW,
    ):
        raise CalcContractError(
            f"{path}: the Phase-4 reservation on {LOCKED_SHEET} is rows "
            f"{LOCKED_PHASE4_FIRST_ROW}-{LOCKED_PHASE4_LAST_ROW}, but the contract declares "
            f"{contract.phase4_first_row}-{contract.phase4_last_row}. Phase 4 is frozen; the "
            "contract records that reservation, it does not redraw it."
        )
    if tuple(contract.phase4_cells) != LOCKED_PHASE4_CELLS:
        raise CalcContractError(
            f"{path}: the reserved Phase-4 counter cells are {list(LOCKED_PHASE4_CELLS)}, but "
            f"the contract declares {list(contract.phase4_cells)}"
        )


def _validate_scalar_blocks(contract: CalcContract, path: Path) -> None:
    required = ("calc_state", "calc_totals")
    missing = [name for name in required if name not in contract.scalar_blocks]
    if missing:
        raise CalcContractError(f"{path}: required Phase-5 scalar block(s) missing: {missing}")
    extra = sorted(set(contract.scalar_blocks) - set(required))
    if extra:
        raise CalcContractError(
            f"{path}: unexpected scalar block(s) {extra}. Phase 5 declares exactly "
            f"{list(required)}; a new block is a layout change and needs review."
        )

    _validate_block_rows(contract.calc_state, LOCKED_CALC_STATE_ROWS, path)
    _validate_block_rows(contract.calc_totals, LOCKED_CALC_TOTALS_ROWS, path)
    _validate_calc_state_fields(contract.calc_state, path)
    _validate_calc_totals_fields(contract.calc_totals, path)


def _validate_block_rows(block: ScalarBlock, locked: tuple[int, int], path: Path) -> None:
    if (block.first_row, block.last_row) != locked:
        raise CalcContractError(
            f"{path}: {block.key} occupies rows {locked[0]}:{locked[1]}, but the contract "
            f"declares {block.first_row}:{block.last_row}"
        )
    expected = locked[1] - locked[0] + 1
    if len(block.fields) != expected:
        raise CalcContractError(
            f"{path}: {block.key} spans {expected} rows but declares {len(block.fields)} fields"
        )
    rows = [f.row for f in block.fields]
    if rows != list(range(locked[0], locked[1] + 1)):
        raise CalcContractError(
            f"{path}: {block.key} field rows must be exactly {locked[0]}..{locked[1]} in order, "
            f"got {rows}"
        )
    keys = [f.key for f in block.fields]
    duplicates = sorted({k for k in keys if keys.count(k) > 1})
    if duplicates:
        raise CalcContractError(f"{path}: {block.key} has duplicate field keys: {duplicates}")
    labels = [f.label for f in block.fields]
    duplicates = sorted({label for label in labels if labels.count(label) > 1})
    if duplicates:
        raise CalcContractError(f"{path}: {block.key} has duplicate labels: {duplicates}")
    if block.label_column == block.value_column or block.value_column == block.note_column:
        raise CalcContractError(
            f"{path}: {block.key} label, value and note columns must be distinct"
        )


def _validate_calc_state_fields(block: ScalarBlock, path: Path) -> None:
    """Every attribute of every row, against the accepted design.

    Row order is load-bearing, and so is everything else on the row: a
    `timestamp` downgraded to `text`, or a date format replaced by `@`, changes
    what the audit trail can be sorted and compared by without moving anything.
    """
    if len(block.fields) != len(LOCKED_CALC_STATE):
        raise CalcContractError(
            f"{path}: calc_state must declare exactly {len(LOCKED_CALC_STATE)} rows, "
            f"got {len(block.fields)}"
        )
    for expected, got in zip(LOCKED_CALC_STATE, block.fields):
        actual = StateRow(
            row=got.row,
            group=got.group,
            key=got.key,
            label=got.label,
            value_type=got.value_type,
            number_format=got.number_format,
            enum=got.enum,
            initial=got.initial,
        )
        if actual != expected:
            differing = [
                f"{name}: expected {getattr(expected, name)!r}, declared {getattr(actual, name)!r}"
                for name in (
                    "row",
                    "group",
                    "key",
                    "label",
                    "value_type",
                    "number_format",
                    "enum",
                    "initial",
                )
                if getattr(expected, name) != getattr(actual, name)
            ]
            raise CalcContractError(
                f"{path}: calc_state row {expected.row} ({expected.key}) does not match the "
                f"accepted Revision-E design:\n  " + "\n  ".join(differing) + "\n"
                "The row order is load-bearing - C13:C16 is the snapshot, C17:C18 the attempt "
                "and C19:C20 the derived reading - and so are the labels, types and formats an "
                "auditor reads them by."
            )

    version_field = block.field_by_key("fingerprint_version")
    if version_field.initial is not None:
        raise CalcContractError(
            f"{path}: calc_state 'Fingerprint Version' must be seeded BLANK, not "
            f"{version_field.initial!r}. Seeding the algorithm version at build time would make "
            "a never-calculated workbook look as though it held a partial successful snapshot."
        )
    for entry in block.fields:
        if entry.value_type == "enum" and entry.enum not in ("derived_status", "attempt_result"):
            raise CalcContractError(
                f"{path}: calc_state field {entry.key!r} draws from unknown label set "
                f"{entry.enum!r}"
            )
        if entry.initial is not None and entry.value_type != "enum":
            raise CalcContractError(
                f"{path}: calc_state field {entry.key!r} seeds {entry.initial!r} into a "
                f"{entry.value_type} cell. Only the two enum cells are seeded; every other cell "
                "is left genuinely blank, never 0 and never an empty string."
            )
    groups = [f.group for f in block.fields]
    if groups != sorted(groups, key=LOCKED_CALC_STATE_GROUPS.index):
        raise CalcContractError(
            f"{path}: calc_state groups must appear contiguously in the order "
            f"{list(LOCKED_CALC_STATE_GROUPS)}, got {groups}"
        )


def _validate_calc_totals_fields(block: ScalarBlock, path: Path) -> None:
    """Every attribute of every row, against the accepted design.

    A relabelled or reformatted headline total is a design change even though the
    number in the cell is unaffected: the label is what says which of the five
    measures an auditor is looking at, and `#,##0.00` is what says it is money.
    """
    if len(block.fields) != len(LOCKED_CALC_TOTALS):
        raise CalcContractError(
            f"{path}: calc_totals must declare exactly {len(LOCKED_CALC_TOTALS)} rows, "
            f"got {len(block.fields)}"
        )
    for expected, got in zip(LOCKED_CALC_TOTALS, block.fields):
        actual = TotalRow(
            row=got.row,
            key=got.key,
            label=got.label,
            value_type=got.value_type,
            number_format=got.number_format,
            units=got.units,
            measure=got.measure,
            basis=got.basis,
            initial=got.initial,
        )
        if actual != expected:
            differing = [
                f"{name}: expected {getattr(expected, name)!r}, declared {getattr(actual, name)!r}"
                for name in (
                    "row",
                    "key",
                    "label",
                    "value_type",
                    "number_format",
                    "units",
                    "measure",
                    "basis",
                    "initial",
                )
                if getattr(expected, name) != getattr(actual, name)
            ]
            raise CalcContractError(
                f"{path}: calc_totals row {expected.row} ({expected.key}) does not match the "
                f"accepted Revision-E design:\n  " + "\n  ".join(differing)
            )

    for entry in block.fields:
        if entry.value_type != "double":
            raise CalcContractError(
                f"{path}: calc_totals field {entry.key!r} must be a double, not "
                f"{entry.value_type!r}"
            )
        if entry.units != "SAR":
            raise CalcContractError(
                f"{path}: calc_totals field {entry.key!r} must be in SAR, not {entry.units!r}. "
                "The reporting currency is a locked model invariant."
            )
        if entry.initial is not None:
            raise CalcContractError(
                f"{path}: calc_totals field {entry.key!r} seeds {entry.initial!r}. Totals are "
                "left blank until a calculation commits; a seeded 0 would be a fabricated "
                "result."
            )


def _validate_tables(contract: CalcContract, path: Path) -> None:
    declared = {t.table_name for t in contract.all_tables}
    missing = sorted(set(LOCKED_TABLE_ANCHORS) - declared)
    if missing:
        raise CalcContractError(f"{path}: required Phase-5 ListObject(s) missing: {missing}")
    extra = sorted(declared - set(LOCKED_TABLE_ANCHORS))
    if extra:
        raise CalcContractError(
            f"{path}: unexpected Phase-5 ListObject(s) {extra}. The accepted layout declares "
            f"exactly {sorted(LOCKED_TABLE_ANCHORS)}."
        )
    names = contract.table_names
    duplicates = sorted({n for n in names if names.count(n) > 1})
    if duplicates:
        raise CalcContractError(f"{path}: duplicate Phase-5 table names: {duplicates}")

    for table in contract.all_tables:
        first, last, header_row, width = LOCKED_TABLE_ANCHORS[table.table_name]
        if (table.first_column, table.last_column, table.header_row) != (first, last, header_row):
            raise CalcContractError(
                f"{path}: {table.table_name} is anchored at {first}:{last} header row "
                f"{header_row}, but the contract declares "
                f"{table.first_column}:{table.last_column} header row {table.header_row}. "
                "The contract encodes the accepted anchors; it does not choose them."
            )
        if len(table.columns) != width:
            raise CalcContractError(
                f"{path}: {table.table_name} declares {len(table.columns)} columns but its "
                f"locked band {first}:{last} is {width} columns wide. A schema that does not "
                "exactly fill its band would either waste a column or run into the gutter."
            )
        if table.band_width != len(table.columns):
            raise CalcContractError(
                f"{path}: {table.table_name} band {table.first_column}:{table.last_column} is "
                f"{table.band_width} columns wide but the schema declares {len(table.columns)}"
            )
        keys = [c.key for c in table.columns]
        duplicates = sorted({k for k in keys if keys.count(k) > 1})
        if duplicates:
            raise CalcContractError(
                f"{path}: {table.table_name} has duplicate column keys: {duplicates}"
            )
        headers = table.headers
        duplicates = sorted({h for h in headers if headers.count(h) > 1})
        if duplicates:
            raise CalcContractError(
                f"{path}: {table.table_name} has duplicate column headers: {duplicates}"
            )

    for table_name in sorted(LOCKED_TABLES):
        _validate_locked_schema(contract, table_name, path)
    _validate_driver_kind_coverage(contract.table_by_name("tblCalcDrivers"), path)


def _validate_locked_schema(contract: CalcContract, table_name: str, path: Path) -> None:
    """The FULL schema of one table, attribute by attribute.

    Headers alone are not the schema. A number format decides whether a discount
    factor is shown to six places or rounded to two; a unit decides whether a
    column of money is SAR or something else; a value type decides whether a
    calendar year is a number or text. Each is a design change that leaves the
    column count and every anchor untouched, so each is compared here.
    """
    locked = LOCKED_TABLES[table_name]
    table = contract.table_by_name(table_name)

    if table.row_rule != locked.row_rule:
        raise CalcContractError(
            f"{path}: {table_name} row rule does not match the accepted design.\n"
            f"  expected: {locked.row_rule!r}\n"
            f"  declared: {table.row_rule!r}\n"
            "The row rule states which rows the table is required to contain; changing it "
            "changes what a complete audit table is."
        )
    if len(table.columns) != len(locked.columns):
        raise CalcContractError(
            f"{path}: {table_name} must declare exactly {len(locked.columns)} columns, "
            f"got {len(table.columns)}"
        )

    for index, (expected, got) in enumerate(zip(locked.columns, table.columns), start=1):
        actual = ColumnSchema(
            key=got.key,
            header=got.header,
            value_type=got.value_type,
            number_format=got.number_format,
            units=got.units,
            applies_to=got.applies_to,
        )
        if actual == expected:
            continue
        differing = [
            f"{name}: expected {getattr(expected, name)!r}, declared {getattr(actual, name)!r}"
            for name in ("key", "header", "value_type", "number_format", "units", "applies_to")
            if getattr(expected, name) != getattr(actual, name)
        ]
        raise CalcContractError(
            f"{path}: {table_name} column {index} ({expected.header!r}) does not match the "
            f"accepted Revision-E design:\n  " + "\n  ".join(differing) + "\n"
            "The contract encodes the accepted schema; it does not get to choose a new one."
        )


def _validate_driver_kind_coverage(table: CalcTable, path: Path) -> None:
    """No audit column may carry two meanings depending on Driver Kind.

    `applies_to` is what makes that machine-checkable: a column that applies to
    neither kind is dead, and every kind must have at least one column of its own,
    or the audit rows could not reconstruct the headline measures by kind.
    """
    for column in table.columns:
        if not column.applies_to:
            raise CalcContractError(
                f"{path}: {table.table_name} column {column.key!r} declares no applies_to. Every "
                "audit column must say which driver kinds populate it, so that 'blank, never "
                "zero' is checkable rather than prose."
            )
    for kind in VALID_DRIVER_KINDS:
        if not any(kind in (c.applies_to or ()) for c in table.columns):
            raise CalcContractError(
                f"{path}: {table.table_name} has no column that applies to {kind!r}"
            )


def _validate_no_overlap(contract: CalcContract, path: Path) -> None:
    """Nothing Phase 5 declares may touch Phase 4, or any other Phase-5 block."""
    reserved_rows = set(contract.phase4_reserved_rows)
    reserved_cells = set(contract.phase4_cells)

    for block in contract.scalar_blocks.values():
        clash = sorted(set(block.rows) & reserved_rows)
        if clash:
            raise CalcContractError(
                f"{path}: {block.key} claims {LOCKED_SHEET} row(s) {clash}, which are frozen "
                f"Phase-4 territory (rows {contract.phase4_first_row}-{contract.phase4_last_row})"
            )
        for entry in block.fields:
            for column in (block.label_column, block.value_column, block.note_column):
                cell = f"{column}{entry.row}"
                if cell in reserved_cells:
                    raise CalcContractError(
                        f"{path}: {block.key} field {entry.key!r} claims reserved Phase-4 cell "
                        f"{LOCKED_SHEET}!{cell}"
                    )

    blocks = sorted(contract.scalar_blocks.values(), key=lambda b: b.first_row)
    for earlier, later in zip(blocks, blocks[1:]):
        if later.first_row <= earlier.last_row:
            raise CalcContractError(
                f"{path}: scalar blocks {earlier.key} (rows {earlier.first_row}-"
                f"{earlier.last_row}) and {later.key} (rows {later.first_row}-{later.last_row}) "
                "overlap. Both are fixed-height, so this is a layout error, not growth."
            )

    for table in contract.all_tables:
        if table.header_row in reserved_rows:
            raise CalcContractError(
                f"{path}: {table.table_name} anchors its header at row {table.header_row}, "
                "inside the frozen Phase-4 reservation"
            )

    ordered = sorted(contract.all_tables, key=lambda t: t.first_column_index)
    for earlier, later in zip(ordered, ordered[1:]):
        if later.first_column_index <= earlier.last_column_index:
            raise CalcContractError(
                f"{path}: column bands {earlier.table_name} "
                f"({earlier.first_column}:{earlier.last_column}) and {later.table_name} "
                f"({later.first_column}:{later.last_column}) overlap. Each dynamic table needs "
                "its own band because every one of them grows downward without bound."
            )

    # The scalar blocks live in B/C/E; no dynamic band may reach back into them.
    scalar_columns = set()
    for block in contract.scalar_blocks.values():
        scalar_columns.update(
            column_index_from_string(c)
            for c in (block.label_column, block.value_column, block.note_column)
        )
    for table in contract.all_tables:
        band = set(range(table.first_column_index, table.last_column_index + 1))
        clash = sorted(get_column_letter(c) for c in band & scalar_columns)
        if clash:
            raise CalcContractError(
                f"{path}: {table.table_name} band {table.first_column}:{table.last_column} "
                f"includes scalar-block column(s) {clash}"
            )


def _validate_state_labels(contract: CalcContract, path: Path) -> None:
    if contract.derived_status_labels != LOCKED_DERIVED_STATUS:
        raise CalcContractError(
            f"{path}: the derived-status axis is exactly {list(LOCKED_DERIVED_STATUS)}, in that "
            f"order; the contract declares {list(contract.derived_status_labels)}"
        )
    if contract.attempt_result_labels != LOCKED_ATTEMPT_RESULT:
        raise CalcContractError(
            f"{path}: the attempt-result axis is exactly {list(LOCKED_ATTEMPT_RESULT)}, in that "
            f"order; the contract declares {list(contract.attempt_result_labels)}"
        )
    # Stated separately from the exact-list checks above, because these two are the
    # specific confusions the orthogonal-axis design exists to prevent, and a
    # reviewer should see them fail by name.
    if "REFUSED" in contract.derived_status_labels:
        raise CalcContractError(
            f"{path}: 'REFUSED' is an ATTEMPT result and must never appear on the derived-status "
            "axis. A model with an invalid input is INVALID whether or not anyone pressed "
            "Calculate; conflating the two loses that distinction."
        )
    if "FAILED" in contract.derived_status_labels:
        raise CalcContractError(
            f"{path}: 'FAILED' is an ATTEMPT result and must never appear on the derived-status "
            "axis. After a rolled-back write the model is STALE - derived from the current "
            "inputs against the restored snapshot - not 'failed'."
        )
    overlap = sorted(set(contract.derived_status_labels) & set(contract.attempt_result_labels))
    if overlap:
        raise CalcContractError(
            f"{path}: the two state axes share label(s) {overlap}. They are orthogonal and must "
            "remain separately readable."
        )


def _validate_fingerprint(contract: CalcContract, path: Path) -> None:
    if contract.fingerprint_version != LOCKED_FP_VERSION:
        raise CalcContractError(
            f"{path}: FP_VERSION is {LOCKED_FP_VERSION}; the contract declares "
            f"{contract.fingerprint_version}. Changing it declares a new canonical encoding and "
            "invalidates every stored digest, which is a design change, not a contract edit."
        )


def _validate_authority_reference_set(contract: CalcContract, path: Path) -> None:
    """The COMPLETE required reference set - none missing, extra, duplicated or renamed.

    Resolving the references that happen to be present is not enough. A reference
    can simply be deleted, and then the boundary it declared stops being declared
    at all: the contract would still load, still cross-validate, and would no
    longer record that the FX convention is owned by `input_contract.yaml`. These
    six boundaries are what Step 1 exists to protect, so the set itself is locked.

    Only the boundary METADATA is locked here. The referenced values stay owned by
    the upstream contracts, and `validate_calc_against` still resolves each locator
    against the file that owns it.
    """
    declared = tuple((r.concept, r.owner, r.locator) for r in contract.authority_references)

    duplicates = sorted({r for r in declared if declared.count(r) > 1})
    if duplicates:
        raise CalcContractError(
            f"{path}: duplicate authority reference(s): {duplicates}. Each boundary is declared "
            "exactly once."
        )
    concepts = [r[0] for r in declared]
    repeated = sorted({c for c in concepts if concepts.count(c) > 1})
    if repeated:
        raise CalcContractError(
            f"{path}: authority concept(s) {repeated} declared more than once, with differing "
            "owners or locators. A concept has exactly one owning authority."
        )

    if declared == LOCKED_AUTHORITY_REFERENCES:
        return

    expected_set = set(LOCKED_AUTHORITY_REFERENCES)
    declared_set = set(declared)
    missing = sorted(expected_set - declared_set)
    unexpected = sorted(declared_set - expected_set)
    detail = []
    if missing:
        detail.append(f"  missing: {missing}")
    if unexpected:
        detail.append(f"  unexpected: {unexpected}")
    if not detail:
        detail.append(
            f"  order differs:\n    expected {list(LOCKED_AUTHORITY_REFERENCES)}\n"
            f"    declared {list(declared)}"
        )
    raise CalcContractError(
        f"{path}: the authority-reference set does not match the accepted Revision-E "
        f"boundaries.\n" + "\n".join(detail) + "\n"
        "Every value the calculation contract borrows must name the specification that owns "
        "it. A deleted reference silently drops that boundary; a renamed concept or a moved "
        "locator silently redirects it."
    )


def _validate_tolerances(contract: CalcContract, path: Path) -> None:
    """Exact equality first, then the structural rules.

    A tolerance edit is a NUMERICAL-DESIGN change. Every one of these values was
    argued for in the plan, and none of them may pass merely because it is still a
    positive number that satisfies the general shape rules below.
    """
    tolerances = contract.tolerances

    for name, expected in LOCKED_TOLERANCES.items():
        declared = getattr(tolerances, name)
        if isinstance(expected, bool) or isinstance(declared, bool):
            matches = declared is expected
        else:
            matches = declared == expected
        if not matches:
            raise CalcContractError(
                f"{path}: tolerance {name} is locked at {expected!r}; the contract declares "
                f"{declared!r}. Loosening a tolerance changes which models the reconciliation "
                "accepts, so it is a design change and cannot pass validation as an edit."
            )

    if tuple(sorted(tolerances.conditioning_terms)) != tuple(sorted(LOCKED_CONDITIONING_TERMS)):
        raise CalcContractError(
            f"{path}: conditioning terms must be declared for exactly "
            f"{sorted(LOCKED_CONDITIONING_TERMS)}, got {sorted(tolerances.conditioning_terms)}"
        )
    for identity, expected_terms in LOCKED_CONDITIONING_TERMS.items():
        declared_terms = tolerances.conditioning_terms[identity]
        if declared_terms != expected_terms:
            raise CalcContractError(
                f"{path}: identity {identity} conditioning terms are locked as "
                f"{list(expected_terms)}; the contract declares {list(declared_terms)}. The scale "
                "must sum the absolute magnitudes THAT IDENTITY accumulates - borrowing another "
                "identity's terms sizes the tolerance by quantities it never touches."
            )

    if tolerances.identity_absolute_floor <= tolerances.identity_relative_coefficient:
        raise CalcContractError(
            f"{path}: the identity absolute floor ({tolerances.identity_absolute_floor}) must "
            f"exceed the relative coefficient ({tolerances.identity_relative_coefficient}); the "
            "floor is what binds for a small model"
        )
    if tolerances.conditioning_scale_floor < 1.0:
        raise CalcContractError(
            f"{path}: the conditioning scale floor must be at least 1.0, got "
            f"{tolerances.conditioning_scale_floor}. Below unity the scale could shrink the "
            "tolerance instead of only ever widening it."
        )
    if not tolerances.fx_rate_strictly_positive or not tolerances.growth_factor_strictly_positive:
        raise CalcContractError(
            f"{path}: FX rates and growth factors are checked STRICTLY positive, with no "
            "epsilon. A zero rate is a refusal, not a rounding question."
        )
    for identity, terms in tolerances.conditioning_terms.items():
        if len(terms) < 2:
            raise CalcContractError(
                f"{path}: identity {identity} declares conditioning terms {list(terms)}. The "
                "scale must sum the ABSOLUTE magnitudes of the terms actually accumulated, or a "
                "model whose large positive and negative contributions cancel would collapse "
                "its own tolerance to the floor and report ordinary accumulation error as a "
                "bookkeeping mismatch."
            )
        duplicates = sorted({t for t in terms if terms.count(t) > 1})
        if duplicates:
            raise CalcContractError(
                f"{path}: identity {identity} repeats conditioning term(s) {duplicates}"
            )


def _validate_excel_bounds(contract: CalcContract, path: Path) -> None:
    for block in contract.scalar_blocks.values():
        for entry in block.fields:
            _checked(check_row, entry.row, f"{path}: {block.key}.{entry.key} row")
            for column in (block.label_column, block.value_column, block.note_column):
                _checked(
                    check_cell, f"{column}{entry.row}", f"{path}: {block.key}.{entry.key} cell"
                )
    for cell in contract.phase4_cells:
        _checked(check_cell, cell, f"{path}: phase4_reservation cell")
    for table in contract.all_tables:
        _checked(check_row, table.header_row, f"{path}: {table.table_name} header_row")
        if table.first_column_index > table.last_column_index:
            raise CalcContractError(
                f"{path}: {table.table_name} band {table.first_column}:{table.last_column} runs "
                "backwards"
            )


# ---------------------------------------------------------------------------
# cross-contract validation
# ---------------------------------------------------------------------------
def validate_calc_against(
    calc: CalcContract,
    spec: WorkbookSpec,
    inputs: InputContract,
    drivers: DriverContract,
    structure: StructureContract,
    path: Path | None = None,
) -> None:
    """Assert the calculation contract agrees with the four accepted authorities.

    Everything the calculation contract borrows is checked to still exist under the
    name it borrows it by, so a rename upstream fails the build here instead of
    leaving a stale reference behind.
    """
    path = path or calc.source_path

    # --- the sheet is the one workbook.yaml declares -------------------------
    sheet = next((s for s in spec.sheets if s.name == calc.sheet), None)
    if sheet is None:
        raise CalcContractError(
            f"{path}: sheet {calc.sheet!r} is not declared by {spec.source_path.name}"
        )
    if sheet.visibility != calc.required_visibility:
        raise CalcContractError(
            f"{path}: {calc.sheet} is {sheet.visibility!r} in {spec.source_path.name} but the "
            f"calculation contract requires {calc.required_visibility!r}"
        )

    # --- the reserved cells are the Phase-4 counters, by their own contract ---
    if structure.identity_sheet != calc.sheet:
        raise CalcContractError(
            f"{path}: the permanent-ID counters live on {structure.identity_sheet!r}, not "
            f"{calc.sheet!r}; the Phase-4 reservation recorded here would guard the wrong sheet"
        )
    counter_cells = tuple(sorted(c.cell for c in structure.counters))
    if tuple(sorted(calc.phase4_cells)) != counter_cells:
        raise CalcContractError(
            f"{path}: the reserved Phase-4 cells {list(calc.phase4_cells)} are not the counter "
            f"cells {list(counter_cells)} declared by {structure.source_path.name}"
        )
    for counter in structure.counters:
        for cell in (counter.cell, counter.label_cell):
            row = int("".join(ch for ch in cell if ch.isdigit()))
            if row not in calc.phase4_reserved_rows:
                raise CalcContractError(
                    f"{path}: Phase-4 counter cell {cell} is on row {row}, outside the "
                    f"reservation rows {calc.phase4_first_row}-{calc.phase4_last_row} this "
                    "contract declares. The reservation must cover everything Phase 4 owns."
                )

    # --- structural blocks already on _Calc must not be overwritten ----------
    identity = structure.identity_block
    for key in ("section_row", "intro_row"):
        row = identity.get(key)
        if row is not None and row not in calc.phase4_reserved_rows:
            raise CalcContractError(
                f"{path}: the Phase-4 identity block claims {calc.sheet} row {row} ({key}), "
                f"outside the declared reservation {calc.phase4_first_row}-"
                f"{calc.phase4_last_row}"
            )

    # --- no table-name collision across the five contracts -------------------
    existing = (
        {t.table_name for t in inputs.all_tables}
        | {r.table_name for r in drivers.all_registers}
        | {g.table_name for g in structure.all_grids}
    )
    for name in calc.table_names:
        if name in existing:
            raise CalcContractError(
                f"{path}: Phase-5 table {name!r} is already declared by another contract"
            )

    # --- every borrowed value still resolves in its owning authority ---------
    owners: dict[str, tuple[Path, dict[str, Any]]] = {}
    for contract_path in (
        spec.source_path,
        inputs.source_path,
        drivers.source_path,
        structure.source_path,
    ):
        with contract_path.open("r", encoding="utf-8") as handle:
            owners[contract_path.name] = (contract_path, yaml.safe_load(handle))

    for reference in calc.authority_references:
        if reference.owner not in owners:
            raise CalcContractError(
                f"{path}: authority reference {reference.concept!r} names owner "
                f"{reference.owner!r}, which is not one of the four accepted specifications "
                f"{sorted(owners)}"
            )
        _, document = owners[reference.owner]
        node: Any = document
        for part in reference.locator.split("."):
            resolved = _step(node, part)
            if resolved is _UNRESOLVED:
                raise CalcContractError(
                    f"{path}: authority reference {reference.concept!r} points at "
                    f"{reference.owner}:{reference.locator}, which does not resolve (failed at "
                    f"{part!r}). The calculation contract references borrowed values instead of "
                    "copying them precisely so a rename upstream breaks here rather than going "
                    "unnoticed."
                )
            node = resolved


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
_UNRESOLVED = object()

_ELEMENT_KEYS = ("key", "name", "table_name", "defined_name")


def _step(node: Any, part: str) -> Any:
    """One step of a dotted authority locator.

    The four accepted contracts express collections both ways - `registers` is a
    mapping keyed by name, `config_tables` and `sheets` are lists of mappings each
    carrying its own identifier - so a locator resolves against either shape. A
    locator names a concept, not a YAML container type.
    """
    if isinstance(node, dict):
        return node[part] if part in node else _UNRESOLVED
    if isinstance(node, list):
        for element in node:
            if isinstance(element, dict) and any(
                element.get(key) == part for key in _ELEMENT_KEYS
            ):
                return element
    return _UNRESOLVED


def _as_mapping(node: Any, where: str) -> dict[str, Any]:
    if not isinstance(node, dict) or not node:
        raise CalcContractError(f"{where}: must be a non-empty mapping")
    return node


def _req(mapping: Any, key: str, where: str) -> Any:
    if not isinstance(mapping, dict) or key not in mapping:
        raise CalcContractError(f"{where}: missing required key {key!r}")
    return mapping[key]


def _req_str(mapping: Any, key: str, where: str) -> str:
    value = _req(mapping, key, where)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        value = str(value)
    if not isinstance(value, str) or not value.strip():
        raise CalcContractError(f"{where}: {key!r} must be a non-empty string, got {value!r}")
    return value


def _req_key(mapping: Any, key: str, where: str) -> str:
    value = _req_str(mapping, key, where)
    if not KEY_RE.match(value):
        raise CalcContractError(f"{where}: {key} {value!r} must be lower_snake_case")
    return value


def _req_bool(mapping: Any, key: str, where: str) -> bool:
    value = _req(mapping, key, where)
    if not isinstance(value, bool):
        raise CalcContractError(f"{where}: {key!r} must be a boolean, got {value!r}")
    return value


def _positive_int(mapping: Any, key: str, where: str) -> int:
    value = _req(mapping, key, where)
    if not isinstance(value, int) or isinstance(value, bool) or value < 1:
        raise CalcContractError(f"{where}: {key!r} must be a positive integer, got {value!r}")
    return value


def _positive_float(mapping: Any, key: str, where: str) -> float:
    value = _req(mapping, key, where)
    if isinstance(value, bool) or not isinstance(value, (int, float)) or value <= 0:
        raise CalcContractError(f"{where}: {key!r} must be a positive number, got {value!r}")
    return float(value)
