"""Create the PCCM Stage A workbook from the manifest and the input contract.

Responsibilities are deliberately split across modules:
    workbook_builder  orchestration: sheets, visibility, active sheet, metadata
    contract_render   Setup and Config bodies (inputs, tables)
    names             defined names
    validation        data validation
    styling           presentation tokens
    spec_loader       structural manifest
    contract_loader   input contract
    verify            structural verification

No business rule or calculation belongs in any of them.
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .contract_loader import InputContract
from .contract_render import render_config, render_setup
from .driver_loader import DriverContract, validate_against_input_contract
from .driver_render import render_register
from .names import apply_defined_names
from .spec_loader import SheetSpec, WorkbookSpec
from .structure_loader import StructureContract, validate_structure_against
from .calc_loader import CalcContract
from .calc_render import render_calc_workspace
from .structure_render import render_applied_timeline, render_grid, render_identity
from .styling import StyleBook
from .validation import apply_validation

BUILDER_VERSION = "0.5.0"
DEFAULT_SHEET_TITLE = "Sheet"
TIMESTAMP_ENV_VAR = "PCCM_BUILD_TIMESTAMP"


@dataclass(frozen=True)
class BuildMetadata:
    """Build-time provenance.

    None of these values is a computational input. They record how the artifact
    was produced and are kept distinct from:
      * model_version   - the version of the model design itself
      * builder_version - the version of this build tooling
      * run metadata    - which belongs to a Monte Carlo run, not to a build
    """

    model_version: str
    build_phase: str
    builder_version: str
    build_timestamp: str
    manifest_version: str
    contract_version: str
    driver_contract_version: str
    structure_contract_version: str

    @classmethod
    def create(
        cls,
        spec: WorkbookSpec,
        contract: InputContract,
        drivers: DriverContract,
        structure: StructureContract,
    ) -> "BuildMetadata":
        return cls(
            model_version=spec.model["model_version"],
            build_phase=spec.model["build_phase"],
            builder_version=BUILDER_VERSION,
            build_timestamp=resolve_build_timestamp(),
            manifest_version=spec.manifest_version,
            contract_version=contract.contract_version,
            driver_contract_version=drivers.version,
            structure_contract_version=structure.version,
        )

    def as_rows(self) -> list[tuple[str, str]]:
        return [
            ("PCCM Model Version", self.model_version),
            ("Build Phase", self.build_phase),
            ("Builder Version", self.builder_version),
            ("Build Timestamp (UTC)", self.build_timestamp),
            ("Source Manifest Version", self.manifest_version),
            ("Input Contract Version", self.contract_version),
            ("Driver Contract Version", self.driver_contract_version),
            ("Structure Contract Version", self.structure_contract_version),
        ]


def resolve_build_timestamp() -> str:
    """UTC build timestamp, overridable for reproducible builds.

    Setting PCCM_BUILD_TIMESTAMP makes two builds of the same source produce
    structurally identical workbooks, which is what the reproducibility test
    relies on.
    """
    override = os.environ.get(TIMESTAMP_ENV_VAR)
    if override:
        return override
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def build_workbook(
    spec: WorkbookSpec,
    contract: InputContract,
    drivers: DriverContract,
    structure: StructureContract,
    calc: CalcContract | None = None,
) -> tuple[Workbook, BuildMetadata]:
    """Create the Stage A workbook from the manifest and every contract.

    `calc` is the Phase-5 calculation contract. It is optional ONLY so that
    isolated Phase 1-4 unit tests can still build the structural workbook they
    were written against; `build_stage_a.py` always supplies it, and the
    production Stage-A path therefore always renders and verifies the Phase-5
    workspace. Passing `None` builds a workbook with no `_Calc` Phase-5 blocks,
    which post-build verification with a contract will reject.
    """
    _assert_consistent(spec, contract, drivers, structure)

    styles = StyleBook(spec.presentation)
    metadata = BuildMetadata.create(spec, contract, drivers, structure)

    workbook = Workbook()
    default_sheet = workbook.active

    for sheet_spec in spec.sheets:
        worksheet = workbook.create_sheet(title=sheet_spec.name)
        _apply_presentation(worksheet, sheet_spec, styles)
        _write_header(worksheet, sheet_spec, styles)
        if sheet_spec.body == "contract":
            if sheet_spec.name == contract.setup_sheet:
                render_setup(worksheet, contract, styles)
                # Setup carries two authorities: the input contract owns the entered
                # inputs and the FX table, the structure contract appends the applied
                # timeline below them. The loaders prove the two areas cannot overlap.
                render_applied_timeline(worksheet, structure, styles)
            else:
                render_config(worksheet, contract, styles)
        elif sheet_spec.body == "drivers":
            render_register(worksheet, drivers.register_for_sheet(sheet_spec.name), styles)
        elif sheet_spec.body == "structure":
            grid = structure.grid_for_sheet(sheet_spec.name)
            if grid is not None:
                render_grid(worksheet, grid, structure, styles)
            else:
                render_identity(worksheet, structure, styles)
                # `_Calc` carries two authorities: Phase 4 owns rows 1-11 above,
                # and the calculation contract owns everything from row 13 down.
                # The calc loader proves the two areas cannot intersect.
                if calc is not None and sheet_spec.name == calc.sheet:
                    render_calc_workspace(worksheet, calc, styles)
        else:
            _populate_blocks(worksheet, sheet_spec, styles, metadata)

    # Remove openpyxl's default sheet only after the real sheets exist, so the
    # workbook is never momentarily empty.
    if default_sheet is not None and default_sheet.title == DEFAULT_SHEET_TITLE:
        workbook.remove(default_sheet)
    if DEFAULT_SHEET_TITLE in workbook.sheetnames:
        raise RuntimeError("the default openpyxl sheet was not removed cleanly")

    # Visibility is applied after creation so the active sheet can be set while
    # every sheet is still visible.
    workbook.active = workbook.sheetnames.index(spec.active_sheet)
    for sheet_spec in spec.sheets:
        workbook[sheet_spec.name].sheet_state = sheet_spec.visibility

    apply_defined_names(workbook, contract, structure)
    apply_validation(
        {name: workbook[name] for name in workbook.sheetnames}, contract, drivers
    )

    _apply_document_properties(workbook, spec, metadata)
    return workbook, metadata


def _assert_consistent(
    spec: WorkbookSpec,
    contract: InputContract,
    drivers: DriverContract | None = None,
    structure: StructureContract | None = None,
) -> None:
    """The specifications must agree before anything is rendered."""
    # Cross-spec: the reporting currency is declared in both files. They must not
    # be allowed to drift apart silently.
    manifest_currency = spec.model["reporting_currency"]
    if manifest_currency != contract.reporting_currency:
        raise RuntimeError(
            "reporting currency disagrees between specifications: "
            f"workbook.yaml model.reporting_currency={manifest_currency!r}, "
            f"input_contract.yaml model_invariants.reporting_currency="
            f"{contract.reporting_currency!r}"
        )

    declared = set(spec.contract_sheets)
    used = contract.contract_sheets
    if declared != used:
        raise RuntimeError(
            "manifest and input contract disagree about contract-bodied sheets: "
            f"manifest says {sorted(declared)}, contract targets {sorted(used)}"
        )
    known = set(spec.sheet_names)
    for table in contract.all_tables:
        if table.sheet not in known:
            raise RuntimeError(
                f"input contract table {table.table_name!r} targets unknown sheet {table.sheet!r}"
            )
    for input_spec in contract.inputs.values():
        if input_spec.sheet not in known:
            raise RuntimeError(
                f"input {input_spec.key!r} targets unknown sheet {input_spec.sheet!r}"
            )

    if drivers is None:
        return

    declared_drivers = set(spec.driver_sheets)
    if declared_drivers != drivers.sheets:
        raise RuntimeError(
            "manifest and driver contract disagree about driver-bodied sheets: "
            f"manifest says {sorted(declared_drivers)}, "
            f"driver contract targets {sorted(drivers.sheets)}"
        )
    for register in drivers.all_registers:
        if register.sheet not in known:
            raise RuntimeError(
                f"driver register {register.table_name!r} targets unknown sheet "
                f"{register.sheet!r}"
            )
    validate_against_input_contract(drivers, contract)

    if structure is None:
        return

    declared_structure = set(spec.structure_sheets)
    if declared_structure != structure.owned_sheets:
        raise RuntimeError(
            "manifest and structure contract disagree about structure-bodied sheets: "
            f"manifest says {sorted(declared_structure)}, "
            f"structure contract owns {sorted(structure.owned_sheets)}"
        )
    for sheet in sorted(structure.owned_sheets | {structure.setup_sheet}):
        if sheet not in known:
            raise RuntimeError(f"structure contract targets unknown sheet {sheet!r}")
    if structure.setup_sheet != contract.setup_sheet:
        raise RuntimeError(
            "manifest and structure contract disagree about the Setup sheet: the applied "
            f"timeline targets {structure.setup_sheet!r}, the input contract owns "
            f"{contract.setup_sheet!r}"
        )
    for button in structure.buttons:
        if button.sheet not in known:
            raise RuntimeError(
                f"button {button.shape_name!r} targets unknown sheet {button.sheet!r}"
            )
    validate_structure_against(structure, contract, drivers)


def _apply_document_properties(
    workbook: Workbook, spec: WorkbookSpec, metadata: BuildMetadata
) -> None:
    properties = workbook.properties
    properties.title = spec.model["name"]
    properties.creator = f"{spec.model['short_name']} builder {metadata.builder_version}"
    properties.lastModifiedBy = properties.creator
    properties.category = metadata.build_phase
    properties.description = (
        f"{spec.model['short_name']} model version {metadata.model_version}; "
        f"manifest {metadata.manifest_version}; built {metadata.build_timestamp}."
    )
    # Deterministic document timestamps keep repeated builds comparable.
    fixed = datetime(2000, 1, 1, tzinfo=timezone.utc).replace(tzinfo=None)
    properties.created = fixed
    properties.modified = fixed


def _apply_presentation(
    worksheet: Worksheet, sheet_spec: SheetSpec, styles: StyleBook
) -> None:
    worksheet.sheet_view.showGridLines = sheet_spec.show_gridlines
    for column, width in sheet_spec.column_widths.items():
        worksheet.column_dimensions[column].width = width
    if sheet_spec.freeze_panes:
        worksheet.freeze_panes = sheet_spec.freeze_panes


def _write_header(worksheet: Worksheet, sheet_spec: SheetSpec, styles: StyleBook) -> None:
    """Title, subtitle and the thin rule. Common to every sheet."""
    layout = styles.layout
    label_col = layout.label_column

    _write(worksheet, f"{label_col}{layout.title_row}", sheet_spec.title, styles.title)
    worksheet.row_dimensions[layout.title_row].height = styles.row_height("title")

    if sheet_spec.subtitle:
        _write(
            worksheet,
            f"{label_col}{layout.subtitle_row}",
            sheet_spec.subtitle,
            styles.subtitle,
        )
        worksheet.row_dimensions[layout.subtitle_row].height = styles.row_height("subtitle")

    # A thin rule under the header, drawn across the used column span.
    for column in _rule_columns(sheet_spec):
        worksheet[f"{column}{layout.rule_row}"].border = styles.rule
    worksheet.row_dimensions[layout.rule_row].height = styles.row_height("spacer")


def _populate_blocks(
    worksheet: Worksheet,
    sheet_spec: SheetSpec,
    styles: StyleBook,
    metadata: BuildMetadata,
) -> None:
    """Sheets whose body comes from the manifest rather than the input contract."""
    layout = styles.layout
    label_col = layout.label_column
    value_col = layout.value_column

    row = layout.body_start_row
    for block in sheet_spec.blocks:
        row = _write_block(worksheet, block, row, styles, metadata, label_col, value_col)
    return None


def _rule_columns(sheet_spec: SheetSpec) -> list[str]:
    if not sheet_spec.column_widths:
        return ["B"]
    return sorted(sheet_spec.column_widths, key=lambda c: (len(c), c))


def _write_block(
    worksheet: Worksheet,
    block: dict[str, Any],
    row: int,
    styles: StyleBook,
    metadata: BuildMetadata,
    label_col: str,
    value_col: str,
) -> int:
    if block["type"] == "note":
        _write(worksheet, f"{label_col}{row}", block["text"], styles.note)
        return row + 2

    _write(worksheet, f"{label_col}{row}", block["title"], styles.section)
    worksheet.row_dimensions[row].height = styles.row_height("section")
    row += 1

    if block.get("note"):
        _write(worksheet, f"{label_col}{row}", block["note"], styles.note)
        row += 1

    for entry in block.get("rows") or []:
        _write(worksheet, f"{label_col}{row}", entry["label"], styles.label)
        if "value" in entry and entry["value"] is not None:
            font = styles.value_locked if entry.get("locked") else styles.value
            _write(worksheet, f"{value_col}{row}", entry["value"], font)
        if entry.get("note"):
            _write(worksheet, f"F{row}", entry["note"], styles.note)
        row += 1

    for item in block.get("list") or []:
        _write(worksheet, f"{value_col}{row}", item, styles.list_item)
        row += 1

    if block.get("metadata"):
        for label, value in metadata.as_rows():
            _write(worksheet, f"{label_col}{row}", label, styles.label)
            _write(worksheet, f"{value_col}{row}", value, styles.value)
            row += 1

    return row + 1


def _write(worksheet: Worksheet, address: str, value: Any, font) -> None:
    cell = worksheet[address]
    cell.value = value
    cell.font = font
