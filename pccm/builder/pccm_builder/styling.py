"""Presentation tokens for the PCCM workbook.

Every font, size and colour used by the builder is resolved here from the
manifest's ``presentation`` block. No other module defines a style literal.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class Layout:
    title_row: int
    subtitle_row: int
    rule_row: int
    body_start_row: int
    label_column: str
    value_column: str


class StyleBook:
    """Resolved fonts, borders and layout positions for one workbook build."""

    def __init__(self, presentation: dict[str, Any]) -> None:
        family = presentation["font_family"]
        sizes = presentation["sizes"]
        colors = presentation["colors"]

        self._row_heights: dict[str, float] = presentation["row_heights"]

        self.title = Font(name=family, size=sizes["title"], bold=True, color=colors["title"])
        self.subtitle = Font(name=family, size=sizes["subtitle"], italic=True, color=colors["subtitle"])
        self.section = Font(name=family, size=sizes["section"], bold=True, color=colors["section"])
        self.label = Font(name=family, size=sizes["label"], color=colors["label"])
        self.value = Font(name=family, size=sizes["value"], color=colors["value"])
        self.value_locked = Font(name=family, size=sizes["value"], bold=True, color=colors["value"])
        self.note = Font(name=family, size=sizes["note"], italic=True, color=colors["note"])
        self.list_item = Font(name=family, size=sizes["value"], color=colors["value"])

        # --- Phase 2: input language -------------------------------------
        self.input_font = Font(name=family, size=sizes["value"], color=colors["value"])
        self.input_fill = PatternFill("solid", fgColor=colors["input_fill"])
        input_side = Side(style="thin", color=colors["input_border"])
        self.input_border = Border(
            left=input_side, right=input_side, top=input_side, bottom=input_side
        )

        self.locked_font = Font(name=family, size=sizes["value"], bold=True, color=colors["locked_text"])
        self.locked_fill = PatternFill("solid", fgColor=colors["locked_fill"])
        locked_side = Side(style="thin", color=colors["locked_border"])
        self.locked_border = Border(
            left=locked_side, right=locked_side, top=locked_side, bottom=locked_side
        )

        # Presentation-only marker: a cell that is not applicable for the current
        # row's distribution. It constrains nothing.
        self.not_applicable_fill = PatternFill("solid", fgColor=colors["not_applicable_fill"])

        self.table_header_font = Font(name=family, size=sizes["value"], bold=True, color=colors["section"])
        self.table_header_fill = PatternFill("solid", fgColor=colors["header_fill"])
        self.table_style_name = presentation["table_style"]

        self.rule = Border(bottom=Side(style="thin", color=colors["rule"]))
        self.left = Alignment(horizontal="left", vertical="center")
        self.left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=False)

        layout = presentation["layout"]
        self.layout = Layout(
            title_row=int(layout["title_row"]),
            subtitle_row=int(layout["subtitle_row"]),
            rule_row=int(layout["rule_row"]),
            body_start_row=int(layout["body_start_row"]),
            label_column=str(layout["label_column"]),
            value_column=str(layout["value_column"]),
        )

    def apply_input(self, cell) -> None:
        """Editable user input: pale fill, thin border."""
        cell.font = self.input_font
        cell.fill = self.input_fill
        cell.border = self.input_border

    def apply_locked(self, cell) -> None:
        """Model-controlled or locked constant: grey fill, bold, distinct border."""
        cell.font = self.locked_font
        cell.fill = self.locked_fill
        cell.border = self.locked_border

    def apply_table_header(self, cell) -> None:
        cell.font = self.table_header_font
        cell.fill = self.table_header_fill
        cell.border = self.rule

    def row_height(self, key: str) -> float:
        try:
            return float(self._row_heights[key])
        except KeyError as exc:  # pragma: no cover - manifest validation covers this
            raise KeyError(f"presentation.row_heights is missing {key!r}") from exc
