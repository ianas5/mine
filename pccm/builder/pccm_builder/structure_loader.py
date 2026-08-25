"""Load and validate the PCCM structure contract.

Owns structural runtime: the applied timeline, derived structural state, the
profiling and inflation grids, permanent-ID identity, the persistent counters and
the Stage-B command surface. Fails loudly; never repairs.

Excel address bounds are checked with the SAME central validators the input and
driver contracts use, not a third implementation.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml
from openpyxl.utils import get_column_letter

from .contract_loader import (
    EXCEL_MAX_COLUMN,
    ContractError,
    InputContract,
    check_cell,
    check_column,
    check_row,
)
from .driver_loader import DriverContract

TABLE_NAME_RE = re.compile(r"^tbl[A-Z][A-Za-z0-9]*$")
KEY_RE = re.compile(r"^[a-z][a-z0-9_]*$")
STRUCTURE_NAME_RE = re.compile(r"^nm[A-Z][A-Za-z0-9_]*$")
ENTRY_POINT_RE = re.compile(r"^PCCM_[A-Z][A-Za-z0-9]*$")
GENERATED_MODULES = ("modConstants", "modCalcContract", "modSimContract")
"""Every VBA module the Stage-A builder emits, and therefore every module the
contract may mark ``generated: true``.

``vba.generated_module`` names the PRIMARY generated module - the constants
projection every phase depends on - and stays ``modConstants``. Phase 5 adds
``modCalcContract``, emitted by ``calc_emit``; Phase 6 Step 5 adds
``modSimContract``, emitted by ``sim_emit``, and Step 6 is the first
implementation step that depends on it, which is when it enters the registry.
All three are build artifacts that must never be hand-edited, and a module
outside this tuple claiming to be generated means the contract and the builder
disagree about what the build produces."""

MODULE_NAME_RE = re.compile(r"^mod[A-Z][A-Za-z0-9]*$")
SHAPE_NAME_RE = re.compile(r"^btnPCCM[A-Z][A-Za-z0-9]*$")
VALID_GRID_KINDS = ("profiling", "inflation")


class StructureContractError(ContractError):
    """Raised when the structure contract is invalid.

    Subclasses ContractError so the build entry point reports every specification
    failure the same way.
    """


def _checked(fn, *args):
    """Run a central bound validator, reporting failures as a structure fault.

    The Excel-bound machinery is shared with the other contracts rather than
    duplicated; only the reported exception type is specialised, so the user is
    told which specification file is at fault.
    """
    try:
        return fn(*args)
    except StructureContractError:
        raise
    except ContractError as error:
        raise StructureContractError(str(error)) from error


# ---------------------------------------------------------------------------
# value types
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class Limits:
    min_year: int
    max_year: int
    max_generated_year_columns: int


@dataclass(frozen=True)
class EnteredAlias:
    defined_name: str
    input_key: str


@dataclass(frozen=True)
class StructuralField:
    """One cell of the applied timeline, or one derived structural display cell."""

    key: str
    label: str
    defined_name: str
    label_cell: str
    cell: str
    number_format: str
    note: str | None
    formula: str | None = None

    @property
    def is_derived(self) -> bool:
        return self.formula is not None


@dataclass(frozen=True)
class StructuralState:
    sheet: str
    label: str
    defined_name: str
    label_cell: str
    cell: str
    number_format: str
    note: str | None
    note_column: str
    labels: dict[str, str]
    formula: str


@dataclass(frozen=True)
class Counter:
    key: str
    label: str
    defined_name: str
    label_cell: str
    cell: str
    number_format: str
    initial: int
    prefix: str
    pad_width: int
    pattern: str
    driver_register: str
    note: str | None

    def format_id(self, sequence: int) -> str:
        """CL-001 ... CL-999, then CL-1000 and beyond. Pad width is a floor."""
        return f"{self.prefix}{sequence:0{self.pad_width}d}"


@dataclass(frozen=True)
class GridColumn:
    key: str
    header: str
    number_format: str
    width: float
    note: str | None
    source_driver_column: str | None = None
    source_list_column: str | None = None


@dataclass(frozen=True)
class YearColumn:
    header_format: str
    number_format: str
    width: float
    initial_value: float | None
    validation: dict[str, Any] | None


@dataclass(frozen=True)
class Grid:
    key: str
    sheet: str
    table_name: str
    kind: str
    section: str
    section_row: int
    note: str | None
    note_row: int | None
    intro: str | None
    intro_row: int | None
    state_message_row: int
    header_row: int
    first_column: str
    reserved_rows: int
    freeze_panes: str | None
    fixed_columns: list[GridColumn]
    year_column: YearColumn
    driver_register: str | None = None
    key_driver_column: str | None = None
    source_list_table: str | None = None

    @property
    def headers(self) -> list[str]:
        return [c.header for c in self.fixed_columns]

    @property
    def first_data_row(self) -> int:
        return self.header_row + 1

    @property
    def last_data_row(self) -> int:
        return self.header_row + self.reserved_rows

    @property
    def first_col_index(self) -> int:
        return _column_index(self.first_column)

    @property
    def last_fixed_column(self) -> str:
        return get_column_letter(self.first_col_index + len(self.fixed_columns) - 1)

    @property
    def ref(self) -> str:
        """Stage-A extent: fixed columns only. Year columns do not exist yet."""
        return f"{self.first_column}{self.header_row}:{self.last_fixed_column}{self.last_data_row}"

    def column_letter(self, index: int) -> str:
        return get_column_letter(self.first_col_index + index)

    def first_year_column_index(self) -> int:
        """Zero-based column offset at which generated year columns begin."""
        return len(self.fixed_columns)


@dataclass(frozen=True)
class ButtonSpec:
    key: str
    sheet: str
    shape_name: str
    caption: str
    entry_point: str
    anchor_cell: str
    width: float
    height: float


@dataclass(frozen=True)
class VbaModule:
    name: str
    generated: bool
    responsibility: str


@dataclass(frozen=True)
class ForbiddenConstruct:
    """One forbidden-construct rule - D6-11.

    Two shapes, one meaning. A bare string is GLOBALLY forbidden. A mapping
    `{construct, allowed_in}` is forbidden everywhere EXCEPT the modules named as
    its owners, which is how a construct can be introduced by the one module that
    is meant to own it without becoming legal everywhere.

    `allowed_in` is empty for the global shape, so a consumer that only cares
    about the construct text keeps working unchanged.
    """

    construct: str
    allowed_in: tuple[str, ...] = ()

    @property
    def is_scoped(self) -> bool:
        return bool(self.allowed_in)

    def forbidden_in(self, module_name: str) -> bool:
        return module_name not in self.allowed_in


@dataclass(frozen=True)
class StructureContract:
    version: str
    limits: Limits
    entered_aliases: list[EnteredAlias]
    applied_block: dict[str, Any]
    applied: list[StructuralField]
    derived: list[StructuralField]
    structural_state: StructuralState
    identity_block: dict[str, Any]
    counters: list[Counter]
    grids: dict[str, Grid]
    state_messages: dict[str, str]
    buttons: list[ButtonSpec]
    vba_modules: list[VbaModule]
    vba_source_dir: str
    vba_generated_dir: str
    vba_generated_module: str
    entry_points: list[str]
    api_procedures: list[str]
    forbidden_constructs: list[str]
    forbidden_construct_rules: list[ForbiddenConstruct]
    structural_checks: list[dict[str, str]]
    source_path: Path

    # --- convenience -------------------------------------------------------
    @property
    def all_grids(self) -> list[Grid]:
        return list(self.grids.values())

    @property
    def profiling_grids(self) -> list[Grid]:
        return [g for g in self.all_grids if g.kind == "profiling"]

    @property
    def inflation_grid(self) -> Grid:
        matches = [g for g in self.all_grids if g.kind == "inflation"]
        if len(matches) != 1:
            raise StructureContractError(
                f"expected exactly one inflation grid, found {len(matches)}"
            )
        return matches[0]

    @property
    def grid_sheets(self) -> set[str]:
        return {g.sheet for g in self.all_grids}

    @property
    def owned_sheets(self) -> set[str]:
        """Sheets whose body this contract renders in full."""
        return self.grid_sheets | {self.identity_sheet}

    @property
    def identity_sheet(self) -> str:
        return self.identity_block["sheet"]

    @property
    def setup_sheet(self) -> str:
        return self.applied_block["sheet"]

    @property
    def structural_fields(self) -> list[StructuralField]:
        return list(self.applied) + list(self.derived)

    @property
    def formula_cells(self) -> dict[str, set[str]]:
        """sheet -> the exact cells this contract is permitted to write a formula into.

        Phases 1-3 forbade every formula. Phase 4 permits structural-state display
        only, so the permitted set is enumerated here and the phase-aware tests
        assert the workbook's formulas are a subset of it.
        """
        allowed: dict[str, set[str]] = {}
        allowed.setdefault(self.setup_sheet, set()).update(
            field_.cell for field_ in self.derived
        )
        allowed.setdefault(self.structural_state.sheet, set()).add(
            self.structural_state.cell
        )
        for grid in self.all_grids:
            allowed.setdefault(grid.sheet, set()).add(f"B{grid.state_message_row}")
        return allowed

    @property
    def defined_names(self) -> dict[str, str]:
        """Every workbook-level name this contract owns, as name -> reference."""
        names: dict[str, str] = {}
        sheet = self.setup_sheet
        for field_ in self.structural_fields:
            names[field_.defined_name] = _absolute(sheet, field_.cell)
        names[self.structural_state.defined_name] = _absolute(
            self.structural_state.sheet, self.structural_state.cell
        )
        for counter in self.counters:
            names[counter.defined_name] = _absolute(self.identity_sheet, counter.cell)
        return names

    def alias_defined_names(self, inputs: InputContract) -> dict[str, str]:
        """The nm*_Entered aliases, resolved against the input contract's own cells."""
        resolved: dict[str, str] = {}
        for alias in self.entered_aliases:
            spec = inputs.inputs[alias.input_key]
            resolved[alias.defined_name] = _absolute(spec.sheet, spec.cell)
        return resolved

    def counter_by_key(self, key: str) -> Counter:
        for counter in self.counters:
            if counter.key == key:
                return counter
        raise StructureContractError(f"no counter with key {key!r}")

    def grid_for_sheet(self, sheet: str) -> Grid | None:
        for grid in self.all_grids:
            if grid.sheet == sheet:
                return grid
        return None

    def button_for(self, entry_point: str) -> ButtonSpec:
        for button in self.buttons:
            if button.entry_point == entry_point:
                return button
        raise StructureContractError(f"no button bound to {entry_point!r}")


# ---------------------------------------------------------------------------
# loading
# ---------------------------------------------------------------------------
def load_structure_contract(path: str | Path) -> StructureContract:
    path = Path(path)
    if not path.is_file():
        raise StructureContractError(f"structure contract not found: {path}")

    with path.open("r", encoding="utf-8") as handle:
        raw = yaml.safe_load(handle)
    if not isinstance(raw, dict):
        raise StructureContractError(f"{path}: contract root must be a mapping")

    where = str(path)
    version = _req_str(raw, "structure_contract_version", where)
    limits = _parse_limits(_req(raw, "limits", where), where)
    timeline = _req(raw, "timeline", where)
    identity = _req(raw, "identity", where)
    raw_grids = _req(raw, "grids", where)
    messages = _req(raw, "state_messages", where)
    raw_buttons = _req(raw, "buttons", where)
    raw_vba = _req(raw, "vba", where)
    checks = _req(raw, "structural_checks", where)

    applied_block = _req(timeline, "applied_block", f"{where}: timeline")
    aliases = [
        EnteredAlias(
            defined_name=_req_str(entry, "defined_name", f"{where}: entered_aliases[{i}]"),
            input_key=_req_str(entry, "input_key", f"{where}: entered_aliases[{i}]"),
        )
        for i, entry in enumerate(_req(timeline, "entered_aliases", f"{where}: timeline"))
    ]
    applied = [
        _parse_field(entry, f"{where}: timeline.applied[{i}]", require_formula=False)
        for i, entry in enumerate(_req(timeline, "applied", f"{where}: timeline"))
    ]
    derived = [
        _parse_field(entry, f"{where}: timeline.derived[{i}]", require_formula=True)
        for i, entry in enumerate(_req(timeline, "derived", f"{where}: timeline"))
    ]

    contract = StructureContract(
        version=version,
        limits=limits,
        entered_aliases=aliases,
        applied_block=applied_block,
        applied=applied,
        derived=derived,
        structural_state=_parse_state(_req(raw, "structural_state", where), where),
        identity_block=identity,
        counters=[
            _parse_counter(entry, f"{where}: identity.counters[{i}]")
            for i, entry in enumerate(_req(identity, "counters", f"{where}: identity"))
        ],
        grids={
            key: _parse_grid(key, entry, f"{where}: grids.{key}")
            for key, entry in raw_grids.items()
        },
        state_messages={str(k): str(v) for k, v in messages.items()},
        buttons=_parse_buttons(raw_buttons, where),
        vba_modules=[
            VbaModule(
                name=_req_str(entry, "name", f"{where}: vba.modules[{i}]"),
                generated=bool(entry.get("generated", False)),
                responsibility=_req_str(entry, "responsibility", f"{where}: vba.modules[{i}]"),
            )
            for i, entry in enumerate(_req(raw_vba, "modules", f"{where}: vba"))
        ],
        vba_source_dir=_req_str(raw_vba, "source_dir", f"{where}: vba"),
        vba_generated_dir=_req_str(raw_vba, "generated_dir", f"{where}: vba"),
        vba_generated_module=_req_str(raw_vba, "generated_module", f"{where}: vba"),
        entry_points=[str(e) for e in _req(raw_vba, "entry_points", f"{where}: vba")],
        # The Phase-5 calculation endpoints. Optional so a contract predating
        # Phase 5 still loads; the static tests are what require them to be
        # declared once they exist.
        api_procedures=[str(e) for e in raw_vba.get("api_procedures", [])],
        forbidden_constructs=[
            _forbidden_construct_text(c, i, f"{where}: vba")
            for i, c in enumerate(_req(raw_vba, "forbidden_constructs", f"{where}: vba"))
        ],
        forbidden_construct_rules=[
            _parse_forbidden_construct(c, i, f"{where}: vba")
            for i, c in enumerate(_req(raw_vba, "forbidden_constructs", f"{where}: vba"))
        ],
        structural_checks=[dict(c) for c in checks],
        source_path=path,
    )

    _validate_limits(contract, path)
    _validate_names(contract, path)
    _validate_grids(contract, path)
    _validate_identity(contract, path)
    _validate_state(contract, path)
    _validate_buttons_and_vba(contract, path)
    _validate_excel_bounds(contract, path)
    _validate_no_overlap(contract, path)
    return contract


# ---------------------------------------------------------------------------
# parsing
# ---------------------------------------------------------------------------
def _parse_limits(entry: Any, where: str) -> Limits:
    if not isinstance(entry, dict):
        raise StructureContractError(f"{where}: limits must be a mapping")
    values = {}
    for key in ("min_year", "max_year", "max_generated_year_columns"):
        value = _req(entry, key, f"{where}: limits")
        if not isinstance(value, int) or isinstance(value, bool) or value < 1:
            raise StructureContractError(
                f"{where}: limits.{key} must be a positive integer, got {value!r}"
            )
        values[key] = value
    return Limits(**values)


def _parse_field(entry: Any, where: str, require_formula: bool) -> StructuralField:
    if not isinstance(entry, dict):
        raise StructureContractError(f"{where}: must be a mapping")
    formula = entry.get("formula")
    if require_formula:
        if not isinstance(formula, str) or not formula.startswith("="):
            raise StructureContractError(
                f"{where}: a derived structural field must declare a formula starting "
                f"with '=', got {formula!r}"
            )
        formula = _normalise_formula(formula)
    elif formula is not None:
        raise StructureContractError(
            f"{where}: an applied-timeline field is model-written state and must not "
            "declare a formula; only derived structural display cells may."
        )
    return StructuralField(
        key=_req_key(entry, "key", where),
        label=_req_str(entry, "label", where),
        defined_name=_req_name(entry, "defined_name", where),
        label_cell=_req_str(entry, "label_cell", where),
        cell=_req_str(entry, "cell", where),
        number_format=_req_str(entry, "number_format", where),
        note=entry.get("note"),
        formula=formula,
    )


def _parse_state(entry: Any, where: str) -> StructuralState:
    where = f"{where}: structural_state"
    if not isinstance(entry, dict):
        raise StructureContractError(f"{where}: must be a mapping")
    labels = _req(entry, "labels", where)
    for key in ("not_applied", "pending", "current"):
        if not isinstance(labels.get(key), str) or not labels[key].strip():
            raise StructureContractError(f"{where}: labels.{key} must be a non-empty string")
    return StructuralState(
        sheet=_req_str(entry, "sheet", where),
        label=_req_str(entry, "label", where),
        defined_name=_req_name(entry, "defined_name", where),
        label_cell=_req_str(entry, "label_cell", where),
        cell=_req_str(entry, "cell", where),
        number_format=_req_str(entry, "number_format", where),
        note=entry.get("note"),
        note_column=_req_str(entry, "note_column", where),
        labels={str(k): str(v) for k, v in labels.items()},
        formula=_normalise_formula(_req_str(entry, "formula", where)),
    )


def _parse_counter(entry: Any, where: str) -> Counter:
    if not isinstance(entry, dict):
        raise StructureContractError(f"{where}: must be a mapping")
    initial = _req(entry, "initial", where)
    if not isinstance(initial, int) or isinstance(initial, bool) or initial < 0:
        raise StructureContractError(
            f"{where}: initial must be a non-negative integer, got {initial!r}"
        )
    pad = _req(entry, "pad_width", where)
    if not isinstance(pad, int) or isinstance(pad, bool) or pad < 1:
        raise StructureContractError(f"{where}: pad_width must be a positive integer")
    return Counter(
        key=_req_key(entry, "key", where),
        label=_req_str(entry, "label", where),
        defined_name=_req_name(entry, "defined_name", where),
        label_cell=_req_str(entry, "label_cell", where),
        cell=_req_str(entry, "cell", where),
        number_format=_req_str(entry, "number_format", where),
        initial=initial,
        prefix=_req_str(entry, "prefix", where),
        pad_width=pad,
        pattern=_req_str(entry, "pattern", where),
        driver_register=_req_str(entry, "driver_register", where),
        note=entry.get("note"),
    )


def _parse_grid(key: str, entry: Any, where: str) -> Grid:
    if not isinstance(entry, dict):
        raise StructureContractError(f"{where}: must be a mapping")

    table_name = _req_str(entry, "table_name", where)
    if not TABLE_NAME_RE.match(table_name):
        raise StructureContractError(
            f"{where}: table_name {table_name!r} must match tbl<PascalCase>"
        )
    kind = _req_str(entry, "kind", where)
    if kind not in VALID_GRID_KINDS:
        raise StructureContractError(f"{where}: kind {kind!r} must be one of {VALID_GRID_KINDS}")

    reserved = _req(entry, "reserved_rows", where)
    if not isinstance(reserved, int) or isinstance(reserved, bool) or reserved < 1:
        raise StructureContractError(
            f"{where}: reserved_rows must be a positive integer (initial capacity, "
            f"not a business maximum), got {reserved!r}"
        )

    raw_columns = _req(entry, "fixed_columns", where)
    if not isinstance(raw_columns, list) or not raw_columns:
        raise StructureContractError(f"{where}: fixed_columns must be a non-empty list")

    return Grid(
        key=key,
        sheet=_req_str(entry, "sheet", where),
        table_name=table_name,
        kind=kind,
        section=_req_str(entry, "section", where),
        section_row=_positive_int(entry, "section_row", where),
        note=entry.get("note"),
        note_row=entry.get("note_row"),
        intro=entry.get("intro"),
        intro_row=entry.get("intro_row"),
        state_message_row=_positive_int(entry, "state_message_row", where),
        header_row=_positive_int(entry, "header_row", where),
        first_column=_checked(
            check_column, _req_str(entry, "first_column", where), f"{where}: first_column"
        ),
        reserved_rows=reserved,
        freeze_panes=entry.get("freeze_panes"),
        fixed_columns=[
            _parse_grid_column(c, f"{where}: fixed_columns[{i}]")
            for i, c in enumerate(raw_columns)
        ],
        year_column=_parse_year_column(_req(entry, "year_column", where), f"{where}: year_column"),
        driver_register=entry.get("driver_register"),
        key_driver_column=entry.get("key_driver_column"),
        source_list_table=entry.get("source_list_table"),
    )


def _parse_grid_column(entry: Any, where: str) -> GridColumn:
    if not isinstance(entry, dict):
        raise StructureContractError(f"{where}: must be a mapping")
    width = entry.get("width")
    if not isinstance(width, (int, float)) or isinstance(width, bool) or width <= 0:
        raise StructureContractError(f"{where}: width must be a positive number")
    return GridColumn(
        key=_req_key(entry, "key", where),
        header=_req_str(entry, "header", where),
        number_format=_req_str(entry, "number_format", where),
        width=float(width),
        note=entry.get("note"),
        source_driver_column=entry.get("source_driver_column"),
        source_list_column=entry.get("source_list_column"),
    )


def _parse_year_column(entry: Any, where: str) -> YearColumn:
    if not isinstance(entry, dict):
        raise StructureContractError(f"{where}: must be a mapping")
    width = entry.get("width")
    if not isinstance(width, (int, float)) or isinstance(width, bool) or width <= 0:
        raise StructureContractError(f"{where}: width must be a positive number")
    initial = entry.get("initial_value")
    if initial is not None and (isinstance(initial, bool) or not isinstance(initial, (int, float))):
        raise StructureContractError(
            f"{where}: initial_value must be a number or null (null means the cell is "
            f"left genuinely blank), got {initial!r}"
        )
    return YearColumn(
        header_format=_req_str(entry, "header_format", where),
        number_format=_req_str(entry, "number_format", where),
        width=float(width),
        initial_value=None if initial is None else float(initial),
        validation=entry.get("validation"),
    )


def _parse_buttons(entry: Any, where: str) -> list[ButtonSpec]:
    where = f"{where}: buttons"
    if not isinstance(entry, dict):
        raise StructureContractError(f"{where}: must be a mapping")
    width = _req(entry, "width", where)
    height = _req(entry, "height", where)
    for label, value in (("width", width), ("height", height)):
        if not isinstance(value, (int, float)) or isinstance(value, bool) or value <= 0:
            raise StructureContractError(f"{where}: {label} must be a positive number")

    buttons: list[ButtonSpec] = []
    for index, item in enumerate(_req(entry, "definitions", where)):
        at = f"{where}: definitions[{index}]"
        if not isinstance(item, dict):
            raise StructureContractError(f"{at}: must be a mapping")
        shape_name = _req_str(item, "shape_name", at)
        if not SHAPE_NAME_RE.match(shape_name):
            raise StructureContractError(
                f"{at}: shape_name {shape_name!r} must match btnPCCM<PascalCase>"
            )
        entry_point = _req_str(item, "entry_point", at)
        if not ENTRY_POINT_RE.match(entry_point):
            raise StructureContractError(
                f"{at}: entry_point {entry_point!r} must match PCCM_<PascalCase>"
            )
        buttons.append(
            ButtonSpec(
                key=_req_key(item, "key", at),
                sheet=_req_str(item, "sheet", at),
                shape_name=shape_name,
                caption=_req_str(item, "caption", at),
                entry_point=entry_point,
                anchor_cell=_req_str(item, "anchor_cell", at),
                width=float(width),
                height=float(height),
            )
        )
    return buttons


# ---------------------------------------------------------------------------
# validation
# ---------------------------------------------------------------------------
ARCHITECTURE_MAX_GENERATED_YEAR_COLUMNS = 200
"""Architecture Lock Revision B: generated project-year column count > 200 = ERROR.

A locked structural constant in its own right. It is deliberately NOT derived from
the calendar-year window: the two protections guard different things -- this one
bounds Duration and the width of the profiling grids, while min_year/max_year bound
Base Year, Start Year, Last Project Year and therefore the inflation span. Deriving
one from the other silently imposed a 301-column project guard and would equally
have imposed a 200-year cap on the inflation span had it been done the other way.
"""


def _validate_limits(contract: StructureContract, path: Path) -> None:
    limits = contract.limits
    if limits.min_year >= limits.max_year:
        raise StructureContractError(
            f"{path}: limits.min_year ({limits.min_year}) must be below "
            f"limits.max_year ({limits.max_year})"
        )
    if limits.max_generated_year_columns != ARCHITECTURE_MAX_GENERATED_YEAR_COLUMNS:
        raise StructureContractError(
            f"{path}: limits.max_generated_year_columns is "
            f"{limits.max_generated_year_columns}, but the Architecture Lock Revision B "
            f"structural protection on generated project-year columns is "
            f"{ARCHITECTURE_MAX_GENERATED_YEAR_COLUMNS} "
            '("Generated column count > 200 = ERROR"). It is an independent locked '
            "constant, not a function of the calendar-year window."
        )
    # The widest grid that guard permits must still fit the Excel column limit.
    for grid in contract.all_grids:
        last = grid.first_col_index + len(grid.fixed_columns) + limits.max_generated_year_columns - 1
        if last > EXCEL_MAX_COLUMN:
            raise StructureContractError(
                f"{path}: {grid.table_name} with {len(grid.fixed_columns)} fixed columns plus "
                f"{limits.max_generated_year_columns} generated year columns would reach column "
                f"index {last}, beyond the Excel maximum of XFD ({EXCEL_MAX_COLUMN})"
            )


def _validate_names(contract: StructureContract, path: Path) -> None:
    names = [f.defined_name for f in contract.structural_fields]
    names.append(contract.structural_state.defined_name)
    names += [c.defined_name for c in contract.counters]
    names += [a.defined_name for a in contract.entered_aliases]
    duplicates = sorted({n for n in names if names.count(n) > 1})
    if duplicates:
        raise StructureContractError(f"{path}: duplicate structural defined names: {duplicates}")
    for name in names:
        if not STRUCTURE_NAME_RE.match(name):
            raise StructureContractError(
                f"{path}: structural defined name {name!r} must match the nm<PascalCase> "
                "convention, so structural names never collide with inp*, lst* or tbl*"
            )

    keys = [f.key for f in contract.structural_fields]
    duplicates = sorted({k for k in keys if keys.count(k) > 1})
    if duplicates:
        raise StructureContractError(f"{path}: duplicate structural field keys: {duplicates}")


def _validate_grids(contract: StructureContract, path: Path) -> None:
    tables = [g.table_name for g in contract.all_grids]
    duplicates = sorted({t for t in tables if tables.count(t) > 1})
    if duplicates:
        raise StructureContractError(f"{path}: duplicate grid table names: {duplicates}")

    sheets = [g.sheet for g in contract.all_grids]
    duplicates = sorted({s for s in sheets if sheets.count(s) > 1})
    if duplicates:
        raise StructureContractError(f"{path}: more than one grid targets sheet(s): {duplicates}")

    for grid in contract.all_grids:
        headers = grid.headers
        duplicates = sorted({h for h in headers if headers.count(h) > 1})
        if duplicates:
            raise StructureContractError(
                f"{path}: {grid.table_name} has duplicate fixed-column headers: {duplicates}"
            )
        keys = [c.key for c in grid.fixed_columns]
        duplicates = sorted({k for k in keys if keys.count(k) > 1})
        if duplicates:
            raise StructureContractError(
                f"{path}: {grid.table_name} has duplicate fixed-column keys: {duplicates}"
            )

        if grid.kind == "profiling":
            if not grid.driver_register or not grid.key_driver_column:
                raise StructureContractError(
                    f"{path}: {grid.table_name} is a profiling grid and must declare both "
                    "driver_register and key_driver_column"
                )
            if grid.fixed_columns[0].key != grid.key_driver_column:
                raise StructureContractError(
                    f"{path}: {grid.table_name} key_driver_column is "
                    f"{grid.key_driver_column!r} but the first fixed column is "
                    f"{grid.fixed_columns[0].key!r}; the permanent key must come first"
                )
            for column in grid.fixed_columns:
                if not column.source_driver_column:
                    raise StructureContractError(
                        f"{path}: {grid.table_name} fixed column {column.key!r} must declare "
                        "source_driver_column; a profiling grid holds trace copies only and "
                        "never originates a driver attribute"
                    )
            if grid.year_column.initial_value is None:
                raise StructureContractError(
                    f"{path}: {grid.table_name} year_column.initial_value is null. A new "
                    "project-year profiling cell is initialised to 0%, so an existing row's "
                    "total is unchanged by a duration increase."
                )
        else:
            if not grid.source_list_table:
                raise StructureContractError(
                    f"{path}: {grid.table_name} is an inflation grid and must declare "
                    "source_list_table"
                )
            if grid.year_column.initial_value is not None:
                raise StructureContractError(
                    f"{path}: {grid.table_name} year_column.initial_value is "
                    f"{grid.year_column.initial_value!r}. A newly required annual escalation "
                    "year must be left BLANK: seeding zero fabricates an assumption the user "
                    "never made and would hide a missing rate from Model Check."
                )


def _validate_identity(contract: StructureContract, path: Path) -> None:
    keys = [c.key for c in contract.counters]
    duplicates = sorted({k for k in keys if keys.count(k) > 1})
    if duplicates:
        raise StructureContractError(f"{path}: duplicate counter keys: {duplicates}")

    prefixes = [c.prefix for c in contract.counters]
    duplicates = sorted({p for p in prefixes if prefixes.count(p) > 1})
    if duplicates:
        raise StructureContractError(
            f"{path}: counters share ID prefix(es) {duplicates}; each register's sequence "
            "must be independently identifiable"
        )

    registers = [c.driver_register for c in contract.counters]
    duplicates = sorted({r for r in registers if registers.count(r) > 1})
    if duplicates:
        raise StructureContractError(
            f"{path}: more than one counter claims driver register(s) {duplicates}"
        )

    for counter in contract.counters:
        try:
            compiled = re.compile(counter.pattern)
        except re.error as error:
            raise StructureContractError(
                f"{path}: counter {counter.key!r} pattern is not a valid regular "
                f"expression: {error}"
            ) from error
        sample = counter.format_id(1)
        if not compiled.match(sample):
            raise StructureContractError(
                f"{path}: counter {counter.key!r} would allocate {sample!r}, which does not "
                f"match its own declared pattern {counter.pattern!r}"
            )
        # A four-digit sequence must remain legal: the pad width is a display floor,
        # never a range limit, so CL-1000 has to pass the same pattern.
        wide = counter.format_id(1000)
        if not compiled.match(wide):
            raise StructureContractError(
                f"{path}: counter {counter.key!r} pattern {counter.pattern!r} rejects "
                f"{wide!r}. Pad width is a minimum display width, not a maximum ID range; "
                "sequences beyond the padded width must remain valid."
            )
        if not counter.prefix.endswith("-"):
            raise StructureContractError(
                f"{path}: counter {counter.key!r} prefix {counter.prefix!r} must end with "
                "'-' so the sequence is unambiguously separable"
            )


def _validate_state(contract: StructureContract, path: Path) -> None:
    state = contract.structural_state
    formula = state.formula
    for key, label in state.labels.items():
        if f'"{label}"' not in formula:
            raise StructureContractError(
                f"{path}: structural_state.labels.{key} is {label!r} but that literal does "
                "not appear in the structural state formula; the declared vocabulary and the "
                "formula that produces it must not drift apart"
            )
    alias_names = {a.defined_name for a in contract.entered_aliases}
    applied_names = {f.defined_name for f in contract.applied}
    for name in alias_names | applied_names:
        if name not in formula:
            raise StructureContractError(
                f"{path}: structural state formula does not reference {name!r}. The pending "
                "flag must compare every entered value against its applied counterpart."
            )

    known = set(contract.defined_names) | alias_names
    for field_ in contract.derived:
        for name in re.findall(r"\bnm[A-Za-z0-9_]+", field_.formula):
            if name not in known:
                raise StructureContractError(
                    f"{path}: derived field {field_.key!r} references unknown structural "
                    f"name {name!r}"
                )

    for message_key in ("not_applied", "inflation_empty_span"):
        if message_key not in contract.state_messages:
            raise StructureContractError(f"{path}: state_messages.{message_key} is required")
    for formula_key in ("profiling_formula", "inflation_formula"):
        value = contract.state_messages.get(formula_key)
        if not isinstance(value, str) or not value.startswith("="):
            raise StructureContractError(
                f"{path}: state_messages.{formula_key} must be a formula starting with '='"
            )
    if contract.state_messages["not_applied"] not in contract.state_messages["profiling_formula"]:
        raise StructureContractError(
            f"{path}: the profiling state formula does not emit the declared "
            "state_messages.not_applied text"
        )
    inflation_formula = contract.state_messages["inflation_formula"]
    for message_key in ("not_applied", "inflation_empty_span"):
        if contract.state_messages[message_key] not in inflation_formula:
            raise StructureContractError(
                f"{path}: the inflation state formula does not emit the declared "
                f"state_messages.{message_key} text"
            )


def _validate_buttons_and_vba(contract: StructureContract, path: Path) -> None:
    names = [b.shape_name for b in contract.buttons]
    duplicates = sorted({n for n in names if names.count(n) > 1})
    if duplicates:
        raise StructureContractError(f"{path}: duplicate button shape names: {duplicates}")

    bound = [b.entry_point for b in contract.buttons]
    duplicates = sorted({e for e in bound if bound.count(e) > 1})
    if duplicates:
        raise StructureContractError(
            f"{path}: entry point(s) {duplicates} are bound to more than one button"
        )
    declared = list(contract.entry_points)
    if sorted(bound) != sorted(declared):
        raise StructureContractError(
            f"{path}: button bindings and vba.entry_points disagree.\n"
            f"  buttons bind : {sorted(bound)}\n"
            f"  vba declares : {sorted(declared)}"
        )
    for name in declared:
        if not ENTRY_POINT_RE.match(name):
            raise StructureContractError(
                f"{path}: entry point {name!r} must match PCCM_<PascalCase>"
            )

    modules = [m.name for m in contract.vba_modules]
    duplicates = sorted({m for m in modules if modules.count(m) > 1})
    if duplicates:
        raise StructureContractError(f"{path}: duplicate VBA module names: {duplicates}")
    for name in modules:
        if not MODULE_NAME_RE.match(name):
            raise StructureContractError(
                f"{path}: VBA module {name!r} must match the mod<PascalCase> convention"
            )
    generated = [m.name for m in contract.vba_modules if m.generated]
    if contract.vba_generated_module not in generated:
        raise StructureContractError(
            f"{path}: vba.generated_module is {contract.vba_generated_module!r} but that "
            f"module is not declared generated; the contract marks {generated}"
        )
    if sorted(generated) != sorted(GENERATED_MODULES):
        # The DEPLOYMENT invariant, generalised exactly as far as the build has
        # gone and no further. Phase 5 emits a second generated module, so "exactly
        # one" is no longer the rule - but "whatever the builder actually emits,
        # and nothing else" still is. The set is statically locked here rather
        # than left open, so a module cannot become generated without a deliberate
        # change to the builder's own list.
        raise StructureContractError(
            f"{path}: the generated module set is locked to {sorted(GENERATED_MODULES)}; "
            f"the contract marks {sorted(generated)}"
        )
    if not contract.forbidden_constructs:
        raise StructureContractError(
            f"{path}: vba.forbidden_constructs must not be empty; the static tests rely on "
            "it to keep later-phase functionality out of Phase 4"
        )
    _validate_forbidden_constructs(contract, path)


def _validate_excel_bounds(contract: StructureContract, path: Path) -> None:
    for field_ in contract.structural_fields:
        _checked(check_cell, field_.label_cell, f"{path}: {field_.key} label_cell")
        _checked(check_cell, field_.cell, f"{path}: {field_.key} cell")
    state = contract.structural_state
    _checked(check_cell, state.label_cell, f"{path}: structural_state label_cell")
    _checked(check_cell, state.cell, f"{path}: structural_state cell")
    _checked(check_column, state.note_column, f"{path}: structural_state note_column")

    for counter in contract.counters:
        _checked(check_cell, counter.label_cell, f"{path}: counter {counter.key} label_cell")
        _checked(check_cell, counter.cell, f"{path}: counter {counter.key} cell")

    for grid in contract.all_grids:
        where = f"{path}: {grid.table_name}"
        _checked(check_row, grid.header_row, f"{where} header_row")
        _checked(check_row, grid.last_data_row, f"{where} last data row")
        _checked(check_row, grid.section_row, f"{where} section_row")
        _checked(check_row, grid.state_message_row, f"{where} state_message_row")
        if grid.note_row is not None:
            _checked(check_row, grid.note_row, f"{where} note_row")
        if grid.intro_row is not None:
            _checked(check_row, grid.intro_row, f"{where} intro_row")

    for button in contract.buttons:
        _checked(check_cell, button.anchor_cell, f"{path}: button {button.key} anchor_cell")


def _validate_no_overlap(contract: StructureContract, path: Path) -> None:
    """Nothing this contract renders may claim the same cell twice."""
    claimed: dict[str, dict[str, str]] = {}

    def claim(sheet: str, address: str, owner: str) -> None:
        cells = claimed.setdefault(sheet, {})
        if address in cells:
            raise StructureContractError(
                f"{path}: cell {sheet}!{address} claimed by both {cells[address]} and {owner}"
            )
        cells[address] = owner

    setup = contract.setup_sheet
    for field_ in contract.structural_fields:
        claim(setup, field_.label_cell, f"label of {field_.key}")
        claim(setup, field_.cell, f"value of {field_.key}")
    state = contract.structural_state
    claim(state.sheet, state.label_cell, "structural state label")
    claim(state.sheet, state.cell, "structural state value")
    block = contract.applied_block
    claim(setup, f"{block['label_column']}{block['section_row']}", "applied timeline section")
    if block.get("note_row"):
        claim(setup, f"{block['label_column']}{block['note_row']}", "applied timeline note")

    identity = contract.identity_block
    sheet = contract.identity_sheet
    claim(sheet, f"{identity['label_column']}{identity['section_row']}", "identity section")
    if identity.get("intro_row"):
        claim(sheet, f"{identity['label_column']}{identity['intro_row']}", "identity intro")
    for counter in contract.counters:
        claim(sheet, counter.label_cell, f"label of counter {counter.key}")
        claim(sheet, counter.cell, f"value of counter {counter.key}")

    for grid in contract.all_grids:
        claim(grid.sheet, f"B{grid.section_row}", f"section of {grid.table_name}")
        claim(grid.sheet, f"B{grid.state_message_row}", f"state message of {grid.table_name}")
        if grid.note_row is not None:
            claim(grid.sheet, f"B{grid.note_row}", f"note of {grid.table_name}")
        if grid.intro_row is not None:
            claim(grid.sheet, f"B{grid.intro_row}", f"intro of {grid.table_name}")
        for index in range(len(grid.fixed_columns)):
            letter = grid.column_letter(index)
            for row in range(grid.header_row, grid.last_data_row + 1):
                claim(grid.sheet, f"{letter}{row}", grid.table_name)


# ---------------------------------------------------------------------------
# cross-contract validation
# ---------------------------------------------------------------------------
def validate_structure_against(
    structure: StructureContract,
    inputs: InputContract,
    drivers: DriverContract,
    path: Path | None = None,
) -> None:
    """Assert the structure contract agrees with the input and driver contracts."""
    path = path or structure.source_path

    # --- entered aliases resolve to the accepted Phase-2 input cells ---------
    for alias in structure.entered_aliases:
        spec = inputs.inputs.get(alias.input_key)
        if spec is None:
            raise StructureContractError(
                f"{path}: entered alias {alias.defined_name} points at input "
                f"{alias.input_key!r}, which the input contract does not declare"
            )
        if spec.sheet != structure.setup_sheet:
            raise StructureContractError(
                f"{path}: entered alias {alias.defined_name} targets input "
                f"{alias.input_key!r} on sheet {spec.sheet!r}, not the applied-timeline "
                f"sheet {structure.setup_sheet!r}"
            )
        if not spec.editable:
            raise StructureContractError(
                f"{path}: entered alias {alias.defined_name} targets input "
                f"{alias.input_key!r}, which is model-controlled. The entered timeline is "
                "user-owned by definition."
            )
        if spec.type != "integer":
            raise StructureContractError(
                f"{path}: entered alias {alias.defined_name} targets input "
                f"{alias.input_key!r} of type {spec.type!r}; a timeline value must be an "
                "integer year or count"
            )

    # The alias must not shadow or replace the accepted inp* name: both point at the
    # same cell, so there is exactly one semantic source for each entered value.
    resolved = structure.alias_defined_names(inputs)
    for alias in structure.entered_aliases:
        spec = inputs.inputs[alias.input_key]
        existing = inputs.input_defined_names[spec.defined_name]
        if resolved[alias.defined_name] != existing:
            raise StructureContractError(
                f"{path}: entered alias {alias.defined_name} resolves to "
                f"{resolved[alias.defined_name]}, but the accepted input name "
                f"{spec.defined_name} resolves to {existing}. An alias must address the "
                "same cell, never a second copy of the value."
            )

    # --- no defined-name collision across the four contracts ----------------
    taken = set(inputs.input_defined_names) | set(inputs.list_defined_names)
    for name in list(structure.defined_names) + list(resolved):
        if name in taken:
            raise StructureContractError(
                f"{path}: structural defined name {name!r} is already declared by "
                f"{inputs.source_path.name}"
            )

    # --- table-name collision ----------------------------------------------
    existing_tables = {t.table_name for t in inputs.all_tables} | {
        r.table_name for r in drivers.all_registers
    }
    for grid in structure.all_grids:
        if grid.table_name in existing_tables:
            raise StructureContractError(
                f"{path}: grid table {grid.table_name!r} is already declared by another "
                "contract"
            )

    # --- profiling capacity must track its driver register ------------------
    for grid in structure.profiling_grids:
        register = drivers.registers.get(grid.driver_register)
        if register is None:
            raise StructureContractError(
                f"{path}: {grid.table_name} mirrors driver register "
                f"{grid.driver_register!r}, which the driver contract does not declare"
            )
        if grid.reserved_rows != register.reserved_rows:
            raise StructureContractError(
                f"{path}: {grid.table_name} reserves {grid.reserved_rows} rows but "
                f"{register.table_name} reserves {register.reserved_rows}. A profiling grid "
                "holds one row per driver, so the two capacities must not drift apart."
            )
        known = {c.key for c in register.columns}
        for column in grid.fixed_columns:
            if column.source_driver_column not in known:
                raise StructureContractError(
                    f"{path}: {grid.table_name} fixed column {column.key!r} traces driver "
                    f"column {column.source_driver_column!r}, which {register.table_name} "
                    f"does not declare. Known: {sorted(known)}"
                )
        identity = register.columns[0]
        if grid.key_driver_column != identity.key:
            raise StructureContractError(
                f"{path}: {grid.table_name} keys on {grid.key_driver_column!r} but "
                f"{register.table_name}'s permanent identifier is {identity.key!r}"
            )

    # --- inflation grid must track the Config profile master ----------------
    inflation = structure.inflation_grid
    source = next(
        (t for t in inputs.all_tables if t.table_name == inflation.source_list_table), None
    )
    if source is None:
        raise StructureContractError(
            f"{path}: {inflation.table_name} sources profile names from "
            f"{inflation.source_list_table!r}, which the input contract does not declare"
        )
    if inflation.reserved_rows != source.data_rows:
        raise StructureContractError(
            f"{path}: {inflation.table_name} reserves {inflation.reserved_rows} rows but "
            f"{source.table_name} reserves {source.data_rows}. The inflation grid holds one "
            "row per profile name, so the two capacities must not drift apart."
        )
    source_headers = {c.header for c in source.columns}
    for column in inflation.fixed_columns:
        if column.source_list_column not in source_headers:
            raise StructureContractError(
                f"{path}: {inflation.table_name} fixed column {column.key!r} sources "
                f"{column.source_list_column!r}, which {source.table_name} does not declare"
            )

    # --- counters must name real registers ----------------------------------
    for counter in structure.counters:
        if counter.driver_register not in drivers.registers:
            raise StructureContractError(
                f"{path}: counter {counter.key!r} claims driver register "
                f"{counter.driver_register!r}, which the driver contract does not declare"
            )

    # --- the applied block must sit clear of the input contract's Setup area -
    setup_rows = inputs.occupied_rows(structure.setup_sheet)
    if setup_rows:
        highest = max(setup_rows)
        structural_rows = [f.label_cell for f in structure.structural_fields]
        structural_rows += [f.cell for f in structure.structural_fields]
        structural_rows += [structure.structural_state.label_cell, structure.structural_state.cell]
        for address in structural_rows:
            row = _row_of(address)
            if row in setup_rows:
                raise StructureContractError(
                    f"{path}: structural cell {structure.setup_sheet}!{address} lands on row "
                    f"{row}, which the input contract already occupies. The applied block must "
                    f"sit below the Phase-2 Setup area (highest occupied row {highest})."
                )
        block = structure.applied_block
        for row in (block["section_row"], block.get("note_row")):
            if row is not None and row in setup_rows:
                raise StructureContractError(
                    f"{path}: the applied timeline block claims Setup row {row}, which the "
                    "input contract already occupies"
                )


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def _normalise_formula(formula: str) -> str:
    """Collapse the whitespace YAML folding introduces, without touching content.

    Contract formulas are wrapped across several lines for readability. YAML folds
    each break into a single space, which Excel tolerates but which makes two
    otherwise identical formulas compare unequal. Runs of whitespace collapse to one
    space and the ends are trimmed; nothing inside the formula is rewritten.
    """
    return re.sub(r"\s+", " ", formula).strip()


def _absolute(sheet: str, cell: str) -> str:
    letter = "".join(ch for ch in cell if ch.isalpha())
    row = "".join(ch for ch in cell if ch.isdigit())
    return f"'{sheet}'!${letter}${row}"


def _row_of(cell: str) -> int:
    return int("".join(ch for ch in cell if ch.isdigit()))


def _column_index(letter: str) -> int:
    index = 0
    for char in letter:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


def _req(mapping: Any, key: str, where: str) -> Any:
    if not isinstance(mapping, dict) or key not in mapping:
        raise StructureContractError(f"{where}: missing required key {key!r}")
    return mapping[key]


def _req_str(mapping: Any, key: str, where: str) -> str:
    value = _req(mapping, key, where)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        value = str(value)
    if not isinstance(value, str) or not value.strip():
        raise StructureContractError(f"{where}: {key!r} must be a non-empty string, got {value!r}")
    return value


def _req_key(mapping: Any, key: str, where: str) -> str:
    value = _req_str(mapping, key, where)
    if not KEY_RE.match(value):
        raise StructureContractError(f"{where}: {key} {value!r} must be lower_snake_case")
    return value


def _req_name(mapping: Any, key: str, where: str) -> str:
    return _req_str(mapping, key, where)


def _positive_int(mapping: dict[str, Any], key: str, where: str) -> int:
    value = _req(mapping, key, where)
    if not isinstance(value, int) or isinstance(value, bool) or value < 1:
        raise StructureContractError(f"{where}: {key!r} must be a positive integer, got {value!r}")
    return value


# ---------------------------------------------------------------------------
# Forbidden constructs - D6-11's mixed scalar-or-scoped schema
# ---------------------------------------------------------------------------
FORBIDDEN_CONSTRUCT_KEYS = frozenset({"construct", "allowed_in"})

FORBIDDEN_OWNER_WILDCARDS = ("*", "**", "all", "any", "all_modules", "*.bas", "?", "-")
"""Owner spellings that mean "everywhere" and therefore mean nothing.

A scoped exception exists to name ONE owner. A wildcard turns the rule into a
deletion of the rule, which is the failure mode this schema is built to make
impossible rather than merely discouraged."""


def _forbidden_construct_text(entry: Any, index: int, where: str) -> str:
    """The construct TEXT, whichever shape declared it.

    Existing consumers - the Stage-B manifest and the Phase-4 static scan - ask
    only "what strings must not appear", so they keep reading a flat list of
    strings and are unaffected by the schema extension.
    """
    return _parse_forbidden_construct(entry, index, where).construct


def _parse_forbidden_construct(entry: Any, index: int, where: str) -> ForbiddenConstruct:
    """Parse one entry in either accepted shape, refusing every ambiguous one."""
    at = f"{where}: forbidden_constructs[{index}]"

    if isinstance(entry, str):
        if not entry.strip():
            raise StructureContractError(f"{at}: must be a non-empty construct string")
        return ForbiddenConstruct(construct=entry)

    if not isinstance(entry, dict):
        raise StructureContractError(
            f"{at}: must be either a construct STRING (globally forbidden) or a MAPPING with "
            f"'construct' and 'allowed_in' (forbidden except in the declared owners), got "
            f"{type(entry).__name__}"
        )

    extra = set(entry) - FORBIDDEN_CONSTRUCT_KEYS
    if extra:
        raise StructureContractError(
            f"{at}: unknown key(s) {sorted(extra)}; the scoped shape is exactly "
            f"{sorted(FORBIDDEN_CONSTRUCT_KEYS)}"
        )
    missing = FORBIDDEN_CONSTRUCT_KEYS - set(entry)
    if missing:
        raise StructureContractError(
            f"{at}: the scoped shape requires {sorted(missing)}. A mapping carrying only one of "
            "the two keys is ambiguous: it reads as a scoped rule but grants or forbids nothing."
        )

    construct = entry["construct"]
    if not isinstance(construct, str) or not construct.strip():
        raise StructureContractError(f"{at}: 'construct' must be a non-empty string")

    owners = entry["allowed_in"]
    if not isinstance(owners, list):
        raise StructureContractError(
            f"{at}: 'allowed_in' must be a list of module names, got {type(owners).__name__}"
        )
    if not owners:
        raise StructureContractError(
            f"{at}: 'allowed_in' must name at least one owner. An empty exception list is the "
            "global shape written ambiguously - declare the bare string instead, so a reader "
            "cannot mistake it for a granted exception."
        )

    seen: set[str] = set()
    for owner in owners:
        if not isinstance(owner, str) or not owner.strip():
            raise StructureContractError(f"{at}: every 'allowed_in' entry must be a module name")
        if owner in seen:
            raise StructureContractError(
                f"{at}: duplicate owner {owner!r} in 'allowed_in'. A module owns a construct "
                "once or not at all."
            )
        seen.add(owner)
        if owner.strip().lower() in FORBIDDEN_OWNER_WILDCARDS:
            raise StructureContractError(
                f"{at}: owner {owner!r} is a wildcard. A scoped exception names ONE owning "
                "module; a wildcard silently deletes the rule it appears to scope."
            )

    return ForbiddenConstruct(construct=construct, allowed_in=tuple(owners))


def _validate_forbidden_constructs(contract: StructureContract, path: Path) -> None:
    """Every scoped owner must be a module the contract actually declares.

    This is what stops a scoped exception from pre-authorising code that does not
    exist yet: a construct cannot be granted to `modSimRng` before `modSimRng` is
    a declared module, so the grant and the module arrive together or not at all.
    """
    declared = {module.name for module in contract.vba_modules}
    seen: set[str] = set()
    for rule in contract.forbidden_construct_rules:
        if rule.construct in seen:
            raise StructureContractError(
                f"{path}: construct {rule.construct!r} is declared twice in "
                "vba.forbidden_constructs; two rules for one construct can disagree"
            )
        seen.add(rule.construct)
        unknown = [owner for owner in rule.allowed_in if owner not in declared]
        if unknown:
            raise StructureContractError(
                f"{path}: forbidden construct {rule.construct!r} is scoped to module(s) "
                f"{unknown}, which the contract does not declare. An exception granted to a "
                "module that does not exist authorises code nobody has written."
            )
