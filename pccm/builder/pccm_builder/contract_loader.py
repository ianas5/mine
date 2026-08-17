"""Load and validate the PCCM input contract.

The contract is the semantic authority for model inputs, list masters, tables,
defined names and validation. Like the structural manifest, it fails loudly and
never repairs an invalid specification silently.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml

from openpyxl.utils import get_column_letter

CELL_RE = re.compile(r"^([A-Z]{1,3})([1-9][0-9]{0,6})$")

# Excel worksheet limits. A reference is not valid merely because it matches the
# regex above: XFE1 and A1048577 both match and are both outside the grid.
EXCEL_MAX_ROW = 1_048_576
EXCEL_MAX_COLUMN = 16_384          # XFD
DEFINED_NAME_RE = re.compile(r"^(inp|lst)[A-Z][A-Za-z0-9]*$")
TABLE_NAME_RE = re.compile(r"^tbl[A-Z][A-Za-z0-9]*$")

VALID_TYPES = ("text", "integer", "decimal", "percentage")
VALID_VALIDATION_KINDS = ("list", "whole", "decimal")


class ContractError(Exception):
    """Raised when the input contract is invalid."""


def check_row(row: Any, where: str) -> int:
    """Central row-bound validator. Raises ContractError outside the Excel grid."""
    if not isinstance(row, int) or isinstance(row, bool) or row < 1:
        raise ContractError(f"{where}: row {row!r} must be a positive integer")
    if row > EXCEL_MAX_ROW:
        raise ContractError(
            f"{where}: row {row} exceeds the Excel maximum of {EXCEL_MAX_ROW}"
        )
    return row


def check_column(letter: Any, where: str) -> str:
    """Central column-bound validator."""
    if not isinstance(letter, str) or not letter.isalpha() or not letter.isupper():
        raise ContractError(f"{where}: column {letter!r} must be an upper-case column letter")
    index = _column_index(letter)
    if index > EXCEL_MAX_COLUMN:
        raise ContractError(
            f"{where}: column {letter} (index {index}) exceeds the Excel maximum of "
            f"XFD ({EXCEL_MAX_COLUMN})"
        )
    return letter


def check_cell(address: Any, where: str) -> str:
    """Central cell-reference validator: syntax AND Excel grid bounds."""
    if not isinstance(address, str) or not CELL_RE.match(address):
        raise ContractError(f"{where}: {address!r} is not a valid cell reference")
    match = CELL_RE.match(address)
    check_column(match.group(1), f"{where} ({address})")
    check_row(int(match.group(2)), f"{where} ({address})")
    return address


@dataclass(frozen=True)
class InputSpec:
    key: str
    sheet: str
    label: str
    defined_name: str
    label_cell: str
    cell: str
    type: str
    required: bool
    editable: bool
    default: Any
    number_format: str
    validation: dict[str, Any] | None
    note: str | None


@dataclass(frozen=True)
class TableColumnSpec:
    header: str
    number_format: str
    validation: dict[str, Any] | None


@dataclass(frozen=True)
class TableSpec:
    key: str
    sheet: str
    table_name: str
    header_row: int
    first_column: str
    data_rows: int
    editable: bool
    columns: list[TableColumnSpec]
    locked_seed_rows: int = 0
    seed_rows: list[list[Any]] = field(default_factory=list)
    defined_name: str | None = None
    section: str | None = None
    section_row: int | None = None
    note: str | None = None
    note_row: int | None = None

    @property
    def first_data_row(self) -> int:
        return self.header_row + 1

    @property
    def last_data_row(self) -> int:
        return self.header_row + self.data_rows

    @property
    def first_col_index(self) -> int:
        return _column_index(self.first_column)

    @property
    def last_column(self) -> str:
        return get_column_letter(self.first_col_index + len(self.columns) - 1)

    @property
    def ref(self) -> str:
        return f"{self.first_column}{self.header_row}:{self.last_column}{self.last_data_row}"

    def column_letter(self, index: int) -> str:
        return get_column_letter(self.first_col_index + index)

    @property
    def first_user_row(self) -> int:
        """First data row the user owns. Locked identity rows come before it."""
        return self.first_data_row + self.locked_seed_rows

    def is_locked_row(self, offset: int) -> bool:
        """offset is 0-based within the data body."""
        return (not self.editable) or offset < self.locked_seed_rows

    def data_range(self, index: int = 0) -> str:
        letter = self.column_letter(index)
        return f"{letter}{self.first_data_row}:{letter}{self.last_data_row}"

    def user_data_range(self, index: int = 0) -> str | None:
        """Range of user-owned rows only, or None if the table has none.

        Data validation targets this, never the locked identity rows.
        """
        if not self.editable or self.first_user_row > self.last_data_row:
            return None
        letter = self.column_letter(index)
        return f"{letter}{self.first_user_row}:{letter}{self.last_data_row}"

    def absolute_data_range(self, index: int = 0) -> str:
        letter = self.column_letter(index)
        return f"'{self.sheet}'!${letter}${self.first_data_row}:${letter}${self.last_data_row}"


@dataclass(frozen=True)
class SectionSpec:
    title: str
    row: int
    inputs: list[str] = field(default_factory=list)
    note: str | None = None
    note_row: int | None = None
    convention_row: int | None = None
    table: str | None = None


@dataclass(frozen=True)
class InputContract:
    contract_version: str
    conventions: dict[str, Any]
    model_invariants: dict[str, Any]
    inputs: dict[str, InputSpec]
    setup_sheet: str
    setup_intro: dict[str, Any]
    setup_sections: list[SectionSpec]
    config_sheet: str
    config_intro: dict[str, Any]
    tables: dict[str, TableSpec]
    config_tables: list[TableSpec]
    source_path: Path

    @property
    def fx_convention(self) -> str:
        return self.conventions["fx_convention"]

    @property
    def reporting_currency(self) -> str:
        return self.model_invariants["reporting_currency"]

    def table_by_name(self, name: str) -> TableSpec | None:
        for table in self.all_tables:
            if table.table_name == name:
                return table
        return None

    @property
    def contract_sheets(self) -> set[str]:
        return {self.setup_sheet, self.config_sheet}

    def inputs_for_sheet(self, sheet: str) -> list[InputSpec]:
        return [i for i in self.inputs.values() if i.sheet == sheet]

    def tables_for_sheet(self, sheet: str) -> list[TableSpec]:
        return [t for t in self.all_tables if t.sheet == sheet]

    @property
    def all_tables(self) -> list[TableSpec]:
        return list(self.tables.values()) + list(self.config_tables)

    @property
    def list_defined_names(self) -> dict[str, str]:
        return {t.defined_name: t.absolute_data_range() for t in self.config_tables if t.defined_name}

    @property
    def input_defined_names(self) -> dict[str, str]:
        return {
            spec.defined_name: f"'{spec.sheet}'!${_col(spec.cell)}${_row(spec.cell)}"
            for spec in self.inputs.values()
        }


# ---------------------------------------------------------------------------
# loading
# ---------------------------------------------------------------------------
def load_contract(path: str | Path) -> InputContract:
    path = Path(path)
    if not path.is_file():
        raise ContractError(f"input contract not found: {path}")

    with path.open("r", encoding="utf-8") as handle:
        raw = yaml.safe_load(handle)
    if not isinstance(raw, dict):
        raise ContractError(f"{path}: contract root must be a mapping")

    where = str(path)
    contract_version = _req_str(raw, "contract_version", where)
    conventions = _req(raw, "conventions", where)
    for key in ("input_prefix", "list_prefix", "table_prefix", "fx_convention"):
        _req_str(conventions, key, f"{where}: conventions")

    invariants = _req(raw, "model_invariants", where)
    for key in ("reporting_currency", "reporting_currency_input", "reporting_currency_defined_name"):
        _req_str(invariants, key, f"{where}: model_invariants")
    identities = _req(invariants, "locked_identities", f"{where}: model_invariants")
    if not isinstance(identities, list) or not identities:
        raise ContractError(f"{where}: model_invariants.locked_identities must be a non-empty list")

    inputs = _parse_inputs(_req(raw, "inputs", where), path)
    tables = _parse_tables(_req(raw, "tables", where), path)
    config_tables = _parse_config_tables(_req(raw, "config_tables", where), path)

    setup_layout = _req(raw, "setup_layout", where)
    config_layout = _req(raw, "config_layout", where)
    setup_sheet = _req_str(setup_layout, "sheet", f"{where}: setup_layout")
    config_sheet = _req_str(config_layout, "sheet", f"{where}: config_layout")
    sections = _parse_sections(_req(setup_layout, "sections", where), path)

    contract = InputContract(
        contract_version=contract_version,
        conventions=conventions,
        model_invariants=invariants,
        inputs=inputs,
        setup_sheet=setup_sheet,
        setup_intro=_req(setup_layout, "intro", f"{where}: setup_layout"),
        setup_sections=sections,
        config_sheet=config_sheet,
        config_intro=_req(config_layout, "intro", f"{where}: config_layout"),
        tables=tables,
        config_tables=config_tables,
        source_path=path,
    )

    _validate_unique_names(contract, path)
    _validate_sections(contract, path)
    _validate_validation_sources(contract, path)
    _validate_no_cell_collisions(contract, path)
    _validate_seed_rows(contract, path)
    _validate_excel_bounds(contract, path)
    _validate_model_invariants(contract, path)
    return contract


def _parse_inputs(raw: Any, path: Path) -> dict[str, InputSpec]:
    if not isinstance(raw, dict) or not raw:
        raise ContractError(f"{path}: 'inputs' must be a non-empty mapping")
    result: dict[str, InputSpec] = {}
    for key, entry in raw.items():
        where = f"{path}: input {key!r}"
        if not isinstance(entry, dict):
            raise ContractError(f"{where}: must be a mapping")

        defined_name = _req_str(entry, "defined_name", where)
        if not DEFINED_NAME_RE.match(defined_name):
            raise ContractError(
                f"{where}: defined_name {defined_name!r} must match inp<PascalCase>"
            )
        type_ = _req_str(entry, "type", where)
        if type_ not in VALID_TYPES:
            raise ContractError(f"{where}: type {type_!r} must be one of {VALID_TYPES}")

        for cell_key in ("label_cell", "cell"):
            check_cell(_req_str(entry, cell_key, where), f"{where}: {cell_key}")

        for flag in ("required", "editable"):
            if not isinstance(entry.get(flag), bool):
                raise ContractError(f"{where}: {flag!r} must be a boolean")

        validation = entry.get("validation")
        if validation is not None:
            _validate_validation_block(validation, where)

        if entry.get("editable") is False and entry.get("default") is None:
            raise ContractError(
                f"{where}: a model-controlled input must declare its locked default value"
            )

        result[key] = InputSpec(
            key=key,
            sheet=_req_str(entry, "sheet", where),
            label=_req_str(entry, "label", where),
            defined_name=defined_name,
            label_cell=entry["label_cell"],
            cell=entry["cell"],
            type=type_,
            required=bool(entry["required"]),
            editable=bool(entry["editable"]),
            default=entry.get("default"),
            number_format=_req_str(entry, "number_format", where),
            validation=validation,
            note=entry.get("note"),
        )
    return result


def _parse_tables(raw: Any, path: Path) -> dict[str, TableSpec]:
    if not isinstance(raw, dict) or not raw:
        raise ContractError(f"{path}: 'tables' must be a non-empty mapping")
    result: dict[str, TableSpec] = {}
    for key, entry in raw.items():
        where = f"{path}: table {key!r}"
        columns = _parse_columns(entry, where)
        result[key] = TableSpec(
            key=key,
            sheet=_req_str(entry, "sheet", where),
            table_name=_table_name(entry, where),
            header_row=_positive_int(entry, "header_row", where),
            first_column=_column(entry, where),
            data_rows=_positive_int(entry, "data_rows", where),
            editable=_bool(entry, "editable", where),
            columns=columns,
            locked_seed_rows=_locked_seed_rows(entry, where),
            seed_rows=entry.get("seed_rows") or [],
        )
    return result


def _parse_config_tables(raw: Any, path: Path) -> list[TableSpec]:
    if not isinstance(raw, list) or not raw:
        raise ContractError(f"{path}: 'config_tables' must be a non-empty list")
    result: list[TableSpec] = []
    for index, entry in enumerate(raw):
        where = f"{path}: config_tables[{index}]"
        if not isinstance(entry, dict):
            raise ContractError(f"{where}: must be a mapping")
        key = _req_str(entry, "key", where)
        where = f"{path}: config table {key!r}"

        defined_name = _req_str(entry, "defined_name", where)
        if not DEFINED_NAME_RE.match(defined_name) or not defined_name.startswith("lst"):
            raise ContractError(
                f"{where}: defined_name {defined_name!r} must match lst<PascalCase>"
            )

        values = entry.get("values") or []
        if not isinstance(values, list):
            raise ContractError(f"{where}: values must be a list")
        data_rows = _positive_int(entry, "data_rows", where)
        if len(values) > data_rows:
            raise ContractError(
                f"{where}: {len(values)} seeded values exceed data_rows={data_rows}"
            )
        editable = _bool(entry, "editable", where)
        if not editable and not values:
            raise ContractError(f"{where}: a locked list must declare its constant values")
        if not editable and len(values) != data_rows:
            raise ContractError(
                f"{where}: a locked list must size data_rows to its value count "
                f"({len(values)} values, data_rows={data_rows})"
            )

        result.append(
            TableSpec(
                key=key,
                sheet="Config",
                table_name=_table_name(entry, where),
                header_row=_positive_int(entry, "header_row", where),
                first_column=_column(entry, where),
                data_rows=data_rows,
                editable=editable,
                columns=[
                    TableColumnSpec(
                        header=_req_str(entry, "column_header", where),
                        number_format=_req_str(entry, "number_format", where),
                        validation=None,
                    )
                ],
                locked_seed_rows=_locked_seed_rows(entry, where),
                seed_rows=[[value] for value in values],
                defined_name=defined_name,
                section=_req_str(entry, "section", where),
                section_row=_positive_int(entry, "section_row", where),
                note=entry.get("note"),
                note_row=entry.get("note_row"),
            )
        )
    return result


def _parse_columns(entry: dict[str, Any], where: str) -> list[TableColumnSpec]:
    raw_columns = _req(entry, "columns", where)
    if not isinstance(raw_columns, list) or not raw_columns:
        raise ContractError(f"{where}: columns must be a non-empty list")
    columns: list[TableColumnSpec] = []
    for position, column in enumerate(raw_columns):
        col_where = f"{where}: columns[{position}]"
        if not isinstance(column, dict):
            raise ContractError(f"{col_where}: must be a mapping")
        validation = column.get("validation")
        if validation is not None:
            _validate_validation_block(validation, col_where)
        columns.append(
            TableColumnSpec(
                header=_req_str(column, "header", col_where),
                number_format=_req_str(column, "number_format", col_where),
                validation=validation,
            )
        )
    headers = [c.header for c in columns]
    if len(headers) != len(set(headers)):
        raise ContractError(f"{where}: duplicate column headers {headers}")
    return columns


def _parse_sections(raw: Any, path: Path) -> list[SectionSpec]:
    if not isinstance(raw, list) or not raw:
        raise ContractError(f"{path}: setup_layout.sections must be a non-empty list")
    sections: list[SectionSpec] = []
    for index, entry in enumerate(raw):
        where = f"{path}: setup_layout.sections[{index}]"
        if not isinstance(entry, dict):
            raise ContractError(f"{where}: must be a mapping")
        sections.append(
            SectionSpec(
                title=_req_str(entry, "title", where),
                row=_positive_int(entry, "row", where),
                inputs=entry.get("inputs") or [],
                note=entry.get("note"),
                note_row=entry.get("note_row"),
                convention_row=entry.get("convention_row"),
                table=entry.get("table"),
            )
        )
    return sections


# ---------------------------------------------------------------------------
# cross-cutting validation
# ---------------------------------------------------------------------------
def _validate_unique_names(contract: InputContract, path: Path) -> None:
    defined = [spec.defined_name for spec in contract.inputs.values()]
    defined += [t.defined_name for t in contract.config_tables if t.defined_name]
    duplicates = sorted({n for n in defined if defined.count(n) > 1})
    if duplicates:
        raise ContractError(f"{path}: duplicate defined names: {duplicates}")

    tables = [t.table_name for t in contract.all_tables]
    duplicates = sorted({n for n in tables if tables.count(n) > 1})
    if duplicates:
        raise ContractError(f"{path}: duplicate table names: {duplicates}")

    collisions = sorted(set(defined) & set(tables))
    if collisions:
        raise ContractError(f"{path}: names used for both a range and a table: {collisions}")


def _validate_sections(contract: InputContract, path: Path) -> None:
    seen: set[str] = set()
    for section in contract.setup_sections:
        for key in section.inputs:
            if key not in contract.inputs:
                raise ContractError(
                    f"{path}: section {section.title!r} references unknown input {key!r}"
                )
            if key in seen:
                raise ContractError(f"{path}: input {key!r} appears in more than one section")
            seen.add(key)
        if section.table is not None and section.table not in contract.tables:
            raise ContractError(
                f"{path}: section {section.title!r} references unknown table {section.table!r}"
            )
    setup_inputs = {k for k, v in contract.inputs.items() if v.sheet == contract.setup_sheet}
    missing = sorted(setup_inputs - seen)
    if missing:
        raise ContractError(f"{path}: Setup inputs not placed in any section: {missing}")


def _validate_validation_sources(contract: InputContract, path: Path) -> None:
    known = set(contract.list_defined_names)
    sources: list[tuple[str, str]] = []
    for spec in contract.inputs.values():
        if spec.validation and spec.validation.get("kind") == "list":
            sources.append((f"input {spec.key!r}", spec.validation["source"]))
    for table in contract.all_tables:
        for column in table.columns:
            if column.validation and column.validation.get("kind") == "list":
                sources.append(
                    (f"table {table.table_name!r} column {column.header!r}", column.validation["source"])
                )
    for owner, source in sources:
        if source not in known:
            raise ContractError(
                f"{path}: {owner} validates against {source!r}, which is not a declared "
                f"list defined name. Known: {sorted(known)}"
            )


def _validate_no_cell_collisions(contract: InputContract, path: Path) -> None:
    """No two contract-owned cells may occupy the same address on a sheet."""
    occupied: dict[str, dict[str, str]] = {}

    def claim(sheet: str, address: str, owner: str) -> None:
        sheet_cells = occupied.setdefault(sheet, {})
        if address in sheet_cells:
            raise ContractError(
                f"{path}: cell {sheet}!{address} claimed by both "
                f"{sheet_cells[address]} and {owner}"
            )
        sheet_cells[address] = owner

    for spec in contract.inputs.values():
        claim(spec.sheet, spec.label_cell, f"label of {spec.key!r}")
        claim(spec.sheet, spec.cell, f"value of {spec.key!r}")

    for section in contract.setup_sections:
        claim(contract.setup_sheet, f"B{section.row}", f"section {section.title!r}")
        if section.note_row:
            claim(contract.setup_sheet, f"B{section.note_row}", f"note of {section.title!r}")
        if section.convention_row:
            claim(contract.setup_sheet, f"B{section.convention_row}", f"convention of {section.title!r}")

    for table in contract.all_tables:
        if table.section_row:
            claim(table.sheet, f"B{table.section_row}", f"section of {table.table_name}")
        if table.note_row:
            claim(table.sheet, f"B{table.note_row}", f"note of {table.table_name}")
        for column_index in range(len(table.columns)):
            letter = table.column_letter(column_index)
            for row in range(table.header_row, table.last_data_row + 1):
                claim(table.sheet, f"{letter}{row}", f"table {table.table_name}")


def _validate_seed_rows(contract: InputContract, path: Path) -> None:
    for table in contract.all_tables:
        for index, row in enumerate(table.seed_rows):
            if not isinstance(row, list):
                raise ContractError(f"{path}: {table.table_name} seed_rows[{index}] must be a list")
            if len(row) != len(table.columns):
                raise ContractError(
                    f"{path}: {table.table_name} seed_rows[{index}] has {len(row)} values "
                    f"but the table has {len(table.columns)} columns"
                )
        if len(table.seed_rows) > table.data_rows:
            raise ContractError(
                f"{path}: {table.table_name} has more seed rows than data_rows"
            )
        if table.locked_seed_rows > len(table.seed_rows):
            raise ContractError(
                f"{path}: {table.table_name} locks {table.locked_seed_rows} seed row(s) "
                f"but declares only {len(table.seed_rows)}; a locked row must carry a "
                "model-declared value"
            )
        if not table.editable and table.locked_seed_rows:
            raise ContractError(
                f"{path}: {table.table_name} is wholly locked (editable: false), so "
                "locked_seed_rows is meaningless and must be omitted"
            )


def _validate_excel_bounds(contract: InputContract, path: Path) -> None:
    """Every contract-owned coordinate must lie inside the Excel grid.

    Runs before any rendering, so an out-of-grid address fails as ContractError
    rather than as an incidental exception from openpyxl.
    """
    for spec in contract.inputs.values():
        check_cell(spec.label_cell, f"{path}: input {spec.key!r} label_cell")
        check_cell(spec.cell, f"{path}: input {spec.key!r} cell")

    check_row(contract.setup_intro.get("row"), f"{path}: setup_layout.intro")
    check_row(contract.config_intro.get("row"), f"{path}: config_layout.intro")

    for section in contract.setup_sections:
        where = f"{path}: setup section {section.title!r}"
        check_row(section.row, where)
        if section.note_row is not None:
            check_row(section.note_row, f"{where} note_row")
        if section.convention_row is not None:
            check_row(section.convention_row, f"{where} convention_row")

    for table in contract.all_tables:
        where = f"{path}: table {table.table_name}"
        check_column(table.first_column, f"{where} first_column")
        check_row(table.header_row, f"{where} header_row")
        # The far corner matters, not just the anchor.
        last_index = table.first_col_index + len(table.columns) - 1
        if last_index > EXCEL_MAX_COLUMN:
            raise ContractError(
                f"{where}: {len(table.columns)} columns starting at {table.first_column} "
                f"reach column index {last_index}, beyond the Excel maximum of "
                f"XFD ({EXCEL_MAX_COLUMN})"
            )
        check_row(table.last_data_row, f"{where} last data row (header_row + data_rows)")
        if table.section_row is not None:
            check_row(table.section_row, f"{where} section_row")
        if table.note_row is not None:
            check_row(table.note_row, f"{where} note_row")


def _validate_model_invariants(contract: InputContract, path: Path) -> None:
    """Values the user does not own must be present, locked and exactly right."""
    currency = contract.reporting_currency
    key = contract.model_invariants["reporting_currency_input"]
    expected_name = contract.model_invariants["reporting_currency_defined_name"]

    # Checked against the NAMED semantic input, not against whichever input
    # happens to hold the value.
    reporting = contract.inputs.get(key)
    if reporting is None:
        raise ContractError(
            f"{path}: model_invariants.reporting_currency_input is {key!r}, "
            f"which is not a declared input"
        )
    if reporting.editable:
        raise ContractError(
            f"{path}: reporting-currency input {key!r} must be model-controlled "
            "(editable: false)"
        )
    if reporting.default != currency:
        raise ContractError(
            f"{path}: reporting-currency input {key!r} defaults to "
            f"{reporting.default!r}, but model_invariants.reporting_currency is {currency!r}"
        )
    if reporting.defined_name != expected_name:
        raise ContractError(
            f"{path}: reporting-currency input {key!r} uses defined name "
            f"{reporting.defined_name!r}, but the invariant requires {expected_name!r}"
        )

    for index, identity in enumerate(contract.model_invariants["locked_identities"]):
        where = f"{path}: model_invariants.locked_identities[{index}]"
        if not isinstance(identity, dict):
            raise ContractError(f"{where}: must be a mapping")
        table_name = _req_str(identity, "table", where)
        row = _req(identity, "row", where)
        values = _req(identity, "values", where)
        if not isinstance(row, int) or isinstance(row, bool) or row < 1:
            raise ContractError(f"{where}: row must be a positive integer")
        if not isinstance(values, list) or not values:
            raise ContractError(f"{where}: values must be a non-empty list")
        if values[0] != currency:
            raise ContractError(
                f"{where}: identity must begin with the reporting currency "
                f"{currency!r}, found {values[0]!r}"
            )

        table = contract.table_by_name(table_name)
        if table is None:
            raise ContractError(f"{where}: unknown table {table_name!r}")
        if table.locked_seed_rows < row:
            raise ContractError(
                f"{where}: {table_name} declares locked_seed_rows="
                f"{table.locked_seed_rows}, so row {row} is not model-controlled"
            )
        if len(table.seed_rows) < row:
            raise ContractError(f"{where}: {table_name} has no seed row {row}")
        actual = table.seed_rows[row - 1]
        if list(actual) != list(values):
            raise ContractError(
                f"{where}: {table_name} row {row} is {actual!r}, "
                f"but the model invariant requires {values!r}"
            )


def _validate_validation_block(validation: Any, where: str) -> None:
    if not isinstance(validation, dict):
        raise ContractError(f"{where}: validation must be a mapping")
    kind = _req_str(validation, "kind", where)
    if kind not in VALID_VALIDATION_KINDS:
        raise ContractError(f"{where}: validation kind {kind!r} must be one of {VALID_VALIDATION_KINDS}")
    if kind == "list":
        _req_str(validation, "source", where)
    else:
        _req_str(validation, "operator", where)
        _req_str(validation, "formula1", where)


# ---------------------------------------------------------------------------
# small helpers
# ---------------------------------------------------------------------------
def _req(mapping: Any, key: str, where: str) -> Any:
    if not isinstance(mapping, dict) or key not in mapping:
        raise ContractError(f"{where}: missing required key {key!r}")
    return mapping[key]


def _req_str(mapping: Any, key: str, where: str) -> str:
    value = _req(mapping, key, where)
    if not isinstance(value, str) or not value.strip():
        raise ContractError(f"{where}: {key!r} must be a non-empty string, got {value!r}")
    return value


def _positive_int(mapping: dict[str, Any], key: str, where: str) -> int:
    value = _req(mapping, key, where)
    if not isinstance(value, int) or isinstance(value, bool) or value < 1:
        raise ContractError(f"{where}: {key!r} must be a positive integer, got {value!r}")
    return value


def _bool(mapping: dict[str, Any], key: str, where: str) -> bool:
    value = _req(mapping, key, where)
    if not isinstance(value, bool):
        raise ContractError(f"{where}: {key!r} must be a boolean, got {value!r}")
    return value


def _locked_seed_rows(mapping: dict[str, Any], where: str) -> int:
    value = mapping.get("locked_seed_rows", 0)
    if not isinstance(value, int) or isinstance(value, bool) or value < 0:
        raise ContractError(f"{where}: locked_seed_rows must be a non-negative integer")
    return value


def _column(mapping: dict[str, Any], where: str) -> str:
    return check_column(_req_str(mapping, "first_column", where), f"{where}: first_column")


def _table_name(mapping: dict[str, Any], where: str) -> str:
    value = _req_str(mapping, "table_name", where)
    if not TABLE_NAME_RE.match(value):
        raise ContractError(f"{where}: table_name {value!r} must match tbl<PascalCase>")
    return value


def _col(cell: str) -> str:
    match = CELL_RE.match(cell)
    if match is None:
        raise ContractError(f"invalid cell reference {cell!r}")
    return match.group(1)


def _row(cell: str) -> int:
    match = CELL_RE.match(cell)
    if match is None:
        raise ContractError(f"invalid cell reference {cell!r}")
    return int(match.group(2))


def _column_index(letter: str) -> int:
    index = 0
    for char in letter:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index
