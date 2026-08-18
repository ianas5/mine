"""Structural verification of a generated PCCM workbook.

This checks the artifact against the manifest that produced it. It is used both
by the builder (as a post-build gate) and by the Phase 1 test suite.
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from .calc_loader import CalcContract
from .calc_render import placeholder_row, table_ref
from .contract_loader import EXCEL_MAX_COLUMN, EXCEL_MAX_ROW, InputContract
from .driver_loader import DriverContract
from .spec_loader import WorkbookSpec
from .structure_loader import StructureContract


@dataclass
class VerificationResult:
    passed: list[str] = field(default_factory=list)
    failures: list[str] = field(default_factory=list)

    def check(self, description: str, condition: bool, detail: str = "") -> bool:
        if condition:
            self.passed.append(description)
        else:
            self.failures.append(f"{description}{f' -- {detail}' if detail else ''}")
        return condition

    @property
    def ok(self) -> bool:
        return not self.failures

    def report(self) -> str:
        lines = [f"  [PASS] {item}" for item in self.passed]
        lines += [f"  [FAIL] {item}" for item in self.failures]
        lines.append("")
        lines.append(f"  {len(self.passed)} passed, {len(self.failures)} failed")
        return "\n".join(lines)


def verify_workbook(
    path: str | Path,
    spec: WorkbookSpec,
    contract: InputContract | None = None,
    drivers: DriverContract | None = None,
    structure: StructureContract | None = None,
    calc: CalcContract | None = None,
) -> VerificationResult:
    """Verify the workbook at *path* against the manifest and every contract.

    `calc` adds the Phase-5 calculation-workspace checks. It is optional for the
    same reason `build_workbook`'s is: isolated Phase 1-4 unit tests still verify
    the artifact they build. `build_stage_a.py` always passes it, so the
    production Stage-A gate always includes them.
    """
    path = Path(path)
    result = VerificationResult()

    if not result.check("artifact exists", path.is_file(), str(path)):
        return result

    result.check(
        "artifact is a genuine OOXML (ZIP) package",
        zipfile.is_zipfile(path),
    )
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
    result.check(
        "no vbaProject.bin is present in this Stage-A .xlsx",
        not any(n.lower().endswith("vbaproject.bin") for n in names),
        ", ".join(n for n in names if "vba" in n.lower()),
    )
    result.check(
        "no external link parts are present",
        not any("externalLink" in n for n in names),
        ", ".join(n for n in names if "externalLink" in n),
    )
    result.check(
        "no data connection parts are present",
        not any("connections" in n.lower() for n in names),
        ", ".join(n for n in names if "connections" in n.lower()),
    )

    workbook = load_workbook(path)
    try:
        expected = spec.sheet_names
        actual = workbook.sheetnames

        result.check(
            f"workbook contains exactly {len(expected)} worksheets",
            len(actual) == len(expected),
            f"found {len(actual)}",
        )
        result.check("sheet names and order match the manifest", actual == expected,
                     f"expected {expected}, found {actual}")
        result.check(
            "no unexpected worksheet exists",
            set(actual) <= set(expected),
            f"unexpected: {sorted(set(actual) - set(expected))}",
        )

        for sheet_spec in spec.sheets:
            if sheet_spec.name not in workbook.sheetnames:
                result.check(f"sheet {sheet_spec.name!r} exists", False)
                continue
            worksheet = workbook[sheet_spec.name]
            result.check(
                f"sheet {sheet_spec.name!r} visibility is {sheet_spec.visibility!r}",
                worksheet.sheet_state == sheet_spec.visibility,
                f"found {worksheet.sheet_state!r}",
            )
            result.check(
                f"sheet {sheet_spec.name!r} shows its title",
                _contains(worksheet, sheet_spec.title),
            )

        active = workbook.active
        result.check(
            f"active sheet is {spec.active_sheet!r}",
            active is not None and active.title == spec.active_sheet,
            f"found {active.title if active is not None else None!r}",
        )
        result.check(
            "the active sheet is visible",
            active is not None and active.sheet_state == "visible",
        )

        # Phase-aware, and deliberately not relaxed to "some formulas are fine".
        # Phase 4 permits structural-state display only, and the permitted cells are
        # enumerated by the structure contract, so any other formula still fails.
        permitted = structure.formula_cells if structure is not None else {}
        unexpected = [
            found
            for found in _formula_cells(workbook)
            if found.split("!", 1)[1] not in permitted.get(found.split("!", 1)[0], set())
        ]
        result.check(
            "no worksheet contains a formula outside the contract-permitted structural cells",
            not unexpected,
            "; ".join(unexpected[:5]),
        )
        expected_names = set(contract.input_defined_names) | set(contract.list_defined_names) \
            if contract is not None else set()
        if structure is not None and contract is not None:
            expected_names |= set(structure.defined_names)
            expected_names |= set(structure.alias_defined_names(contract))
        found_names = set(workbook.defined_names)
        result.check(
            "only contract-declared defined names exist",
            found_names == expected_names,
            f"unexpected {sorted(found_names - expected_names)}, "
            f"missing {sorted(expected_names - found_names)}",
        )
        expected_tables = (
            {t.table_name for t in contract.all_tables} if contract is not None else set()
        )
        if drivers is not None:
            expected_tables |= {r.table_name for r in drivers.all_registers}
        if structure is not None:
            expected_tables |= {g.table_name for g in structure.all_grids}
        # EXTENDED, NOT RELAXED. The gate still says "only contract-declared Tables
        # exist"; Phase 5 adds five declared tables to the expected set.
        if calc is not None:
            expected_tables |= set(calc.table_names)
        found_tables = {name for _, name in _tables(workbook)}
        result.check(
            "only contract-declared Excel Tables exist",
            found_tables == expected_tables,
            f"expected {sorted(expected_tables)}, found {sorted(found_tables)}",
        )

        if contract is not None:
            _verify_contract(result, workbook, contract)
        if drivers is not None:
            _verify_drivers(result, workbook, drivers)
        if structure is not None and contract is not None:
            _verify_structure(result, workbook, structure, contract)
        if calc is not None:
            _verify_calc(result, workbook, calc)
    finally:
        workbook.close()

    return result


def _contains(worksheet, text: str) -> bool:
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value == text:
                return True
    return False


def _formula_cells(workbook) -> list[str]:
    found: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    found.append(f"{worksheet.title}!{cell.coordinate}")
    return found


def _tables(workbook) -> list[tuple[str, str]]:
    found: list[tuple[str, str]] = []
    for worksheet in workbook.worksheets:
        for name in getattr(worksheet, "tables", {}):
            found.append((worksheet.title, name))
    return found


# ---------------------------------------------------------------------------
# Data-validation coverage
#
# Whether a validation touches a protected cell is a question about Excel
# rectangles, not about text. Comparing range strings answers a different and
# much weaker question: 'B12:B36' not in {'B12', 'B12:B20', 'B20:B36'} is true
# while every one of those ranges covers protected identity cells. Ranges are
# therefore reduced to integer bounds and tested for rectangle overlap.
# ---------------------------------------------------------------------------
def _rect(reference) -> tuple[int, int, int, int]:
    """(min_col, min_row, max_col, max_row) for a range string or CellRange.

    An open reference such as ``B:B`` or ``12:12`` parses with a missing bound;
    those are widened to the sheet edge so an unbounded range still intersects
    everything it actually covers.
    """
    bounds = getattr(reference, "bounds", None)
    if bounds is None:
        bounds = range_boundaries(str(reference))
    min_col, min_row, max_col, max_row = bounds
    return (
        min_col or 1,
        min_row or 1,
        max_col or EXCEL_MAX_COLUMN,
        max_row or EXCEL_MAX_ROW,
    )


def _overlaps(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> bool:
    return a[0] <= b[2] and a[2] >= b[0] and a[1] <= b[3] and a[3] >= b[1]


def intersecting_validation_ranges(worksheet, target_range) -> list[str]:
    """Every data-validation range on *worksheet* that overlaps *target_range*.

    Iterates the ranges of each ``DataValidation.sqref`` -- so a multi-area
    validation is handled area by area -- and compares bounds. It never
    enumerates individual cells, so testing a large target costs nothing.
    """
    target = _rect(target_range)
    found: list[str] = []
    for dv in worksheet.data_validations.dataValidation:
        for cell_range in dv.sqref.ranges:
            if _overlaps(_rect(cell_range), target):
                found.append(str(cell_range))
    return found


def data_validation_intersects(worksheet, target_range) -> bool:
    """True if ANY data validation on *worksheet* covers ANY cell of *target_range*.

    Accepts a single cell (``B12``), a contiguous range (``B12:B36``) or a
    ``CellRange``, and handles a validation whose sqref holds several areas.
    """
    return bool(intersecting_validation_ranges(worksheet, target_range))


def _verify_contract(result: VerificationResult, workbook, contract: InputContract) -> None:
    """Every contract promise must be observable in the built artifact."""
    for spec in contract.inputs.values():
        worksheet = workbook[spec.sheet]
        result.check(
            f"input {spec.key!r} label at {spec.sheet}!{spec.label_cell}",
            worksheet[spec.label_cell].value == spec.label,
            f"found {worksheet[spec.label_cell].value!r}",
        )
        result.check(
            f"input {spec.key!r} default at {spec.sheet}!{spec.cell}",
            worksheet[spec.cell].value == spec.default,
            f"expected {spec.default!r}, found {worksheet[spec.cell].value!r}",
        )
        reference = contract.input_defined_names[spec.defined_name]
        result.check(
            f"defined name {spec.defined_name} -> {spec.sheet}!{spec.cell}",
            workbook.defined_names[spec.defined_name].attr_text == reference,
            f"found {workbook.defined_names[spec.defined_name].attr_text!r}",
        )

    for table in contract.all_tables:
        worksheet = workbook[table.sheet]
        tables = getattr(worksheet, "tables", {})
        if not result.check(
            f"table {table.table_name} exists on {table.sheet}", table.table_name in tables
        ):
            continue
        result.check(
            f"table {table.table_name} ref is {table.ref}",
            tables[table.table_name].ref == table.ref,
            f"found {tables[table.table_name].ref}",
        )
        for index, column in enumerate(table.columns):
            address = f"{table.column_letter(index)}{table.header_row}"
            result.check(
                f"table {table.table_name} header {column.header!r} at {address}",
                worksheet[address].value == column.header,
                f"found {worksheet[address].value!r}",
            )
        for offset, seed in enumerate(table.seed_rows):
            for index, value in enumerate(seed):
                address = f"{table.column_letter(index)}{table.first_data_row + offset}"
                result.check(
                    f"table {table.table_name} seed {address} = {value!r}",
                    worksheet[address].value == value,
                    f"found {worksheet[address].value!r}",
                )
        if table.defined_name:
            result.check(
                f"defined name {table.defined_name} covers {table.table_name} data body",
                workbook.defined_names[table.defined_name].attr_text
                == table.absolute_data_range(),
                f"found {workbook.defined_names[table.defined_name].attr_text!r}",
            )

    # Model invariants: the declared identity values must be exactly right.
    for identity in contract.model_invariants["locked_identities"]:
        table = contract.table_by_name(identity["table"])
        worksheet = workbook[table.sheet]
        row = table.first_data_row + identity["row"] - 1
        for index, value in enumerate(identity["values"]):
            address = f"{table.column_letter(index)}{row}"
            result.check(
                f"model invariant {table.table_name} {address} = {value!r}",
                worksheet[address].value == value,
                f"found {worksheet[address].value!r}",
            )

    validations = {
        f"{ws.title}!{dv.sqref}": dv
        for ws in workbook.worksheets
        for dv in ws.data_validations.dataValidation
    }
    # No validation may touch ANY cell of a locked identity row. Partial and
    # multi-area coverage counts: attaching a user input rule to a cell the user
    # does not own misrepresents ownership however small the overlap.
    for table in contract.all_tables:
        if not table.locked_seed_rows:
            continue
        worksheet = workbook[table.sheet]
        locked = (
            f"{table.column_letter(0)}{table.first_data_row}:"
            f"{table.column_letter(len(table.columns) - 1)}{table.first_user_row - 1}"
        )
        offenders = intersecting_validation_ranges(worksheet, locked)
        result.check(
            f"no data validation intersects {table.table_name} locked identity rows ({locked})",
            not offenders,
            ", ".join(offenders),
        )
    result.check(
        "data validation rules were created",
        len(validations) > 0,
        f"found {len(validations)}",
    )


def _verify_drivers(result: VerificationResult, workbook, drivers: DriverContract) -> None:
    """Every driver-schema promise must be observable in the artifact."""
    for register in drivers.all_registers:
        worksheet = workbook[register.sheet]
        tables = getattr(worksheet, "tables", {})
        if not result.check(
            f"table {register.table_name} exists on {register.sheet}",
            register.table_name in tables,
        ):
            continue
        result.check(
            f"table {register.table_name} ref is {register.ref}",
            tables[register.table_name].ref == register.ref,
            f"found {tables[register.table_name].ref}",
        )
        result.check(
            f"table {register.table_name} has {len(register.columns)} columns in locked order",
            [
                worksheet[f"{register.column_letter(i)}{register.header_row}"].value
                for i in range(len(register.columns))
            ]
            == register.headers,
        )
        # Identity columns must be genuinely blank: no pre-seeded IDs, no formula.
        identity = register.columns[0]
        blanks = [
            f"{register.column_letter(0)}{row}"
            for row in range(register.first_data_row, register.last_data_row + 1)
            if worksheet[f"{register.column_letter(0)}{row}"].value is not None
        ]
        result.check(
            f"{register.table_name} {identity.header!r} column is entirely blank "
            "(no IDs are allocated in Stage A)",
            not blanks,
            ", ".join(blanks[:5]),
        )
        # Every data cell must be empty: a driver register carries no seeded data.
        populated = [
            f"{register.column_letter(i)}{row}"
            for i in range(len(register.columns))
            for row in range(register.first_data_row, register.last_data_row + 1)
            if worksheet[f"{register.column_letter(i)}{row}"].value is not None
        ]
        result.check(
            f"{register.table_name} reserved rows are blank",
            not populated,
            ", ".join(populated[:5]),
        )

        # Any overlap fails: a single ID cell, a partial run of the column, or one
        # area of a multi-area sqref is as much a breach of ownership as the whole
        # column would be.
        identity_range = register.data_range(0)
        offenders = intersecting_validation_ranges(worksheet, identity_range)
        result.check(
            f"no data validation intersects {register.table_name} identity column "
            f"({identity_range})",
            not offenders,
            ", ".join(offenders),
        )


def _verify_structure(
    result: VerificationResult,
    workbook,
    structure: StructureContract,
    contract: InputContract,
) -> None:
    """Every structural-runtime promise must be observable in the artifact."""
    setup = workbook[structure.setup_sheet]

    # --- applied timeline: blank state, derived formulas ---------------------
    for field_ in structure.applied:
        cell = setup[field_.cell]
        result.check(
            f"applied field {field_.key!r} at {structure.setup_sheet}!{field_.cell} is blank "
            "(no timeline is applied in Stage A)",
            cell.value is None,
            f"found {cell.value!r}",
        )
        result.check(
            f"applied field {field_.key!r} label at {field_.label_cell}",
            setup[field_.label_cell].value == field_.label,
            f"found {setup[field_.label_cell].value!r}",
        )
    for field_ in structure.derived:
        cell = setup[field_.cell]
        result.check(
            f"derived field {field_.key!r} carries its structural formula",
            cell.value == field_.formula,
            f"found {cell.value!r}",
        )

    state = structure.structural_state
    result.check(
        "structural state indicator carries its formula",
        workbook[state.sheet][state.cell].value == state.formula,
        f"found {workbook[state.sheet][state.cell].value!r}",
    )
    result.check(
        "no macro is required to maintain the structural state indicator",
        str(workbook[state.sheet][state.cell].value).startswith("="),
    )

    # --- defined names -------------------------------------------------------
    for name, reference in structure.defined_names.items():
        result.check(
            f"structural defined name {name} -> {reference}",
            name in workbook.defined_names
            and workbook.defined_names[name].attr_text == reference,
            f"found {workbook.defined_names[name].attr_text!r}"
            if name in workbook.defined_names
            else "missing",
        )
    for name, reference in structure.alias_defined_names(contract).items():
        result.check(
            f"entered alias {name} addresses the same cell as its inp* name ({reference})",
            name in workbook.defined_names
            and workbook.defined_names[name].attr_text == reference,
            f"found {workbook.defined_names[name].attr_text!r}"
            if name in workbook.defined_names
            else "missing",
        )
    for alias in structure.entered_aliases:
        spec = contract.inputs[alias.input_key]
        result.check(
            f"the accepted input name {spec.defined_name} still exists alongside "
            f"{alias.defined_name}",
            spec.defined_name in workbook.defined_names,
        )

    # --- permanent-ID counters ----------------------------------------------
    identity = workbook[structure.identity_sheet]
    for counter in structure.counters:
        result.check(
            f"counter {counter.key!r} seeded at {structure.identity_sheet}!{counter.cell}",
            identity[counter.cell].value == counter.initial,
            f"expected {counter.initial!r}, found {identity[counter.cell].value!r}",
        )

    # --- grids ---------------------------------------------------------------
    for grid in structure.all_grids:
        worksheet = workbook[grid.sheet]
        tables = getattr(worksheet, "tables", {})
        if not result.check(
            f"grid table {grid.table_name} exists on {grid.sheet}", grid.table_name in tables
        ):
            continue
        result.check(
            f"grid table {grid.table_name} ref is {grid.ref} (fixed columns only)",
            tables[grid.table_name].ref == grid.ref,
            f"found {tables[grid.table_name].ref}",
        )
        result.check(
            f"grid {grid.table_name} has {len(grid.fixed_columns)} fixed columns in order",
            [
                worksheet[f"{grid.column_letter(i)}{grid.header_row}"].value
                for i in range(len(grid.fixed_columns))
            ]
            == grid.headers,
        )
        # No year column may exist before a timeline is applied: a generated year
        # column would assert a timeline the user has not entered.
        beyond = grid.column_letter(len(grid.fixed_columns))
        result.check(
            f"grid {grid.table_name} has no generated year column yet (first free column "
            f"{beyond} is empty)",
            worksheet[f"{beyond}{grid.header_row}"].value is None,
            f"found {worksheet[f'{beyond}{grid.header_row}'].value!r}",
        )
        populated = [
            f"{grid.column_letter(i)}{row}"
            for i in range(len(grid.fixed_columns))
            for row in range(grid.first_data_row, grid.last_data_row + 1)
            if worksheet[f"{grid.column_letter(i)}{row}"].value is not None
        ]
        result.check(
            f"grid {grid.table_name} reserved rows are blank",
            not populated,
            ", ".join(populated[:5]),
        )
        offenders = intersecting_validation_ranges(
            worksheet, f"{grid.first_column}{grid.first_data_row}:"
            f"{grid.last_fixed_column}{grid.last_data_row}"
        )
        result.check(
            f"no data validation intersects {grid.table_name}, whose fixed columns are all "
            "model-controlled",
            not offenders,
            ", ".join(offenders),
        )
        key = "inflation_formula" if grid.kind == "inflation" else "profiling_formula"
        result.check(
            f"grid {grid.sheet} shows its structural state message",
            worksheet[f"B{grid.state_message_row}"].value == structure.state_messages[key],
            f"found {worksheet[f'B{grid.state_message_row}'].value!r}",
        )


def _verify_calc(result: VerificationResult, workbook, calc: CalcContract) -> None:
    """The Phase-5 calculation workspace, checked against the GENERATED ARTIFACT.

    Not against the contract's own statements, and not against the fact that the
    renderer was called: every assertion below reads the workbook that was just
    written, so a renderer that drifts from the contract is caught by the artifact
    rather than excused by it.
    """
    sheet = calc.sheet
    if sheet not in workbook.sheetnames:
        result.check(f"calculation sheet {sheet!r} exists", False)
        return
    worksheet = workbook[sheet]
    result.check(f"calculation sheet {sheet!r} exists", True)
    result.check(
        f"calculation sheet {sheet!r} visibility is {calc.required_visibility!r}",
        worksheet.sheet_state == calc.required_visibility,
        f"found {worksheet.sheet_state!r}",
    )

    # --- Phase-4 territory is intact -----------------------------------------
    # The counters are Phase-4 state. Phase 5 renders below them and must not have
    # overwritten, cleared or reformatted any of it.
    for address in calc.phase4_cells:
        cell = worksheet[address]
        result.check(
            f"Phase-4 counter {sheet}!{address} still holds its integer counter",
            cell.value == 0 and cell.number_format == "0",
            f"found {cell.value!r} / {cell.number_format!r}",
        )
    result.check(
        f"Phase-4 reserved rows {calc.phase4_first_row}-{calc.phase4_last_row} carry no "
        "Phase-5 content",
        all(
            worksheet.cell(row=row, column=column).value is None
            for row in calc.phase4_reserved_rows
            for column in range(6, EXCEL_MAX_COLUMN + 1)
            if column <= 60
        ),
    )

    # --- scalar blocks --------------------------------------------------------
    for block in (calc.calc_state, calc.calc_totals):
        for entry in block.fields:
            label = worksheet[f"{block.label_column}{entry.row}"]
            result.check(
                f"{block.key} row {entry.row} label is {entry.label!r}",
                label.value == entry.label,
                f"found {label.value!r}",
            )
            cell = worksheet[f"{block.value_column}{entry.row}"]
            result.check(
                f"{block.key} value cell {cell.coordinate} carries number format "
                f"{entry.number_format!r}",
                cell.number_format == entry.number_format,
                f"found {cell.number_format!r}",
            )
            result.check(
                f"{block.key} value cell {cell.coordinate} holds its initial state "
                f"{'BLANK' if entry.initial is None else repr(entry.initial)}",
                cell.value == entry.initial if entry.initial is not None
                else cell.value is None,
                f"found {cell.value!r}",
            )
            if entry.note:
                note = worksheet[f"{block.note_column}{entry.row}"]
                result.check(
                    f"{block.key} row {entry.row} carries its contract note",
                    note.value == entry.note,
                    f"found {note.value!r}",
                )

    # Stated separately because it is the one that is easiest to get wrong: the
    # contract declares FP_VERSION = 1, and C15 must still ship blank.
    version_entry = calc.calc_state.field_by_key("fingerprint_version")
    version_cell = f"{calc.calc_state.value_column}{version_entry.row}"
    result.check(
        f"fingerprint version cell {sheet}!{version_cell} is BLANK, not seeded with "
        f"FP_VERSION={calc.fingerprint_version}",
        worksheet[version_cell].value is None,
        f"found {worksheet[version_cell].value!r}",
    )
    result.check(
        "no calc_totals cell is seeded with zero",
        all(
            worksheet[f"{calc.calc_totals.value_column}{entry.row}"].value is None
            for entry in calc.calc_totals.fields
        ),
        "blank means no calculation has committed; zero would mean a total of zero",
    )

    # --- the five tables ------------------------------------------------------
    tables = getattr(worksheet, "tables", {})
    for table in calc.all_tables:
        name = table.table_name
        if name not in tables:
            result.check(f"calculation table {name} exists on {sheet}", False)
            continue
        result.check(f"calculation table {name} exists on {sheet}", True)
        expected_ref = table_ref(table)
        result.check(
            f"calculation table {name} spans {expected_ref}",
            tables[name].ref == expected_ref,
            f"found {tables[name].ref}",
        )
        left, top, right, bottom = _rect(tables[name].ref)
        result.check(
            f"calculation table {name} has its header on row {table.header_row}",
            top == table.header_row,
            f"found row {top}",
        )
        result.check(
            f"calculation table {name} occupies columns "
            f"{table.first_column}:{table.last_column}",
            left == table.first_column_index and right == table.last_column_index,
            f"found {get_column_letter(left)}:{get_column_letter(right)}",
        )
        result.check(
            f"calculation table {name} is {len(table.columns)} columns wide",
            right - left + 1 == len(table.columns),
            f"found {right - left + 1}",
        )
        headers = [
            worksheet.cell(row=table.header_row, column=left + index).value
            for index in range(len(table.columns))
        ]
        result.check(
            f"calculation table {name} header spelling and order match the contract",
            headers == table.headers,
            f"expected {table.headers}, found {headers}",
        )
        for index, column in enumerate(table.columns):
            cell = worksheet.cell(row=placeholder_row(table), column=left + index)
            result.check(
                f"{name}.{column.key} body carries number format "
                f"{column.number_format!r}",
                cell.number_format == column.number_format,
                f"found {cell.number_format!r}",
            )
        result.check(
            f"calculation table {name} has ZERO semantic rows: its one physical body "
            "row is blank",
            all(
                worksheet.cell(row=row, column=col).value is None
                for row in range(table.header_row + 1, bottom + 1)
                for col in range(left, right + 1)
            ),
            "a build-time row would be a fabricated calculation output",
        )
        result.check(
            f"calculation table {name} carries no formula",
            not any(
                isinstance(worksheet.cell(row=row, column=col).value, str)
                and str(worksheet.cell(row=row, column=col).value).startswith("=")
                for row in range(table.header_row, bottom + 1)
                for col in range(left, right + 1)
            ),
        )
        result.check(
            f"calculation table {name} carries no data validation",
            not data_validation_intersects(worksheet, tables[name].ref),
            "the calculation workspace is model-controlled, not a user input surface",
        )

    rects = {t.table_name: _rect(tables[t.table_name].ref)
             for t in calc.all_tables if t.table_name in tables}
    names = sorted(rects)
    for i, left_name in enumerate(names):
        for right_name in names[i + 1:]:
            result.check(
                f"calculation tables {left_name} and {right_name} do not overlap",
                not _overlaps(rects[left_name], rects[right_name]),
                f"{tables[left_name].ref} vs {tables[right_name].ref}",
            )
    for name, rect in rects.items():
        result.check(
            f"calculation table {name} clears the Phase-4 reservation "
            f"(rows {calc.phase4_first_row}-{calc.phase4_last_row})",
            rect[1] > calc.phase4_last_row,
            f"starts on row {rect[1]}",
        )

    for block in (calc.calc_state, calc.calc_totals):
        result.check(
            f"{block.key} value column carries no data validation",
            not data_validation_intersects(worksheet, block.value_range()),
            "these are model-controlled audit cells, not user inputs",
        )


def structural_digest(path: str | Path) -> str:
    """A normalised, order-sensitive description of workbook structure.

    Deliberately excludes ZIP metadata and file timestamps, so two builds of the
    same source compare equal without requiring byte-identical archives.
    """
    workbook = load_workbook(Path(path))
    try:
        parts: list[str] = []
        for worksheet in workbook.worksheets:
            parts.append(
                f"SHEET|{worksheet.title}|{worksheet.sheet_state}|"
                f"grid={worksheet.sheet_view.showGridLines}|freeze={worksheet.freeze_panes}"
            )
            for column, dimension in sorted(worksheet.column_dimensions.items()):
                parts.append(f"COL|{worksheet.title}|{column}|{dimension.width}")
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        parts.append(f"CELL|{worksheet.title}|{cell.coordinate}|{cell.value!r}")
        active = workbook.active
        parts.append(f"ACTIVE|{active.title if active is not None else None}")
        return "\n".join(parts)
    finally:
        workbook.close()
