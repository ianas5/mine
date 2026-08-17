#!/usr/bin/env python3
"""PCCM Phase 1 structural tests.

Runs standalone (``python pccm/tests/test_phase1_structure.py``) and is also
collectable by pytest if it is available.

The locked sheet order, visibility and intended CodeNames are duplicated here on
purpose. The manifest is the builder's structural authority; this file is the
independent architecture-conformance check, so a manifest that drifts from the
Architecture Lock fails here even though the build itself succeeded.
"""

from __future__ import annotations

import os
import sys
import tempfile
import zipfile
from pathlib import Path

PCCM_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PCCM_ROOT / "builder"))

from openpyxl import load_workbook  # noqa: E402

from pccm_builder import (  # noqa: E402
    build_workbook,
    load_contract,
    load_driver_contract,
    load_structure_contract,
    load_spec,
    structural_digest,
)

SPEC_PATH = PCCM_ROOT / "spec" / "workbook.yaml"
CONTRACT_PATH = PCCM_ROOT / "spec" / "input_contract.yaml"
DRIVERS_PATH = PCCM_ROOT / "spec" / "driver_contract.yaml"
STRUCTURE_PATH = PCCM_ROOT / "spec" / "structure_contract.yaml"

# --- Architecture Lock Revision B -------------------------------------------
LOCKED_SHEET_ORDER = [
    "Dashboard",
    "Setup",
    "Config",
    "Cost Lines",
    "Risk Register",
    "Inflation",
    "Cost Profiling",
    "Risk Profiling",
    "Model Check",
    "Results",
    "Sensitivity",
    "Methodology",
    "_Calc",
    "_SimData",
]

LOCKED_CODENAMES = {
    "Dashboard": "shDashboard",
    "Setup": "shSetup",
    "Config": "shConfig",
    "Cost Lines": "shCostLines",
    "Risk Register": "shRiskRegister",
    "Inflation": "shInflation",
    "Cost Profiling": "shCostProfiling",
    "Risk Profiling": "shRiskProfiling",
    "Model Check": "shModelCheck",
    "Results": "shResults",
    "Sensitivity": "shSensitivity",
    "Methodology": "shMethodology",
    "_Calc": "shCalc",
    "_SimData": "shSimData",
}

HIDDEN_SHEET = "_Calc"
VERY_HIDDEN_SHEET = "_SimData"
INTERNAL_SHEETS = {HIDDEN_SHEET, VERY_HIDDEN_SHEET}
ACTIVE_SHEET = "Dashboard"

REQUIRED_SETUP_LABELS = [
    "Project Name",
    "Project Duration (Years)",
    "Project Start Year",
    "Base Year",
    "Reporting Currency",
    "Selected Confidence Level",
    "Monte Carlo Iterations",
    "Discount Rate",
    "Random Seed",
]

REQUIRED_CONFIG_SECTIONS = [
    "Categories",
    "Currencies",
    "Units of Measure",
    "Inflation Profile Names",
    "Distribution Names",
    "Confidence Levels",
]

LOCKED_DISTRIBUTIONS = ["Triangular", "Beta-PERT", "Uniform"]
LOCKED_CONFIDENCE_LEVELS = [f"P{n}" for n in range(50, 100, 5)]

_ARTIFACT_CACHE: dict[str, Path] = {}
_TEMPDIR: tempfile.TemporaryDirectory | None = None


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def _build(name: str, timestamp: str = "1970-01-01 00:00:00 UTC") -> Path:
    """Build the skeleton into a temp directory and return its path."""
    global _TEMPDIR
    if name in _ARTIFACT_CACHE:
        return _ARTIFACT_CACHE[name]
    if _TEMPDIR is None:
        _TEMPDIR = tempfile.TemporaryDirectory(prefix="pccm-phase1-")

    previous = os.environ.get("PCCM_BUILD_TIMESTAMP")
    os.environ["PCCM_BUILD_TIMESTAMP"] = timestamp
    try:
        spec = load_spec(SPEC_PATH)
        contract = load_contract(CONTRACT_PATH)
        drivers = load_driver_contract(DRIVERS_PATH)
        structure = load_structure_contract(STRUCTURE_PATH)
        workbook, _ = build_workbook(spec, contract, drivers, structure)
        path = Path(_TEMPDIR.name) / f"{name}.xlsx"
        workbook.save(path)
        workbook.close()
    finally:
        if previous is None:
            os.environ.pop("PCCM_BUILD_TIMESTAMP", None)
        else:
            os.environ["PCCM_BUILD_TIMESTAMP"] = previous

    _ARTIFACT_CACHE[name] = path
    return path


def _artifact() -> Path:
    return _build("primary")


def _strings(worksheet) -> set[str]:
    values: set[str] = set()
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                values.add(cell.value.strip())
    return values


# ---------------------------------------------------------------------------
# 1-3, 8. sheet inventory
# ---------------------------------------------------------------------------
def _permitted_formula_cells() -> dict:
    """The exact cells the structure contract is allowed to write a formula into.

    Phases 1-3 forbade every formula. Phase 4 permits structural-state display only
    -- Structure Change Pending, the derived applied-timeline cells and each grid's
    "timeline not yet applied" message -- and the contract enumerates them. The
    regression intent is unchanged: any formula outside this enumerated set is still
    a failure, and a business calculation would land outside it.
    """
    structure = load_structure_contract(STRUCTURE_PATH)
    return structure.formula_cells


def _unexpected_formulas(workbook) -> list[str]:
    permitted = _permitted_formula_cells()
    return [
        f"{ws.title}!{cell.coordinate}"
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
        and cell.value.startswith("=")
        and cell.coordinate not in permitted.get(ws.title, set())
    ]


def test_01_exactly_fourteen_worksheets() -> None:
    workbook = load_workbook(_artifact())
    assert len(workbook.sheetnames) == 14, f"found {len(workbook.sheetnames)}"


def test_02_sheet_names_match_locked_list() -> None:
    workbook = load_workbook(_artifact())
    assert set(workbook.sheetnames) == set(LOCKED_SHEET_ORDER)


def test_03_sheet_order_matches_locked_list() -> None:
    workbook = load_workbook(_artifact())
    assert workbook.sheetnames == LOCKED_SHEET_ORDER


def test_08_no_unexpected_worksheet() -> None:
    workbook = load_workbook(_artifact())
    unexpected = set(workbook.sheetnames) - set(LOCKED_SHEET_ORDER)
    assert not unexpected, f"unexpected sheets: {sorted(unexpected)}"
    assert "Sheet" not in workbook.sheetnames, "default openpyxl sheet was not removed"


# ---------------------------------------------------------------------------
# 4-7, 18. visibility and active sheet
# ---------------------------------------------------------------------------
def test_04_dashboard_is_active() -> None:
    workbook = load_workbook(_artifact())
    assert workbook.active is not None
    assert workbook.active.title == ACTIVE_SHEET


def test_05_calc_is_hidden() -> None:
    workbook = load_workbook(_artifact())
    assert workbook[HIDDEN_SHEET].sheet_state == "hidden"


def test_06_simdata_is_very_hidden() -> None:
    workbook = load_workbook(_artifact())
    assert workbook[VERY_HIDDEN_SHEET].sheet_state == "veryHidden"


def test_07_all_other_sheets_are_visible() -> None:
    workbook = load_workbook(_artifact())
    for name in LOCKED_SHEET_ORDER:
        if name in INTERNAL_SHEETS:
            continue
        assert workbook[name].sheet_state == "visible", (
            f"{name} is {workbook[name].sheet_state}"
        )


def test_18_internal_sheets_are_not_active() -> None:
    workbook = load_workbook(_artifact())
    assert workbook.active.title not in INTERNAL_SHEETS
    assert workbook.active.sheet_state == "visible"


# ---------------------------------------------------------------------------
# 9-10. intended CodeNames (manifest only; Stage B is the runtime authority)
# ---------------------------------------------------------------------------
def test_09_manifest_has_one_intended_codename_per_sheet() -> None:
    spec = load_spec(SPEC_PATH)
    assert len(spec.sheets) == len(LOCKED_SHEET_ORDER)
    for sheet in spec.sheets:
        assert sheet.codename, f"{sheet.name} has no intended CodeName"
        assert sheet.codename == LOCKED_CODENAMES[sheet.name], (
            f"{sheet.name}: manifest says {sheet.codename}, "
            f"lock says {LOCKED_CODENAMES[sheet.name]}"
        )


def test_10_intended_codenames_are_unique() -> None:
    spec = load_spec(SPEC_PATH)
    codenames = [s.codename for s in spec.sheets]
    assert len(codenames) == len(set(codenames)), "duplicate intended CodeNames"


# ---------------------------------------------------------------------------
# 11-12. artifact integrity
# ---------------------------------------------------------------------------
def test_11_workbook_reopens_with_openpyxl() -> None:
    workbook = load_workbook(_artifact())
    assert workbook.sheetnames == LOCKED_SHEET_ORDER
    workbook.close()


def test_12_artifact_is_a_genuine_ooxml_package() -> None:
    path = _artifact()
    assert path.suffix == ".xlsx"
    assert zipfile.is_zipfile(path), "artifact is not a ZIP container"
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
    assert "[Content_Types].xml" in names
    assert "xl/workbook.xml" in names


# ---------------------------------------------------------------------------
# 13-14. Phase 1 shells
# ---------------------------------------------------------------------------
def test_13_setup_shell_contains_required_labels() -> None:
    workbook = load_workbook(_artifact())
    values = _strings(workbook["Setup"])
    missing = [label for label in REQUIRED_SETUP_LABELS if label not in values]
    assert not missing, f"missing Setup labels: {missing}"
    assert "SAR" in values, "Setup does not show the locked reporting currency"


def test_14_config_shell_contains_reserved_sections() -> None:
    workbook = load_workbook(_artifact())
    values = _strings(workbook["Config"])
    missing = [name for name in REQUIRED_CONFIG_SECTIONS if name not in values]
    assert not missing, f"missing Config sections: {missing}"
    # Phase 2: the reserved sections are now backed by real Excel Tables.
    tables = set(getattr(workbook["Config"], "tables", {}))
    assert tables, "Config sections are not backed by Excel Tables"
    for item in LOCKED_DISTRIBUTIONS + LOCKED_CONFIDENCE_LEVELS:
        assert item in values, f"Config is missing locked constant {item!r}"


# ---------------------------------------------------------------------------
# 15-17. nothing from a later phase has leaked in
# ---------------------------------------------------------------------------
def test_15_no_vba_project_present() -> None:
    with zipfile.ZipFile(_artifact()) as archive:
        names = [n.lower() for n in archive.namelist()]
    assert not any(n.endswith("vbaproject.bin") for n in names)
    assert not any(n.endswith(".xlsm") for n in names)


def test_16_no_calculation_formulas_introduced() -> None:
    """Phase-aware. Tables and defined names are asserted to be *exactly* the
    contracts', not merely absent, and the only formulas tolerated are the
    structural-state displays the structure contract enumerates."""
    contract = load_contract(CONTRACT_PATH)
    structure = load_structure_contract(STRUCTURE_PATH)
    workbook = load_workbook(_artifact())

    offenders = _unexpected_formulas(workbook)
    assert not offenders, f"formulas outside the permitted structural cells: {offenders[:10]}"

    expected_names = set(contract.input_defined_names) | set(contract.list_defined_names)
    expected_names |= set(structure.defined_names)
    expected_names |= set(structure.alias_defined_names(contract))
    assert set(workbook.defined_names) == expected_names, (
        f"unexpected {sorted(set(workbook.defined_names) - expected_names)}, "
        f"missing {sorted(expected_names - set(workbook.defined_names))}"
    )

    drivers = load_driver_contract(DRIVERS_PATH)
    expected_tables = {t.table_name for t in contract.all_tables}
    expected_tables |= {r.table_name for r in drivers.all_registers}
    expected_tables |= {g.table_name for g in structure.all_grids}
    found_tables = {
        name for ws in workbook.worksheets for name in getattr(ws, "tables", {})
    }
    assert found_tables == expected_tables, (
        f"unexpected {sorted(found_tables - expected_tables)}, "
        f"missing {sorted(expected_tables - found_tables)}"
    )


def test_17_no_external_links_or_connections() -> None:
    with zipfile.ZipFile(_artifact()) as archive:
        names = archive.namelist()
    assert not [n for n in names if "externalLink" in n]
    assert not [n for n in names if "connections" in n.lower()]
    assert not [n for n in names if "queryTable" in n]


# ---------------------------------------------------------------------------
# structural reproducibility (logical, not byte-identical)
# ---------------------------------------------------------------------------
def test_19_structurally_reproducible() -> None:
    first = _build("repro_a", timestamp="2000-01-01 00:00:00 UTC")
    second = _build("repro_b", timestamp="2000-01-01 00:00:00 UTC")
    assert structural_digest(first) == structural_digest(second), (
        "two builds of the same source produced different workbook structure"
    )


def test_21_fx_single_source_of_truth() -> None:
    """Exactly one FX-rate owner: Setup. Config owns the currency master only.

    Asserted at the architectural boundary - which table holds a rate column -
    rather than by a blanket rule such as "Config may contain no number", which
    would wrongly block legitimate future configuration metadata.
    """
    contract = load_contract(CONTRACT_PATH)
    workbook = load_workbook(_artifact())

    config_tables = getattr(workbook["Config"], "tables", {})
    setup_tables = getattr(workbook["Setup"], "tables", {})

    # Config owns the currency master list...
    assert "tblCurrencies" in config_tables, "Config is missing tblCurrencies"
    currencies = next(t for t in contract.all_tables if t.table_name == "tblCurrencies")
    assert currencies.sheet == "Config"

    # ...with no FX or rate column of any kind.
    rate_columns = [
        c.header for c in currencies.columns
        if "fx" in c.header.lower() or "rate" in c.header.lower() or "sar" in c.header.lower()
    ]
    assert not rate_columns, f"tblCurrencies has rate-like columns: {rate_columns}"
    assert len(currencies.columns) == 1, "tblCurrencies must hold the currency code only"

    # Setup owns the FX rates.
    assert "tblFXRates" in setup_tables, "Setup is missing tblFXRates"
    fx = next(t for t in contract.all_tables if t.table_name == "tblFXRates")
    assert fx.sheet == "Setup"
    assert "FX to SAR" in [c.header for c in fx.columns], "tblFXRates lacks the 'FX to SAR' column"

    # Exactly one rate-bearing table exists anywhere in the workbook.
    rate_tables = [
        t.table_name for t in contract.all_tables
        if any("fx" in c.header.lower() or "rate" in c.header.lower() for c in t.columns)
    ]
    assert rate_tables == ["tblFXRates"], f"more than one FX-rate table: {rate_tables}"

    # The convention is stated on Setup.
    setup_values = _strings(workbook["Setup"])
    assert any(contract.fx_convention in v for v in setup_values), (
        f"Setup does not state the FX convention {contract.fx_convention!r}"
    )


def test_20_manifest_locked_order_matches_architecture() -> None:
    spec = load_spec(SPEC_PATH)
    assert spec.workbook["locked_sheet_order"] == LOCKED_SHEET_ORDER
    assert spec.sheet_names == LOCKED_SHEET_ORDER
    assert spec.active_sheet == ACTIVE_SHEET


# ---------------------------------------------------------------------------
# standalone runner
# ---------------------------------------------------------------------------
def _run_all() -> int:
    tests = sorted(
        (name, fn)
        for name, fn in globals().items()
        if name.startswith("test_") and callable(fn)
    )
    failures: list[tuple[str, str]] = []
    print("PCCM Phase 1 structural tests")
    print("=" * 66)
    for name, fn in tests:
        try:
            fn()
        except AssertionError as error:
            failures.append((name, str(error) or "assertion failed"))
            print(f"  [FAIL] {name}")
            print(f"         {error}")
        except Exception as error:  # noqa: BLE001 - report any unexpected error
            failures.append((name, f"{type(error).__name__}: {error}"))
            print(f"  [ERROR] {name}")
            print(f"          {type(error).__name__}: {error}")
        else:
            print(f"  [PASS] {name}")
    print("=" * 66)
    print(f"  {len(tests) - len(failures)} passed, {len(failures)} failed")
    if _TEMPDIR is not None:
        _TEMPDIR.cleanup()
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(_run_all())
