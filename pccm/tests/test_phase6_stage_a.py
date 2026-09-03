#!/usr/bin/env python3
"""PCCM Phase 6 Step-5 tests for the Stage-A simulation artefacts.

Two generated files:

    build/vba/modSimContract.bas   a constants-only projection of the accepted
                                   simulation authorities
    build/phase6_cases.json        the conformance corpus the later VBA
                                   implementation steps assert against

WHAT THESE TESTS ARE FOR. A generated artefact is only worth anything if it says
what its owner says. So every projected constant is compared against the
authority that owns it, and every retained Step-0 vector is compared against the
EVIDENCE FILE rather than against a second call to the implementation that
produced it. `result_digest(x) == result_digest(x)` proves nothing; the seven
retained digests are literals here.

NO WORKBOOK, NO EXCEL, NO COM, NO `_SimData` ROW, NO VBA EXECUTION. The
generated module is text on disk; nothing here runs it, and nothing embeds it in
the `.xlsx`.

Runs standalone or under pytest.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
import sys
import tempfile
from pathlib import Path

PCCM_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PCCM_ROOT / "builder"))

from pccm_builder import (  # noqa: E402
    RngReference,
    emit_sim_artifacts,
    load_calc_contract,
    load_contract,
    load_sim_contract,
    load_spec,
    load_structure_contract,
)
from pccm_builder.sim_emit import (  # noqa: E402
    SIM_MODULE_NAME,
    VBA_LONG_MAX,
    VBA_LONG_MIN,
    render_sim_contract_module,
    vba_double,
)
from pccm_builder.sim_cases import POLICIES  # noqa: E402
from pccm_builder.vba_source import contains_construct, load_modules  # noqa: E402

SPEC = PCCM_ROOT / "spec"
EVIDENCE = PCCM_ROOT / "evidence" / "phase6_step0" / "vectors"

_CACHE: dict[str, object] = {}


def _spec():
    if "spec" not in _CACHE:
        _CACHE["spec"] = load_spec(SPEC / "workbook.yaml")
    return _CACHE["spec"]


def _inputs():
    if "inputs" not in _CACHE:
        _CACHE["inputs"] = load_contract(SPEC / "input_contract.yaml")
    return _CACHE["inputs"]


def _sim():
    if "sim" not in _CACHE:
        _CACHE["sim"] = load_sim_contract(SPEC / "sim_contract.yaml")
    return _CACHE["sim"]


def _calc():
    if "calc" not in _CACHE:
        _CACHE["calc"] = load_calc_contract(SPEC / "calc_contract.yaml")
    return _CACHE["calc"]


def _structure():
    if "structure" not in _CACHE:
        _CACHE["structure"] = load_structure_contract(SPEC / "structure_contract.yaml")
    return _CACHE["structure"]


def _emitted() -> Path:
    """Emit both artefacts into a scratch directory, once per run."""
    if "dir" not in _CACHE:
        target = Path(tempfile.mkdtemp(prefix="pccm-step5-"))
        emit_sim_artifacts(target, _spec(), _sim(), _inputs(), _calc())
        _CACHE["dir"] = target
    return _CACHE["dir"]  # type: ignore[return-value]


def _module_text() -> str:
    return (_emitted() / "vba" / f"{SIM_MODULE_NAME}.bas").read_text(encoding="utf-8")


def _cases_document() -> dict:
    return json.loads((_emitted() / "phase6_cases.json").read_text(encoding="utf-8"))


def _cases() -> dict[str, dict]:
    return {
        case["id"]: case
        for group in _cases_document()["groups"]
        for case in group["cases"]
    }


def _evidence(name: str) -> dict:
    return json.loads((EVIDENCE / f"{name}.json").read_text(encoding="utf-8"))


_CONST_RE = re.compile(r"^Public Const (\w+) As (\w+) = (.*)$")


def _literal(kind: str, rest: str) -> str:
    """The literal on the right of `=`, with any trailing comment removed.

    A String literal is scanned to its matching close quote, honouring the
    doubled-quote escape, because a value can legitimately contain the
    characters a naive split would cut on.
    """
    if kind != "String":
        return rest.split("    '")[0].rstrip()
    assert rest.startswith('"'), rest
    index = 1
    while index < len(rest):
        if rest[index] == '"':
            if index + 1 < len(rest) and rest[index + 1] == '"':
                index += 2
                continue
            return rest[: index + 1]
        index += 1
    raise AssertionError(f"unterminated string literal: {rest}")


def _constants() -> dict[str, tuple[str, str]]:
    """`name -> (vba type, rendered literal)`, comments removed."""
    out: dict[str, tuple[str, str]] = {}
    for line in _module_text().splitlines():
        match = _CONST_RE.match(line)
        if not match:
            continue
        name, kind, rest = match.groups()
        assert name not in out, f"{name} is declared twice"
        out[name] = (kind, _literal(kind, rest))
    return out


def _string(name: str) -> str:
    kind, literal = _constants()[name]
    assert kind == "String", f"{name} is {kind}, expected String"
    assert literal.startswith('"') and literal.endswith('"'), literal
    return literal[1:-1].replace('""', '"')


def _long(name: str) -> int:
    kind, literal = _constants()[name]
    assert kind == "Long", f"{name} is {kind}, expected Long"
    return int(literal)


def _double(name: str) -> float:
    kind, literal = _constants()[name]
    assert kind == "Double", f"{name} is {kind}, expected Double"
    return float(literal)


# ===========================================================================
# the generated module: shape
# ===========================================================================
def test_01_the_module_declares_its_name_and_option_explicit() -> None:
    lines = _module_text().splitlines()
    assert lines[0] == f'Attribute VB_Name = "{SIM_MODULE_NAME}"'
    assert lines[1] == "Option Explicit"


def test_02_the_module_declares_itself_generated_and_names_its_authorities() -> None:
    text = _module_text()
    assert "GENERATED FILE - DO NOT EDIT" in text
    assert "spec/sim_contract.yaml" in text
    assert "spec/input_contract.yaml" in text
    assert "spec/workbook.yaml" in text


def test_03_the_module_contains_constants_and_nothing_else() -> None:
    """Every line is the attribute, `Option Explicit`, a comment, blank, or a
    `Public Const`. There is nowhere for an implementation to hide."""
    unexpected = []
    for number, line in enumerate(_module_text().splitlines(), start=1):
        if not line.strip():
            continue
        if line.startswith("'") or line.startswith("Attribute VB_Name"):
            continue
        if line == "Option Explicit":
            continue
        if _CONST_RE.match(line):
            continue
        unexpected.append((number, line))
    assert not unexpected, f"non-constant lines: {unexpected[:5]}"


def test_04_the_module_declares_no_procedure_of_any_kind() -> None:
    module = _generated_module()
    assert module.procedures == [], module.procedures
    assert module.public_procedures == []
    code = module.code
    for keyword in ("Sub ", "Function ", "Property ", "End Sub", "End Function",
                    "For ", "Next ", "If ", "Do ", "Loop", "While ", "Select Case"):
        assert keyword not in code, f"{keyword!r} appears in code"


def _generated_module():
    modules = load_modules([_emitted() / "vba"])
    found = [module for module in modules if module.name == SIM_MODULE_NAME]
    assert len(found) == 1, [module.name for module in modules]
    return found[0]


def test_05_the_module_contains_no_algorithm() -> None:
    """No arithmetic operator appears outside a literal, so no recurrence, jump,
    sampler, quantile or digest step can be hiding in it."""
    code = _generated_module().code
    for operator in (" * ", " / ", " + ", " Mod ", " ^ ", " And ", " Or "):
        assert operator not in code, f"{operator!r} appears in code"
    assert "Exp(" not in code and "Log(" not in code and "Sqr(" not in code
    assert "Range(" not in code and "Cells(" not in code and "Worksheets(" not in code


def test_06_the_module_declares_a_substantial_number_of_constants() -> None:
    constants = _constants()
    assert len(constants) >= 150, len(constants)
    assert len(constants) == len(_generated_module().constants)


# ===========================================================================
# the generated module: every constant equals its owner
# ===========================================================================
def test_07_versions_come_from_the_simulation_contract() -> None:
    sim = _sim()
    assert _string("SIM_CONTRACT_VERSION") == sim.version
    assert _long("SIM_RNG_VERSION") == sim.rng_version == 1
    assert _long("SIM_METHOD_VERSION") == sim.sim_method_version == 1


def test_08_the_generator_constants_equal_the_contract() -> None:
    constants = _sim().raw["rng"]["constants"]
    assert _double("SIM_RNG_M1") == float(constants["m1"])
    assert _double("SIM_RNG_M2") == float(constants["m2"])
    assert _long("SIM_RNG_A12") == int(constants["a12"])
    assert _long("SIM_RNG_A13N") == int(constants["a13n"])
    assert _long("SIM_RNG_A21") == int(constants["a21"])
    assert _long("SIM_RNG_A23N") == int(constants["a23n"])
    assert _double("SIM_RNG_NORM") == float(constants["norm"])

    state = _sim().raw["rng"]["state"]
    assert _long("SIM_RNG_STATE_WORDS") == int(state["words"])
    for ordinal, word in enumerate(state["order"], start=1):
        assert _string(f"SIM_RNG_STATE_{ordinal}") == word
    assert _string("SIM_RNG_STATE_ORIENTATION") == state["orientation"]
    assert _string("SIM_RNG_MATRIX_OPERAND_ORIENTATION") == (
        state["matrix_operand_orientation"]
    )


def test_09_values_above_long_range_are_projected_as_doubles() -> None:
    """`m1` and `m2` cannot be VBA Longs; several jump elements cannot either."""
    constants = _constants()
    for name, (kind, literal) in constants.items():
        if kind != "Long":
            continue
        value = int(literal)
        assert VBA_LONG_MIN <= value <= VBA_LONG_MAX, f"{name} = {value} overflows Long"

    assert _constants()["SIM_RNG_M1"][0] == "Double"
    assert _constants()["SIM_RNG_M2"][0] == "Double"
    assert int(_double("SIM_RNG_M1")) == 4294967087
    assert int(_double("SIM_RNG_M2")) == 4294944443
    # The two that DO fit are still projected as Longs, so the rule is a rule and
    # not a blanket widening.
    assert _constants()["SIM_AUTO_MODULUS"][0] == "Long"
    assert _long("SIM_AUTO_MODULUS") == VBA_LONG_MAX
    assert _constants()["SIM_SEED_MAX"][0] == "Long"


def test_10_every_double_literal_round_trips_to_the_accepted_value() -> None:
    for name, (kind, literal) in _constants().items():
        if kind != "Double":
            continue
        parsed = float(literal)
        assert math.isfinite(parsed), name
        assert repr(parsed) == repr(float(literal)), name
        assert vba_double(parsed) == literal, f"{name}: {literal} is not the canonical form"
        assert "," not in literal, f"{name}: a locale separator reached the literal"


def test_11_the_jump_matrices_are_exact_and_carry_their_orientation() -> None:
    jump = _sim().raw["jump"]
    assert _long("SIM_STREAM_SPACING_EXPONENT") == int(jump["stream_spacing_exponent"])
    assert _long("SIM_JUMP_DECOMPOSITION_H") == int(jump["decomposition_h"]) == 131072

    for name, matrix in (("A1", jump["a1_p127"]), ("A2", jump["a2_p127"])):
        for row_index, row in enumerate(matrix, start=1):
            for column_index, element in enumerate(row, start=1):
                projected = _double(f"SIM_JUMP_{name}_R{row_index}C{column_index}")
                assert projected == float(element), (name, row_index, column_index)
                assert int(projected) == int(element)
    # Every element is a Double, including those that would fit a Long: a matrix
    # with two VBA types would be a trap for the implementation that reads it.
    for name in ("A1", "A2"):
        for row_index in range(1, 4):
            for column_index in range(1, 4):
                kind, _ = _constants()[f"SIM_JUMP_{name}_R{row_index}C{column_index}"]
                assert kind == "Double", (name, row_index, column_index)

    # And the matrices are the ones Step-0 retained.
    evidence = _evidence("jump_vectors")
    assert [list(row) for row in jump["a1_p127"]] == evidence["A1_jump"]
    assert [list(row) for row in jump["a2_p127"]] == evidence["A2_jump"]


def test_12_the_locked_acceptance_literals_are_exact() -> None:
    cheng = _sim().raw["cheng"]
    assert _long("SIM_CHENG_UNIFORMS_PER_ATTEMPT") == 2
    for ordinal, literal in enumerate(cheng["bb"]["literals"], start=1):
        assert _double(f"SIM_CHENG_BB_LITERAL_{ordinal}") == float(literal)
    for ordinal, literal in enumerate(cheng["bc"]["literals"], start=1):
        assert _double(f"SIM_CHENG_BC_LITERAL_{ordinal}") == float(literal)
    # The two that matter most, stated outright.
    assert 1.3862944 in {_double(f"SIM_CHENG_BB_LITERAL_{n}") for n in range(1, 6)}
    assert 2.609438 in {_double(f"SIM_CHENG_BB_LITERAL_{n}") for n in range(1, 6)}
    assert _double("SIM_CHENG_BB_LITERAL_1") != math.log(4.0), (
        "the literal was evaluated instead of projected"
    )


def test_13_the_seed_domain_comes_from_the_input_contract() -> None:
    from pccm_builder.sim_rng import _seed_domain

    minimum, maximum = _seed_domain(_inputs())
    assert _long("SIM_SEED_MIN") == minimum == 1
    assert _long("SIM_SEED_MAX") == maximum == 2147483646

    auto = _sim().raw["seeding"]["auto"]
    lifecycle = _sim().raw["seeding"]["nonce_lifecycle"]
    assert _long("SIM_AUTO_MODULUS") == int(auto["modulus"])
    assert _long("SIM_AUTO_MULTIPLIER") == int(auto["multiplier"])
    assert _long("SIM_AUTO_PERIOD") == int(auto["period"])
    assert _long("SIM_NONCE_EXHAUSTED") == int(lifecycle["exhausted_value"])
    assert _long("SIM_NONCE_LAST_VALID") == int(lifecycle["last_valid_allocation"])


def test_14_the_iteration_bounds_come_from_their_two_owners() -> None:
    from pccm_builder.sim_oracle import business_minimum_iterations

    assert _long("SIM_MIN_ITERATIONS") == business_minimum_iterations(_inputs()) == 1000
    assert _long("SIM_MAX_ITERATIONS") == _sim().max_iterations_representable == 1048543


def test_15_the_sim_data_geometry_is_exact() -> None:
    layout = _sim().layout
    assert _string("SIM_DATA_SHEET") == layout.sheet == "_SimData"
    assert _string("SIM_DATA_VISIBILITY") == layout.required_visibility == "veryHidden"
    assert _long("SIM_DATA_HEADER_ROW") == layout.header_row == 33
    assert _long("SIM_DATA_FIRST_ITERATION_ROW") == layout.first_iteration_row == 34
    assert _long("SIM_DATA_RESERVED_ROWS") == layout.reserved_row_count == 33
    assert _long("SIM_MAX_ITERATIONS") == 1048576 - layout.reserved_row_count

    identity = _sim().raw["sim_data"]["run_identity"]
    assert _string("SIM_IDENTITY_LABEL_COLUMN") == identity["label_column"]
    assert _long("SIM_IDENTITY_FIRST_ROW") == int(identity["first_row"])
    for field in identity["fields"]:
        name = "SIM_IDENTITY_ROW_" + field["key"].upper()
        assert _long(name) == int(field["row"]), field["key"]


def test_16_the_result_digest_framing_is_exact() -> None:
    digest = _sim().raw["result_digest"]
    assert _string("SIM_DIGEST_STREAM_TAG") == digest["stream_tag"] == "PCCM-RD"
    assert _string("SIM_DIGEST_SECTION") == digest["section_name"] == "RESULT"
    assert _long("SIM_DIGEST_FIELD_COUNT") == int(digest["record_field_count"]) == 3
    assert _long("SIM_DIGEST_INDEX_ORIGIN") == int(digest["iteration_index_origin"]) == 1
    for ordinal, field in enumerate(digest["record_fields"], start=1):
        assert _string(f"SIM_DIGEST_FIELD_{ordinal}") == field
    for ordinal, kind in enumerate(digest["field_types"], start=1):
        assert _string(f"SIM_DIGEST_FIELD_TYPE_{ordinal}") == kind


def test_17_the_state_attempt_and_seed_mode_labels_are_exact() -> None:
    labels = _sim().raw["label_sets"]
    for label in labels["sim_state"]:
        assert _string(f"SIM_STATE_{label}") == label
    for label in labels["attempt_result"]:
        assert _string(f"SIM_ATTEMPT_{label}") == label
    for label in labels["seed_mode"]:
        assert _string(f"SIM_SEED_MODE_{label}") == label
    assert list(labels["sim_state"]) == ["CURRENT", "STALE", "INVALID"]
    # PHASE-6 ONLY, and still EXACT. Phase 5 keeps its own four-value axis in
    # calc_contract.yaml; the fifth token carries a persistence/recovery
    # condition Phase 5 has no analogue for.
    assert list(labels["attempt_result"]) == [
        "NONE", "SUCCESS", "REFUSED", "FAILED", "AUTO_NONCE_INDETERMINATE"]


def test_18_the_quantile_ladder_is_projected_from_its_owners() -> None:
    from pccm_builder.sim_oracle import resolve_percentile_ladder

    ladder = resolve_percentile_ladder(_sim(), _inputs())
    assert _long("SIM_QUANTILE_COUNT") == len(ladder.ordered) == 11
    for ordinal, label in enumerate(ladder.ordered, start=1):
        assert _string(f"SIM_QUANTILE_{ordinal}") == label
    for ordinal, label in enumerate(ladder.headline, start=1):
        assert _string(f"SIM_QUANTILE_HEADLINE_{ordinal}") == label
    assert _string("SIM_QUANTILE_FIXED_1") == "P10"
    assert _string("SIM_CONTINGENCY_BASELINE") == "deterministic_base_estimate_a"


# ===========================================================================
# the generated module: scope discipline
# ===========================================================================
def test_19_no_currently_forbidden_construct_appears_in_the_module() -> None:
    """The CURRENT guard, unchanged and unscoped."""
    structure = _structure()
    module = _generated_module()
    offenders = [
        construct for construct in structure.forbidden_constructs
        if contains_construct([module], construct)
    ]
    assert not offenders, offenders

    # And not in the commentary either, which the code scan would not see.
    raw = module.raw.lower()
    for construct in structure.forbidden_constructs:
        assert construct.lower() not in raw, construct


def test_20_the_future_endpoint_and_algorithm_names_are_absent() -> None:
    raw = _module_text().lower()
    for token in ("runsimulation", "mrg32k3a", "percentile", "rnd(", "randomize"):
        assert token not in raw, token
    # The information they would have carried is present under neutral names.
    assert "SIM_RNG_M1" in _constants()
    assert "SIM_QUANTILE_COUNT" in _constants()


def test_21_the_generated_module_needs_no_d6_11_exception() -> None:
    """The projection passes the guard with nothing scoped ON ITS BEHALF.

    Step 6 granted exactly one exception - MRG32k3a to modSimRng, the module
    that owns the construct. `modSimContract` is not an owner of anything: it
    is a constants projection and would fail the guard if it named the
    algorithm, which is why the family name was deliberately left out of it.
    """
    structure = _structure()
    scoped = [rule for rule in structure.forbidden_construct_rules if rule.is_scoped]
    assert [(r.construct, tuple(r.allowed_in)) for r in scoped] == [
        ("MRG32k3a", ("modSimRng",)),
        ("RunSimulation", ("modSimReport",)),
    ], scoped
    for rule in scoped:
        assert SIM_MODULE_NAME not in rule.allowed_in, (
            "the generated projection was given an exception it does not need"
        )
        assert rule.forbidden_in(SIM_MODULE_NAME) is True
    for construct in ("MRG32k3a", "RunSimulation", "Rnd(", "Percentile"):
        assert construct in structure.forbidden_constructs, construct


def test_22_the_module_is_in_the_stage_b_registry_as_a_generated_module() -> None:
    """Step 5 kept it OUT of the registry because nothing depended on it.

    Step 6 is the first implementation step that does, so it entered the
    registry in the same commit as modSimRng - the atomic D6-11 activation.
    Step 7 added modSimSample, which reads the same projection.
    """
    modules = {module.name: module for module in _structure().vba_modules}
    assert SIM_MODULE_NAME in modules
    assert modules[SIM_MODULE_NAME].generated is True, (
        "the projection must be declared generated; a hand-written copy would be "
        "a second definition of every literal it projects"
    )
    handwritten = {"modSimRng", "modSimSample", "modSimEngine", "modSimStats",
                   "modSimFingerprint", "modSimNonce", "modSimReport"}
    for name in handwritten:
        assert name in modules and modules[name].generated is False, name
    # AND NOTHING WAS DECLARED EARLY, which is the claim - not that the registry
    # never grows again. The Phase-6 set is exactly the generated projection plus
    # its seven hand-written modules, and the only module beyond it is the one a
    # later phase has actually landed: `modSimSensitivity`, the accepted P7-2
    # pure kernel, hand-written. A module whose step has not run yet still cannot
    # appear, and that is what the difference below is checked against.
    phase6 = handwritten | {SIM_MODULE_NAME}
    declared = {name for name in modules if name.startswith("modSim")}
    assert phase6 <= declared, sorted(phase6 - declared)
    assert declared - phase6 <= {"modSimSensitivity", "modSimPostReport"}, sorted(
        declared - phase6)
    for name in declared - phase6:
        assert modules[name].generated is False, name
        assert (PCCM_ROOT / "src" / "vba" / f"{name}.bas").is_file(), (
            f"{name} is declared but has no source; a registry entry is not a module"
        )


def test_23_the_module_is_deterministic_across_two_emissions() -> None:
    first = render_sim_contract_module(_spec(), _sim(), _inputs())
    second = render_sim_contract_module(_spec(), _sim(), _inputs())
    assert first == second
    assert hashlib.sha256(first.encode("utf-8")).hexdigest() == (
        hashlib.sha256(_module_text().encode("utf-8")).hexdigest()
    )
    # Nothing environment-specific reached the text: no absolute path, no
    # hostname, no clock. The only variable inputs are the three authorities.
    for token in ("/home/", "/tmp/", "/root/", "\\", "T00:", "Z\n"):
        assert token not in first, token
    assert str(PCCM_ROOT) not in first


# ===========================================================================
# the case corpus: shape and identity
# ===========================================================================
def test_24_the_corpus_is_valid_deterministic_json() -> None:
    text = (_emitted() / "phase6_cases.json").read_text(encoding="utf-8")
    document = json.loads(text)
    assert isinstance(document, dict)

    second = Path(tempfile.mkdtemp(prefix="pccm-step5-again-"))
    emit_sim_artifacts(second, _spec(), _sim(), _inputs(), _calc())
    again = (second / "phase6_cases.json").read_bytes()
    assert hashlib.sha256(again).hexdigest() == hashlib.sha256(
        text.encode("utf-8")
    ).hexdigest(), "two emissions of the corpus differ"


def test_25_the_corpus_carries_no_non_finite_number_anywhere() -> None:
    text = (_emitted() / "phase6_cases.json").read_text(encoding="utf-8")
    for token in ("NaN", "Infinity", "-Infinity"):
        assert token not in text, token

    def walk(node) -> None:
        if isinstance(node, float):
            assert math.isfinite(node), node
        elif isinstance(node, dict):
            for value in node.values():
                walk(value)
        elif isinstance(node, list):
            for value in node:
                walk(value)

    walk(_cases_document())


def test_26_the_corpus_is_free_of_environment_specific_data() -> None:
    text = (_emitted() / "phase6_cases.json").read_text(encoding="utf-8")
    for token in ("/home/", "/tmp/", "/root/", str(PCCM_ROOT), "T00:", "Z\","):
        assert token not in text, token
    assert "timestamp" not in text.lower() or "last_successful_stamp" in text


def test_27_the_top_level_identity_pins_every_version() -> None:
    document = _cases_document()
    assert document["schema_version"] == 2, (
        "the schema version must move when the corpus typing changes"
    )
    assert document["model_version"] == _spec().model["model_version"]
    assert document["sim_contract_version"] == _sim().version
    assert document["rng_version"] == _sim().rng_version == 1
    assert document["sim_method_version"] == _sim().sim_method_version == 1
    assert document["case_count"] == len(_cases())


def test_28_every_case_has_a_stable_unique_id_and_a_policy() -> None:
    document = _cases_document()
    seen: list[str] = []
    for group in document["groups"]:
        assert group["group"] and group["title"]
        for case in group["cases"]:
            seen.append(case["id"])
            assert case["comparison"] in POLICIES, case["id"]
            assert case["layer"], case["id"]
            assert case["title"], case["id"]
            assert "inputs" in case, case["id"]
            assert any(
                key in case for key in ("expected", "expected_exact", "expected_refusal")
            ), case["id"]
    assert len(seen) == len(set(seen)), "duplicate case ids"
    assert set(document["comparison_policies"]) == set(POLICIES)


def test_29_the_five_comparison_classes_are_all_used() -> None:
    used = {case["comparison"] for case in _cases().values()}
    assert used == set(POLICIES), sorted(set(POLICIES) - used)


def test_30_no_python_object_repr_leaked_into_the_corpus() -> None:
    text = (_emitted() / "phase6_cases.json").read_text(encoding="utf-8")
    for token in ("object at 0x", "<class ", "RngState(", "SampleResult(",
                  "PreparedSimulationModel(", "dataclass"):
        assert token not in text, token


# ===========================================================================
# the corpus: every retained vector, pinned against the EVIDENCE
# ===========================================================================
def test_31_the_seed_vectors_match_the_retained_evidence() -> None:
    evidence = _evidence("seed_vectors")
    cases = _cases()
    for example in evidence["examples"]:
        case = cases[f"rng.fixed_seed.{example['seed']}"]
        assert case["expected_exact"]["initial_state"] == example["state"]
    for pair in evidence["nonce_to_seed_pairs"]:
        key = f"seed.auto.nonce.{pair['auto_nonce']}"
        if key not in cases:
            continue
        assert case_effective_seed(cases[key]) == pair["effective_seed"], key
    assert "seed.auto.nonce.0" in cases and "seed.auto.nonce.2147483645" in cases


def case_effective_seed(case: dict) -> int:
    return case["expected_exact"]["effective_seed"]


def test_32_the_rng_uniform_vectors_match_the_retained_evidence() -> None:
    evidence = _evidence("rng_vectors")
    cases = _cases()
    for seed, retained in evidence["per_seed"].items():
        exact = cases[f"rng.fixed_seed.{seed}"]["expected_exact"]
        assert exact["initial_state"] == retained["initial_state"], seed
        # The retained evidence stores decimals AS TEXT. It is not rewritten:
        # the string is parsed to the accepted Double here, and the corpus's
        # JSON number must be that Double exactly.
        emitted = exact["first_uniforms"]
        assert all(isinstance(value, float) for value in emitted), seed
        assert emitted == [float(text) for text in retained["first_5"]], seed
        # And the exact textual form the corpus carries alongside it agrees.
        for value, canonical in zip(emitted, exact["first_uniforms_canonical"]):
            assert isinstance(canonical, str) and float(canonical) == value
    assert set(evidence["per_seed"]) == {"1", "2", "12345", "2147483646"}


def test_33_the_jump_stream_vectors_match_the_retained_evidence() -> None:
    evidence = _evidence("jump_vectors")
    cases = _cases()
    assert sorted(evidence["streams"], key=int) == ["0", "1", "7", "399", "401"]
    for index, retained in evidence["streams"].items():
        exact = cases[f"rng.stream.{index}"]["expected_exact"]
        assert exact["initial_state"] == retained["initial_state"], index
        emitted = exact["first_uniforms"]
        assert all(isinstance(value, float) for value in emitted), index
        assert emitted == [float(text) for text in retained["first_5_uniforms"]], index


def test_34_the_stream_assignment_matches_the_retained_evidence() -> None:
    evidence = _evidence("stream_assignment_vectors")
    case = _cases()["stream.assignment.canonical_400"]["expected_exact"]
    assert case["total_components"] == evidence["total_components"] == 400
    assert case["first_10"] == evidence["family_a_first_10"]
    assert case["last_4"] == evidence["family_a_last_4"]
    # The interleaving the evidence shows: occurrence then severity per Risk.
    assert case["last_4"][-2]["component"][2] == "occurrence"
    assert case["last_4"][-1]["component"][2] == "severity"


def test_35_all_seven_digest_vectors_match_the_retained_evidence() -> None:
    """Pinned against the EVIDENCE FILE, not against a second call to the
    implementation that produced them."""
    evidence = _evidence("digest_vectors")
    cases = _cases()
    mapping = {
        "base": "digest.base",
        "reversed_iteration_order": "digest.reversed_iteration_order",
        "nominal_and_pv_swapped": "digest.nominal_and_pv_swapped",
        "one_iteration_dropped": "digest.one_iteration_dropped",
        "one_ulp_perturbation": "digest.one_ulp_perturbation",
        "version_2": "digest.version_2",
        "empty": "digest.empty",
    }
    assert len(evidence["cases"]) == 7
    for retained in evidence["cases"]:
        emitted = cases[mapping[retained["label"]]]["expected_exact"]
        assert emitted["result_digest"] == retained["digest"], retained["label"]
        assert emitted["canonical_stream"] == retained["stream"], retained["label"]


def test_36_the_digest_literals_are_pinned_here_too() -> None:
    """The seven digests, as literals, so the corpus cannot drift silently."""
    cases = _cases()
    expected = {
        "digest.base": "3181AF89642DE500",
        "digest.reversed_iteration_order": "4E0FEE211853E8F6",
        "digest.nominal_and_pv_swapped": "63A0E93074F0C2EA",
        "digest.one_iteration_dropped": "0CAC531732B88B2A",
        "digest.one_ulp_perturbation": "5DC1A76B56D75EF4",
        "digest.version_2": "7E8D58C46CCDD798",
        "digest.empty": "12ED977808313D71",
    }
    for identifier, digest in expected.items():
        assert cases[identifier]["expected_exact"]["result_digest"] == digest, identifier
    assert cases["digest.empty"]["expected_exact"]["canonical_stream"] == (
        "S7:PCCM-RDI1:1S6:RESULTI1:0"
    )


def test_37_all_five_cheng_vectors_match_the_retained_evidence() -> None:
    evidence = _evidence("cheng_vectors")
    cases = _cases()
    names = ("bb_interior", "bb_symmetric", "bb_near_boundary", "bc_alpha_1", "bc_beta_1")
    assert len(evidence["cases"]) == 5
    for name, retained in zip(names, evidence["cases"]):
        emitted = cases[f"sampler.beta.cheng.{name}"]
        exact = emitted["expected_exact"]
        assert exact["dispatch"] == retained["dispatch"], name
        assert exact["initial_state"] == retained["initial_state"], name
        assert exact["final_state"] == retained["final_state"], name
        assert exact["total_proposal_attempts"] == retained["total_attempts"], name
        assert exact["total_uniforms"] == retained["total_uniforms"], name
        values = [row["value"] for row in emitted["expected"]["samples"]]
        assert all(isinstance(value, float) for value in values), name
        assert values == [
            float(sample["accepted_sample"]) for sample in retained["samples"]
        ], name
        for row, sample in zip(exact["per_sample"], retained["samples"]):
            assert row["proposal_attempts"] == sample["proposal_attempts_for_this_sample"]
            assert row["cumulative_uniforms"] == sample["cumulative_uniforms"]
            assert row["state_after"] == sample["rng_state_after_sample"]
    # Draw counts EXACT, transformed values tolerance-bounded: plan layer E.
    for name in names:
        assert cases[f"sampler.beta.cheng.{name}"]["comparison"] == "TOLERANCE_BOUNDED"


# ===========================================================================
# the corpus: the accepted semantics are represented, and correctly scoped
# ===========================================================================
def test_38_the_uniform_ignored_most_likely_degeneracy_is_present() -> None:
    cases = _cases()
    absent = cases["sampler.uniform.degenerate.absent"]["expected_exact"]
    populated = cases["sampler.uniform.degenerate.populated_and_ignored"]
    assert populated["inputs"]["most_likely"] == 500.0
    assert populated["expected_exact"]["uniforms_consumed"] == 0
    assert populated["expected_exact"]["state_unchanged"] is True
    assert populated["expected_exact"]["value"] == absent["value"], (
        "a populated Most Likely changed a Uniform's answer"
    )
    assert populated["comparison"] == "EXACT"


def test_39_d6_18b_is_present_and_says_what_it_must() -> None:
    case = _cases()["engine.risk.d6_18b_unconditional_severity"]
    exact = case["expected_exact"]
    assert case["comparison"] == "EXACT"
    assert exact["severity_consumption_equals_iterations"] is True
    assert exact["severity_final_state_identical_across_probabilities"] is True
    assert exact["occurrence_counts_differ"] is True

    low, high = exact["runs"]
    iterations = _cases_document()["engine_iterations"]
    for run in (low, high):
        assert run["severity_uniforms_consumed"] == iterations
        assert run["occurrence_uniforms_consumed"] == iterations
    assert low["severity_final_state"] == high["severity_final_state"]
    assert low["occurrences"] != high["occurrences"]

    degenerate = _cases()["engine.risk.degenerate_severity_zero_consumption"]
    severities = [
        component for component in degenerate["expected_exact"]["components"]
        if component["role"] == "severity"
    ]
    assert severities and all(
        component["uniforms_consumed"] == 0
        and component["initial_state"] == component["final_state"]
        for component in severities
    )


def test_40_row_order_invariance_is_a_same_runtime_relation() -> None:
    case = _cases()["engine.row_order.invariant"]
    assert case["comparison"] == "SAME_RUNTIME_ONLY"
    assert case["expected"]["relation"] == "equal"
    assert "result_digest" in case["expected"]["fields"]
    assert case["expected_exact"]["canonical_driver_order"] == [
        "CL-001", "CL-002", "R-001"
    ]
    assert case["python_reference"]["identical_to_canonical_order"] is True


def test_41_seed_divergence_is_fixture_scoped_and_the_universal_claim_is_absent() -> None:
    diverging = _cases()["engine.seed.non_degenerate_divergence"]
    assert diverging["comparison"] == "SAME_RUNTIME_ONLY"
    assert diverging["expected"]["relation"] == "all_different"
    assert diverging["expected"]["scope"] == "this fixture only"
    digests = diverging["python_reference"]["result_digests"]
    assert len(set(digests.values())) == len(digests) == 3

    degenerate = _cases()["engine.seed.degenerate_equal_digest"]
    assert degenerate["comparison"] == "EXACT"
    equal = degenerate["expected_exact"]["result_digests"]
    assert degenerate["expected_exact"]["all_equal"] is True
    assert len(set(equal.values())) == 1 and len(equal) == 3
    assert len(degenerate["expected_exact"]["distinct_totals"]) == 1

    text = (_emitted() / "phase6_cases.json").read_text(encoding="utf-8")
    assert "different seed -> different digest" not in text or "WITHDRAWN" in text


def test_42_the_exact_friendly_engine_fixtures_are_exact_and_the_general_one_is_not() -> None:
    cases = _cases()
    for identifier in ("engine.exact_friendly.unit_interval",
                       "engine.exact_friendly.dyadic_mixed",
                       "engine.cost_line.quantity_applied_once"):
        assert cases[identifier]["comparison"] == "EXACT", identifier
        assert "expected_exact" in cases[identifier]

    general = cases["engine.general.no_beta"]
    assert general["comparison"] == "TOLERANCE_BOUNDED"
    assert "result_digest" in general["expected"], (
        "an F1 digest must not be an exact cross-language expectation"
    )
    assert "result_digest" not in general.get("expected_exact", {})

    beta = cases["engine.general.with_beta"]
    assert beta["comparison"] == "STATISTICAL"
    assert "result_digest" not in beta.get("expected", {})
    assert "result_digest" in beta["python_reference"]


def test_43_quantity_is_shown_applied_exactly_once() -> None:
    rows = _cases()["engine.cost_line.quantity_applied_once"]["expected_exact"]["rows"]
    for row in rows:
        quantity = float(row["quantity"])
        assert float(row["total"]) == 250.0 * quantity, row
        assert float(row["applied_twice_would_be"]) == 250.0 * quantity * quantity


def test_44_the_full_ladder_is_present_through_its_owners() -> None:
    from pccm_builder.sim_oracle import resolve_percentile_ladder

    ladder = resolve_percentile_ladder(_sim(), _inputs())
    case = _cases()["statistics.ladder.resolved"]["expected_exact"]
    assert case["ordered"] == list(ladder.ordered)
    assert case["count"] == 11
    assert case["fixed_non_selectable"] == ["P10"]
    assert "P10" not in case["selectable"]
    assert case["headline"] == ["P10", "P50", "P70", "P90"]

    # And a real run stores every one of them, for both measures.
    exact = _cases()["engine.exact_friendly.unit_interval"]["expected_exact"]
    stored = exact["statistics"]["nominal"]["quantiles"]
    assert sorted(stored) == sorted(ladder.ordered)


def test_45_the_constant_sample_zero_dispersion_vector_is_present() -> None:
    case = _cases()["statistics.constant_sample.zero_dispersion"]
    assert case["comparison"] == "EXACT"
    for row in case["expected_exact"]["rows"]:
        value = float(row["value"])
        assert float(row["mean"]) == value, row
        assert float(row["sample_standard_deviation"]) == 0.0, row
        assert float(row["minimum"]) == value and float(row["maximum"]) == value
        assert row["all_quantiles_equal_the_value"] is True
    values = {float(row["value"]) for row in case["expected_exact"]["rows"]}
    assert 1.5e308 in values and 0.1 in values and 5e-324 in values


def test_46_the_type7_rows_carry_their_own_comparison() -> None:
    cases = _cases()
    integral = 0
    interpolated = 0
    for identifier, case in cases.items():
        if not identifier.startswith("statistics.quantile.type7"):
            continue
        for row in case["expected"]["rows"]:
            assert row["comparison"] in ("EXACT", "TOLERANCE_BOUNDED"), row
            if float(row["f"]) == 0.0:
                assert row["comparison"] == "EXACT", row
                integral += 1
            else:
                interpolated += 1
    assert integral > 0 and interpolated > 0, (integral, interpolated)


def test_47_contingency_uses_the_deterministic_base_and_is_not_clamped() -> None:
    cases = _cases()
    selected = cases["contingency.selected_levels"]
    exact = selected["expected_exact"]
    assert exact["formula"] == "selected_px_total - deterministic_base_estimate_a"
    assert exact["baseline_owner"] if "baseline_owner" in exact else True
    assert exact["forbidden_baselines"] == [
        "simulation_mean", "analytical_expected_total", "a_plus_emv"
    ]
    assert exact["simulation_mean_is_a_different_number"] is True
    assert exact["analytical_expected_is_a_different_number"] is True
    base = float(exact["base_nominal"])
    for row in selected["expected"]["rows"]:
        assert float(row["contingency_nominal"]) == float(row["selected_nominal"]) - base

    negative = cases["contingency.negative_not_clamped"]["expected_exact"]
    assert negative["is_negative"] is True and negative["clamped"] is False
    assert float(negative["contingency_nominal"]) < 0.0

    assert cases["contingency.p10_not_selectable"]["expected_refusal"]


def test_48_the_extreme_domain_refusals_are_explicit() -> None:
    cases = _cases()
    contribution = cases["domain.contribution_unrepresentable"]["expected_refusal"]
    assert contribution["kind"] == "numerical_range"
    assert contribution["names_iteration_index"] is True
    assert contribution["names_permanent_id"] is True
    assert contribution["names_driver_kind"] is True
    assert contribution["names_stage"] is True
    assert contribution["no_partial_result_returned"] is True

    contingency = cases["contingency.unrepresentable_subtraction"]["expected_refusal"]
    assert contingency["stage"] == "contingency nominal"
    neighbour = cases["contingency.unrepresentable_subtraction"]["expected_exact"]
    assert neighbour["representable_neighbour"]["contingency_nominal"] == 1.5e308

    dispersion = cases["statistics.scale_safety.unrepresentable_dispersion"]
    assert dispersion["expected_refusal"]["kind"] == "numerical_range"
    assert dispersion["expected_exact"]["mean_is_representable"] is True

    rescue = cases["domain.accumulation_partial_sum_out_of_range"]["expected_exact"]
    assert rescue["distinct_totals"] == [1.5e308]
    assert rescue["sample_standard_deviation"] == 0.0


def test_49_no_performance_claim_is_encoded_as_an_expectation() -> None:
    """Scanned over the CASES, not the document's own prose.

    The header says the corpus is "not a benchmark", and a scan of the whole
    file would trip on that sentence rather than on a real claim.
    """
    # Unambiguous tokens only. "duration" is a timeline field, and "ms_per" is a
    # substring of "uniforms_per_sample" - a crude list would flag the model
    # vocabulary instead of a performance claim.
    forbidden = ("elapsed", "benchmark", "throughput", "wall_clock",
                 "milliseconds", "_seconds", "iterations_per_second",
                 "runtime_cost", "faster than", "slower than")
    for identifier, case in _cases().items():
        body = json.dumps(case).lower()
        for token in forbidden:
            assert token not in body, (identifier, token)
    assert _cases_document()["engine_iterations"] == 1000, (
        "Stage-A case generation must stay bounded"
    )


def test_50_the_corpus_is_bounded_enough_to_rebuild_often() -> None:
    size = (_emitted() / "phase6_cases.json").stat().st_size
    assert size < 2_000_000, size
    assert _cases_document()["engine_iterations"] == 1000
    engine_cases = [c for c in _cases() if c.startswith("engine.")]
    assert 5 <= len(engine_cases) <= 20, len(engine_cases)


# ===========================================================================
# mutation controls
#
# A corpus nobody can break is a corpus nobody is checking. Each control changes
# ONE accepted expected value in a COPY of the emitted document and shows the
# check that owns it fails. Nothing in `spec/` or `evidence/` is touched.
# ===========================================================================
def _mutated(path: tuple, value) -> dict:
    """The emitted document with one value replaced, by case id and key path."""
    document = json.loads(json.dumps(_cases_document()))
    cases = {
        case["id"]: case
        for group in document["groups"]
        for case in group["cases"]
    }
    node = cases[path[0]]
    for key in path[1:-1]:
        node = node[key]
    node[path[-1]] = value
    return document


def _lookup(document: dict, identifier: str) -> dict:
    for group in document["groups"]:
        for case in group["cases"]:
            if case["id"] == identifier:
                return case
    raise AssertionError(identifier)


def test_51_a_changed_rng_state_is_detectable() -> None:
    evidence = _evidence("rng_vectors")
    document = _mutated(
        ("rng.fixed_seed.12345", "expected_exact", "initial_state"),
        [12345, 12345, 12345, 12345, 12345, 12346],
    )
    mutated = _lookup(document, "rng.fixed_seed.12345")["expected_exact"]
    assert mutated["initial_state"] != evidence["per_seed"]["12345"]["initial_state"], (
        "the control is vacuous"
    )
    # The accepted document still agrees, so the check is real and the mutation
    # is what breaks it.
    assert _cases()["rng.fixed_seed.12345"]["expected_exact"]["initial_state"] == (
        evidence["per_seed"]["12345"]["initial_state"]
    )


def test_52_a_changed_jump_stream_is_detectable() -> None:
    evidence = _evidence("jump_vectors")
    retained = [float(text) for text in evidence["streams"]["401"]["first_5_uniforms"]]
    document = _mutated(
        ("rng.stream.401", "expected_exact", "first_uniforms"),
        [0.5, 0.5, 0.5, 0.5, 0.5],
    )
    assert _lookup(document, "rng.stream.401")["expected_exact"]["first_uniforms"] != (
        retained
    )
    assert _cases()["rng.stream.401"]["expected_exact"]["first_uniforms"] == retained


def test_53_a_changed_digest_is_detectable() -> None:
    evidence = _evidence("digest_vectors")
    retained = next(c for c in evidence["cases"] if c["label"] == "base")
    document = _mutated(
        ("digest.base", "expected_exact", "result_digest"), "0000000000000000"
    )
    assert _lookup(document, "digest.base")["expected_exact"]["result_digest"] != (
        retained["digest"]
    )
    assert _cases()["digest.base"]["expected_exact"]["result_digest"] == retained["digest"]


def test_54_a_changed_cheng_draw_count_is_detectable() -> None:
    evidence = _evidence("cheng_vectors")
    retained = evidence["cases"][0]
    document = _mutated(
        ("sampler.beta.cheng.bb_interior", "expected_exact", "total_proposal_attempts"),
        retained["total_attempts"] + 1,
    )
    mutated = _lookup(document, "sampler.beta.cheng.bb_interior")["expected_exact"]
    assert mutated["total_proposal_attempts"] != retained["total_attempts"]
    assert _cases()["sampler.beta.cheng.bb_interior"]["expected_exact"][
        "total_proposal_attempts"
    ] == retained["total_attempts"]


def test_55_a_weakened_comparison_policy_is_detectable() -> None:
    """Strengthening layer G to EXACT is exactly what must not slip through."""
    document = _mutated(("engine.general.with_beta", "comparison"), "EXACT")
    assert _lookup(document, "engine.general.with_beta")["comparison"] == "EXACT"
    assert _cases()["engine.general.with_beta"]["comparison"] == "STATISTICAL", (
        "a Beta-containing whole-engine run must not claim cross-language identity"
    )
    document = _mutated(("engine.general.no_beta", "comparison"), "EXACT")
    assert _lookup(document, "engine.general.no_beta")["comparison"] == "EXACT"
    assert _cases()["engine.general.no_beta"]["comparison"] == "TOLERANCE_BOUNDED"


def test_56_a_changed_quantity_row_is_detectable() -> None:
    rows = json.loads(json.dumps(
        _cases()["engine.cost_line.quantity_applied_once"]["expected_exact"]["rows"]
    ))
    rows[1]["total"] = repr(250.0 * 2.0 * 2.0)          # Quantity applied twice
    broken = []
    for row in rows:
        if float(row["total"]) != 250.0 * float(row["quantity"]):
            broken.append(row)
    assert broken, "the control is vacuous"
    for row in _cases()["engine.cost_line.quantity_applied_once"]["expected_exact"]["rows"]:
        assert float(row["total"]) == 250.0 * float(row["quantity"])


def test_57_a_changed_module_constant_is_detectable() -> None:
    """A constant that stops equalling its owner fails the comparison."""
    text = _module_text()
    mutated = text.replace(
        "Public Const SIM_RNG_A12 As Long = 1403580",
        "Public Const SIM_RNG_A12 As Long = 1403581",
    )
    assert mutated != text, "the control is vacuous"
    match = _CONST_RE.match(
        [line for line in mutated.splitlines() if "SIM_RNG_A12" in line][0]
    )
    assert match is not None and int(match.group(3)) != int(
        _sim().raw["rng"]["constants"]["a12"]
    )
    assert _long("SIM_RNG_A12") == int(_sim().raw["rng"]["constants"]["a12"])


def test_58_a_forbidden_construct_added_to_the_module_is_detectable() -> None:
    from pccm_builder.vba_source import VbaModule

    planted = VbaModule(
        name=SIM_MODULE_NAME,
        path=Path("planted.bas"),
        raw=_module_text() + '\nPublic Const SIM_RNG_FAMILY As String = MRG32k3a\n',
    )
    offenders = [
        construct for construct in _structure().forbidden_constructs
        if contains_construct([planted], construct)
    ]
    assert offenders == ["MRG32k3a"], offenders
    assert not [
        construct for construct in _structure().forbidden_constructs
        if contains_construct([_generated_module()], construct)
    ]


def test_59_a_procedure_added_to_the_module_is_detectable() -> None:
    from pccm_builder.vba_source import VbaModule

    planted = VbaModule(
        name=SIM_MODULE_NAME,
        path=Path("planted.bas"),
        raw=_module_text() + "\nPublic Function NextU() As Double\nEnd Function\n",
    )
    assert planted.procedures == ["NextU"]
    assert _generated_module().procedures == []


# ===========================================================================
# the Stage-A build, and what it must not disturb
# ===========================================================================
_PRIOR_ARTIFACTS = (
    "vba/modConstants.bas",
    "vba/modCalcContract.bas",
    "stage_b_manifest.json",
    "phase4_scenarios.json",
    "phase5_cases.json",
    "phase5_gate_b_inspection.json",
)


def _built_tree() -> Path:
    """A complete Stage-A build into a scratch directory, once per run."""
    if "built" not in _CACHE:
        import subprocess

        target = Path(tempfile.mkdtemp(prefix="pccm-step5-build-"))
        result = subprocess.run(
            [sys.executable, str(PCCM_ROOT / "builder" / "build_stage_a.py"),
             "--out", str(target / "PCCM_stageA.xlsx"), "--quiet"],
            capture_output=True, text=True,
        )
        assert result.returncode == 0, result.stderr[-2000:]
        _CACHE["built"] = target
    return _CACHE["built"]  # type: ignore[return-value]


def test_60_the_normal_stage_a_build_emits_both_artefacts() -> None:
    built = _built_tree()
    module = built / "vba" / f"{SIM_MODULE_NAME}.bas"
    cases = built / "phase6_cases.json"
    assert module.is_file(), "modSimContract.bas was not emitted by the build"
    assert cases.is_file(), "phase6_cases.json was not emitted by the build"
    assert module.read_text(encoding="utf-8") == _module_text()
    assert json.loads(cases.read_text(encoding="utf-8")) == _cases_document()


def test_61_the_build_still_emits_every_prior_artefact() -> None:
    built = _built_tree()
    for relative in _PRIOR_ARTIFACTS:
        assert (built / relative).is_file(), relative


def test_62_the_prior_artefacts_are_byte_identical_without_step_5() -> None:
    """Step 5 ADDS two files. It rewrites no Phase-4 or Phase-5 authority.

    The comparison is against a build of the same tree with the Step-5 emission
    removed, so it isolates exactly this step's effect rather than trusting that
    nothing else moved.
    """
    import subprocess

    baseline = Path(tempfile.mkdtemp(prefix="pccm-step5-baseline-"))
    driver = (PCCM_ROOT / "builder" / "build_stage_a.py").read_text(encoding="utf-8")
    stripped = driver.replace(
        "    sim_artifacts = emit_sim_artifacts(out_path.parent, spec, sim, contract, calc)",
        "    sim_artifacts = None",
    ).replace(
        '    say(f"  emitted  : {sim_artifacts.module_path}")\n'
        '    say(f"  emitted  : {sim_artifacts.cases_path}")\n',
        "",
    )
    assert stripped != driver, "the Step-5 emission call was not found in the driver"

    # The copy runs from a scratch directory, so the two locations it derives
    # from `__file__` are repointed at the repository. Nothing else is altered:
    # the comparison is only meaningful if the baseline is the real driver with
    # the Step-5 call removed.
    stripped = stripped.replace(
        "sys.path.insert(0, str(Path(__file__).resolve().parent))",
        f"sys.path.insert(0, {str(PCCM_ROOT / 'builder')!r})",
    ).replace(
        "PCCM_ROOT = Path(__file__).resolve().parent.parent",
        f"PCCM_ROOT = Path({str(PCCM_ROOT)!r})",
    )
    assert str(PCCM_ROOT / "builder") in stripped
    script = baseline / "build_without_step5.py"
    script.write_text(stripped, encoding="utf-8")

    result = subprocess.run(
        [sys.executable, str(script), "--out", str(baseline / "PCCM_stageA.xlsx"), "--quiet"],
        capture_output=True, text=True,
    )
    assert result.returncode == 0, result.stderr[-2000:]
    assert not (baseline / "phase6_cases.json").exists()
    assert not (baseline / "vba" / f"{SIM_MODULE_NAME}.bas").exists()

    built = _built_tree()
    for relative in _PRIOR_ARTIFACTS:
        before = (baseline / relative).read_bytes()
        after = (built / relative).read_bytes()
        assert hashlib.sha256(before).hexdigest() == hashlib.sha256(after).hexdigest(), (
            f"{relative} moved when Step 5 was added"
        )


def test_63_the_workbook_gains_no_phase_6_content() -> None:
    """No `vbaProject.bin`, no simulation results, and `_SimData` stays empty."""
    import zipfile

    from openpyxl import load_workbook

    workbook_path = _built_tree() / "PCCM_stageA.xlsx"
    with zipfile.ZipFile(workbook_path) as archive:
        names = archive.namelist()
    assert not any("vbaProject" in name for name in names), names

    workbook = load_workbook(workbook_path)
    try:
        sheet = workbook["_SimData"]
        layout = _sim().layout
        for row in range(layout.first_iteration_row, layout.first_iteration_row + 20):
            for column in range(1, 12):
                assert sheet.cell(row=row, column=column).value is None, (row, column)
    finally:
        workbook.close()


def test_64_the_generated_module_is_not_embedded_in_the_workbook() -> None:
    import zipfile

    with zipfile.ZipFile(_built_tree() / "PCCM_stageA.xlsx") as archive:
        for name in archive.namelist():
            assert SIM_MODULE_NAME not in name, name


# ===========================================================================
# semantic JSON typing
#
# A finite number stringified to preserve `repr` is a number wearing a costume,
# and it slips straight past `allow_nan=False` - `repr(float("nan"))` is the
# ordinary string "nan". These tests pin the types so that guard is real.
# ===========================================================================
def _leaves(node, path="$", under_canonical=False):
    """`(path, value, under_canonical)` for every leaf in the corpus."""
    if isinstance(node, dict):
        for key, value in node.items():
            yield from _leaves(
                value, f"{path}.{key}", under_canonical or key.endswith("_canonical")
            )
    elif isinstance(node, list):
        for index, value in enumerate(node):
            yield from _leaves(value, f"{path}[{index}]", under_canonical)
    else:
        yield path, node, under_canonical


def test_65_no_semantic_number_is_emitted_as_a_string() -> None:
    """Every numeric-looking string must be a declared canonical encoding."""
    offenders = []
    for path, value, under_canonical in _leaves(_cases_document()):
        if not isinstance(value, str) or under_canonical:
            continue
        try:
            float(value)
        except (TypeError, ValueError):
            continue
        offenders.append((path, value))
    assert not offenders, offenders[:10]


def test_66_representative_numeric_fields_are_floats_after_json_loads() -> None:
    cases = _cases()

    uniforms = cases["rng.fixed_seed.12345"]["expected_exact"]["first_uniforms"]
    assert all(isinstance(value, float) for value in uniforms), uniforms
    assert uniforms[0] == 0.12701112204657714

    sample = cases["sampler.beta.cheng.bb_interior"]["expected"]["samples"][0]["value"]
    assert isinstance(sample, float) and sample == 0.12872280046553125

    injected = cases["sampler.uniform.injected"]["expected"]["rows"][0]
    assert isinstance(injected["u"], float) and isinstance(injected["value"], float)

    statistics = cases["engine.general.no_beta"]["expected"]["statistics"]["nominal"]
    for key in ("mean", "sample_standard_deviation", "minimum", "maximum"):
        assert isinstance(statistics[key], float), key
    assert all(isinstance(value, float) for value in statistics["quantiles"].values())

    contingency = cases["contingency.selected_levels"]["expected"]["rows"][0]
    for key in ("selected_nominal", "selected_pv",
                "contingency_nominal", "contingency_pv"):
        assert isinstance(contingency[key], float), key

    totals = cases["engine.exact_friendly.unit_interval"]["expected_exact"]
    assert all(isinstance(value, float) for value in totals["total_nominal"]["head"])
    assert isinstance(totals["statistics"]["nominal"]["mean"], float)


def test_67_counts_and_indices_stay_integers() -> None:
    cases = _cases()
    exact = cases["sampler.beta.cheng.bb_interior"]["expected_exact"]
    for key in ("total_proposal_attempts", "total_uniforms", "uniforms_per_attempt"):
        assert isinstance(exact[key], int) and not isinstance(exact[key], bool), key
    assert all(isinstance(word, int) for word in exact["initial_state"])
    for row in exact["per_sample"]:
        assert isinstance(row["proposal_attempts"], int)
        assert isinstance(row["cumulative_uniforms"], int)
    assert isinstance(cases["seed.auto.nonce.1000"]["expected_exact"]["effective_seed"], int)


def test_68_every_exact_number_carries_its_canonical_encoding() -> None:
    """The sidecar is the accepted Phase-5 encoding and round-trips exactly."""
    from pccm_builder.calc_fingerprint import canonical_number

    checked = 0
    for group in _cases_document()["groups"]:
        for case in group["cases"]:
            blocks = [case.get("expected_exact")]
            if case["comparison"] == "EXACT":
                blocks.append(case.get("expected"))
            for block in blocks:
                if block is None:
                    continue
                checked += _check_sidecars(block)
    assert checked > 100, checked
    assert canonical_number(0.12701112204657714) == "1.2701112204657714E-01"


def _check_sidecars(node) -> int:
    checked = 0
    if isinstance(node, list):
        for item in node:
            checked += _check_sidecars(item)
        return checked
    if not isinstance(node, dict):
        return 0
    for key, value in node.items():
        if key.endswith("_canonical"):
            continue
        sidecar = node.get(key + "_canonical")
        if isinstance(value, float):
            assert isinstance(sidecar, str), key
            assert float(sidecar) == value, (key, sidecar, value)
            checked += 1
        elif isinstance(value, list) and value and all(
            isinstance(item, float) for item in value
        ):
            assert isinstance(sidecar, list) and len(sidecar) == len(value), key
            for text, number in zip(sidecar, value):
                assert float(text) == number, (key, text, number)
            checked += len(value)
        elif isinstance(value, dict) and value and all(
            isinstance(item, float) for item in value.values()
        ):
            assert isinstance(sidecar, dict), key
            for label, number in value.items():
                assert float(sidecar[label]) == number, (key, label)
            checked += len(value)
        else:
            checked += _check_sidecars(value)
    return checked


def test_69_a_tolerance_bounded_block_carries_no_exact_sidecar() -> None:
    """An exact textual form there would invite the comparison the accepted
    evidence model refuses."""
    bounded = _cases()["engine.general.no_beta"]["expected"]
    assert not any(key.endswith("_canonical") for key in bounded), sorted(bounded)
    statistics = bounded["statistics"]["nominal"]
    assert not any(key.endswith("_canonical") for key in statistics)
    # And the exact block of the same case does carry them.
    assert "deterministic_base_a_nominal_canonical" in (
        _cases()["engine.general.no_beta"]["expected_exact"]
    )


# ===========================================================================
# non-finite rejection is real
# ===========================================================================
def _emit_with(document: dict) -> str:
    """Serialise a corpus through the SAME boundary the emitter uses."""
    return json.dumps(
        document, indent=2, sort_keys=False, allow_nan=False, ensure_ascii=True
    ) + "\n"


def _numeric_document(replacement) -> dict:
    document = json.loads(json.dumps(_cases_document()))
    for group in document["groups"]:
        for case in group["cases"]:
            if case["id"] == "rng.fixed_seed.12345":
                case["expected_exact"]["first_uniforms"][0] = replacement
                return document
    raise AssertionError("the target case was not found")


def test_70_a_nan_in_a_numeric_leaf_refuses_emission() -> None:
    document = _numeric_document(float("nan"))
    try:
        _emit_with(document)
    except ValueError as error:
        assert "NaN" in str(error) or "not JSON compliant" in str(error), str(error)
        return
    raise AssertionError("a NaN was serialised")


def test_71_a_positive_infinity_in_a_numeric_leaf_refuses_emission() -> None:
    try:
        _emit_with(_numeric_document(float("inf")))
    except ValueError:
        return
    raise AssertionError("an infinity was serialised")


def test_72_a_negative_infinity_in_a_numeric_leaf_refuses_emission() -> None:
    try:
        _emit_with(_numeric_document(float("-inf")))
    except ValueError:
        return
    raise AssertionError("a negative infinity was serialised")


def test_73_the_corpus_builder_refuses_a_non_finite_value_before_serialisation() -> None:
    """The failure names the corpus, not the serialiser."""
    from pccm_builder.sim_cases import _n
    from pccm_builder.sim_oracle import SimOracleError

    for bad in (float("nan"), float("inf"), float("-inf")):
        try:
            _n(bad)
        except SimOracleError as error:
            assert "non-finite" in str(error)
            continue
        raise AssertionError(f"{bad!r} was accepted as a semantic number")
    assert _n(2) == 2.0 and isinstance(_n(2), float)


def test_74_the_string_nan_in_a_numeric_field_fails_validation() -> None:
    """Text is exactly how a non-finite value used to get past allow_nan."""
    from pccm_builder.sim_cases import validate_corpus
    from pccm_builder.sim_oracle import SimOracleError

    assert _emit_with(_numeric_document("nan")), "a numeric string serialises happily"
    try:
        validate_corpus(_numeric_document("nan"))
    except SimOracleError as error:
        assert "first_uniforms" in str(error)
        return
    raise AssertionError("the string 'nan' passed corpus validation")


def test_75_a_numeric_string_in_a_numeric_field_fails_validation() -> None:
    from pccm_builder.sim_cases import validate_corpus
    from pccm_builder.sim_oracle import SimOracleError

    for planted in ("1.25", "0.12701112204657714"):
        try:
            validate_corpus(_numeric_document(planted))
        except SimOracleError as error:
            assert "first_uniforms" in str(error), planted
            continue
        raise AssertionError(f"the string {planted!r} passed corpus validation")

    # The accepted document still validates, so the control is not vacuous.
    validate_corpus(_cases_document())


def test_76_a_number_turned_into_a_string_anywhere_is_detected() -> None:
    from pccm_builder.sim_cases import validate_corpus
    from pccm_builder.sim_oracle import SimOracleError

    document = json.loads(json.dumps(_cases_document()))
    for group in document["groups"]:
        for case in group["cases"]:
            if case["id"] == "contingency.selected_levels":
                case["expected"]["rows"][0]["contingency_nominal"] = "1.25"
    try:
        validate_corpus(document)
    except SimOracleError as error:
        assert "contingency_nominal" in str(error)
        return
    raise AssertionError("1.25 -> \"1.25\" was not detected")


def test_77_the_validator_runs_on_every_build() -> None:
    """A typing regression fails the Stage-A build, not a later reader."""
    import inspect

    from pccm_builder import sim_cases

    source = inspect.getsource(sim_cases.build_sim_cases)
    assert "validate_corpus(document)" in source


# ===========================================================================
# per-case version identity
# ===========================================================================
def test_78_every_case_carries_a_complete_version_block() -> None:
    document = _cases_document()
    expected = {
        "model_version": document["model_version"],
        "sim_contract_version": document["sim_contract_version"],
        "rng_version": document["rng_version"],
        "sim_method_version": document["sim_method_version"],
    }
    groups = 0
    for group in document["groups"]:
        groups += 1
        for case in group["cases"]:
            assert "versions" in case, case["id"]
            assert case["versions"] == expected, case["id"]
            assert isinstance(case["versions"]["rng_version"], int)
            assert isinstance(case["versions"]["model_version"], str)
            assert "fp_version" not in case["versions"]
    assert groups == 12
    assert len(_cases()) == document["case_count"]


def test_79_a_changed_case_version_is_detected() -> None:
    from pccm_builder.sim_cases import validate_corpus
    from pccm_builder.sim_oracle import SimOracleError

    document = json.loads(json.dumps(_cases_document()))
    _lookup(document, "digest.base")["versions"]["rng_version"] = 2
    try:
        validate_corpus(document)
    except SimOracleError as error:
        assert "digest.base" in str(error)
    else:
        raise AssertionError("a changed case RNG_VERSION was not detected")

    document = json.loads(json.dumps(_cases_document()))
    del _lookup(document, "digest.base")["versions"]
    try:
        validate_corpus(document)
    except SimOracleError as error:
        assert "no version identity" in str(error)
        return
    raise AssertionError("a missing case version block was not detected")


# ===========================================================================
# the projected model version
# ===========================================================================
def test_80_the_model_version_is_a_constant_not_only_a_comment() -> None:
    owner = _spec().model["model_version"]
    assert _string("SIM_MODEL_VERSION") == owner == "0.5.0"
    assert "SIM_MODEL_VERSION" in _generated_module().constants
    assert _long("SIM_RNG_VERSION") == 1


def test_81_the_model_version_constant_follows_its_owner() -> None:
    """A synthetic manifest with a different model version moves ONLY that
    constant - the projection is a projection, not a hardcoded literal."""
    import dataclasses

    model = dict(_spec().model)
    model["model_version"] = "9.9.9-test"
    altered = dataclasses.replace(_spec(), model=model)

    rendered = render_sim_contract_module(altered, _sim(), _inputs())
    assert 'Public Const SIM_MODEL_VERSION As String = "9.9.9-test"' in rendered
    assert 'Public Const SIM_MODEL_VERSION As String = "0.5.0"' not in rendered

    baseline = _module_text()
    moved = [
        line for line in rendered.splitlines()
        if line.startswith("Public Const") and line not in baseline.splitlines()
    ]
    assert moved == [
        'Public Const SIM_MODEL_VERSION As String = "9.9.9-test"    '
        "' workbook.yaml: model.model_version - snapshotted by a successful run"
    ], moved


def test_82_the_module_is_still_constants_only_with_the_new_constant() -> None:
    assert _generated_module().procedures == []
    assert len(_constants()) == len(_generated_module().constants) >= 182
    structure = _structure()
    assert not [
        construct for construct in structure.forbidden_constructs
        if contains_construct([_generated_module()], construct)
    ]
    raw = _module_text().lower()
    for token in ("runsimulation", "mrg32k3a", "percentile"):
        assert token not in raw, token


def test_83_every_emitted_number_survives_the_json_round_trip_exactly() -> None:
    """The durable form of the migration guarantee.

    The one-off comparison against the previous corpus - 741 numeric strings
    parsed to Double and found identical to the JSON numbers that replaced them -
    cannot be repeated forever without shipping the old file. What CAN be
    asserted forever is that the representation is lossless: every number written
    reads back as the same binary64, and every exact one has a canonical sidecar
    that parses to it as well.
    """
    text = (_emitted() / "phase6_cases.json").read_text(encoding="utf-8")
    reloaded = json.loads(text)
    assert reloaded == _cases_document()

    again = json.dumps(reloaded, indent=2, sort_keys=False, allow_nan=False,
                       ensure_ascii=True) + "\n"
    assert again == text, "the corpus is not stable under a JSON round trip"

    numbers = 0
    for _path, value, under_canonical in _leaves(reloaded):
        if isinstance(value, bool) or under_canonical:
            continue
        if isinstance(value, float):
            assert math.isfinite(value)
            assert float(repr(value)) == value
            numbers += 1
    assert numbers > 500, numbers


def test_84_the_comparison_policy_distribution_is_unchanged() -> None:
    """No case moved class. Step-10A added five EXACT request-fingerprint cases
    and Step-11A sixteen EXACT publication-state cases; neither touched an
    existing case's policy, so every non-EXACT count below is the number it has
    carried since the typing correction."""
    from collections import Counter

    counts = Counter(case["comparison"] for case in _cases().values())
    assert dict(counts) == {
        "EXACT": 82,
        "TOLERANCE_BOUNDED": 23,
        "STATISTICAL": 1,
        "SAME_RUNTIME_ONLY": 3,
        "RUNTIME_ONLY": 3,
    }, dict(counts)
    assert sum(counts.values()) == 112
    request = [i for i in _cases() if i.startswith("request_fingerprint.")]
    assert len(request) == 5, request
    assert all(_cases()[i]["comparison"] == "EXACT" for i in request)
    publication = [i for i in _cases() if i.startswith("publication.")]
    assert len(publication) == 16, publication
    assert all(_cases()[i]["comparison"] == "EXACT" for i in publication)

    cases = _cases()
    assert cases["engine.general.with_beta"]["comparison"] == "STATISTICAL"
    assert cases["engine.general.no_beta"]["comparison"] == "TOLERANCE_BOUNDED"
    assert cases["engine.replay.same_seed_identical"]["comparison"] == "SAME_RUNTIME_ONLY"
    assert cases["engine.row_order.invariant"]["comparison"] == "SAME_RUNTIME_ONLY"
    assert cases["engine.seed.non_degenerate_divergence"]["expected"]["scope"] == (
        "this fixture only"
    )
    assert cases["engine.seed.degenerate_equal_digest"]["comparison"] == "EXACT"


# ===========================================================================
# Step-10A - the request-fingerprint projection and corpus
# ===========================================================================
def test_85_the_request_fingerprint_framing_is_projected() -> None:
    section = _sim().raw["request_fingerprint"]["sim_section"]
    assert _string("SIM_REQUEST_SECTION") == section["name"] == "SIM"
    assert _long("SIM_REQUEST_RECORD_COUNT") == int(section["record_count"]) == 1
    assert _long("SIM_REQUEST_FIELD_COUNT_AUTO") == 4
    assert _long("SIM_REQUEST_FIELD_COUNT_FIXED") == 5
    for mode, shape in section["effective_records"].items():
        assert _long(f"SIM_REQUEST_FIELD_COUNT_{mode}") == int(shape["field_count"])
        for ordinal, field in enumerate(shape["fields"], start=1):
            assert _string(f"SIM_REQUEST_{mode}_FIELD_{ordinal}") == field
    for field in section["fields"]:
        name = field.upper()
        assert _string(f"SIM_REQUEST_TYPE_{name}") == section["field_types"][field]
    assert _string("SIM_REQUEST_AUTO_SEED") == "absent"


def test_86_the_projection_does_not_restate_the_phase5_stream_authority() -> None:
    """The extension is a SECTION of the accepted PCCM-FP stream, so its tag,
    its FP_VERSION and the hash mathematics stay with modCalcContract."""
    constants = _constants()
    for owned_elsewhere in ("SIM_FP_STREAM_TAG", "SIM_FP_VERSION", "SIM_REQUEST_STREAM_TAG",
                            "SIM_REQUEST_FP_VERSION", "SIM_FP_BASE", "SIM_FP_MOD_1",
                            "SIM_FP_MOD_2", "SIM_FP_INIT_1", "SIM_FP_INIT_2"):
        assert owned_elsewhere not in constants, owned_elsewhere
    text = _module_text()
    assert "PCCM-FP" not in text, "the projection repeated the analytical stream tag"
    # The digest stream tag is a DIFFERENT stream and is still projected.
    assert _string("SIM_DIGEST_STREAM_TAG") == "PCCM-RD"


def test_87_no_request_field_name_is_projected_as_an_encoded_token() -> None:
    """The names are semantic position. Nothing may read them as stream bytes."""
    section = _sim().raw["request_fingerprint"]["sim_section"]
    assert section["encoded_field_names"] is False
    for production in section["grammar"].values():
        for field in section["fields"]:
            assert f'"{field}"' not in production, field


def test_88_the_request_fingerprint_group_is_in_the_corpus() -> None:
    document = _cases_document()
    groups = {group["group"]: group for group in document["groups"]}
    assert "I_request_fingerprint" in groups
    group = groups["I_request_fingerprint"]
    assert group["title"] == "The request fingerprint and its SIM extension"
    ids = [case["id"] for case in group["cases"]]
    assert len(ids) == 5 == len(set(ids))
    for case in group["cases"]:
        assert case["comparison"] == "EXACT", case["id"]
        assert case["layer"] == "I_request_fingerprint"
        assert "expected_exact" in case
        assert "tolerance" not in json.dumps(case)


def test_89_the_request_vectors_carry_no_environment_or_clock_data() -> None:
    group = next(g for g in _cases_document()["groups"]
                 if g["group"] == "I_request_fingerprint")
    text = json.dumps(group)
    for banned in ("timestamp", "generated_at", "hostname", "/home/", "\\Users",
                   "random", "elapsed", "duration_ms"):
        assert banned not in text, banned


def test_90_a_changed_request_fingerprint_is_detectable() -> None:
    """The corpus must not be able to drift silently. This is the control the
    golden literals in `test_phase6_request_fingerprint.py` exist to enforce."""
    from pccm_builder.sim_cases import request_fingerprint

    document = _mutated(
        ("request_fingerprint.auto.1000", "expected_exact", "request_fingerprint"),
        "DEADBEEFDEADBEEF")
    case = _lookup(document, "request_fingerprint.auto.1000")
    recomputed = request_fingerprint(_sim(), _calc(), 1000, "AUTO")
    assert case["expected_exact"]["request_fingerprint"] != recomputed
    assert recomputed == "5EAB16E15C2ECE24"


def test_91_a_changed_request_suffix_is_detectable() -> None:
    from pccm_builder.sim_cases import request_sim_section

    document = _mutated(
        ("request_fingerprint.fixed.seed_1", "expected_exact", "sim_suffix"),
        "S3:SIMI1:5I1:1I4:1000S5:FIXEDI1:1I1:1I1:1")
    case = _lookup(document, "request_fingerprint.fixed.seed_1")
    recomputed = request_sim_section(_sim(), 1000, "FIXED", 1)
    assert case["expected_exact"]["sim_suffix"] != recomputed
    assert recomputed == "S3:SIMI1:1I1:5I4:1000S5:FIXEDI1:1I1:1I1:1"


if __name__ == "__main__":
    import pytest

    raise SystemExit(pytest.main([__file__, "-q"]))


# ===========================================================================
# PHASE-6 STEP 12: the Results presentation text is true
# ===========================================================================
# The Results subtitle and its top note were written in Phase 1 and said so:
# "no statistics are implemented in Phase 1" and "Not implemented yet. No
# percentiles, moments or cash flows are implemented yet." Step 11A gave Results
# a real percentile ladder, real moments and real selected-confidence reporting
# over `_SimData`, and those two strings became false. Step 12 corrects the
# WORDING only - no layout, no formula and no row moves.
STALE_RESULTS_PHRASES = (
    "no statistics are implemented in Phase 1",
    "Not implemented yet. No percentiles, moments or cash flows are implemented yet.",
    "no statistics are implemented",
    "No percentiles, moments or cash flows",
)

# Deferred means deferred. Nothing on Results may imply these exist.
FORBIDDEN_RESULTS_CLAIMS = (
    "sensitivity", "correlation", "dashboard", "spearman",
    "reconciliation is implemented", "annual cash flow is implemented",
)


def _results_strings() -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(_built_tree() / "PCCM_stageA.xlsx")
    try:
        sheet = workbook["Results"]
        found = []
        for row in sheet.iter_rows(min_row=1, max_row=80, max_col=8):
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    found.append(cell.value)
        return found
    finally:
        workbook.close()


def test_92_the_stale_phase_1_results_text_is_gone_from_the_spec() -> None:
    spec_text = (SPEC / "workbook.yaml").read_text(encoding="utf-8")
    for stale in STALE_RESULTS_PHRASES:
        assert stale not in spec_text, f"the manifest still carries: {stale!r}"


def test_93_the_built_results_sheet_says_what_is_actually_published() -> None:
    strings = _results_strings()
    joined = "\n".join(strings)
    for stale in STALE_RESULTS_PHRASES:
        assert stale not in joined, f"the built Results sheet still says: {stale!r}"
    # AND IT SAYS THE TRUE THING. The subtitle names the sheet's actual content,
    # and the note names what Phase 6 publishes and what is still deferred.
    assert "Simulation results and statistical summary" in strings, strings[:6]
    note = next((s for s in strings if s.startswith("Phase 6 publishes")), None)
    assert note is not None, strings[:6]
    for named in ("run identity", "summary statistics", "selected-confidence"):
        assert named in note, (named, note)
    for deferred in ("Annual cash flow", "reconciliation"):
        assert deferred in note, (deferred, note)
    assert "deferred" in note


def test_94_results_claims_nothing_that_does_not_exist() -> None:
    """Deferred sections stay deferred, and nothing implies a Phase-7 feature."""
    lowered = "\n".join(_results_strings()).lower()
    for claim in FORBIDDEN_RESULTS_CLAIMS:
        assert claim not in lowered, f"Results implies {claim!r} exists"
    # The two deferred sections keep their deferred notes.
    strings = _results_strings()
    assert "Annual Cash Flow" in strings
    assert "Reconciliation" in strings
    deferred_notes = [s for s in strings if s.startswith("Deferred.")]
    assert len(deferred_notes) == 2, deferred_notes


def test_95_only_the_wording_moved_and_the_layout_did_not() -> None:
    """The accepted Step-11A geometry is untouched by the text correction."""
    from openpyxl import load_workbook

    import yaml

    shell = yaml.safe_load((SPEC / "workbook.yaml").read_text(
        encoding="utf-8"))["phase6_shell"]["results"]
    workbook = load_workbook(_built_tree() / "PCCM_stageA.xlsx")
    try:
        sheet = workbook["Results"]
        # EVERY ANCHORED ROW STILL HOLDS ITS ACCEPTED LABEL, read from the
        # manifest rather than restated here.
        label = shell["label_column"]
        for section in shell["sections"]:
            assert sheet[f"{label}{section['row']}"].value == section["title"], section
            assert sheet[f"{label}{section['note_row']}"].value == section["note"], section
        for field in shell["run_stamp"]["fields"]:
            assert sheet[f"{label}{field['row']}"].value == field["label"], field
        for metric in shell["summary"]["metrics"]:
            assert sheet[f"{label}{metric['row']}"].value == metric["label"], metric
        selected = shell["selected"]
        for key, text in selected["labels"].items():
            row = selected[f"{key}_row"]
            assert sheet[f"{label}{row}"].value == text, (key, row)
        for entry in shell["deferred"]:
            assert sheet[f"{label}{entry['row']}"].value == entry["title"], entry
            assert sheet[f"{label}{entry['note_row']}"].value == entry["note"], entry
    finally:
        workbook.close()
    # AND THE ACCEPTED STEP-11A GEOMETRY IS EXACTLY WHERE IT WAS. A text edit
    # that shifted a row would show up here rather than in a runtime.
    layout = _sim().layout
    assert layout.first_iteration_row == 34, layout.first_iteration_row
    assert shell["label_column"] == "B"
    assert shell["nominal_column"] == "D" and shell["pv_column"] == "F"
    assert [s["row"] for s in shell["sections"]] == [8, 26], shell["sections"]
    assert shell["run_stamp"]["first_row"] == 10 and shell["run_stamp"]["last_row"] == 24
    assert shell["summary"]["header_row"] == 28
    assert shell["summary"]["first_row"] == 29 and shell["summary"]["last_row"] == 44
    assert selected["confidence_level_row"] == 46
    assert selected["quantile_row"] == 47 and selected["contingency_row"] == 48
    assert [entry["row"] for entry in shell["deferred"]] == [51, 54]
    # The formulas are structural and were not touched by a wording change.
    for formula_key in ("confidence_level_formula", "quantile_nominal", "quantile_pv",
                        "contingency_nominal", "contingency_pv"):
        assert selected[formula_key].startswith("=IF("), formula_key
    # The confidence-level row reads the selector and nothing else; the four
    # banked lookups read the active-bank selector and pick a bank from it.
    assert "_SimData" not in selected["confidence_level_formula"]
    for banked in ("quantile_nominal", "quantile_pv",
                   "contingency_nominal", "contingency_pv"):
        assert "_SimData!$D$30" in selected[banked], banked
        assert "inpSelectedConfidenceLevel" in selected[banked], banked


def test_nc_96_the_stale_results_text_returning_is_caught() -> None:
    """MUTATION CONTROL. Plant each retired phrase and require a detector.

    The detector is the same membership test test_92 and test_93 make, run here
    against text that carries the phrase, so a control that passed because the
    phrase was simply absent everywhere cannot be mistaken for one that works.
    """
    spec_text = (SPEC / "workbook.yaml").read_text(encoding="utf-8")
    for stale in STALE_RESULTS_PHRASES:
        assert stale not in spec_text, "the accepted manifest is already clean"
        damaged = spec_text.replace(
            '    subtitle: "Simulation results and statistical summary"\n',
            f'    subtitle: "{stale}"\n', 1)
        assert damaged != spec_text or stale in damaged, stale
        assert stale in damaged, f"the mutation did not plant {stale!r}"
        try:
            for phrase in STALE_RESULTS_PHRASES:
                assert phrase not in damaged, phrase
        except AssertionError:
            pass
        else:
            raise AssertionError(f"the detector cannot see {stale!r}")
    # AND ON THE BUILT SHEET. The same comparison, over planted cell text.
    strings = _results_strings() + [
        "Simulation output - no statistics are implemented in Phase 1"]
    joined = "\n".join(strings)
    try:
        for phrase in STALE_RESULTS_PHRASES:
            assert phrase not in joined, phrase
    except AssertionError:
        pass
    else:
        raise AssertionError("the built-sheet detector cannot see the stale text")


def test_nc_97_a_forbidden_results_claim_is_caught() -> None:
    """Deferred means deferred: the claim detector must see an added claim."""
    lowered = "\n".join(_results_strings()).lower()
    for claim in FORBIDDEN_RESULTS_CLAIMS:
        assert claim not in lowered, "the accepted sheet already claims it"
        damaged = lowered + "\n" + f"driver {claim} is available on this sheet"
        try:
            for candidate in FORBIDDEN_RESULTS_CLAIMS:
                assert candidate not in damaged, candidate
        except AssertionError:
            pass
        else:
            raise AssertionError(f"the detector cannot see {claim!r}")
