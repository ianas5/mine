#!/usr/bin/env python3
"""PCCM Phase 3 tests: Cost Lines and Risk Register input layer.

Runs standalone or under pytest.

The locked column schemas are re-declared here independently of the driver
contract. The contract is the builder's authority; this file is the
architecture-conformance check, so schema drift fails here even when the build
itself succeeds.
"""

from __future__ import annotations

import os
import sys
import tempfile
import zipfile
from pathlib import Path

import yaml

PCCM_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PCCM_ROOT / "builder"))

from openpyxl import load_workbook  # noqa: E402

from pccm_builder import (  # noqa: E402
    build_workbook,
    load_contract,
    load_driver_contract,
    load_spec,
)

SPEC_PATH = PCCM_ROOT / "spec" / "workbook.yaml"
CONTRACT_PATH = PCCM_ROOT / "spec" / "input_contract.yaml"
DRIVERS_PATH = PCCM_ROOT / "spec" / "driver_contract.yaml"

# --- Architecture Lock Revision B: locked driver schemas --------------------
COST_LINE_COLUMNS = [
    "Cost Line ID",
    "Category",
    "Description",
    "UOM",
    "Quantity",
    "Currency",
    "Inflation Profile",
    "Unit Cost Min",
    "Unit Cost Most Likely",
    "Unit Cost Max",
    "Distribution",
]

RISK_COLUMNS = [
    "Risk ID",
    "Risk Name",
    "Description",
    "Category",
    "Probability",
    "Currency",
    "Inflation Profile",
    "Impact Min",
    "Impact Most Likely",
    "Impact Max",
    "Distribution",
    "Risk Owner",
]

COST_TABLE = "tblCostLines"
RISK_TABLE = "tblRiskRegister"
COST_SHEET = "Cost Lines"
RISK_SHEET = "Risk Register"

EXPECTED_LIST_SOURCES = {
    ("cost_lines", "Category"): "lstCategories",
    ("cost_lines", "UOM"): "lstUOM",
    ("cost_lines", "Currency"): "lstCurrencies",
    ("cost_lines", "Inflation Profile"): "lstInflationProfiles",
    ("cost_lines", "Distribution"): "lstDistributions",
    ("risk_register", "Category"): "lstCategories",
    ("risk_register", "Currency"): "lstCurrencies",
    ("risk_register", "Inflation Profile"): "lstInflationProfiles",
    ("risk_register", "Distribution"): "lstDistributions",
}

_CACHE: dict[str, Path] = {}
_TEMPDIR: tempfile.TemporaryDirectory | None = None


# ---------------------------------------------------------------------------
def _artifact() -> Path:
    global _TEMPDIR
    if "primary" in _CACHE:
        return _CACHE["primary"]
    if _TEMPDIR is None:
        _TEMPDIR = tempfile.TemporaryDirectory(prefix="pccm-phase3-")
    previous = os.environ.get("PCCM_BUILD_TIMESTAMP")
    os.environ["PCCM_BUILD_TIMESTAMP"] = "1970-01-01 00:00:00 UTC"
    try:
        workbook, _ = build_workbook(
            load_spec(SPEC_PATH), load_contract(CONTRACT_PATH), load_driver_contract(DRIVERS_PATH)
        )
        path = Path(_TEMPDIR.name) / "stage_a.xlsx"
        workbook.save(path)
        workbook.close()
    finally:
        if previous is None:
            os.environ.pop("PCCM_BUILD_TIMESTAMP", None)
        else:
            os.environ["PCCM_BUILD_TIMESTAMP"] = previous
    _CACHE["primary"] = path
    return path


def _wb():
    return load_workbook(_artifact())


def _drivers():
    return load_driver_contract(DRIVERS_PATH)


def _register(key: str):
    return _drivers().registers[key]


def _headers_in_sheet(worksheet, register) -> list[str]:
    return [
        worksheet[f"{register.column_letter(i)}{register.header_row}"].value
        for i in range(len(register.columns))
    ]


def _fill_rgb(cell) -> str:
    rgb = getattr(cell.fill.fgColor, "rgb", None)
    return str(rgb)[-6:].upper() if rgb else ""


def _tokens() -> dict:
    with SPEC_PATH.open(encoding="utf-8") as handle:
        return yaml.safe_load(handle)["presentation"]["colors"]


def _dv_covering(worksheet, target: str):
    for dv in worksheet.data_validations.dataValidation:
        if target in str(dv.sqref):
            return dv
    return None


# --- 1-4. tables and schemas ------------------------------------------------
def test_01_cost_lines_table_exists() -> None:
    assert COST_TABLE in getattr(_wb()[COST_SHEET], "tables", {})


def test_02_cost_line_schema_and_order_are_exact() -> None:
    register = _register("cost_lines")
    assert register.headers == COST_LINE_COLUMNS, f"contract schema: {register.headers}"
    assert _headers_in_sheet(_wb()[COST_SHEET], register) == COST_LINE_COLUMNS
    assert len(register.columns) == 11


def test_03_risk_register_table_exists() -> None:
    assert RISK_TABLE in getattr(_wb()[RISK_SHEET], "tables", {})


def test_04_risk_schema_and_order_are_exact() -> None:
    register = _register("risk_register")
    assert register.headers == RISK_COLUMNS, f"contract schema: {register.headers}"
    assert _headers_in_sheet(_wb()[RISK_SHEET], register) == RISK_COLUMNS
    assert len(register.columns) == 12


def test_05_no_included_column_on_risks() -> None:
    register = _register("risk_register")
    lowered = [h.lower() for h in register.headers]
    assert "included" not in lowered and "include" not in lowered, (
        "every entered risk is simulated; there must be no Included column"
    )


# --- 6-9. row and column ownership ------------------------------------------
def test_06_id_columns_are_blank() -> None:
    workbook = _wb()
    for key, sheet in (("cost_lines", COST_SHEET), ("risk_register", RISK_SHEET)):
        register = _register(key)
        worksheet = workbook[sheet]
        letter = register.column_letter(0)
        values = [
            worksheet[f"{letter}{row}"].value
            for row in range(register.first_data_row, register.last_data_row + 1)
        ]
        assert all(v is None for v in values), f"{register.table_name} has pre-seeded IDs: {values}"


def test_07_id_columns_are_model_controlled_styled() -> None:
    workbook = _wb()
    locked = _tokens()["locked_fill"].upper()
    for key, sheet in (("cost_lines", COST_SHEET), ("risk_register", RISK_SHEET)):
        register = _register(key)
        assert register.columns[0].editable is False
        worksheet = workbook[sheet]
        letter = register.column_letter(0)
        for row in range(register.first_data_row, register.last_data_row + 1):
            assert _fill_rgb(worksheet[f"{letter}{row}"]) == locked, (
                f"{register.table_name} {letter}{row} is not locked-styled"
            )


def test_08_id_columns_carry_no_data_validation() -> None:
    workbook = _wb()
    for key, sheet in (("cost_lines", COST_SHEET), ("risk_register", RISK_SHEET)):
        register = _register(key)
        assert register.columns[0].validation is None
        assert _dv_covering(workbook[sheet], register.data_range(0)) is None, (
            f"{register.table_name} identity column has data validation"
        )


def test_09_user_input_columns_use_editable_styling() -> None:
    workbook = _wb()
    editable = _tokens()["input_fill"].upper()
    for key, sheet in (("cost_lines", COST_SHEET), ("risk_register", RISK_SHEET)):
        register = _register(key)
        worksheet = workbook[sheet]
        for index, column in enumerate(register.columns):
            if not column.editable:
                continue
            letter = register.column_letter(index)
            for row in range(register.first_data_row, register.last_data_row + 1):
                assert _fill_rgb(worksheet[f"{letter}{row}"]) == editable, (
                    f"{register.table_name}.{column.header} {letter}{row} is not editable-styled"
                )


# --- 10-15. validation sources ----------------------------------------------
def test_10_to_14_list_validations_use_the_config_masters() -> None:
    workbook = _wb()
    drivers = _drivers()
    for (register_key, header), source in EXPECTED_LIST_SOURCES.items():
        register = drivers.registers[register_key]
        index = next(i for i, c in enumerate(register.columns) if c.header == header)
        column = register.columns[index]
        assert column.validation is not None, f"{register.table_name}.{header} has no validation"
        assert column.validation["source"] == source, (
            f"{register.table_name}.{header} validates against "
            f"{column.validation['source']!r}, expected {source!r}"
        )
        dv = _dv_covering(workbook[register.sheet], register.data_range(index))
        assert dv is not None, f"{register.table_name}.{header} validation not applied"
        assert dv.type == "list"
        assert dv.formula1 == f"={source}", f"formula1 is {dv.formula1!r}"


def test_15_risk_probability_is_bounded_zero_to_one() -> None:
    register = _register("risk_register")
    index = register.column_index_of("probability")
    dv = _dv_covering(_wb()[RISK_SHEET], register.data_range(index))
    assert dv is not None, "Probability has no data validation"
    assert dv.type == "decimal"
    assert dv.operator == "between"
    assert dv.formula1 == "0" and dv.formula2 == "1"


def test_16_all_driver_validations_allow_blanks() -> None:
    workbook = _wb()
    for register in _drivers().all_registers:
        for dv in workbook[register.sheet].data_validations.dataValidation:
            assert dv.allow_blank, f"{register.sheet} validation {dv.sqref} disallows blanks"


# --- 17-19. rules deliberately NOT imposed ----------------------------------
def test_17_no_validation_imposes_min_ml_max_ordering() -> None:
    """Three-point ordering is a Model Check rule, not a keystroke restriction."""
    workbook = _wb()
    for key, columns in (
        ("cost_lines", ("unit_cost_min", "unit_cost_most_likely", "unit_cost_max")),
        ("risk_register", ("impact_min", "impact_most_likely", "impact_max")),
    ):
        register = _register(key)
        for column_key in columns:
            index = register.column_index_of(column_key)
            assert register.columns[index].validation is None, (
                f"{register.table_name}.{column_key} declares validation"
            )
            assert _dv_covering(workbook[register.sheet], register.data_range(index)) is None


def test_18_no_quantity_constraint_invented() -> None:
    register = _register("cost_lines")
    index = register.column_index_of("quantity")
    assert register.columns[index].validation is None
    assert _dv_covering(_wb()[COST_SHEET], register.data_range(index)) is None


def test_19_no_cost_total_user_input_column() -> None:
    register = _register("cost_lines")
    forbidden = {"total", "total cost", "line total", "cost total", "amount"}
    offenders = [h for h in register.headers if h.strip().lower() in forbidden]
    assert not offenders, f"a user-input total exists: {offenders}"


# --- 20-23. no faked identity, no calculation, no VBA ------------------------
def test_20_no_ids_are_pre_generated() -> None:
    workbook = _wb()
    for register in _drivers().all_registers:
        worksheet = workbook[register.sheet]
        letter = register.column_letter(0)
        for row in range(register.first_data_row, register.last_data_row + 1):
            value = worksheet[f"{letter}{row}"].value
            assert value is None, f"{letter}{row} holds {value!r}"


def test_21_no_row_number_formula_used_for_identity() -> None:
    workbook = _wb()
    for register in _drivers().all_registers:
        worksheet = workbook[register.sheet]
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    upper = cell.value.upper()
                    assert not upper.startswith("=ROW"), f"{cell.coordinate} uses ROW()"
                    assert "ROW()" not in upper, f"{cell.coordinate} references ROW()"


def test_22_no_formulas_or_business_calculations() -> None:
    workbook = _wb()
    offenders = [
        f"{ws.title}!{c.coordinate}"
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert not offenders, f"formulas present: {offenders[:10]}"


def test_23_no_vba_project() -> None:
    with zipfile.ZipFile(_artifact()) as archive:
        names = [n.lower() for n in archive.namelist()]
    assert not any(n.endswith("vbaproject.bin") for n in names)
    assert _artifact().suffix == ".xlsx"


# --- 24-27. prior phases and preflight --------------------------------------
def test_24_phase2_setup_and_config_still_intact() -> None:
    workbook = _wb()
    contract = load_contract(CONTRACT_PATH)
    assert workbook["Setup"][contract.inputs["reporting_currency"].cell].value == "SAR"
    assert workbook["Setup"][contract.inputs["selected_confidence_level"].cell].value == "P50"
    assert workbook["Setup"][contract.inputs["monte_carlo_iterations"].cell].value == 10000
    for name in ("tblCategories", "tblCurrencies", "tblUOM", "tblInflationProfiles",
                 "tblDistributions", "tblConfidenceLevels"):
        assert name in getattr(workbook["Config"], "tables", {})
    assert "tblFXRates" in getattr(workbook["Setup"], "tables", {})


def test_25_sar_model_invariants_still_hold() -> None:
    contract = load_contract(CONTRACT_PATH)
    workbook = _wb()
    for identity in contract.model_invariants["locked_identities"]:
        table = contract.table_by_name(identity["table"])
        worksheet = workbook[table.sheet]
        row = table.first_data_row + identity["row"] - 1
        for index, value in enumerate(identity["values"]):
            assert worksheet[f"{table.column_letter(index)}{row}"].value == value


def test_26_manifest_and_input_contract_currencies_agree() -> None:
    spec = load_spec(SPEC_PATH)
    contract = load_contract(CONTRACT_PATH)
    assert spec.model["reporting_currency"] == contract.reporting_currency


def test_27_version_file_matches_model_version() -> None:
    version = (PCCM_ROOT / "VERSION").read_text(encoding="utf-8").strip()
    spec = load_spec(SPEC_PATH)
    assert version == spec.model["model_version"], (
        f"pccm/VERSION is {version}, manifest model_version is {spec.model['model_version']}"
    )


# --- 28-30. contract integrity ----------------------------------------------
def test_28_driver_tables_stay_inside_excel_bounds() -> None:
    from pccm_builder.contract_loader import EXCEL_MAX_COLUMN, EXCEL_MAX_ROW
    for register in _drivers().all_registers:
        assert register.last_data_row <= EXCEL_MAX_ROW
        assert register.first_col_index + len(register.columns) - 1 <= EXCEL_MAX_COLUMN


def test_29_reserved_rows_are_capacity_not_a_declared_maximum() -> None:
    """The reserved row count must not encode a business limit.

    The Architecture design targets (200 cost lines, 100 risks) are benchmark
    figures and must appear nowhere in the contract as a cap.
    """
    drivers = _drivers()
    raw = DRIVERS_PATH.read_text(encoding="utf-8")
    for register in drivers.all_registers:
        assert register.reserved_rows > 0
        assert register.reserved_rows not in (200, 100), (
            f"{register.table_name} reserved_rows equals an Architecture design target, "
            "which would read as a hard cap"
        )
    for token in ("max_rows", "maximum_rows", "max_cost_lines", "max_risks", "row_limit"):
        assert token not in raw, f"driver contract declares a hard limit key {token!r}"


def test_30_conditional_formatting_is_presentation_only() -> None:
    """The Uniform/Most-Likely greying must not act as input enforcement."""
    workbook = _wb()
    drivers = _drivers()
    for register in drivers.all_registers:
        worksheet = workbook[register.sheet]
        rules = list(worksheet.conditional_formatting)
        assert rules, f"{register.sheet} has no conditional formatting"
        for cf in register.conditional_formatting:
            index = register.column_index_of(cf.target_column)
            column = register.columns[index]
            # It greys the cell; it does not restrict what may be typed.
            assert column.validation is None, (
                f"{register.table_name}.{cf.target_column} has data validation; the "
                "Uniform greying must remain presentation-only"
            )
            assert _dv_covering(worksheet, register.data_range(index)) is None
            assert column.editable, "the Most Likely cell must remain user-editable"


def _run_all() -> int:
    tests = sorted(
        (name, fn) for name, fn in globals().items()
        if name.startswith("test_") and callable(fn)
    )
    failures = 0
    print("PCCM Phase 3 tests - Cost Lines & Risk Register input layer")
    print("=" * 70)
    for name, fn in tests:
        try:
            fn()
        except AssertionError as error:
            failures += 1
            print(f"  [FAIL] {name}\n         {error}")
        except Exception as error:  # noqa: BLE001
            failures += 1
            print(f"  [ERROR] {name}\n          {type(error).__name__}: {error}")
        else:
            print(f"  [PASS] {name}")
    print("=" * 70)
    print(f"  {len(tests) - failures} passed, {failures} failed")
    if _TEMPDIR is not None:
        _TEMPDIR.cleanup()
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(_run_all())
