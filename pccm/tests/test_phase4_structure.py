#!/usr/bin/env python3
"""PCCM Phase 4 tests: structural runtime surfaces in the Stage-A artifact.

Runs standalone or under pytest.

The expected names, cells and grid schemas are re-declared here independently of
the structure contract, so a contract change that alters the structural surface
fails here even when the build itself succeeds.

Stage A carries the STRUCTURE. It carries no applied timeline, no generated year
column and no permanent identifier, because none of those exist until Stage-B
runtime creates them. Every assertion below is about that resting state.
"""

from __future__ import annotations

import os
import sys
import tempfile
import zipfile
from pathlib import Path

PCCM_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PCCM_ROOT / "builder"))

from openpyxl import load_workbook  # noqa: E402

from pccm_builder import (  # noqa: E402
    BUILDER_VERSION,
    build_workbook,
    load_contract,
    load_driver_contract,
    load_spec,
    load_structure_contract,
)

SPEC_PATH = PCCM_ROOT / "spec" / "workbook.yaml"
CONTRACT_PATH = PCCM_ROOT / "spec" / "input_contract.yaml"
DRIVERS_PATH = PCCM_ROOT / "spec" / "driver_contract.yaml"
STRUCTURE_PATH = PCCM_ROOT / "spec" / "structure_contract.yaml"

# --- independent structural lock --------------------------------------------
# The entered triple keeps its accepted Phase-2 inp* names; the nm*_Entered names
# are aliases over the SAME cells, never a second copy of the value.
ENTERED_ALIASES = {
    "nmBaseYear_Entered": ("inpBaseYear", "C12"),
    "nmStartYear_Entered": ("inpProjectStartYear", "C11"),
    "nmDuration_Entered": ("inpDurationYears", "C10"),
}

APPLIED_NAMES = ["nmBaseYear_Applied", "nmStartYear_Applied", "nmDuration_Applied"]
DERIVED_NAMES = [
    "nmLastYear_Applied",
    "nmYearCount_Applied",
    "nmInflFirstYear",
    "nmInflLastYear",
]
STATE_NAME = "nmStructuralState"
COUNTER_NAMES = ["nmCounterCostLine", "nmCounterRisk"]

GRID_SCHEMA = {
    "tblCostProfiling": ("Cost Profiling", ["Cost Line ID", "Description"]),
    "tblRiskProfiling": ("Risk Profiling", ["Risk ID", "Risk Name"]),
    "tblInflation": ("Inflation", ["Inflation Profile"]),
}

ID_FORMATS = {"nmCounterCostLine": "CL-", "nmCounterRisk": "R-"}

_CACHE: dict[str, Path] = {}
_TEMPDIR: tempfile.TemporaryDirectory | None = None


# ---------------------------------------------------------------------------
def _artifact() -> Path:
    global _TEMPDIR
    if "primary" in _CACHE:
        return _CACHE["primary"]
    if _TEMPDIR is None:
        _TEMPDIR = tempfile.TemporaryDirectory(prefix="pccm-phase4-")
    previous = os.environ.get("PCCM_BUILD_TIMESTAMP")
    os.environ["PCCM_BUILD_TIMESTAMP"] = "1970-01-01 00:00:00 UTC"
    try:
        workbook, _ = build_workbook(
            load_spec(SPEC_PATH),
            load_contract(CONTRACT_PATH),
            load_driver_contract(DRIVERS_PATH),
            load_structure_contract(STRUCTURE_PATH),
        )
        path = Path(_TEMPDIR.name) / "stage_a.xlsx"
        workbook.save(path)
        workbook.close()
    finally:
        if previous is None:
            os.environ.pop("PCCM_BUILD_TIMESTAMP", None)
        else:
            os.environ["PCCM_BUILD_TIMESTAMP"] = previous
    _CACHE["primary"] = path
    return path


def _wb():
    return load_workbook(_artifact())


def _structure():
    return load_structure_contract(STRUCTURE_PATH)


def _ref(workbook, name: str) -> str:
    assert name in workbook.defined_names, f"defined name {name} is missing"
    return workbook.defined_names[name].attr_text


def _cell_of(reference: str) -> str:
    return reference.split("!", 1)[1].replace("$", "")


def _fill_rgb(cell) -> str:
    rgb = getattr(cell.fill.fgColor, "rgb", None)
    return str(rgb)[-6:].upper() if rgb else ""


# ===========================================================================
# entered vs applied
# ===========================================================================
def test_01_the_accepted_phase2_input_names_still_exist() -> None:
    """Phase 4 must not remove or rename an accepted Phase-2 defined name."""
    workbook = _wb()
    for _alias, (inp_name, cell) in ENTERED_ALIASES.items():
        assert inp_name in workbook.defined_names, f"{inp_name} was removed"
        assert _cell_of(_ref(workbook, inp_name)) == cell


def test_02_entered_aliases_address_the_same_cells_not_copies() -> None:
    workbook = _wb()
    for alias, (inp_name, _cell) in ENTERED_ALIASES.items():
        assert alias in workbook.defined_names, f"{alias} is missing"
        assert _ref(workbook, alias) == _ref(workbook, inp_name), (
            f"{alias} and {inp_name} must resolve to the same cell; there is exactly "
            "one semantic source for each entered value"
        )


def test_03_the_applied_triple_exists_and_is_separate_from_the_entered_triple() -> None:
    workbook = _wb()
    entered_cells = {_ref(workbook, n) for n in ENTERED_ALIASES}
    for name in APPLIED_NAMES:
        assert name in workbook.defined_names, f"{name} is missing"
        assert _ref(workbook, name) not in entered_cells, (
            f"{name} points at an entered cell; applied state must be separate"
        )


def test_04_the_applied_triple_is_blank_on_a_new_workbook() -> None:
    """No timeline is applied, and no year is fabricated."""
    workbook = _wb()
    for name in APPLIED_NAMES:
        cell = workbook["Setup"][_cell_of(_ref(workbook, name))]
        assert cell.value is None, f"{name} holds {cell.value!r}"


def test_05_the_entered_triple_is_blank_on_a_new_workbook() -> None:
    workbook = _wb()
    for _alias, (inp_name, _cell) in ENTERED_ALIASES.items():
        assert workbook["Setup"][_cell_of(_ref(workbook, inp_name))].value is None


def test_06_derived_structural_names_exist_and_carry_formulas() -> None:
    workbook = _wb()
    for name in DERIVED_NAMES:
        cell = workbook["Setup"][_cell_of(_ref(workbook, name))]
        assert isinstance(cell.value, str) and cell.value.startswith("="), (
            f"{name} holds {cell.value!r}; a derived structural value must be a formula "
            "so it cannot go stale against the applied triple"
        )


def test_07_derived_formulas_reference_only_structural_names() -> None:
    """A derived cell may display structural state. It may not calculate cost."""
    import re
    for field in _structure().derived:
        tokens = set(re.findall(r"[A-Za-z_][A-Za-z0-9_.]*", field.formula))
        allowed = {"IF", "OR", "AND"} | set(APPLIED_NAMES) | set(DERIVED_NAMES) | set(ENTERED_ALIASES)
        assert tokens <= allowed, f"{field.key} references {sorted(tokens - allowed)}"


def test_08_applied_last_year_is_start_plus_duration_minus_one() -> None:
    formula = next(f for f in _structure().derived if f.key == "applied_last_year").formula
    assert "nmStartYear_Applied+nmDuration_Applied-1" in formula.replace(" ", "")


def test_09_inflation_first_year_is_the_year_after_the_base_year() -> None:
    formula = next(f for f in _structure().derived if f.key == "inflation_first_year").formula
    assert "nmBaseYear_Applied+1" in formula.replace(" ", "")


# ===========================================================================
# structural state indicator
# ===========================================================================
def test_10_the_structural_state_indicator_exists_and_is_a_formula() -> None:
    workbook = _wb()
    cell = workbook["Setup"][_cell_of(_ref(workbook, STATE_NAME))]
    assert isinstance(cell.value, str) and cell.value.startswith("=")


def test_11_the_indicator_compares_every_entered_value_to_its_applied_counterpart() -> None:
    formula = _structure().structural_state.formula
    for name in list(ENTERED_ALIASES) + APPLIED_NAMES:
        assert name in formula, f"the pending flag does not consider {name}"


def test_12_the_indicator_declares_the_pending_wording() -> None:
    state = _structure().structural_state
    assert state.labels["pending"] == "STRUCTURE CHANGE PENDING"
    assert f'"{state.labels["pending"]}"' in state.formula


def test_13_a_new_workbook_reads_as_no_timeline_applied_not_current() -> None:
    """Blank == blank must not be reported as agreement."""
    state = _structure().structural_state
    assert state.labels["not_applied"] != state.labels["current"]
    assert f'"{state.labels["not_applied"]}"' in state.formula


def test_14_no_worksheet_event_handler_maintains_the_indicator() -> None:
    """The flag is a worksheet formula, never a Worksheet_Change side effect.

    Scanned over CODE only: a comment explaining why the handler is absent is not
    a declaration of it.
    """
    from pccm_builder.vba_source import contains_construct, load_modules
    modules = load_modules([PCCM_ROOT / "src" / "vba"])
    assert modules, "no VBA modules were found"
    for construct in ("Worksheet_Change", "Workbook_SheetChange"):
        offenders = contains_construct(modules, construct)
        assert not offenders, f"{construct} appears in code in {offenders}"


def test_15_the_indicator_is_model_controlled_not_an_input() -> None:
    workbook = _wb()
    structure = _structure()
    cell = workbook["Setup"][structure.structural_state.cell]
    tokens_locked = "EFF1F4"
    assert _fill_rgb(cell) == tokens_locked, "the indicator must not look like an input"


# ===========================================================================
# structural grids
# ===========================================================================
def test_16_every_structural_grid_exists_with_its_fixed_columns() -> None:
    workbook = _wb()
    structure = _structure()
    for table_name, (sheet, headers) in GRID_SCHEMA.items():
        assert table_name in getattr(workbook[sheet], "tables", {}), f"{table_name} is missing"
        grid = next(g for g in structure.all_grids if g.table_name == table_name)
        found = [
            workbook[sheet][f"{grid.column_letter(i)}{grid.header_row}"].value
            for i in range(len(headers))
        ]
        assert found == headers, f"{table_name} fixed columns are {found}"


def test_17_no_year_column_exists_before_a_timeline_is_applied() -> None:
    """Generating a year column would assert a timeline the user never entered."""
    workbook = _wb()
    structure = _structure()
    for grid in structure.all_grids:
        worksheet = workbook[grid.sheet]
        beyond = grid.column_letter(len(grid.fixed_columns))
        assert worksheet[f"{beyond}{grid.header_row}"].value is None, (
            f"{grid.table_name} already has a generated year column"
        )
        table = worksheet.tables[grid.table_name]
        assert table.ref == grid.ref, f"{grid.table_name} ref is {table.ref}"


def test_18_every_grid_shows_the_timeline_not_yet_applied_message() -> None:
    workbook = _wb()
    structure = _structure()
    for grid in structure.all_grids:
        cell = workbook[grid.sheet][f"B{grid.state_message_row}"]
        assert isinstance(cell.value, str) and cell.value.startswith("="), (
            f"{grid.sheet} state message is {cell.value!r}; a formula clears itself once "
            "a timeline is applied, with no macro maintaining it"
        )
        assert "Timeline not yet applied" in cell.value


def test_19_the_inflation_message_covers_the_legitimate_empty_span() -> None:
    structure = _structure()
    formula = structure.state_messages["inflation_formula"]
    assert "nmInflFirstYear>nmInflLastYear" in formula.replace(" ", "")
    assert structure.state_messages["inflation_empty_span"] in formula


def test_20_grid_rows_are_blank_and_model_controlled() -> None:
    workbook = _wb()
    structure = _structure()
    for grid in structure.all_grids:
        worksheet = workbook[grid.sheet]
        for index in range(len(grid.fixed_columns)):
            letter = grid.column_letter(index)
            for row in range(grid.first_data_row, grid.last_data_row + 1):
                cell = worksheet[f"{letter}{row}"]
                assert cell.value is None, f"{grid.table_name} {letter}{row} holds {cell.value!r}"
                assert _fill_rgb(cell) == "EFF1F4", (
                    f"{grid.table_name} {letter}{row} is not model-controlled; a trace "
                    "column is owned by the register, not by the user"
                )


def test_21_profiling_capacity_tracks_its_driver_register() -> None:
    structure = _structure()
    drivers = load_driver_contract(DRIVERS_PATH)
    for grid in structure.profiling_grids:
        register = drivers.registers[grid.driver_register]
        assert grid.reserved_rows == register.reserved_rows, (
            f"{grid.table_name} reserves {grid.reserved_rows}, "
            f"{register.table_name} reserves {register.reserved_rows}"
        )


def test_22_inflation_capacity_tracks_the_config_profile_master() -> None:
    structure = _structure()
    contract = load_contract(CONTRACT_PATH)
    grid = structure.inflation_grid
    source = contract.table_by_name(grid.source_list_table)
    assert source is not None
    assert grid.reserved_rows == source.data_rows


def test_23_a_new_inflation_year_is_blank_and_a_new_profile_cell_is_zero() -> None:
    """The two initial values are deliberately different and both load-bearing."""
    structure = _structure()
    assert structure.inflation_grid.year_column.initial_value is None, (
        "an escalation assumption the user never made must not be fabricated as zero"
    )
    for grid in structure.profiling_grids:
        assert grid.year_column.initial_value == 0.0, (
            "a new project-year cell is 0% so an existing row's total is unchanged"
        )


def test_24_profiling_year_cells_carry_no_invented_validation() -> None:
    structure = _structure()
    for grid in structure.all_grids:
        assert grid.year_column.validation is None, (
            f"{grid.table_name} imposes a cell rule on year values; the 100% requirement "
            "is a Model Check rule and a per-cell bound would block partial entry"
        )


# ===========================================================================
# permanent identity
# ===========================================================================
def test_25_the_id_counters_exist_and_are_seeded_at_zero() -> None:
    workbook = _wb()
    structure = _structure()
    for counter in structure.counters:
        assert counter.defined_name in COUNTER_NAMES
        cell = workbook[structure.identity_sheet][counter.cell]
        assert cell.value == 0, f"{counter.defined_name} is seeded {cell.value!r}"


def test_26_the_counters_live_on_the_hidden_calc_sheet() -> None:
    structure = _structure()
    spec = load_spec(SPEC_PATH)
    sheet = spec.sheet(structure.identity_sheet)
    assert sheet.visibility == "hidden", (
        "the counters are internal state, but must stay inspectable by an auditor"
    )


def test_27_id_prefixes_and_patterns_are_declared_and_consistent() -> None:
    import re
    structure = _structure()
    for counter in structure.counters:
        expected = ID_FORMATS[counter.defined_name]
        assert counter.prefix == expected, f"{counter.key} prefix is {counter.prefix!r}"
        assert re.match(counter.pattern, counter.format_id(1))
        assert counter.format_id(1) == f"{expected}001"


def test_28_no_artificial_id_maximum_is_imposed() -> None:
    import re
    structure = _structure()
    for counter in structure.counters:
        wide = counter.format_id(1000)
        assert re.match(counter.pattern, wide), (
            f"{counter.key} would reject {wide}; pad width is a display floor, never a cap"
        )


def test_29_no_permanent_id_is_allocated_in_stage_a() -> None:
    """Stage A still allocates nothing: the registers' ID columns stay blank."""
    workbook = _wb()
    drivers = load_driver_contract(DRIVERS_PATH)
    for register in drivers.all_registers:
        worksheet = workbook[register.sheet]
        letter = register.column_letter(0)
        for row in range(register.first_data_row, register.last_data_row + 1):
            assert worksheet[f"{letter}{row}"].value is None


def test_30_no_profiling_row_is_pre_seeded() -> None:
    workbook = _wb()
    for table_name, (sheet, _headers) in GRID_SCHEMA.items():
        grid = next(g for g in _structure().all_grids if g.table_name == table_name)
        letter = grid.column_letter(0)
        for row in range(grid.first_data_row, grid.last_data_row + 1):
            assert workbook[sheet][f"{letter}{row}"].value is None


# ===========================================================================
# scope discipline
# ===========================================================================
def test_31_only_structural_state_formulas_exist() -> None:
    workbook = _wb()
    permitted = _structure().formula_cells
    offenders = [
        f"{ws.title}!{c.coordinate}"
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str)
        and c.value.startswith("=")
        and c.coordinate not in permitted.get(ws.title, set())
    ]
    assert not offenders, f"formulas outside the permitted structural cells: {offenders}"


def test_32_no_business_calculation_function_appears_in_any_formula() -> None:
    workbook = _wb()
    forbidden = ("SUMPRODUCT", "NPV", "PERCENTILE", "NORM", "RAND", "VLOOKUP", "SUMIF")
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    upper = cell.value.upper()
                    for token in forbidden:
                        assert token not in upper, f"{ws.title}!{cell.coordinate} uses {token}"


def test_33_stage_a_still_contains_no_vba() -> None:
    with zipfile.ZipFile(_artifact()) as archive:
        names = [n.lower() for n in archive.namelist()]
    assert not any(n.endswith("vbaproject.bin") for n in names)
    assert _artifact().suffix == ".xlsx"


def test_34_the_fourteen_sheet_structure_is_unchanged() -> None:
    workbook = _wb()
    spec = load_spec(SPEC_PATH)
    assert workbook.sheetnames == spec.sheet_names
    assert len(workbook.sheetnames) == 14


def test_35_phase_2_and_3_surfaces_are_intact() -> None:
    workbook = _wb()
    contract = load_contract(CONTRACT_PATH)
    assert workbook["Setup"][contract.inputs["reporting_currency"].cell].value == "SAR"
    for identity in contract.model_invariants["locked_identities"]:
        table = contract.table_by_name(identity["table"])
        row = table.first_data_row + identity["row"] - 1
        for index, value in enumerate(identity["values"]):
            assert workbook[table.sheet][f"{table.column_letter(index)}{row}"].value == value
    for name in ("tblCostLines", "tblRiskRegister"):
        assert any(name in getattr(ws, "tables", {}) for ws in workbook.worksheets)


def test_36_version_file_matches_the_model_version() -> None:
    """The repository's phase-version convention: `VERSION`, the manifest's
    `model_version` and `BUILDER_VERSION` advance together, once per phase.

    The literal moved from `0.4.0` to `0.5.0` when Phase 5 Gate-A Step 3 put real
    Phase-5 blocks in the generated workbook. What the test proves — that all
    three agree and none drifts on its own — is unchanged, and `BUILDER_VERSION`
    is now checked here too rather than only implied.
    """
    version = (PCCM_ROOT / "VERSION").read_text(encoding="utf-8").strip()
    spec = load_spec(SPEC_PATH)
    assert version == spec.model["model_version"] == BUILDER_VERSION == "0.5.0", (
        f"VERSION={version}, model_version={spec.model['model_version']}, "
        f"BUILDER_VERSION={BUILDER_VERSION}"
    )


def test_37_the_applied_block_sits_clear_of_the_phase2_setup_area() -> None:
    contract = load_contract(CONTRACT_PATH)
    structure = _structure()
    occupied = contract.occupied_rows("Setup")
    for field in structure.structural_fields:
        row = int("".join(ch for ch in field.cell if ch.isdigit()))
        assert row not in occupied, f"{field.key} lands on occupied Setup row {row}"


def test_38_no_worksheet_protection_is_applied() -> None:
    workbook = _wb()
    for worksheet in workbook.worksheets:
        assert not worksheet.protection.sheet, f"{worksheet.title} is protected"


def test_39_no_later_phase_sheet_gained_a_table() -> None:
    workbook = _wb()
    for sheet in ("Model Check", "Results", "Sensitivity", "Dashboard", "_SimData"):
        assert not getattr(workbook[sheet], "tables", {}), f"{sheet} declares a table"


def test_40_the_structural_limits_are_not_business_maxima() -> None:
    """25 years must appear nowhere as a limit."""
    raw = STRUCTURE_PATH.read_text(encoding="utf-8")
    structure = _structure()
    assert structure.limits.max_generated_year_columns != 25
    for token in ("max_duration", "max_years", "max_cost_lines", "max_risks", "duration_cap"):
        assert token not in raw, f"the structure contract declares a hard limit key {token!r}"


def test_41_the_two_structural_protections_are_independent() -> None:
    """The calendar-year window and the project-year column guard bound different things.

    Architecture Lock Revision B fixes "Generated column count > 200 = ERROR" for
    generated PROJECT-YEAR columns. The 1900-2200 window separately bounds Base Year,
    Start Year, Last Project Year and therefore the inflation span. Deriving either
    from the other is what produced the wrong 301-column guard.
    """
    limits = _structure().limits
    assert limits.max_generated_year_columns == 200
    assert limits.min_year == 1900 and limits.max_year == 2200
    assert limits.max_generated_year_columns != limits.max_year - limits.min_year + 1


def test_42_the_contract_rejects_a_derived_generation_guard() -> None:
    """The loader must refuse any value other than the locked 200."""
    import copy
    import tempfile
    import yaml
    from pccm_builder import StructureContractError
    data = yaml.safe_load(STRUCTURE_PATH.read_text(encoding="utf-8"))
    for wrong in (301, 25, 199, 201):
        broken = copy.deepcopy(data)
        broken["limits"]["max_generated_year_columns"] = wrong
        with tempfile.TemporaryDirectory(prefix="pccm-limit-") as tmp:
            path = Path(tmp) / "structure.yaml"
            path.write_text(yaml.safe_dump(broken, sort_keys=False), encoding="utf-8")
            try:
                load_structure_contract(path)
            except StructureContractError as error:
                assert "200" in str(error)
                continue
        raise AssertionError(f"a generation guard of {wrong} was accepted")


def test_43_the_emitted_constants_carry_the_locked_two_hundred() -> None:
    import tempfile
    from pccm_builder import emit_stage_b
    with tempfile.TemporaryDirectory(prefix="pccm-emit-") as tmp:
        artifacts = emit_stage_b(
            Path(tmp),
            load_spec(SPEC_PATH),
            load_contract(CONTRACT_PATH),
            load_driver_contract(DRIVERS_PATH),
            _structure(),
        )
        text = artifacts.module_path.read_text(encoding="utf-8")
    assert "LIMIT_MAX_YEAR_COLUMNS As Long = 200" in text
    assert "301" not in text, "the derived guard must be gone from the generated module"


def _run_all() -> int:
    tests = sorted(
        (name, fn) for name, fn in globals().items()
        if name.startswith("test_") and callable(fn)
    )
    failures = 0
    print("PCCM Phase 4 tests - structural runtime surfaces")
    print("=" * 70)
    for name, fn in tests:
        try:
            fn()
        except AssertionError as error:
            failures += 1
            print(f"  [FAIL] {name}\n         {error}")
        except Exception as error:  # noqa: BLE001
            failures += 1
            print(f"  [ERROR] {name}\n          {type(error).__name__}: {error}")
        else:
            print(f"  [PASS] {name}")
    print("=" * 70)
    print(f"  {len(tests) - failures} passed, {failures} failed")
    if _TEMPDIR is not None:
        _TEMPDIR.cleanup()
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(_run_all())
