#!/usr/bin/env python3
"""PCCM Phase 2 tests: functional Setup and Configuration layer.

Runs standalone or under pytest.

The locked input keys, defaults, table names and constants are re-declared here
independently of the input contract. The contract is the builder's authority;
this file is the conformance check, so contract drift fails here even when the
build itself succeeds.
"""

from __future__ import annotations

import os
import sys
import tempfile
import zipfile
from pathlib import Path

PCCM_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PCCM_ROOT / "builder"))

from openpyxl import load_workbook  # noqa: E402

from pccm_builder import (  # noqa: E402
    build_workbook,
    load_contract,
    load_driver_contract,
    load_structure_contract,
    load_spec,
)

SPEC_PATH = PCCM_ROOT / "spec" / "workbook.yaml"
CONTRACT_PATH = PCCM_ROOT / "spec" / "input_contract.yaml"
DRIVERS_PATH = PCCM_ROOT / "spec" / "driver_contract.yaml"
STRUCTURE_PATH = PCCM_ROOT / "spec" / "structure_contract.yaml"

# --- locked Phase 2 decisions ----------------------------------------------
SETUP_INPUTS = {
    "project_name": ("Project Name", "inpProjectName", None),
    "duration_years": ("Project Duration (Years)", "inpDurationYears", None),
    "project_start_year": ("Project Start Year", "inpProjectStartYear", None),
    "base_year": ("Base Year", "inpBaseYear", None),
    "reporting_currency": ("Reporting Currency", "inpReportingCurrency", "SAR"),
    "selected_confidence_level": ("Selected Confidence Level", "inpSelectedConfidenceLevel", "P50"),
    "monte_carlo_iterations": ("Monte Carlo Iterations", "inpMonteCarloIterations", 10000),
    "discount_rate": ("Discount Rate", "inpDiscountRate", None),
    "random_seed": ("Random Seed", "inpRandomSeed", None),
}

CONFIG_TABLES = {
    "tblCategories": "lstCategories",
    "tblCurrencies": "lstCurrencies",
    "tblUOM": "lstUOM",
    "tblInflationProfiles": "lstInflationProfiles",
    "tblDistributions": "lstDistributions",
    "tblConfidenceLevels": "lstConfidenceLevels",
}

FX_TABLE = "tblFXRates"
FX_COLUMNS = ["Currency", "FX to SAR"]
LOCKED_DISTRIBUTIONS = ["Triangular", "Beta-PERT", "Uniform"]
LOCKED_CONFIDENCE_LEVELS = [f"P{n}" for n in range(50, 100, 5)]
MINIMUM_ITERATIONS = "1000"

_CACHE: dict[str, Path] = {}
_TEMPDIR: tempfile.TemporaryDirectory | None = None


# ---------------------------------------------------------------------------
def _artifact() -> Path:
    global _TEMPDIR
    if "primary" in _CACHE:
        return _CACHE["primary"]
    if _TEMPDIR is None:
        _TEMPDIR = tempfile.TemporaryDirectory(prefix="pccm-phase2-")
    previous = os.environ.get("PCCM_BUILD_TIMESTAMP")
    os.environ["PCCM_BUILD_TIMESTAMP"] = "1970-01-01 00:00:00 UTC"
    try:
        workbook, _ = build_workbook(
            load_spec(SPEC_PATH),
            load_contract(CONTRACT_PATH),
            load_driver_contract(DRIVERS_PATH),
            load_structure_contract(STRUCTURE_PATH),
        )
        path = Path(_TEMPDIR.name) / "stage_a.xlsx"
        workbook.save(path)
        workbook.close()
    finally:
        if previous is None:
            os.environ.pop("PCCM_BUILD_TIMESTAMP", None)
        else:
            os.environ["PCCM_BUILD_TIMESTAMP"] = previous
    _CACHE["primary"] = path
    return path


def _wb():
    return load_workbook(_artifact())


def _validations(worksheet) -> list:
    return list(worksheet.data_validations.dataValidation)


def _dv_for(worksheet, address: str):
    for dv in _validations(worksheet):
        if address in str(dv.sqref):
            return dv
    return None


def _permitted_formula_cells() -> dict:
    """The exact cells the structure contract is allowed to write a formula into.

    Phases 1-3 forbade every formula. Phase 4 permits structural-state display only
    -- Structure Change Pending, the derived applied-timeline cells and each grid's
    "timeline not yet applied" message -- and the contract enumerates them. The
    regression intent is unchanged: any formula outside this enumerated set is still
    a failure, and a business calculation would land outside it.
    """
    structure = load_structure_contract(STRUCTURE_PATH)
    return structure.formula_cells


def _unexpected_formulas(workbook) -> list[str]:
    permitted = _permitted_formula_cells()
    return [
        f"{ws.title}!{cell.coordinate}"
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
        and cell.value.startswith("=")
        and cell.coordinate not in permitted.get(ws.title, set())
    ]


# --- 1-2. Setup inputs and defined names -----------------------------------
def test_01_all_nine_setup_inputs_exist() -> None:
    contract = load_contract(CONTRACT_PATH)
    workbook = _wb()
    setup = workbook["Setup"]
    assert len(contract.inputs) == 9, f"expected 9 Setup inputs, found {len(contract.inputs)}"
    for key, (label, _, _) in SETUP_INPUTS.items():
        assert key in contract.inputs, f"contract is missing input {key!r}"
        spec = contract.inputs[key]
        assert spec.label == label, f"{key}: label is {spec.label!r}, expected {label!r}"
        assert setup[spec.label_cell].value == label, f"{key}: label not written to {spec.label_cell}"


def test_02_setup_defined_names_resolve_to_intended_cells() -> None:
    contract = load_contract(CONTRACT_PATH)
    workbook = _wb()
    for key, (_, defined_name, _) in SETUP_INPUTS.items():
        spec = contract.inputs[key]
        assert spec.defined_name == defined_name, (
            f"{key}: contract name {spec.defined_name!r}, lock says {defined_name!r}"
        )
        assert defined_name in workbook.defined_names, f"missing defined name {defined_name}"
        expected = f"'{spec.sheet}'!${spec.cell[0]}${spec.cell[1:]}"
        assert workbook.defined_names[defined_name].attr_text == expected, (
            f"{defined_name} -> {workbook.defined_names[defined_name].attr_text}, expected {expected}"
        )


# --- 3-7. locked defaults ---------------------------------------------------
def _default_of(key: str):
    contract = load_contract(CONTRACT_PATH)
    return _wb()["Setup"][contract.inputs[key].cell].value


def test_03_reporting_currency_is_sar_and_model_controlled() -> None:
    contract = load_contract(CONTRACT_PATH)
    spec = contract.inputs["reporting_currency"]
    assert spec.default == "SAR"
    assert spec.editable is False, "Reporting Currency must be model-controlled"
    assert spec.validation is None, "Reporting Currency must not have a dropdown"
    assert _default_of("reporting_currency") == "SAR"


def test_04_selected_confidence_level_defaults_to_p50() -> None:
    assert _default_of("selected_confidence_level") == "P50"


def test_05_monte_carlo_iterations_default_is_10000() -> None:
    assert _default_of("monte_carlo_iterations") == 10000


def test_06_random_seed_is_blank() -> None:
    assert _default_of("random_seed") is None


def test_07_discount_rate_is_blank() -> None:
    assert _default_of("discount_rate") is None


# --- 8-12. Config tables and constants --------------------------------------
def test_08_all_six_config_tables_exist() -> None:
    tables = getattr(_wb()["Config"], "tables", {})
    for name in CONFIG_TABLES:
        assert name in tables, f"Config is missing Excel Table {name}"
    assert set(tables) == set(CONFIG_TABLES), f"unexpected Config tables: {set(tables) - set(CONFIG_TABLES)}"


def test_09_currency_table_contains_sar() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = next(t for t in contract.all_tables if t.table_name == "tblCurrencies")
    values = _column_values(_wb()["Config"], table)
    assert "SAR" in values, f"tblCurrencies does not contain SAR: {values}"


def test_10_currency_table_has_no_fx_rate_field() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = next(t for t in contract.all_tables if t.table_name == "tblCurrencies")
    headers = [c.header for c in table.columns]
    assert headers == ["Currency"], f"tblCurrencies schema is {headers}, expected ['Currency']"


def test_11_distribution_constants_are_exact() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = next(t for t in contract.all_tables if t.table_name == "tblDistributions")
    assert _column_values(_wb()["Config"], table) == LOCKED_DISTRIBUTIONS
    assert table.editable is False, "distribution list must be a locked constant"


def test_12_confidence_level_constants_are_exact() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = next(t for t in contract.all_tables if t.table_name == "tblConfidenceLevels")
    assert _column_values(_wb()["Config"], table) == LOCKED_CONFIDENCE_LEVELS
    assert table.editable is False, "confidence level list must be a locked constant"


def _column_values(worksheet, table, index: int = 0) -> list:
    letter = table.column_letter(index)
    values = [
        worksheet[f"{letter}{row}"].value
        for row in range(table.first_data_row, table.last_data_row + 1)
    ]
    return [v for v in values if v is not None]


# --- 13-14. FX table --------------------------------------------------------
def test_13_setup_fx_table_exists_with_exact_schema() -> None:
    contract = load_contract(CONTRACT_PATH)
    assert FX_TABLE in getattr(_wb()["Setup"], "tables", {}), "Setup is missing tblFXRates"
    table = next(t for t in contract.all_tables if t.table_name == FX_TABLE)
    assert [c.header for c in table.columns] == FX_COLUMNS


def test_14_sar_fx_identity_row_is_one() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = next(t for t in contract.all_tables if t.table_name == FX_TABLE)
    setup = _wb()["Setup"]
    currency = setup[f"{table.column_letter(0)}{table.first_data_row}"].value
    rate = setup[f"{table.column_letter(1)}{table.first_data_row}"].value
    assert currency == "SAR", f"first FX row is {currency!r}, expected SAR"
    assert rate == 1, f"SAR identity rate is {rate!r}, expected 1"


# --- 15. list defined names -------------------------------------------------
def test_15_required_list_defined_names_exist() -> None:
    workbook = _wb()
    contract = load_contract(CONTRACT_PATH)
    for table_name, list_name in CONFIG_TABLES.items():
        assert list_name in workbook.defined_names, f"missing defined name {list_name}"
        table = next(t for t in contract.all_tables if t.table_name == table_name)
        assert workbook.defined_names[list_name].attr_text == table.absolute_data_range(), (
            f"{list_name} does not cover the {table_name} data body"
        )


# --- 16-19. data validation -------------------------------------------------
def test_16_confidence_level_validation_uses_the_list() -> None:
    contract = load_contract(CONTRACT_PATH)
    setup = _wb()["Setup"]
    dv = _dv_for(setup, contract.inputs["selected_confidence_level"].cell)
    assert dv is not None, "Selected Confidence Level has no data validation"
    assert dv.type == "list"
    assert dv.formula1 == "=lstConfidenceLevels", f"formula1 is {dv.formula1!r}"


def test_17_iterations_validation_enforces_minimum_1000() -> None:
    contract = load_contract(CONTRACT_PATH)
    setup = _wb()["Setup"]
    dv = _dv_for(setup, contract.inputs["monte_carlo_iterations"].cell)
    assert dv is not None, "Monte Carlo Iterations has no data validation"
    assert dv.type == "whole"
    assert dv.operator == "greaterThanOrEqual"
    assert dv.formula1 == MINIMUM_ITERATIONS, f"formula1 is {dv.formula1!r}"


def test_18_fx_currency_validation_uses_the_currency_master() -> None:
    """Targets the user-owned rows only; the locked SAR identity row is excluded."""
    contract = load_contract(CONTRACT_PATH)
    table = next(t for t in contract.all_tables if t.table_name == FX_TABLE)
    dv = _dv_for(_wb()["Setup"], table.user_data_range(0))
    assert dv is not None, "FX Currency user rows have no data validation"
    assert dv.type == "list"
    assert dv.formula1 == "=lstCurrencies", f"formula1 is {dv.formula1!r}"


def test_19_fx_rate_validation_requires_positive_when_populated() -> None:
    """Targets the user-owned rows only; the locked SAR = 1 identity is excluded."""
    contract = load_contract(CONTRACT_PATH)
    table = next(t for t in contract.all_tables if t.table_name == FX_TABLE)
    dv = _dv_for(_wb()["Setup"], table.user_data_range(1))
    assert dv is not None, "FX rate user rows have no data validation"
    assert dv.type == "decimal"
    assert dv.operator == "greaterThan"
    assert dv.formula1 == "0"
    assert dv.allow_blank, "blank unused FX rows must remain acceptable"


def test_19b_duration_validation_is_positive_and_uncapped() -> None:
    contract = load_contract(CONTRACT_PATH)
    dv = _dv_for(_wb()["Setup"], contract.inputs["duration_years"].cell)
    assert dv is not None, "Project Duration has no data validation"
    assert dv.type == "whole"
    assert dv.operator == "greaterThanOrEqual"
    assert dv.formula1 == "1"
    assert dv.formula2 in (None, ""), "Project Duration must not be capped"


def test_19c_no_unsupported_validation_invented() -> None:
    """Start Year, Base Year, Discount Rate still have no business limits.

    Random Seed WAS in this list, with the docstring "no business limits YET" and
    the contract note "the admissible domain is fixed when the RNG is
    implemented". Phase-6 D6-19 and D6-20 discharged that deferral, so the seed
    now HAS an accepted domain and is checked by test_19d below instead. This is
    a deferral being resolved, not a limit being dropped: the seed moved from
    "unvalidated" to "validated against an accepted authority", and the
    protection here is strictly greater than before.
    """
    contract = load_contract(CONTRACT_PATH)
    setup = _wb()["Setup"]
    for key in ("project_start_year", "base_year", "discount_rate", "project_name"):
        assert contract.inputs[key].validation is None, f"{key} declares a validation rule"
        assert _dv_for(setup, contract.inputs[key].cell) is None, f"{key} has an applied validation"


def test_19d_random_seed_carries_the_accepted_domain() -> None:
    """D6-20: blank is AUTO; populated is FIXED over 1 .. 2147483646 inclusive."""
    contract = load_contract(CONTRACT_PATH)
    seed = contract.inputs["random_seed"]
    assert seed.required is False, "blank must stay legal - it is the AUTO request"
    assert seed.default is None
    validation = seed.validation
    assert validation is not None, "the admissible domain is no longer deferred"
    assert validation["kind"] == "whole", "fractional seeds are refused"
    assert validation["operator"] == "between"
    assert validation["formula1"] == "1", "0 and negatives are refused"
    assert validation["formula2"] == "2147483646"

    dv = _dv_for(_wb()["Setup"], seed.cell)
    assert dv is not None, "the accepted domain is not applied to the sheet"
    assert dv.type == "whole"
    assert dv.operator == "between"
    assert (dv.formula1, dv.formula2) == ("1", "2147483646")
    assert dv.allow_blank is True, "blank means AUTO and must remain enterable"
    assert dv.errorStyle == "stop"


# --- 20-23. nothing from a later phase --------------------------------------
def test_20_only_expected_tables_exist() -> None:
    """Phase-aware. Phase 3 added the driver registers, Phase 4 the structural
    grids. Sheets belonging to still-unimplemented phases must remain bare."""
    drivers = load_driver_contract(DRIVERS_PATH)
    structure = load_structure_contract(STRUCTURE_PATH)
    workbook = _wb()
    for sheet in ("Model Check", "Results", "Sensitivity", "Dashboard",
                  "Methodology", "_Calc", "_SimData"):
        assert not getattr(workbook[sheet], "tables", {}), f"{sheet} already declares a table"
    all_tables = {n for ws in workbook.worksheets for n in getattr(ws, "tables", {})}
    expected = (
        set(CONFIG_TABLES)
        | {FX_TABLE}
        | {r.table_name for r in drivers.all_registers}
        | {g.table_name for g in structure.all_grids}
    )
    assert all_tables == expected, f"unexpected tables: {all_tables ^ expected}"


def test_21_no_formulas_or_business_calculations() -> None:
    workbook = _wb()
    offenders = _unexpected_formulas(workbook)
    assert not offenders, f"formulas outside the permitted structural cells: {offenders[:10]}"


def test_22_no_vba_project() -> None:
    with zipfile.ZipFile(_artifact()) as archive:
        names = [n.lower() for n in archive.namelist()]
    assert not any(n.endswith("vbaproject.bin") for n in names)
    assert _artifact().suffix == ".xlsx"


def test_23_no_external_links_or_connections() -> None:
    with zipfile.ZipFile(_artifact()) as archive:
        names = archive.namelist()
    assert not [n for n in names if "externalLink" in n]
    assert not [n for n in names if "connections" in n.lower()]
    assert not [n for n in names if "queryTable" in n]


# --- 24-25. Phase 1 guarantees still hold ------------------------------------
def test_24_phase1_sheet_order_and_visibility_still_hold() -> None:
    workbook = _wb()
    spec = load_spec(SPEC_PATH)
    assert workbook.sheetnames == spec.workbook["locked_sheet_order"]
    assert workbook.active.title == "Dashboard"
    assert workbook["_Calc"].sheet_state == "hidden"
    assert workbook["_SimData"].sheet_state == "veryHidden"
    for name in workbook.sheetnames:
        if name not in ("_Calc", "_SimData"):
            assert workbook[name].sheet_state == "visible", f"{name} is not visible"


def test_25_artifact_reopens() -> None:
    workbook = load_workbook(_artifact())
    assert len(workbook.sheetnames) == 14
    workbook.close()


# --- 26-28. contract integrity ----------------------------------------------
def test_26_contract_defined_names_are_unique() -> None:
    contract = load_contract(CONTRACT_PATH)
    names = list(contract.input_defined_names) + list(contract.list_defined_names)
    assert len(names) == len(set(names)), "duplicate defined names in the contract"
    tables = [t.table_name for t in contract.all_tables]
    assert len(tables) == len(set(tables)), "duplicate table names in the contract"
    assert not set(names) & set(tables), "a name is used for both a range and a table"


def test_27_contract_targets_are_valid() -> None:
    contract = load_contract(CONTRACT_PATH)
    spec = load_spec(SPEC_PATH)
    known = set(spec.sheet_names)
    workbook = _wb()
    for input_spec in contract.inputs.values():
        assert input_spec.sheet in known, f"{input_spec.key} targets unknown sheet"
        assert workbook[input_spec.sheet][input_spec.cell] is not None
    for table in contract.all_tables:
        assert table.sheet in known, f"{table.table_name} targets unknown sheet"
        assert table.table_name in getattr(workbook[table.sheet], "tables", {})


def test_28_setup_is_the_only_fx_rate_owner() -> None:
    contract = load_contract(CONTRACT_PATH)
    rate_tables = [
        (t.sheet, t.table_name)
        for t in contract.all_tables
        if any(
            "fx" in c.header.lower() or "rate" in c.header.lower() for c in t.columns
        )
    ]
    assert rate_tables == [("Setup", FX_TABLE)], f"FX-rate owners: {rate_tables}"
    config_headers = [
        c.header
        for t in contract.all_tables
        if t.sheet == "Config"
        for c in t.columns
    ]
    assert not [h for h in config_headers if "fx" in h.lower() or "rate" in h.lower()], (
        f"Config declares a rate-like column: {config_headers}"
    )


# ===========================================================================
# SAR identity ownership. These values are MODEL invariants, not user data.
# ===========================================================================
def _fill_rgb(cell) -> str:
    rgb = getattr(cell.fill.fgColor, "rgb", None)
    return str(rgb)[-6:].upper() if rgb else ""


def _tokens() -> dict:
    import yaml
    with (PCCM_ROOT / "spec" / "workbook.yaml").open(encoding="utf-8") as handle:
        return yaml.safe_load(handle)["presentation"]["colors"]


def test_29_reporting_currency_remains_model_controlled_sar() -> None:
    contract = load_contract(CONTRACT_PATH)
    spec = contract.inputs["reporting_currency"]
    assert contract.reporting_currency == "SAR"
    assert spec.default == "SAR" and spec.editable is False
    cell = _wb()["Setup"][spec.cell]
    assert cell.value == "SAR"
    assert _fill_rgb(cell) == _tokens()["locked_fill"].upper()


def test_30_currency_master_has_locked_sar_identity() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = contract.table_by_name("tblCurrencies")
    assert table.locked_seed_rows == 1, "SAR must be a locked identity row"
    assert table.seed_rows[0] == ["SAR"]
    cell = _wb()["Config"][f"{table.column_letter(0)}{table.first_data_row}"]
    assert cell.value == "SAR"
    assert _fill_rgb(cell) == _tokens()["locked_fill"].upper(), "SAR identity is not locked-styled"


def test_31_user_currency_rows_remain_editable() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = contract.table_by_name("tblCurrencies")
    config = _wb()["Config"]
    assert table.editable is True
    assert table.first_user_row == table.first_data_row + 1
    assert table.first_user_row <= table.last_data_row, "no user-maintainable rows remain"
    for row in range(table.first_user_row, table.last_data_row + 1):
        cell = config[f"{table.column_letter(0)}{row}"]
        assert _fill_rgb(cell) == _tokens()["input_fill"].upper(), f"row {row} is not editable-styled"


def test_32_fx_table_has_locked_sar_identity_of_one() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = contract.table_by_name("tblFXRates")
    assert table.locked_seed_rows == 1
    assert table.seed_rows[0] == ["SAR", 1]
    setup = _wb()["Setup"]
    row = table.first_data_row
    assert setup[f"{table.column_letter(0)}{row}"].value == "SAR"
    assert setup[f"{table.column_letter(1)}{row}"].value == 1


def test_33_user_fx_rows_begin_after_the_locked_seed_row() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = contract.table_by_name("tblFXRates")
    assert table.first_user_row == table.first_data_row + table.locked_seed_rows
    assert table.user_data_range(0) == f"B{table.first_user_row}:B{table.last_data_row}"
    assert table.user_data_range(1) == f"C{table.first_user_row}:C{table.last_data_row}"


def test_34_fx_validation_does_not_target_the_locked_identity_row() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = contract.table_by_name("tblFXRates")
    setup = _wb()["Setup"]
    locked = {
        f"{table.column_letter(i)}{table.first_data_row}" for i in range(len(table.columns))
    }
    targeted = {
        str(cell.coord) if hasattr(cell, "coord") else str(cell)
        for dv in setup.data_validations.dataValidation
        for rng in dv.sqref.ranges
        for cell in rng.cells
    }
    targeted = {f"{c[0]}{c[1]}" if isinstance(c, tuple) else c for c in targeted}
    overlap = locked & _normalised(targeted)
    assert not overlap, f"validation targets locked identity cells: {sorted(overlap)}"


def _normalised(targets: set) -> set:
    out = set()
    for item in targets:
        text = str(item).replace("$", "").replace("(", "").replace(")", "").replace(" ", "")
        if "," in text:
            row, col = text.split(",")[0], text.split(",")[1]
            from openpyxl.utils import get_column_letter
            out.add(f"{get_column_letter(int(col))}{int(row)}")
        else:
            out.add(text)
    return out


def test_35_fx_validation_targets_every_user_row() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = contract.table_by_name("tblFXRates")
    setup = _wb()["Setup"]
    for index in range(len(table.columns)):
        expected = table.user_data_range(index)
        dv = _dv_for(setup, expected)
        assert dv is not None, f"no validation covering {expected}"
        assert expected in str(dv.sqref), f"validation sqref is {dv.sqref}, expected {expected}"


def test_36_currency_validation_targets_user_rows_only() -> None:
    contract = load_contract(CONTRACT_PATH)
    table = contract.table_by_name("tblFXRates")
    dv = _dv_for(_wb()["Setup"], table.user_data_range(0))
    assert dv is not None and dv.formula1 == "=lstCurrencies"
    assert str(table.first_data_row) not in str(dv.sqref).split(":")[0], (
        "currency validation begins on the locked identity row"
    )


def test_37_locked_cells_use_locked_treatment_not_input_treatment() -> None:
    contract = load_contract(CONTRACT_PATH)
    workbook = _wb()
    tokens = _tokens()
    for table in contract.all_tables:   # input-contract tables only
        worksheet = workbook[table.sheet]
        for offset in range(table.data_rows):
            row = table.first_data_row + offset
            for index in range(len(table.columns)):
                cell = worksheet[f"{table.column_letter(index)}{row}"]
                expected = (
                    tokens["locked_fill"] if table.is_locked_row(offset) else tokens["input_fill"]
                )
                assert _fill_rgb(cell) == expected.upper(), (
                    f"{table.table_name} {cell.coordinate}: fill {_fill_rgb(cell)}, "
                    f"expected {expected.upper()}"
                )


def test_38_no_model_check_logic_or_formulas_introduced() -> None:
    workbook = _wb()
    offenders = _unexpected_formulas(workbook)
    assert not offenders, f"formulas outside the permitted structural cells: {offenders[:10]}"
    assert not getattr(workbook["Model Check"], "tables", {}), "Model Check has a table"
    assert not list(workbook["Model Check"].data_validations.dataValidation), (
        "Model Check has data validation"
    )


def _run_all() -> int:
    tests = sorted(
        (name, fn) for name, fn in globals().items()
        if name.startswith("test_") and callable(fn)
    )
    failures = 0
    print("PCCM Phase 2 tests - functional Setup & Configuration layer")
    print("=" * 70)
    for name, fn in tests:
        try:
            fn()
        except AssertionError as error:
            failures += 1
            print(f"  [FAIL] {name}\n         {error}")
        except Exception as error:  # noqa: BLE001
            failures += 1
            print(f"  [ERROR] {name}\n          {type(error).__name__}: {error}")
        else:
            print(f"  [PASS] {name}")
    print("=" * 70)
    print(f"  {len(tests) - failures} passed, {failures} failed")
    if _TEMPDIR is not None:
        _TEMPDIR.cleanup()
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(_run_all())
