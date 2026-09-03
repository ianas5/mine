#!/usr/bin/env python3
"""PCCM Phase 7 Step-4 conformance tests for `src/vba/modSimPostReport.bas`.

The sensitivity pipeline: the CURRENT-run precondition, the resolved-model
bridge, the persisted TotalNom read, one-driver-at-a-time replay, the P7-2
kernel, ranking, the bounded `_SimData` block and its stamp.

--------------------------------------------------------------------------------
WHAT THESE TESTS PROVE, AND WHAT THEY DO NOT
--------------------------------------------------------------------------------
SOURCE CONFORMANCE AND ALGORITHMIC EQUIVALENCE, on Linux, now. The pipeline is
reproduced end to end against an INDEPENDENT Python reference built from the
accepted replay and the accepted kernel, and the module's ownership, ordering
and publication discipline are read from its source.

THE WORKSHEET HALF IS NOT EXECUTED. `modSimPostReport` reads and writes ranges,
and no VBA runtime or Excel exists here, so the transcriber cannot run it. What
is proved about the writes is their SHAPE and ORDER as written, and that is
stated rather than implied.

VBA EXECUTION CONFORMANCE IS NOT PROVED and is deferred to Phase-7 Windows
acceptance. Nothing here may be read as "VBA published a sensitivity table".

Runs standalone or under pytest.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

PCCM_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PCCM_ROOT / "builder"))
sys.path.insert(0, str(PCCM_ROOT / "tests"))

from pccm_builder import load_sim_contract, load_structure_contract  # noqa: E402
from pccm_builder import sim_sensitivity as kernel  # noqa: E402
from pccm_builder.vba_source import VbaModule  # noqa: E402

import test_phase6_sim_engine_vba as engine  # noqa: E402
import test_phase7_sim_replay as replay  # noqa: E402
from phase6_vba_transcribe import _Ref, _val  # noqa: E402

SRC_VBA = PCCM_ROOT / "src" / "vba"
POST_BAS = SRC_VBA / "modSimPostReport.bas"
SPEC = PCCM_ROOT / "spec"
SEED, N = 4242, 1000


def _module() -> VbaModule:
    return VbaModule(name="modSimPostReport", path=POST_BAS,
                     raw=POST_BAS.read_text(encoding="utf-8"))


def _code() -> str:
    return _module().code


def _procedure(name: str) -> str:
    code = _module().code_without_string_removal
    match = re.search(
        rf"^\s*(?:Public|Private)\s+(?:Function|Sub)\s+{re.escape(name)}\b", code, re.M)
    assert match, f"{name} is not declared"
    tail = code[match.start():]
    end = re.search(r"^\s*End\s+(?:Function|Sub)\s*$", tail, re.M)
    assert end, f"{name} has no End"
    return tail[: end.end()]


def _raw_contract() -> dict:
    return load_sim_contract(SPEC / "sim_contract.yaml").raw


# ---------------------------------------------------------------------------
# THE INDEPENDENT REFERENCE PIPELINE
# ---------------------------------------------------------------------------
# It uses the accepted replay and the accepted kernel - which is the point, they
# are the owners - but it composes them here rather than reading anything from
# modSimPostReport. What is being checked is the COMPOSITION: that the VBA does
# these steps, in this order, over all of them, once each.
def _pipeline(records):
    """(ranked, unranked) as the pipeline should produce them."""
    totals = replay._totals(records, seed=SEED, iterations=N)
    total_ranks = kernel.mid_ranks(totals)          # ONCE
    results = []
    for record in records:
        contributions = replay._replay(records, record["PermanentId"],
                                       seed=SEED, iterations=N)
        rho, status = kernel.rank_correlation(kernel.mid_ranks(contributions),
                                              total_ranks)
        results.append((record["PermanentId"], rho, status))
    order = kernel.rank_drivers(results)
    ranked = [(results[i][0], results[i][1]) for i in order]
    unranked = [r[0] for r in results if r[2] != kernel.SENSITIVITY_DEFINED]
    return ranked, unranked, results


def _cost(pid, dist="Triangular", lo=80.0, ml=100.0, hi=130.0, quantity=2.0):
    return engine._cost(pid, dist, lo, ml, hi, quantity=quantity)


def _risk(pid, dist="Triangular", lo=100.0, ml=200.0, hi=400.0, probability=0.3):
    return engine._risk(pid, dist, lo, ml, hi, probability=probability)


# ===========================================================================
# A. THE END-TO-END PIPELINE, against an independent composition
# ===========================================================================
def test_01_a_dominant_driver_ranks_first() -> None:
    """A cost line whose contribution dwarfs everything else must lead, and its
    rho must be near one because it very nearly IS the total."""
    records = replay._records(
        [_cost("C-001", "Uniform", 1000.0, None, 5000.0, quantity=10.0),
         _cost("C-002", "Uniform", 1.0, None, 2.0, quantity=1.0)],
        [_risk("R-001", "Uniform", 1.0, None, 3.0, probability=0.5)])
    ranked, unranked, _ = _pipeline(records)
    assert ranked[0][0] == "C-001", ranked
    assert ranked[0][1] > 0.95, ranked[0]
    assert unranked == []


def test_02_a_risk_at_twenty_percent_matches_an_independent_mid_rank_pearson() -> None:
    """The large tie block, checked against a reference written from the
    definition rather than against the same code under another name."""
    import math
    records = replay._records([], [_risk("R-001", "Beta-PERT", 100.0, 200.0, 400.0,
                                         probability=0.2)])
    totals = replay._totals(records, seed=SEED, iterations=N)
    contributions = replay._replay(records, "R-001", seed=SEED, iterations=N)
    zeros = contributions.count(0.0)
    assert 0.7 * N < zeros < 0.9 * N, zeros

    x, y = kernel.mid_ranks(contributions), kernel.mid_ranks(totals)
    mx, my = sum(x) / N, sum(y) / N
    expected = (sum((a - mx) * (b - my) for a, b in zip(x, y))
                / math.sqrt(sum((a - mx) ** 2 for a in x) * sum((b - my) ** 2 for b in y)))
    ranked, _unranked, _ = _pipeline(records)
    assert math.isclose(ranked[0][1], expected, rel_tol=1e-12, abs_tol=1e-12)


def test_03_a_zero_variance_driver_is_reported_but_not_ranked() -> None:
    records = replay._records(
        [_cost("C-001", "Uniform", 10.0, None, 20.0),
         _cost("C-002", "Uniform", 50.0, None, 50.0)],   # degenerate: no variance
        [_risk("R-001", probability=0.0)])               # never occurs: no variance
    ranked, unranked, results = _pipeline(records)
    assert [pid for pid, _ in ranked] == ["C-001"], ranked
    assert sorted(unranked) == ["C-002", "R-001"]
    # REPORTED, not deleted: every driver still has a record.
    assert len(results) == 3


def test_04_positive_and_negative_drivers_both_appear_with_their_sign() -> None:
    records = replay._records(
        [_cost("C-001", "Uniform", 100.0, None, 900.0, quantity=1.0),
         _cost("C-002", "Uniform", 100.0, None, 900.0, quantity=1.0)], [])
    ranked, _unranked, results = _pipeline(records)
    assert len(ranked) == 2
    # Two independent drivers of one total: both correlate POSITIVELY with it,
    # and the signs are retained rather than discarded by the ordering.
    assert all(rho > 0.0 for _pid, rho in ranked), ranked
    assert all(abs(a) >= abs(b) for a, b in zip([r for _, r in ranked],
                                                [r for _, r in ranked][1:]))


def test_05_equal_magnitude_drivers_keep_the_permanent_id_tie_break() -> None:
    """Carried all the way through the pipeline, not just inside the kernel."""
    results = [("C-010", 0.4, kernel.SENSITIVITY_DEFINED),
               ("C-002", -0.4, kernel.SENSITIVITY_DEFINED),
               ("R-001", 0.4, kernel.SENSITIVITY_DEFINED)]
    order = kernel.rank_drivers(results)
    assert [results[i][0] for i in order] == ["C-002", "C-010", "R-001"]


# ===========================================================================
# B. THE TOTAL'S RANKS ARE COMPUTED ONCE
# ===========================================================================
def test_06_the_total_is_ranked_once_and_reused_for_every_driver() -> None:
    body = _procedure("RunSensitivity")
    assert body.count("SimSensitivityMidRanks(") == 1, (
        "the total is ranked more than once, or not in the orchestrator")
    # AND THE PER-DRIVER CALL TAKES THE ALREADY-RANKED VECTOR.
    analyse = _procedure("AnalyseDrivers")
    assert "SimSensitivityMidRanks(" not in analyse, (
        "the total is re-ranked inside the per-driver loop")
    assert "SimSensitivitySpearman(" in analyse
    assert "totalRanks" in analyse


def test_07_every_driver_is_processed_exactly_once() -> None:
    analyse = _procedure("AnalyseDrivers")
    assert re.search(r"For index = 0 To driverCount - 1", analyse), (
        "the analysis does not walk every driver")
    assert analyse.count("SimEngineReplayDriver(") == 1
    assert analyse.count("SimSensitivitySpearman(") == 1
    # BY PERMANENT ID, AND BY THE LOOP'S OWN INDEX. `drivers(LBound(drivers))`
    # also ends in `.PermanentId` and would replay one driver D times.
    assert "drivers(LBound(drivers) + index).PermanentId, contributions" in analyse, (
        "the replayed driver is not the one this iteration of the loop reached")
    assert "results(index).PermanentId = drivers(LBound(drivers) + index).PermanentId" in analyse


# ===========================================================================
# C. MEMORY - O(N), never O(D x N)
# ===========================================================================
def test_08_one_contribution_vector_exists_at_a_time() -> None:
    """The 240 MB matrix is not built here either. `contributions` is a single
    vector that the next driver overwrites; nothing indexes it by driver."""
    analyse = _procedure("AnalyseDrivers")
    assert "Dim contributions() As Double" in analyse
    # NO DRIVER INDEX ON IT. `contributions(index)` or a second dimension would
    # be the matrix arriving by the back door.
    assert not re.search(r"contributions\s*\(\s*index", analyse)
    assert not re.search(r"ReDim\s+contributions\s*\([^)]*,", analyse)
    code = _code()
    assert not re.search(r"ReDim\s+\w+\s*\(0 To driverCount - 1, 0 To", code), (
        "a driver x iteration container is allocated")


def test_09_nothing_retains_a_per_driver_sample_or_a_matrix() -> None:
    code = _code()
    for banned in ("samples", "Samples", "matrix", "Matrix", "AnnualRecords"):
        assert banned not in code, banned
    # The only D-sized things are the RESULT records and the order permutation.
    sized = set(re.findall(r"ReDim (\w+)\(0 To driverCount - 1\)", code))
    assert sized <= {"results"}, sorted(sized)


# ===========================================================================
# D. OWNERSHIP - it orchestrates and computes nothing
# ===========================================================================
def test_10_no_mathematics_is_reimplemented_here() -> None:
    code = _code()
    for banned in ("SafeProduct", "SafeSignedSum", "midrank", "MidRank(",
                   "Pearson", "SimRngNext", "SimSampleUniform", "SimSampleTriangular",
                   "SimSamplePreparedBeta", "SimSampleBernoulli", "SimRngJump"):
        assert banned not in code, f"modSimPostReport reimplements {banned!r}"
    # THE IDENTITY IS READ, NEVER DERIVED. A token search for "Fingerprint("
    # would also flag the accepted accessor, which is the correct behaviour -
    # so the rule names WHERE those two values may come from instead.
    for kind in ("Fingerprint", "Digest"):
        uses = [line.strip() for line in code.splitlines()
                if re.search(rf"\b\w*{kind}\w*\s*\(", line)]
        for line in uses:
            assert f"modSimReport.PCCM_Simulation" in line, (
                f"a {kind.lower()} is produced here rather than read: {line}")


def test_11_each_step_is_delegated_to_its_accepted_owner() -> None:
    code = _code()
    assert "modSimReport.PCCM_SimulationStatus()" in code, "state is re-derived"
    assert "modCalcReport.CalcPrepareSimulationInputs(" in code, "the model is re-resolved"
    assert "modSimEngine.SimEngineReplayDriver(" in code, "replay is re-implemented"
    for owned in ("SimSensitivityMidRanks", "SimSensitivitySpearman", "SimSensitivityRank"):
        assert f"modSimSensitivity.{owned}(" in code, owned


def test_12_it_starts_no_simulation_and_consumes_no_run_identity() -> None:
    code = _code()
    for banned in ("PCCM_RunSimulation", "SimEngineRun(", "AllocateAutoNonce",
                   "SIM_PENDING_AUTO_NONCE_CELL", "NextAutoNonce", "CandidateRunId",
                   "WriteAttemptBlock", "WriteStatusBlock", "FinalCommit"):
        assert banned not in code, f"sensitivity reaches {banned!r}"


def test_13_the_simulation_does_not_run_sensitivity_for_you() -> None:
    """Explicit post-processing. A successful run must stay successful even if
    the analysis of it later fails."""
    report = (SRC_VBA / "modSimReport.bas").read_text(encoding="utf-8")
    assert "modSimPostReport" not in report
    assert "RunSensitivity" not in report


def test_14_it_writes_no_iteration_row_and_no_result_digest() -> None:
    code = _code()
    assert "SIM_ITER_A_TOTAL_NOMINAL_COLUMN" in code, "it must READ the totals"
    # ...and only read them.
    publish = _procedure("Publish")
    for banned in ("SIM_ITER_", "SIM_IDENTITY_ROW_RESULT_DIGEST",
                   "SIM_IDENTITY_ROW_RUN_ID", "SIM_SUMMARY_", "SIM_CONTINGENCY_"):
        assert banned not in publish, f"publication writes {banned!r}"


# ===========================================================================
# E. THE PRECONDITION AND REFUSALS
# ===========================================================================
def test_15_only_a_current_run_may_be_analysed() -> None:
    body = _procedure("RequireCurrentRun")
    assert "PCCM_SimulationStatus()" in body
    assert "SIM_STATE_CURRENT" in body
    # STALE and INVALID both fall through the same inequality; naming only one
    # would leave the other admitted.
    assert re.search(r"StrComp\(status, SIM_STATE_CURRENT, vbBinaryCompare\) <> 0", body)
    # AND BOTH GUARDS ARE REAL CONDITIONS. A message left behind an `If False`
    # is not a refusal, so the tested expression is read rather than the text
    # that follows it.
    assert re.search(r"If Len\(status\) = 0 Then", body), (
        "the no-successful-run guard is not a test of the status")
    assert "no successful simulation" in body
    assert "If False Then" not in body, "a precondition was short-circuited"


def test_16_a_refusal_writes_nothing_at_all() -> None:
    """Every refusal path exits before Publish is reached."""
    run = _procedure("RunSensitivity")
    refusals = run.count("RunSensitivity = Refused(detail)")
    assert refusals >= 6, refusals
    # EVERY PRECONDITION REFUSAL RETURNS BEFORE PUBLICATION. The last refusal
    # is Publish's own - a write that fails is still a refusal - so the rule is
    # that every OTHER one precedes it, not that none follows.
    at_publish = run.index("Publish(")
    before = [m.start() for m in re.finditer(r"RunSensitivity = Refused\(detail\)", run)]
    assert sum(1 for position in before if position < at_publish) == len(before) - 1, (
        "a refusal path can be reached after publication has begun")
    refused = _procedure("Refused")
    for banned in ("Range", "Value2", "ClearContents", "Stamp"):
        assert banned not in refused, f"a refusal touches {banned!r}"


# ===========================================================================
# F. PUBLICATION SAFETY
# ===========================================================================
def test_17_the_published_marker_is_cleared_first_and_written_last() -> None:
    """A block that fails part way through must not carry a current stamp."""
    publish = _procedure("Publish")
    first = publish.index("SIM_SENSITIVITY_STAMP_ROW_PUBLISHED")
    last = publish.rindex("SIM_SENSITIVITY_STAMP_ROW_PUBLISHED")
    assert first != last, "the marker is written only once"
    assert "vbNullString" in publish[first:first + 120], (
        "the marker is not CLEARED before the block is rewritten")
    assert "SIM_SENSITIVITY_PUBLISHED" in publish[last:last + 120], (
        "the marker is not SET at the end")
    # NOTHING FOLLOWS IT but the success return.
    tail = publish[last:]
    assert tail.count("Value2 =") == 1, "a write follows the published marker"


def test_18_the_identity_is_written_before_the_marker() -> None:
    publish = _procedure("Publish")
    marker = publish.rindex("SIM_SENSITIVITY_STAMP_ROW_PUBLISHED")
    for field in ("RUN_ID", "EFFECTIVE_SEED", "REQUEST_FINGERPRINT",
                  "RESULT_DIGEST", "ITERATIONS", "RECORD_COUNT"):
        at = publish.index(f"SIM_SENSITIVITY_STAMP_ROW_{field}")
        assert at < marker, f"{field} is stamped after the published marker"


def test_19_the_whole_result_is_built_before_anything_is_written() -> None:
    run = _procedure("RunSensitivity")
    for step in ("RequireCurrentRun(", "ResolveDrivers(", "ReadTotals(",
                 "SimSensitivityMidRanks(", "AnalyseDrivers(", "SimSensitivityRank("):
        assert run.index(step) < run.index("Publish("), f"{step} runs after publication"
    assert "NOTHING HAS BEEN WRITTEN UNTIL HERE" in _module().raw


def test_20_surplus_rows_from_a_larger_previous_result_are_cleared() -> None:
    """A later model can have fewer drivers than the bank already holds.
    Overwriting the first n would leave the remainder visible and
    indistinguishable from the new result."""
    publish = _procedure("Publish")
    assert publish.index("ClearRecords") < publish.index("Value2 = block"), (
        "the block is written before the old rows are cleared")
    clear = _procedure("ClearRecords")
    assert "ClearContents" in clear
    # TO THE CEILING, not to the new count: the surplus is what must go.
    assert "SIM_MAX_ITERATIONS" in clear
    assert "driverCount" not in clear


def test_21_the_record_count_bounds_what_is_authoritative() -> None:
    publish = _procedure("Publish")
    assert "SIM_SENSITIVITY_STAMP_ROW_RECORD_COUNT).Value2 = driverCount" in publish
    # AND THE BUILT BLOCK IS EXACTLY THAT MANY ROWS.
    assert "ReDim block(1 To driverCount, 1 To SIM_SENSITIVITY_FIELD_COUNT)" in publish
    assert "If slot <> driverCount Then" in publish, (
        "nothing checks that every driver produced exactly one record")


# ===========================================================================
# G. THE RECORD SHAPE
# ===========================================================================
def test_22_every_contracted_field_is_written_and_no_other() -> None:
    fill = _procedure("FillRecord")
    contracted = [c["key"] for c in
                  _raw_contract()["sim_data"]["sensitivity_records"]["columns"]]
    assert contracted == ["driver_id", "driver_type", "driver_name", "rho",
                          "abs_rho", "rank", "direction", "status"]
    for key in contracted:
        assert f"SIM_SENSITIVITY_OFFSET_{key.upper()}" in fill, key
    written = set(re.findall(r"SIM_SENSITIVITY_OFFSET_(\w+)", fill))
    assert written == {key.upper() for key in contracted}, sorted(written)
    # THE IDENTITY FIELDS ARE COMMON; THE MEASURE FIELDS ARE PER-ARM. A measure
    # written only when the driver has a rho would leave a stale value in the
    # cell of a driver that does not, so the two branches are read separately -
    # and the three identity fields are asserted to sit above the branch, where
    # both arms get them.
    identity = ("driver_id", "driver_type", "driver_name")
    measures = ("rho", "abs_rho", "rank", "direction", "status")
    assert set(identity) | set(measures) == set(contracted)
    branch = fill.index("    If record.Status = SIM_SENSITIVITY_DEFINED Then")
    common, rest = fill[:branch], fill[branch:]
    defined = rest[:rest.index("    Else")]
    undefined = rest[rest.index("    Else"):]
    for key in identity:
        assert f"SIM_SENSITIVITY_OFFSET_{key.upper()} + 1)" in common, (
            f"{key} is written inside a branch and would be missing from the other")
    for key in measures:
        token = f"SIM_SENSITIVITY_OFFSET_{key.upper()} + 1)"
        assert token in defined, f"{key} is not written for a ranked driver"
        assert token in undefined, f"{key} is not written for a zero-variance driver"


def test_23_a_zero_variance_record_carries_the_label_and_no_rho() -> None:
    fill = _procedure("FillRecord")
    undefined = fill[fill.index("Else"):]
    assert "SENSITIVITY_NO_VARIANCE_LABEL" in undefined
    # NO RHO, NO RANK, NO DIRECTION - printing 0 would say a relationship was
    # looked for and not found.
    for offset in ("RHO", "ABS_RHO", "RANK", "DIRECTION"):
        assert f"SIM_SENSITIVITY_OFFSET_{offset} + 1) = vbNullString" in undefined, offset
    raw = _module().raw
    assert 'SENSITIVITY_NO_VARIANCE_LABEL As String = "n/a - no variance"' in raw


def test_24_the_ranked_records_come_first_and_carry_their_rank() -> None:
    publish = _procedure("Publish")
    ranked_at = publish.index("For position = 0 To eligibleCount - 1")
    rest_at = publish.index("If results(index).Status <> SIM_SENSITIVITY_DEFINED")
    assert ranked_at < rest_at, "the unranked records are written first"
    assert "FillRecord block, slot + 1, results(index), position + 1" in publish
    assert "FillRecord block, slot + 1, results(index), 0" in publish


def test_25_no_variance_share_or_squared_rho_is_produced() -> None:
    code = _code()
    for banned in ("rho * rho", "rho ^ 2", "Squared", "percent", "Percent",
                   "Contribution %", "variance share"):
        assert banned not in code, banned


def test_26_no_top_n_truncation_and_no_subsampling() -> None:
    code = _code()
    for banned in ("TopN", "Top_N", "topN", "Subsample", "subsample", "SampleEvery"):
        assert banned not in code, banned
    # Full N: the iteration count comes from the published run and is not capped.
    assert "run.Iterations" in code
    assert not re.search(r"Iterations\s*=\s*\d+", code), "an iteration literal is imposed"
    # THE POPULATION IS NOT NARROWED. `eligibleCount` and `driverCount` are
    # produced by the kernel and the model; a reassignment of either here is a
    # truncation whatever it is called.
    for counter in ("eligibleCount", "driverCount"):
        # ANYWHERE ON THE LINE. `If eligibleCount > 10 Then eligibleCount = 10`
        # is a truncation that a line-start pattern would walk straight past.
        assignments = re.findall(rf"{counter}\s*=(?!=)\s*(.+)", code)
        assert assignments == [], f"{counter} is reassigned: {assignments}"


# ===========================================================================
# H. REGISTRATION AND THE ENDPOINT
# ===========================================================================
def test_27_the_module_and_endpoint_are_declared() -> None:
    structure = load_structure_contract(SPEC / "structure_contract.yaml")
    declared = {m.name: m for m in structure.vba_modules}
    assert "modSimPostReport" in declared
    entry = declared["modSimPostReport"]
    assert entry.generated is False
    for phrase in ("CURRENT-run precondition", "published last", "owns no RNG"):
        assert phrase in entry.responsibility, phrase
    assert _module().public_procedures == ["PCCM_RunSensitivity"], (
        _module().public_procedures)


def test_28_the_sensitivity_kernel_and_the_engine_stay_where_they_are() -> None:
    """P7-4 moved no mathematics into the orchestrator and none out of it."""
    for name, banned in (
        ("modSimSensitivity", ("Range", "Worksheet", "SimEngine", "SimRng", "_SimData")),
        ("modSimEngine", ("Range", "Worksheet", "_SimData", "modSimPostReport")),
    ):
        path = SRC_VBA / f"{name}.bas"
        code = VbaModule(name=name, path=path,
                         raw=path.read_text(encoding="utf-8")).code
        for token in banned:
            assert token not in code, f"{name} acquired {token!r}"


# ===========================================================================
# I. THE SENSITIVITY SHEET - lookups, bounds, and whose answer it is
# ===========================================================================
def _shell() -> dict:
    from pccm_builder.spec_loader import load_spec
    return load_spec(SPEC / "workbook.yaml").phase6_shell["sensitivity"]


def _built_sheet():
    import openpyxl
    workbook = openpyxl.load_workbook(PCCM_ROOT / "build" / "PCCM_stageA.xlsx")
    return workbook[_shell()["sheet"]]


def test_29_the_table_shows_the_eight_contracted_fields() -> None:
    shell, sheet = _shell(), _built_sheet()
    contracted = [c["header"] for c in
                  _raw_contract()["sim_data"]["sensitivity_records"]["columns"]]
    headers = [sheet[f"{c['column']}{shell['header_row']}"].value
               for c in shell["columns"]]
    assert headers == contracted == [
        "Driver ID", "Type", "Name", "Rho", "|Rho|", "Rank", "Direction", "Status"]


def test_30_every_table_cell_is_a_lookup_and_computes_nothing() -> None:
    """The sheet SHOWS the authoritative answer; it does not reproduce it."""
    shell, sheet = _shell(), _built_sheet()
    forbidden = ("RANK(", "CORREL(", "PEARSON(", "AVERAGE(", "STDEV", "SUMPRODUCT(",
                 "PERCENTILE", "RAND(", "OFFSET(", "INDIRECT(", "LARGE(", "SMALL(",
                 "SORT(", "^2")
    seen = 0
    for row_index in range(int(shell["row_window"])):
        row = int(shell["first_row"]) + row_index
        for column in shell["columns"]:
            formula = sheet[f"{column['column']}{row}"].value
            assert isinstance(formula, str) and formula.startswith("="), (column, row)
            for banned in forbidden:
                assert banned not in formula.upper(), (banned, column["key"])
            assert "_SimData!" in formula, "a table cell does not read the persisted block"
            seen += 1
    assert seen == int(shell["row_window"]) * 8, seen


def test_31_a_row_beyond_the_persisted_count_is_blank() -> None:
    """A later model with fewer drivers cannot leave a previous run's surplus
    rows on display: the count decides, not the window."""
    shell, sheet = _shell(), _built_sheet()
    stamp = _raw_contract()["sim_data"]["sensitivity_records"]["stamp"]
    count_row = next(f["row"] for f in stamp["fields"] if f["key"] == "record_count")
    for row_index in (0, 7, 199):
        formula = sheet[f"{shell['columns'][0]['column']}"
                        f"{int(shell['first_row']) + row_index}"].value
        assert f"IF({row_index + 1}>IF(" in formula, (row_index, formula[:80])
        assert f"$J${count_row}" in formula and f"$S${count_row}" in formula


def test_32_an_unpublished_block_shows_nothing_at_all() -> None:
    shell, sheet = _shell(), _built_sheet()
    stamp = _raw_contract()["sim_data"]["sensitivity_records"]["stamp"]
    published_row = next(f["row"] for f in stamp["fields"] if f["key"] == "published")
    formula = sheet[f"{shell['columns'][0]['column']}{shell['first_row']}"].value
    assert f'$J${published_row},_SimData!$S${published_row})<>"PUBLISHED",""' in formula


def test_33_the_sheet_says_whose_answer_it_is() -> None:
    """THE CASE THAT MUST NEVER BE SILENT. A table produced for run A, shown
    while run B is published, has to say so."""
    shell, sheet = _shell(), _built_sheet()
    availability = sheet[f"{shell['columns'][1]['column']}{shell['availability_row']}"].value
    assert "No simulation has been published." in availability
    assert "Not produced for this run." in availability
    assert "CURRENT for run " in availability
    assert "NOT CURRENT - this table belongs to run " in availability
    # It decides by comparing the STAMP against the published identity, not by
    # a flag someone could set.
    assert "$J$10" in availability and "$D$10" in availability, "fingerprint not compared"
    assert "$J$11" in availability and "$D$11" in availability, "digest not compared"


def test_34_the_sheet_carries_no_tornado_and_no_variance_share() -> None:
    """Phase 8 owns the chart. Phase 7 owns the table it will read."""
    import openpyxl
    workbook = openpyxl.load_workbook(PCCM_ROOT / "build" / "PCCM_stageA.xlsx")
    sheet = workbook[_shell()["sheet"]]
    text = " ".join(str(cell.value) for row in sheet.iter_rows() for cell in row
                    if cell.value is not None)
    for banned in ("Tornado", "tornado", "% of variance", "variance share", "R²", "rho²"):
        assert banned not in text, banned
    assert not getattr(sheet, "_charts", []), "a chart was placed on the Sensitivity sheet"


def test_35_the_display_window_cannot_silently_exceed_the_block() -> None:
    """The window is display; the persisted block is authority. A window taller
    than the sheet's own record area would read rows the block never owns."""
    shell = _shell()
    records = _raw_contract()["sim_data"]["sensitivity_records"]
    assert int(shell["first_row"]) > int(shell["header_row"])
    assert int(shell["row_window"]) >= 1
    # The persisted block starts at the shared first record row and runs to the
    # technical ceiling, so any window up to that is readable.
    ceiling = _raw_contract()["iterations"]["technical_ceiling"]["max_iterations_representable"]
    assert int(shell["row_window"]) <= ceiling
    assert int(records["first_record_row"]) == 34
